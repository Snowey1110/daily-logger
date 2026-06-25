from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
import base64
import contextlib
import ctypes
import html
import hmac
import io
import importlib
import importlib.util
import atexit
import json
import os
from pathlib import Path, PurePosixPath
import queue
import re
import runpy
import secrets
import shutil
import socket
import subprocess
import sys
import webbrowser
import tempfile
import threading
import textwrap
import time
import traceback
import types
import unicodedata
import uuid
import wave
from typing import Any, Callable, Dict, List, Optional, Tuple, Union
from urllib import error, request
from urllib.parse import parse_qs, unquote, urlparse
import zipfile
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

from journal_i18n import UI_LANGUAGE_PREF_KEY, normalize_ui_language, ui_translate

_journal_ui_language_changed_hook: Optional[Callable[[str], None]] = None


def set_journal_ui_language_changed_hook(hook: Optional[Callable[[str], None]]) -> None:
    global _journal_ui_language_changed_hook
    _journal_ui_language_changed_hook = hook


try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    ttk = None
try:
    from tkcalendar import Calendar, DateEntry
except Exception:
    Calendar = None  # type: ignore[assignment, misc]
    DateEntry = None  # type: ignore[assignment]
try:
    import msvcrt
except Exception:
    msvcrt = None
try:
    import readline as _readline  # type: ignore[assignment]
except ImportError:
    _readline = None


if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

def get_user_data_root() -> Path:
    """Return a stable per-user storage root shared across EXE and source runs."""
    appdata = os.getenv("APPDATA", "").strip()
    if appdata:
        return Path(appdata) / "DailyLogger"
    return BASE_DIR


USER_DATA_ROOT = get_user_data_root()
DATA_DIR = USER_DATA_ROOT / "daily_logs"
RECORDING_DIR = DATA_DIR / "Recording"
IPHONE_INBOX_DIR = RECORDING_DIR / "iPhone Inbox"
IPHONE_INCOMING_DIR = IPHONE_INBOX_DIR / "Incoming"
IPHONE_DECLINED_DIR = IPHONE_INBOX_DIR / "Declined"
BACKUP_DIR = DATA_DIR / "backup"
SETTINGS_DIR = USER_DATA_ROOT / "settings"
LOCAL_TRANSCRIPTION_ADDON_DIR = USER_DATA_ROOT / "addons" / "local_transcription"
MEDIA_TOOLS_ADDON_DIR = USER_DATA_ROOT / "addons" / "media_tools"
LOCAL_TRANSCRIPTION_MODEL_DIR = USER_DATA_ROOT / "models" / "whisper"
ADDON_DOWNLOAD_DIR = USER_DATA_ROOT / "downloads"
COMPANION_ROOT_DIR = USER_DATA_ROOT / "Companion"
COMPANION_CURRENT_DIR = COMPANION_ROOT_DIR / "Current"
COMPANION_INCOMING_DIR = COMPANION_ROOT_DIR / "_incoming"
LEGACY_DESKTOP_PET_ASSET_DIR = USER_DATA_ROOT / "desktop_pet"
DESKTOP_PET_ASSET_DIR = COMPANION_ROOT_DIR
DESKTOP_PET_MODULE_DIR = COMPANION_ROOT_DIR / "modules"
COMPANION_FEATURE_ENABLED = False
LEGACY_DATA_DIR = BASE_DIR / "daily_logs"
LEGACY_SETTINGS_DIR = BASE_DIR / "settings"
MASTER_JOURNAL_SHEET = "Master Journal"
JOURNAL_HEADERS_LEGACY = ["Date", "Time", "Journal"]
OPENAI_CHAT_COMPLETIONS_URL = "https://api.openai.com/v1/chat/completions"
OPENAI_IMAGE_GENERATIONS_URL = os.getenv(
    "OPENAI_IMAGE_GENERATIONS_URL", "https://api.openai.com/v1/images/generations"
).strip()
OPENAI_IMAGE_EDITS_URL = os.getenv(
    "OPENAI_IMAGE_EDITS_URL", "https://api.openai.com/v1/images/edits"
).strip()
OPENAI_IMAGE_MODEL = (
    os.getenv("OPENAI_IMAGE_MODEL", "gpt-image-1-mini").strip()
    or "gpt-image-1-mini"
)
OPENAI_TRANSCRIPTION_URL = os.getenv(
    "OPENAI_TRANSCRIPTION_URL", "https://api.openai.com/v1/audio/transcriptions"
).strip()
OPENAI_TRANSCRIPTION_MODEL = (
    os.getenv("OPENAI_TRANSCRIPTION_MODEL", "gpt-4o-mini-transcribe").strip()
    or "gpt-4o-mini-transcribe"
)
LIVE_STT_CHUNK_INTERVAL_SEC = 5.0
LIVE_STT_MIN_CHUNK_SAMPLES = int(16000 * 0.4)
# Journal waveform: int16 PCM RMS soft noise floor and display scale.
WAVEFORM_RMS_NOISE_FLOOR = 40.0
WAVEFORM_MAX_DRAW_SAMPLES = 4000
WAVEFORM_RMS_NORM = 6000.0
# Smaller input blocks when metering so the canvas updates often enough to feel live.
WAVEFORM_INPUT_BLOCK_SAMPLES = 512
APP_VERSION = "1.0.6"
APP_GITHUB_REPO = "Snowey1110/daily-logger"
APP_RELEASE_API_URL = f"https://api.github.com/repos/{APP_GITHUB_REPO}/releases/latest"
APP_RELEASE_PAGE_URL = f"https://github.com/{APP_GITHUB_REPO}/releases/latest"
APP_PORTABLE_ZIP_NAME = "DailyLoggerPortable.zip"
UPDATE_CHECK_ENABLED_PREF_KEY = "update_check_enabled"
UPDATE_LAST_CHECK_DATE_PREF_KEY = "last_update_check_date"
UPDATE_LAST_SEEN_RELEASE_PREF_KEY = "last_seen_release_tag"
UPDATE_DISMISSED_RELEASE_PREF_KEY = "dismissed_release_tag"
# Journal STT / AI report: same button width (text units) and grid min width so text areas align.
JOURNAL_SIDE_ACTION_BTN_WIDTH_CH = 16
JOURNAL_SIDE_ACTION_GRID_MINSIZE = 190
JOURNAL_TEXT_SCROLLBAR_WIDTH = 11
# Whisper list price per audio minute (USD); verify at https://openai.com/pricing
WHISPER_USD_PER_MIN = 0.006
# Pre-send WAV cleanup (RMS on int16-scale, same ballpark as WAVEFORM_RMS_NOISE_FLOOR).
WHISPER_PRE_FRAME_MS = 25
WHISPER_PRE_SILENCE_RMS = 32.0
WHISPER_PRE_NOISE_PERCENTILE = 10.0
WHISPER_PRE_NOISE_MULTIPLIER = 2.5
WHISPER_PRE_EDGE_PAD_MS = 120
WHISPER_PRE_MIN_SPEECH_MS = 350
WHISPER_PRE_MAX_INTERNAL_SILENCE_SEC = 1.25
WHISPER_PRE_KEEP_INTERNAL_SILENCE_SEC = 0.35
WHISPER_TRANSCRIBE_CHUNK_SEC = 8 * 60
# OpenAI Whisper multipart limit is ~25 MiB total; keep each mono int16 chunk smaller than that.
WHISPER_SAFE_CHUNK_PCM_BYTES = 20 * 1024 * 1024
WHISPER_SKIP_SINGLE_FILE_BYTES = 22 * 1024 * 1024
TRANSCRIPTION_DIRECT_UPLOAD_MAX_BYTES = 24 * 1024 * 1024
TRANSCRIPTION_AUDIO_CHUNK_SEC = 8 * 60
TRANSCRIPTION_CONVERTED_AUDIO_BITRATE = "96k"
WHISPER_TRANSCRIBE_PROMPT_CHAR_LIMIT = 600
WHISPER_REPEAT_SENTENCE_KEEP = 1
WHISPER_UNSUPPORTED_SCRIPT_RATIO = 0.25
WHISPER_UNSUPPORTED_SCRIPT_MIN_LETTERS = 8
TRANSCRIPTION_DIRECT_SUFFIXES = {
    ".aac",
    ".caf",
    ".flac",
    ".m4a",
    ".mp3",
    ".mp4",
    ".mpeg",
    ".mpga",
    ".ogg",
    ".wav",
    ".webm",
}
TRANSCRIPTION_VIDEO_SUFFIXES = {".mov", ".mp4", ".qt", ".webm"}
TRANSCRIPTION_MEDIA_SUFFIXES = TRANSCRIPTION_DIRECT_SUFFIXES | {".mov", ".qt"}
TRANSCRIPTION_FORCE_CONVERT_SUFFIXES = {".aac", ".caf"}
TRANSCRIPTION_CONTENT_TYPES = {
    ".aac": "audio/aac",
    ".caf": "audio/x-caf",
    ".flac": "audio/flac",
    ".m4a": "audio/mp4",
    ".mov": "audio/mp4",
    ".mp3": "audio/mpeg",
    ".mp4": "audio/mp4",
    ".mpeg": "audio/mpeg",
    ".mpga": "audio/mpeg",
    ".ogg": "audio/ogg",
    ".qt": "audio/mp4",
    ".wav": "audio/wav",
    ".webm": "audio/webm",
}
IPHONE_IMPORT_TOKEN_PREF_KEY = "iphone_import_token"
IPHONE_PASSIVE_RECEIVE_PREF_KEY = "iphone_passive_receive_enabled"
TRANSCRIPTION_MODEL_PREF_KEY = "transcription_model"
RECORD_SOURCE_PREF_KEY = "record_source_mode"
RECORD_SOURCE_MIC = "mic"
RECORD_SOURCE_COMPUTER = "computer"
RECORD_SOURCE_BOTH = "both"
RECORD_SOURCE_CHOICES = (
    RECORD_SOURCE_BOTH,
    RECORD_SOURCE_MIC,
    RECORD_SOURCE_COMPUTER,
)
TRANSCRIPTION_MODEL_CLOUD = "cloud"
TRANSCRIPTION_DEFAULT_MODEL = TRANSCRIPTION_MODEL_CLOUD
_TRANSCRIPTION_CLOUD_MODEL_ORDER = [
    "gpt-4o-mini-transcribe",
    "gpt-4o-transcribe",
]
TRANSCRIPTION_CLOUD_MODEL_NAMES = tuple(_TRANSCRIPTION_CLOUD_MODEL_ORDER)
TRANSCRIPTION_DEFAULT_CLOUD_MODEL = TRANSCRIPTION_CLOUD_MODEL_NAMES[0]
TRANSCRIPTION_LOCAL_MODEL_NAMES = (
    "tiny",
    "base",
    "small",
    "medium",
    "large-v3-turbo",
    "large-v3",
)
TRANSCRIPTION_SUGGESTED_LOCAL_MODEL = "small"
TRANSCRIPTION_LOCAL_MODEL_STATS = {
    "tiny": {
        "disk": "~75 MB",
        "bytes": 75 * 1024 * 1024,
        "speed": "Fastest",
        "quality": "Basic",
        "note": "Good for quick rough drafts.",
    },
    "base": {
        "disk": "~150 MB",
        "bytes": 150 * 1024 * 1024,
        "speed": "Very fast",
        "quality": "Better",
        "note": "A light upgrade when tiny is too rough.",
    },
    "small": {
        "disk": "~500 MB",
        "bytes": 500 * 1024 * 1024,
        "speed": "Balanced",
        "quality": "Recommended",
        "note": "Best default for this PC without using API tokens.",
    },
    "medium": {
        "disk": "~1.5 GB",
        "bytes": int(1.5 * 1024 * 1024 * 1024),
        "speed": "Slower",
        "quality": "Higher accuracy",
        "note": "Use when quality matters more than speed.",
    },
    "large-v3-turbo": {
        "disk": "~1.6 GB",
        "bytes": int(1.6 * 1024 * 1024 * 1024),
        "speed": "Medium-heavy",
        "quality": "Turbo large",
        "note": "Faster large-v3 family option, still a bigger download.",
    },
    "large-v3": {
        "disk": "~3 GB",
        "bytes": 3 * 1024 * 1024 * 1024,
        "speed": "Heavy",
        "quality": "Most powerful",
        "note": "Best quality class, but can impact this laptop a lot.",
    },
}
TRANSCRIPTION_CLOUD_MODEL_STATS = {
    "gpt-4o-mini-transcribe": {
        "quality": "Good",
        "speed": "Fast",
        "cost": "Lower cloud cost",
        "note": "Best cloud default for normal daily notes.",
    },
    "gpt-4o-transcribe": {
        "quality": "Higher",
        "speed": "Fast",
        "cost": "Higher cloud cost",
        "note": "Use when accuracy matters more than cost.",
    },
}
ADDON_RELEASE_BASE_URL = "https://github.com/Snowey1110/daily-logger/releases/latest/download"
LOCAL_TRANSCRIPTION_ADDON_ZIP_NAME = "DailyLoggerLocalTranscriptionAddon.zip"
MEDIA_TOOLS_ADDON_ZIP_NAME = "DailyLoggerMediaToolsAddon.zip"
UPDATE_RELEASE_ASSET_NAMES = (
    APP_PORTABLE_ZIP_NAME,
    LOCAL_TRANSCRIPTION_ADDON_ZIP_NAME,
    MEDIA_TOOLS_ADDON_ZIP_NAME,
)
LOCAL_TRANSCRIPTION_ADDON_ESTIMATED_BYTES = 70 * 1024 * 1024
MEDIA_TOOLS_ADDON_ESTIMATED_BYTES = 85 * 1024 * 1024
TRANSCRIPTION_LOCAL_CPU_THREADS = 4
TRANSCRIPTION_LOCAL_COMPUTE_TYPE = "int8"
IPHONE_IMPORT_DEFAULT_PORT = 8768
IPHONE_IMPORT_MAX_UPLOAD_BYTES = 8 * 1024 * 1024 * 1024
IPHONE_IMPORT_CHUNK_BYTES = 1024 * 1024
IPHONE_IMPORT_BROWSER_CHUNK_BYTES = 64 * 1024 * 1024
IPHONE_IMPORT_CHUNK_MAX_BYTES = 128 * 1024 * 1024
IPHONE_IMPORT_CONTENT_TYPE_SUFFIXES = {
    "application/quicktime": ".mov",
    "application/x-caf": ".caf",
    "audio/aac": ".aac",
    "audio/aiff": ".caf",
    "audio/caf": ".caf",
    "audio/x-caf": ".caf",
    "audio/flac": ".flac",
    "audio/m4a": ".m4a",
    "audio/mp4": ".m4a",
    "audio/mpeg": ".mp3",
    "audio/mpga": ".mpga",
    "audio/ogg": ".ogg",
    "audio/wav": ".wav",
    "audio/webm": ".webm",
    "video/mp4": ".mp4",
    "video/quicktime": ".mov",
    "video/webm": ".webm",
    "video/x-msvideo": ".mp4",
}
# Hover tooltips: narrow wrap -> shorter line length, more lines (taller block).
TOOLTIP_WRAP_PX = 220
TOOLTIP_WRAP_PX_MAX = 280
JOURNAL_PREF_THEME_KEY = "journal_window_theme"
JOURNAL_PREF_FULLSCREEN_KEY = "journal_window_fullscreen"
SIDEBAR_MODULE_ORDER_PREF_KEY = "sidebar_module_order"
SIDEBAR_DESKTOP_PET_MIGRATION_PREF_KEY = "sidebar_desktop_pet_migration_v2_done"
DESKTOP_PET_STATE_PREF_KEY = "desktop_pet_state"
DESKTOP_PET_MEMORY_UNLOCK_LEVEL = 2
DESKTOP_PET_TTS_UNLOCK_LEVEL = 3
DESKTOP_PET_CHAT_UNLOCK_LEVEL = 5
SIDEBAR_DEFAULT_MODULE_ORDER = (
    "journal",
    "ai_recap",
    "chatbot",
    "console",
    "reader",
    "settings",
)
SIDEBAR_TOP_SLOT_COUNT = 8
SIDEBAR_BOTTOM_SLOT_COUNT = 2
SIDEBAR_TOTAL_SLOT_COUNT = SIDEBAR_TOP_SLOT_COUNT + SIDEBAR_BOTTOM_SLOT_COUNT
JOURNAL_TEXT_FONT_FAMILY = "Microsoft YaHei UI"


def run_hidden_subprocess(args: List[str], **kwargs: Any) -> subprocess.CompletedProcess:
    """Run console tools from the GUI without flashing a Windows terminal."""
    if os.name == "nt":
        creationflags = int(kwargs.pop("creationflags", 0) or 0)
        creationflags |= int(getattr(subprocess, "CREATE_NO_WINDOW", 0) or 0)
        kwargs["creationflags"] = creationflags
        if "startupinfo" not in kwargs and hasattr(subprocess, "STARTUPINFO"):
            startupinfo = subprocess.STARTUPINFO()  # type: ignore[attr-defined]
            startupinfo.dwFlags |= getattr(subprocess, "STARTF_USESHOWWINDOW", 1)
            startupinfo.wShowWindow = 0
            kwargs["startupinfo"] = startupinfo
    return subprocess.run(args, **kwargs)


def migrate_legacy_storage_if_needed() -> None:
    """One-time best-effort migration from legacy BASE_DIR storage to USER_DATA_ROOT."""
    if USER_DATA_ROOT.resolve() == BASE_DIR.resolve():
        return
    try:
        USER_DATA_ROOT.mkdir(parents=True, exist_ok=True)
    except OSError:
        return

    def _journal_has_entries(path: Path) -> bool:
        if not path.exists():
            return False
        if load_workbook is None:
            # Fallback heuristic when openpyxl is unavailable.
            return path.stat().st_size > 16 * 1024
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[MASTER_JOURNAL_SHEET] if MASTER_JOURNAL_SHEET in wb.sheetnames else wb.active
            max_row = int(ws.max_row or 0)
            if max_row <= 1:
                wb.close()
                return False
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any((str(cell).strip() if cell is not None else "") for cell in row):
                    wb.close()
                    return True
            wb.close()
            return False
        except Exception:
            return path.stat().st_size > 16 * 1024

    # Migrate daily logs. If new journal exists but is empty, replace it with legacy journal.
    try:
        legacy_journal = LEGACY_DATA_DIR / "Journal.xlsx"
        new_journal = DATA_DIR / "Journal.xlsx"
        if legacy_journal.exists() and (
            not new_journal.exists()
            or (not _journal_has_entries(new_journal) and _journal_has_entries(legacy_journal))
        ):
            DATA_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(legacy_journal, new_journal)
        legacy_backup = LEGACY_DATA_DIR / "backup"
        new_backup = DATA_DIR / "backup"
        if legacy_backup.exists() and not new_backup.exists():
            shutil.copytree(legacy_backup, new_backup)
        legacy_recording = LEGACY_DATA_DIR / "Recording"
        new_recording = DATA_DIR / "Recording"
        if legacy_recording.exists() and not new_recording.exists():
            shutil.copytree(legacy_recording, new_recording)
    except OSError:
        pass

    # Migrate settings files only when target files are missing.
    try:
        if LEGACY_SETTINGS_DIR.exists():
            SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
            for src in LEGACY_SETTINGS_DIR.glob("*"):
                if not src.is_file():
                    continue
                dst = SETTINGS_DIR / src.name
                if not dst.exists():
                    shutil.copy2(src, dst)
    except OSError:
        pass


@dataclass(frozen=True)
class JournalWindowThemeSpec:
    """Colors and layout for the journal Tk window (light vs dark)."""

    id: str
    toggle_label: str
    surface: str
    panel: str
    field: str
    text: str
    muted: str
    accent: str
    border: str
    waveform: str
    btn_secondary: str
    btn_disabled: str
    disabled_fg: str
    hover_primary: str
    hover_save: str
    secondary_hover: str
    pad_outer: int
    pad_top_y: Tuple[int, int]
    pad_center_y: int
    pad_button_y: int
    date_label_font: Tuple[Any, ...]
    section_label_font: Tuple[Any, ...]
    is_dark: bool

    def toolbar_btn_config(self) -> Tuple[str, str, str, str]:
        """bg, fg, activebackground, activeforeground for Update Time / Open."""
        if self.is_dark:
            return (self.btn_secondary, self.text, self.accent, "white")
        return (self.btn_secondary, self.text, self.secondary_hover, self.text)

    def toolbar_hover(self) -> Tuple[str, str]:
        if self.is_dark:
            return (self.hover_primary, "white")
        return (self.secondary_hover, self.text)

    def toolbar_bind_rest(self) -> Tuple[str, str, str, str, str]:
        bg, fg, abg, afg = self.toolbar_btn_config()
        return ("normal", bg, fg, abg, afg)

    def side_action_config(self) -> Tuple[str, str, str, str]:
        """bg, fg, activebackground, activeforeground when action is enabled."""
        if self.is_dark:
            return (self.btn_secondary, self.text, self.accent, "white")
        return (self.accent, "white", self.hover_primary, "white")

    def side_action_bind_rest(self) -> Tuple[str, str, str, str, str]:
        bg, fg, abg, afg = self.side_action_config()
        return ("normal", bg, fg, abg, afg)

    def side_action_disabled(self) -> Tuple[str, str, str, str, str]:
        if self.is_dark:
            return (
                "disabled",
                self.btn_disabled,
                self.muted,
                self.btn_secondary,
                self.text,
            )
        return (
            "disabled",
            self.btn_disabled,
            self.disabled_fg,
            self.hover_primary,
            "white",
        )

    def transcribe_busy_config(self) -> Tuple[str, str, str, str, str]:
        if self.is_dark:
            return (
                self.btn_disabled,
                self.muted,
                self.btn_secondary,
                self.text,
                self.muted,
            )
        return (
            self.btn_disabled,
            self.disabled_fg,
            self.secondary_hover,
            self.text,
            self.disabled_fg,
        )

    def transcribe_idle_disabled_config(self) -> Tuple[str, str, str, str, str]:
        return self.transcribe_busy_config()

    def gen_bind_rest(self) -> Tuple[str, str, str, str, str]:
        if self.is_dark:
            bg, fg, abg, afg = self.side_action_config()
            return ("normal", bg, fg, abg, afg)
        return ("normal", self.accent, "white", self.hover_primary, "white")

    def gen_bind_disabled(self) -> Tuple[str, str, str, str, str]:
        if self.is_dark:
            return (
                "disabled",
                self.btn_disabled,
                self.muted,
                self.btn_secondary,
                self.text,
            )
        return (
            "disabled",
            self.btn_disabled,
            self.disabled_fg,
            self.hover_primary,
            "white",
        )

    def save_bind_disabled(self) -> Tuple[str, str, str, str, str]:
        return self.gen_bind_disabled()

    def ttk_combobox_kwargs(self) -> Dict[str, Any]:
        if self.is_dark:
            return {
                "fieldbackground": self.field,
                "background": self.panel,
                "foreground": self.text,
                "bordercolor": self.border,
                "lightcolor": self.panel,
                "darkcolor": self.field,
                "arrowcolor": self.muted,
                "padding": 4,
            }
        return {
            "fieldbackground": self.field,
            "background": self.btn_secondary,
            "foreground": self.text,
        }


JOURNAL_THEME_LIGHT = JournalWindowThemeSpec(
    id="light",
    toggle_label="Journal mode",
    surface="#F2F2F7",
    panel="#FFFFFF",
    field="#FFFFFF",
    text="#1D1D1F",
    muted="#6E6E73",
    accent="#0071E3",
    border="#D2D2D7",
    waveform="#0071E3",
    btn_secondary="#E8E8ED",
    btn_disabled="#E5E5EA",
    disabled_fg="#AEAEB2",
    hover_primary="#0077ED",
    hover_save="#0077ED",
    secondary_hover="#DCDCE0",
    pad_outer=14,
    pad_top_y=(14, 10),
    pad_center_y=10,
    pad_button_y=14,
    date_label_font=("Segoe UI", 10, "bold"),
    section_label_font=("Segoe UI", 10, "bold"),
    is_dark=False,
)

JOURNAL_THEME_DARK = JournalWindowThemeSpec(
    id="dark",
    toggle_label="Light mode",
    surface="#06060C",
    panel="#14141E",
    field="#0A0A12",
    text="#F5F5F7",
    muted="#98989D",
    accent="#0A84FF",
    border="#2C2C38",
    waveform="#64D2FF",
    btn_secondary="#24243A",
    btn_disabled="#101018",
    disabled_fg="#98989D",
    hover_primary="#339CFF",
    hover_save="#5CB0FF",
    secondary_hover="#339CFF",
    # Keep geometry/font metrics identical to light mode to avoid text reflow/shift
    # when toggling themes; only colors should differ between modes.
    pad_outer=14,
    pad_top_y=(14, 10),
    pad_center_y=10,
    pad_button_y=14,
    date_label_font=("Segoe UI", 10, "bold"),
    section_label_font=("Segoe UI", 10, "bold"),
    is_dark=True,
)

JOURNAL_THEME_BOOK = JournalWindowThemeSpec(
    id="journal",
    toggle_label="Dark mode",
    surface="#E7D7BF",
    panel="#F8F0E1",
    field="#FDFAF2",
    text="#241A12",
    muted="#7C6147",
    accent="#7A4B2A",
    border="#C9B897",
    waveform="#9B5C2E",
    btn_secondary="#E8DCCB",
    btn_disabled="#D8CCBB",
    disabled_fg="#A18B70",
    hover_primary="#8B5A34",
    hover_save="#8B5A34",
    secondary_hover="#DCCDB6",
    pad_outer=14,
    pad_top_y=(14, 10),
    pad_center_y=10,
    pad_button_y=14,
    date_label_font=("Segoe UI", 10, "bold"),
    section_label_font=("Segoe UI", 10, "bold"),
    is_dark=False,
)


def normalize_journal_window_theme_key(raw: str) -> str:
    k = (raw or "").strip().lower()
    if k in {"journal", "book", "reader", "parchment"}:
        return "journal"
    if k == "light":
        return "light"
    return "dark"


def journal_window_theme_spec_for_key(key: str) -> JournalWindowThemeSpec:
    normalized = normalize_journal_window_theme_key(key)
    if normalized == "light":
        return JOURNAL_THEME_LIGHT
    if normalized == "journal":
        return JOURNAL_THEME_BOOK
    return JOURNAL_THEME_DARK


def next_journal_window_theme_key(current_key: str) -> str:
    current = normalize_journal_window_theme_key(current_key)
    if current == "dark":
        return "light"
    if current == "light":
        return "journal"
    return "dark"


def journal_window_theme_action_label_key(current_key: str) -> str:
    next_key = next_journal_window_theme_key(current_key)
    if next_key == "light":
        return "theme.light"
    if next_key == "journal":
        return "theme.journal"
    return "theme.dark"


def journal_window_theme_current_label_key(current_key: str) -> str:
    current = normalize_journal_window_theme_key(current_key)
    if current == "light":
        return "theme.light"
    if current == "journal":
        return "theme.journal"
    return "theme.dark"


def load_journal_window_theme_spec() -> JournalWindowThemeSpec:
    prefs = load_preferences()
    return journal_window_theme_spec_for_key(str(prefs.get(JOURNAL_PREF_THEME_KEY, "dark")))


if tk is not None:

    class JournalSlimScrollbar(tk.Canvas):
        """Small themed scrollbar for the main journal editor panes."""

        def __init__(
            self,
            parent: Any,
            command: Callable[..., Any],
            theme: JournalWindowThemeSpec,
            width: int = JOURNAL_TEXT_SCROLLBAR_WIDTH,
        ) -> None:
            super().__init__(
                parent,
                width=width,
                bd=0,
                highlightthickness=0,
                relief="flat",
                takefocus=0,
                bg=theme.field,
            )
            self._command = command
            self._theme = theme
            self._first = 0.0
            self._last = 1.0
            self._drag_offset = 0
            self._width = width
            self._hidden = False
            self._hidden_column: Optional[int] = None
            self._track_id = self.create_rectangle(0, 0, width, 1, width=0)
            self._thumb_id = self.create_rectangle(2, 2, width - 2, 24, width=0)
            self.configure_theme(theme)
            self.bind("<Configure>", lambda _event: self._redraw())
            self.bind("<Button-1>", self._on_press)
            self.bind("<B1-Motion>", self._on_drag)
            self.bind("<ButtonRelease-1>", self._on_release)

        def configure_theme(self, theme: JournalWindowThemeSpec) -> None:
            self._theme = theme
            track = theme.panel
            thumb = theme.muted if theme.is_dark else theme.border
            self.configure(bg=track)
            self.itemconfigure(self._track_id, fill=track, outline="")
            self.itemconfigure(self._thumb_id, fill=thumb, outline="")
            self._redraw()

        def set(self, first: Any, last: Any) -> None:
            try:
                self._first = max(0.0, min(1.0, float(first)))
                self._last = max(0.0, min(1.0, float(last)))
            except (TypeError, ValueError):
                self._first = 0.0
                self._last = 1.0
            self._set_hidden((self._last - self._first) >= 0.999)
            self._redraw()

        def _set_hidden(self, should_hide: bool) -> None:
            if should_hide == self._hidden:
                return
            parent = self.master
            info = self.grid_info()
            if should_hide:
                if not info:
                    return
                try:
                    column = int(info.get("column", self._hidden_column or 0))
                    self._hidden_column = column
                    parent.grid_columnconfigure(column, minsize=0)
                except (tk.TclError, TypeError, ValueError):
                    pass
                self.configure(width=0)
                self.itemconfigure(self._track_id, state="hidden")
                self.itemconfigure(self._thumb_id, state="hidden")
                self._hidden = True
                return
            column = self._hidden_column
            if column is not None:
                try:
                    parent.grid_columnconfigure(column, minsize=self._width)
                except tk.TclError:
                    pass
            self.configure(width=self._width)
            self.itemconfigure(self._track_id, state="normal")
            self._hidden = False

        def _thumb_bounds(self) -> Tuple[int, int]:
            height = max(1, int(self.winfo_height()))
            pad = 2
            track_height = max(1, height - (pad * 2))
            page = max(0.0, min(1.0, self._last - self._first))
            if page >= 0.999:
                return (0, 0)
            thumb_height = max(28, int(track_height * page))
            thumb_height = min(track_height, thumb_height)
            max_first = max(0.001, 1.0 - page)
            top = pad + int((track_height - thumb_height) * (self._first / max_first))
            bottom = min(height - pad, top + thumb_height)
            return (top, bottom)

        def _redraw(self) -> None:
            if self._hidden:
                self.itemconfigure(self._thumb_id, state="hidden")
                return
            width = max(1, int(self.winfo_width() or self._width))
            height = max(1, int(self.winfo_height()))
            self.coords(self._track_id, 0, 0, width, height)
            top, bottom = self._thumb_bounds()
            if bottom <= top:
                self.itemconfigure(self._thumb_id, state="hidden")
                return
            self.itemconfigure(self._thumb_id, state="normal")
            self.coords(self._thumb_id, 2, top, max(3, width - 2), bottom)

        def _on_press(self, event: Any) -> None:
            top, bottom = self._thumb_bounds()
            if bottom > top and top <= int(event.y) <= bottom:
                self._drag_offset = int(event.y) - top
                self.itemconfigure(self._thumb_id, fill=self._theme.accent)
                return
            if int(event.y) < top:
                self._command("scroll", -1, "pages")
            else:
                self._command("scroll", 1, "pages")

        def _on_drag(self, event: Any) -> None:
            top, bottom = self._thumb_bounds()
            if bottom <= top:
                return
            height = max(1, int(self.winfo_height()))
            pad = 2
            track_height = max(1, height - (pad * 2))
            thumb_height = bottom - top
            movable = max(1, track_height - thumb_height)
            page = max(0.0, min(1.0, self._last - self._first))
            max_first = max(0.0, 1.0 - page)
            requested_top = int(event.y) - self._drag_offset
            clamped_top = max(pad, min(pad + movable, requested_top))
            new_first = ((clamped_top - pad) / movable) * max_first
            self._command("moveto", f"{new_first:.6f}")

        def _on_release(self, _event: Any) -> None:
            thumb = self._theme.muted if self._theme.is_dark else self._theme.border
            self.itemconfigure(self._thumb_id, fill=thumb)


OPENAI_MODEL = "gpt-4o-mini"
OPENAI_THINKING_MODEL = "gpt-5.5"
API_KEY_FILE = SETTINGS_DIR / "daily_logger_api_key.txt"
PREFS_FILE = SETTINGS_DIR / "daily_logger_prefs.json"
WIFI_WARN_FILE = SETTINGS_DIR / "wifi_warn_list.json"
JOURNAL_WINDOW_DRAFT_FILE = SETTINGS_DIR / "journal_window_draft.json"
JOURNAL_WINDOW_CONSOLE_RESERVE_BOTTOM = 56
SCREENSHOT_DIR = DATA_DIR / "chat_screenshots"
STARTUP_SHORTCUT_NAME = "Daily Logger.lnk"
VIRTUAL_READER_DIR_NAME = "virtual-journal-reader"
VIRTUAL_READER_ZIP_NAME = "virtual-journal-reader.zip"
VIRTUAL_READER_SERVER_ARG = "--serve-virtual-reader"
AUTO_BACKUP_START_DELAY_SEC = 2.0
BACKUP_COMPRESSION_LEVEL = 1


@dataclass
class ModuleConfig:
    name: str
    workbook_name: str
    sheet_name: str
    headers: List[str]
    prompt_builder: Callable[[], Optional[List[str]]]


PENDING_UNINSTALL_CONFIRM = False
_backup_lock = threading.Lock()


def bind_openpyxl_symbols() -> bool:
    global Workbook, load_workbook
    try:
        openpyxl_module = importlib.import_module("openpyxl")
        Workbook = openpyxl_module.Workbook
        load_workbook = openpyxl_module.load_workbook
        return True
    except Exception:
        Workbook = None  # type: ignore[assignment]
        load_workbook = None  # type: ignore[assignment]
        return False


def _pip_install_packages(packages: List[str]) -> bool:
    if not packages:
        return True
    print("Installing:", ", ".join(packages))
    result = subprocess.run(
        [sys.executable, "-m", "pip", "install", *packages],
        capture_output=False,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        print("Package installation failed. Try installing manually:")
        print(f"  {sys.executable} -m pip install {' '.join(packages)}")
        return False
    return True


def _missing_modules(specs: List[Tuple[str, str]]) -> List[str]:
    """Return pip package names whose import modules are not available."""
    return [
        pip_name
        for module_name, pip_name in specs
        if importlib.util.find_spec(module_name) is None
    ]


def ensure_runtime_dependencies() -> bool:
    core_specs: List[Tuple[str, str]] = [
        ("openpyxl", "openpyxl"),
        ("mss", "mss"),
    ]
    optional_specs: List[Tuple[str, str, str]] = [
        ("sounddevice", "sounddevice", "microphone recording for journal speech-to-text"),
        ("soundcard", "soundcard", "Windows computer-audio recording for meetings"),
        ("numpy", "numpy", "audio buffers for journal speech-to-text"),
        ("tkcalendar", "tkcalendar", "calendar popup on the journal date field"),
    ]

    missing_core = _missing_modules(core_specs)
    optional_missing = [
        (pip_name, blurb)
        for module_name, pip_name, blurb in optional_specs
        if importlib.util.find_spec(module_name) is None
    ]

    if missing_core or optional_missing:
        if missing_core:
            print("Required packages:")
            for pip_name in missing_core:
                print(f"  - {pip_name}")
        if optional_missing:
            print("Optional packages for full journal window features:")
            for pip_name, blurb in optional_missing:
                print(f"  - {pip_name}: {blurb}")
        print("Install missing packages now? (y/N): ", end="")
        answer = input().strip().lower()
        if answer in ("y", "yes"):
            install_list = list(missing_core)
            install_list.extend([pip_name for pip_name, _blurb in optional_missing])
            if install_list and not _pip_install_packages(install_list):
                print("You can install them later with:")
                print(f"  {sys.executable} -m pip install {' '.join(install_list)}")
                if missing_core:
                    return False
        else:
            if missing_core:
                print("Skipped installation of required packages.")
            if optional_missing:
                print("Skipped optional packages. Speech-to-text needs sounddevice and numpy.")

        missing_core = _missing_modules(core_specs)
        if missing_core:
            print(
                "Cannot start: still missing "
                + ", ".join(missing_core)
                + ". Install them, then run the app again."
            )
            return False

    if not bind_openpyxl_symbols():
        print("openpyxl is required to run this app. Please install it and retry.")
        return False

    return True


def is_enter_equivalent(value: str) -> bool:
    return not value or value.upper() == "X"


def _normalize_journal_header_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def migrate_journal_workbook_columns_if_needed(wb, new_headers: List[str]) -> bool:
    """Expand legacy 3-column journal sheets to five columns without deleting data."""
    if len(new_headers) != 5:
        return False
    legacy = JOURNAL_HEADERS_LEGACY
    modified = False
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        first_three = [_normalize_journal_header_cell(ws.cell(row=1, column=col).value) for col in (1, 2, 3)]
        if first_three != legacy:
            continue
        if ws.max_column == 3:
            ws.insert_cols(4, amount=2)
            ws.cell(row=1, column=4, value=new_headers[3])
            ws.cell(row=1, column=5, value=new_headers[4])
            modified = True
            continue
        d1 = _normalize_journal_header_cell(ws.cell(row=1, column=4).value)
        e1 = _normalize_journal_header_cell(ws.cell(row=1, column=5).value)
        want_d = new_headers[3].strip()
        want_e = new_headers[4].strip()
        if d1 != want_d or e1 != want_e:
            ws.cell(row=1, column=4, value=new_headers[3])
            ws.cell(row=1, column=5, value=new_headers[4])
            modified = True
    return modified


def red_text(value: str) -> str:
    return f"\033[31m{value}\033[0m"


def save_workbook_with_retry(wb, workbook_path: Path) -> None:
    while True:
        try:
            wb.save(workbook_path)
            return
        except PermissionError:
            print(
                f"Cannot save '{workbook_path.name}' because it is open in another program."
            )
            input("Close the file and press Enter to retry saving...")


def load_workbook_with_retry(workbook_path: Path):
    while True:
        try:
            return load_workbook(workbook_path)
        except PermissionError:
            print(
                f"Cannot open '{workbook_path.name}' because access is blocked (open/locked by another program or sync)."
            )
            input("Close the file or wait for sync, then press Enter to retry...")


def ensure_workbook(module: ModuleConfig) -> Path:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    workbook_path = DATA_DIR / module.workbook_name

    if not workbook_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = module.sheet_name
        ws.append(module.headers)
        save_workbook_with_retry(wb, workbook_path)
        return workbook_path

    wb = load_workbook_with_retry(workbook_path)
    if module.name == "Journal":
        if migrate_journal_workbook_columns_if_needed(wb, module.headers):
            save_workbook_with_retry(wb, workbook_path)
    if module.sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(module.sheet_name)
        ws.append(module.headers)
        save_workbook_with_retry(wb, workbook_path)
    else:
        ws = wb[module.sheet_name]
        if ws.max_row == 1:
            # Keep file resilient in case user removed headers manually.
            first_row = [cell.value for cell in ws[1]]
            if first_row != module.headers:
                ws.delete_rows(1, ws.max_row)
                ws.append(module.headers)
                save_workbook_with_retry(wb, workbook_path)

    return workbook_path


def append_row(module: ModuleConfig, row: List[str]) -> None:
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)

    if module.name == "Journal":
        row_list = list(row)
        while len(row_list) < len(module.headers):
            row_list.append("")
        row = row_list[: len(module.headers)]
        daily_ws = get_or_create_journal_daily_sheet(wb, module, row[0])
        target_row = find_first_empty_data_row(daily_ws, len(module.headers))
        for col_index, value in enumerate(row, start=1):
            daily_ws.cell(row=target_row, column=col_index, value=value)

        rebuild_master_journal_from_daily_pages(wb, module)
        reorder_journal_sheets(wb)
        save_workbook_with_retry(wb, workbook_path)
        return

    ws = wb[module.sheet_name]
    target_row = find_first_empty_data_row(ws, len(module.headers))
    for col_index, value in enumerate(row, start=1):
        ws.cell(row=target_row, column=col_index, value=value)
    save_workbook_with_retry(wb, workbook_path)


def ensure_master_journal_sheet(wb, module: ModuleConfig):
    if MASTER_JOURNAL_SHEET in wb.sheetnames:
        ws = wb[MASTER_JOURNAL_SHEET]
    elif module.sheet_name in wb.sheetnames:
        ws = wb[module.sheet_name]
        ws.title = MASTER_JOURNAL_SHEET
    else:
        ws = wb.create_sheet(MASTER_JOURNAL_SHEET, 0)

    ensure_headers(ws, module.headers)
    return ws


def get_or_create_journal_daily_sheet(wb, module: ModuleConfig, date_value: str):
    date_obj = datetime.strptime(date_value, "%m/%d/%Y")
    daily_sheet_name = date_obj.strftime("%Y-%m-%d")
    if daily_sheet_name in wb.sheetnames:
        ws = wb[daily_sheet_name]
    else:
        ws = wb.create_sheet(daily_sheet_name)
    ensure_headers(ws, module.headers)
    return ws


def ensure_headers(ws, headers: List[str]) -> None:
    first_row = [ws.cell(row=1, column=index).value for index in range(1, len(headers) + 1)]
    normalized = [cell.strip() if isinstance(cell, str) else cell for cell in first_row]
    if normalized != headers:
        if ws.max_row > 0:
            ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def reorder_journal_sheets(wb) -> None:
    if MASTER_JOURNAL_SHEET not in wb.sheetnames:
        return

    master = wb[MASTER_JOURNAL_SHEET]
    dated_sheets = []
    for sheet in wb.worksheets:
        if sheet.title == MASTER_JOURNAL_SHEET:
            continue
        try:
            sheet_date = datetime.strptime(sheet.title, "%Y-%m-%d")
            dated_sheets.append((sheet_date, sheet))
        except ValueError:
            continue

    ordered = sorted(dated_sheets, key=lambda item: item[0], reverse=True)
    ordered_daily = [item[1] for item in ordered]
    remaining = [
        sheet
        for sheet in wb._sheets
        if sheet is not master and sheet not in ordered_daily
    ]
    wb._sheets = [master] + ordered_daily + remaining


def rebuild_master_journal_from_daily_pages(wb, module: ModuleConfig) -> None:
    master_ws = ensure_master_journal_sheet(wb, module)
    entries: List[Tuple[datetime, int, List[str]]] = []

    for sheet in wb.worksheets:
        if sheet.title == MASTER_JOURNAL_SHEET:
            continue
        try:
            sheet_date = datetime.strptime(sheet.title, "%Y-%m-%d")
        except ValueError:
            continue

        for row_index in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row=row_index, column=col).value
                for col in range(1, len(module.headers) + 1)
            ]
            if is_row_empty(values):
                continue
            normalized = ["" if value is None else str(value) for value in values]
            entries.append((sheet_date, row_index, normalized))

    entries.sort(key=lambda item: (item[0], item[1]), reverse=True)

    if master_ws.max_row > 1:
        master_ws.delete_rows(2, master_ws.max_row - 1)

    for _, _, row_values in entries:
        master_ws.append(row_values)


def delete_latest_journal_entry() -> bool:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)

    latest_sheet = None
    latest_date = None
    for sheet in wb.worksheets:
        if sheet.title == MASTER_JOURNAL_SHEET:
            continue
        try:
            sheet_date = datetime.strptime(sheet.title, "%Y-%m-%d")
        except ValueError:
            continue
        if latest_date is None or sheet_date > latest_date:
            latest_date = sheet_date
            latest_sheet = sheet

    if latest_sheet is None:
        return False

    latest_row = None
    for row_index in range(latest_sheet.max_row, 1, -1):
        values = [
            latest_sheet.cell(row=row_index, column=col).value
            for col in range(1, len(module.headers) + 1)
        ]
        if not is_row_empty(values):
            latest_row = row_index
            break

    if latest_row is None:
        return False

    latest_sheet.delete_rows(latest_row, 1)

    has_remaining_data = False
    for row_index in range(2, latest_sheet.max_row + 1):
        values = [
            latest_sheet.cell(row=row_index, column=col).value
            for col in range(1, len(module.headers) + 1)
        ]
        if not is_row_empty(values):
            has_remaining_data = True
            break

    if not has_remaining_data:
        wb.remove(latest_sheet)

    rebuild_master_journal_from_daily_pages(wb, module)
    reorder_journal_sheets(wb)
    save_workbook_with_retry(wb, workbook_path)
    return True


def get_latest_journal_entry_for_edit() -> Optional[Dict[str, object]]:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)

    latest_sheet = None
    latest_date = None
    for sheet in wb.worksheets:
        if sheet.title == MASTER_JOURNAL_SHEET:
            continue
        try:
            sheet_date = datetime.strptime(sheet.title, "%Y-%m-%d")
        except ValueError:
            continue
        if latest_date is None or sheet_date > latest_date:
            latest_date = sheet_date
            latest_sheet = sheet

    if latest_sheet is None:
        return None

    latest_row = None
    latest_values: Optional[List[object]] = None
    for row_index in range(latest_sheet.max_row, 1, -1):
        values = [
            latest_sheet.cell(row=row_index, column=col).value
            for col in range(1, len(module.headers) + 1)
        ]
        if not is_row_empty(values):
            latest_row = row_index
            latest_values = values
            break

    if latest_row is None or latest_values is None:
        return None

    date_value = "" if latest_values[0] is None else str(latest_values[0])
    time_value = "" if latest_values[1] is None else str(latest_values[1])
    journal_value = "" if latest_values[2] is None else str(latest_values[2])
    speech_value = ""
    report_value = ""
    if len(latest_values) > 3 and latest_values[3] is not None:
        speech_value = str(latest_values[3])
    if len(latest_values) > 4 and latest_values[4] is not None:
        report_value = str(latest_values[4])
    return {
        "sheet_name": latest_sheet.title,
        "row_index": latest_row,
        "date": date_value,
        "time": time_value,
        "text": journal_value,
        "speech_transcript": speech_value,
        "ai_report": report_value,
        "images": [],
    }


def get_latest_journal_entry_for_delete() -> Optional[Dict[str, object]]:
    return get_latest_journal_entry_for_edit()


def update_journal_entry_at(sheet_name: str, row_index: int, row_values: List[str]) -> bool:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    if row_index < 2:
        return False

    row_list = list(row_values)
    while len(row_list) < len(module.headers):
        row_list.append("")
    row_values = row_list[: len(module.headers)]

    for col_index, value in enumerate(row_values, start=1):
        ws.cell(row=row_index, column=col_index, value=value)

    rebuild_master_journal_from_daily_pages(wb, module)
    reorder_journal_sheets(wb)
    save_workbook_with_retry(wb, workbook_path)
    return True


def _journal_cell_to_display_string(col_index_zero_based: int, value: object) -> str:
    if value is None:
        return ""
    if col_index_zero_based == 0:
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")
        if isinstance(value, date):
            return value.strftime("%m/%d/%Y")
        return str(value).strip()
    if col_index_zero_based == 1:
        if isinstance(value, datetime):
            return value.strftime("%I:%M%p").lstrip("0")
        return str(value).strip()
    return "" if value is None else str(value)


def load_journal_reader_entries() -> Tuple[List[Dict[str, object]], Optional[str]]:
    """Load journal rows for the Virtual Reader API (non-interactive; no input() on lock)."""
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    try:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
    except PermissionError:
        return [], "Journal.xlsx is locked or unavailable. Close it in Excel and try again."
    out: List[Dict[str, object]] = []
    try:
        for sheet in wb.worksheets:
            if sheet.title == MASTER_JOURNAL_SHEET:
                continue
            try:
                datetime.strptime(sheet.title, "%Y-%m-%d")
            except ValueError:
                continue
            for row_index in range(2, sheet.max_row + 1):
                values = [
                    sheet.cell(row=row_index, column=col).value
                    for col in range(1, len(module.headers) + 1)
                ]
                if is_row_empty(values):
                    continue
                date_s = _journal_cell_to_display_string(0, values[0])
                if not date_s:
                    # Derive date from sheet name (YYYY-MM-DD) when cell is empty
                    try:
                        date_s = datetime.strptime(sheet.title, "%Y-%m-%d").strftime("%m/%d/%Y")
                    except ValueError:
                        pass
                time_s = _journal_cell_to_display_string(1, values[1])
                journal_s = "" if values[2] is None else str(values[2])
                speech_s = "" if len(values) <= 3 or values[3] is None else str(values[3])
                report_s = "" if len(values) <= 4 or values[4] is None else str(values[4])
                out.append(
                    {
                        "id": f"{sheet.title}|{row_index}",
                        "sheetName": sheet.title,
                        "rowIndex": row_index,
                        "isoDate": sheet.title,
                        "date": date_s,
                        "time": time_s,
                        "journal": journal_s,
                        "speechToText": speech_s,
                        "aiReport": report_s,
                    }
                )
    finally:
        wb.close()
    out.sort(key=lambda item: (str(item.get("isoDate", "")), int(item.get("rowIndex", 0))))
    return out, None


def patch_journal_reader_entry(
    sheet_name: str,
    row_index: int,
    *,
    date: Optional[str] = None,
    time: Optional[str] = None,
    journal: Optional[str] = None,
    speech_to_text: Optional[str] = None,
    ai_report: Optional[str] = None,
) -> Tuple[bool, str]:
    """Update selected columns on one journal row. Non-interactive (no input() on lock)."""
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    try:
        wb = load_workbook(workbook_path)
    except PermissionError:
        return False, "Journal.xlsx is locked or unavailable. Close it in Excel and try again."
    try:
        if sheet_name not in wb.sheetnames:
            return False, "Worksheet not found."
        ws = wb[sheet_name]
        if row_index < 2 or row_index > ws.max_row:
            return False, "Row not found."
        values = [ws.cell(row=row_index, column=col).value for col in range(1, len(module.headers) + 1)]
        row_strs = [_journal_cell_to_display_string(i, v) for i, v in enumerate(values)]
        while len(row_strs) < len(module.headers):
            row_strs.append("")
        row_strs = row_strs[: len(module.headers)]
        if date is not None:
            row_strs[0] = date
        if time is not None:
            row_strs[1] = time
        if journal is not None:
            row_strs[2] = journal
        if speech_to_text is not None:
            row_strs[3] = speech_to_text
        if ai_report is not None:
            row_strs[4] = ai_report
        for col_index, value in enumerate(row_strs, start=1):
            ws.cell(row=row_index, column=col_index, value=value)
        rebuild_master_journal_from_daily_pages(wb, module)
        reorder_journal_sheets(wb)
        try:
            wb.save(workbook_path)
        except PermissionError:
            return False, "Cannot save: Journal.xlsx is locked. Close it in Excel and try again."
        return True, ""
    finally:
        wb.close()


def create_journal_reader_entry(
    date_str: str,
    time_str: str,
) -> Tuple[bool, str, Optional[str]]:
    """Create a blank journal entry. Returns (ok, error_msg, new_entry_id)."""
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    try:
        wb = load_workbook(workbook_path)
    except PermissionError:
        return False, "Journal.xlsx is locked or unavailable. Close it in Excel and try again.", None
    try:
        daily_ws = get_or_create_journal_daily_sheet(wb, module, date_str)
        target_row = find_first_empty_data_row(daily_ws, len(module.headers))
        row_data = [date_str, time_str, "", "", ""]
        for col_index, value in enumerate(row_data, start=1):
            daily_ws.cell(row=target_row, column=col_index, value=value)
        rebuild_master_journal_from_daily_pages(wb, module)
        reorder_journal_sheets(wb)
        try:
            wb.save(workbook_path)
        except PermissionError:
            return False, "Cannot save: Journal.xlsx is locked. Close it in Excel and try again.", None
        sheet_name = daily_ws.title
        entry_id = f"{sheet_name}|{target_row}"
        return True, "", entry_id
    finally:
        wb.close()


def delete_journal_reader_entry(
    sheet_name: str,
    row_index: int,
) -> Tuple[bool, str]:
    """Delete a journal entry row. Returns (ok, error_msg)."""
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    try:
        wb = load_workbook(workbook_path)
    except PermissionError:
        return False, "Journal.xlsx is locked or unavailable. Close it in Excel and try again."
    try:
        if sheet_name not in wb.sheetnames:
            return False, "Worksheet not found."
        ws = wb[sheet_name]
        if row_index < 2 or row_index > ws.max_row:
            return False, "Row not found."
        ws.delete_rows(row_index, 1)
        # If the sheet is now empty (only header), remove it
        if ws.max_row <= 1:
            del wb[sheet_name]
        rebuild_master_journal_from_daily_pages(wb, module)
        reorder_journal_sheets(wb)
        try:
            wb.save(workbook_path)
        except PermissionError:
            return False, "Cannot save: Journal.xlsx is locked. Close it in Excel and try again."
        return True, ""
    finally:
        wb.close()


def _virtual_reader_paths_from_root(root: Path) -> Optional[Tuple[Path, Path]]:
    script = root / "serve_reader.py"
    dist = root / "dist"
    if script.is_file() and (dist / "index.html").is_file():
        return script, dist
    return None


def _safe_virtual_reader_zip_parts(name: str) -> Optional[Tuple[str, ...]]:
    pure = PurePosixPath(name)
    if pure.is_absolute():
        return None
    parts = tuple(part for part in pure.parts if part not in ("", "."))
    if not parts or any(part == ".." for part in parts):
        return None
    return parts


def _virtual_reader_zip_member_relpath(
    parts: Tuple[str, ...],
    archive_has_root_dir: bool,
) -> Optional[Path]:
    if archive_has_root_dir:
        if parts[0] != VIRTUAL_READER_DIR_NAME or len(parts) == 1:
            return None
        parts = parts[1:]
    return Path(*parts)


def _extract_virtual_reader_zip(zip_path: Path, parent_dir: Path) -> Optional[Tuple[Path, Path]]:
    """Extract a packaged Virtual Reader zip when the folder addon is not present."""
    try:
        with zipfile.ZipFile(zip_path) as archive:
            member_parts = [
                parts
                for info in archive.infolist()
                if not info.is_dir()
                for parts in [_safe_virtual_reader_zip_parts(info.filename)]
                if parts is not None
            ]
            if not member_parts:
                return None
            archive_has_root_dir = any(parts[0] == VIRTUAL_READER_DIR_NAME for parts in member_parts)
            relpaths = [
                rel
                for parts in member_parts
                for rel in [_virtual_reader_zip_member_relpath(parts, archive_has_root_dir)]
                if rel is not None
            ]
            relpath_strings = {rel.as_posix() for rel in relpaths}
            if "serve_reader.py" not in relpath_strings or "dist/index.html" not in relpath_strings:
                return None

            target_root = parent_dir / VIRTUAL_READER_DIR_NAME
            target_root.mkdir(parents=True, exist_ok=True)
            target_root_resolved = target_root.resolve()
            for info in archive.infolist():
                if info.is_dir():
                    continue
                parts = _safe_virtual_reader_zip_parts(info.filename)
                if parts is None:
                    return None
                relpath = _virtual_reader_zip_member_relpath(parts, archive_has_root_dir)
                if relpath is None:
                    continue
                dest = target_root / relpath
                try:
                    dest.resolve().relative_to(target_root_resolved)
                except ValueError:
                    return None
                dest.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as src, dest.open("wb") as dst:
                    shutil.copyfileobj(src, dst)
            return _virtual_reader_paths_from_root(target_root)
    except (OSError, zipfile.BadZipFile):
        return None


def virtual_journal_reader_addon_paths() -> Optional[Tuple[Path, Path]]:
    """Return (serve_reader.py, dist_dir) if the Virtual Reader addon is present."""
    roots: List[Path] = []
    if getattr(sys, "frozen", False):
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            roots.append(Path(meipass) / VIRTUAL_READER_DIR_NAME)
    roots.append(BASE_DIR / VIRTUAL_READER_DIR_NAME)
    if getattr(sys, "frozen", False):
        roots.append(Path(sys.executable).resolve().parent / VIRTUAL_READER_DIR_NAME)
    for root in roots:
        found = _virtual_reader_paths_from_root(root)
        if found is not None:
            return found

    zip_parents: List[Path] = [BASE_DIR]
    if getattr(sys, "frozen", False):
        exe_parent = Path(sys.executable).resolve().parent
        if exe_parent not in zip_parents:
            zip_parents.append(exe_parent)
    for parent in zip_parents:
        found = _extract_virtual_reader_zip(parent / VIRTUAL_READER_ZIP_NAME, parent)
        if found is not None:
            return found
    return None


VIRTUAL_READER_REQUIRED_BUILD = 13

_virtual_reader_child_proc: Optional[subprocess.Popen] = None


def shutdown_virtual_reader_child_server() -> None:
    """Stop the Virtual Reader HTTP server process we started from this Daily Logger session."""
    global _virtual_reader_child_proc
    proc = _virtual_reader_child_proc
    _virtual_reader_child_proc = None
    if proc is None:
        return
    try:
        if proc.poll() is None:
            proc.terminate()
            try:
                proc.wait(timeout=5)
            except subprocess.TimeoutExpired:
                proc.kill()
    except Exception:
        try:
            proc.kill()
        except Exception:
            pass


atexit.register(shutdown_virtual_reader_child_server)


def _virtual_reader_health_info(port: int, timeout_sec: float = 0.35) -> Optional[Dict[str, Any]]:
    try:
        with request.urlopen(f"http://127.0.0.1:{port}/api/health", timeout=timeout_sec) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))
    except Exception:
        return None


def _virtual_reader_ui_lang() -> str:
    return normalize_ui_language(str(load_preferences().get(UI_LANGUAGE_PREF_KEY, "en")))


def _virtual_reader_browser_lang_param() -> str:
    """Query param for Virtual Reader web UI (matches journal UI language)."""
    return "zh" if _virtual_reader_ui_lang() == "zh" else "en"


def _virtual_reader_tr(key: str, **kwargs: object) -> str:
    return ui_translate(_virtual_reader_ui_lang(), key, **kwargs)


def open_virtual_journal_reader_in_browser() -> Tuple[bool, str]:
    """Start the Virtual Reader local server if needed and open the default browser."""
    paths = virtual_journal_reader_addon_paths()
    if paths is None:
        return False, _virtual_reader_tr("msg.virtual_reader_addon_missing")
    script_path, dist_dir = paths
    port = 8765

    info = _virtual_reader_health_info(port)
    if info and info.get("ok") is True:
        bld = int(info.get("readerBuild", 0) or 0)
        if bld < VIRTUAL_READER_REQUIRED_BUILD:
            return (False, _virtual_reader_tr("msg.virtual_reader_stale_server"))
        _vr_lang = _virtual_reader_browser_lang_param()
        url = f"http://127.0.0.1:{port}/?lang={_vr_lang}&_cb={int(time.time() * 1000)}"
        webbrowser.open(url)
        return True, ""
    env = os.environ.copy()
    env["VIRTUAL_READER_DIST"] = str(dist_dir.resolve())
    creationflags = 0
    if sys.platform == "win32":
        creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    global _virtual_reader_child_proc
    cmd = [sys.executable]
    if getattr(sys, "frozen", False):
        cmd.append(VIRTUAL_READER_SERVER_ARG)
    else:
        cmd.append(str(script_path))
    cmd.extend(["--port", str(port), "--no-browser"])
    try:
        _virtual_reader_child_proc = subprocess.Popen(
            cmd,
            cwd=str(script_path.resolve().parent),
            env=env,
            close_fds=sys.platform != "win32",
            creationflags=creationflags,
        )
    except OSError as exc:
        _virtual_reader_child_proc = None
        return False, _virtual_reader_tr("msg.virtual_reader_server_start_fail", exc=str(exc))
    for _ in range(50):
        time.sleep(0.1)
        info2 = _virtual_reader_health_info(port, timeout_sec=0.5)
        if (
            info2
            and info2.get("ok") is True
            and int(info2.get("readerBuild", 0) or 0) >= VIRTUAL_READER_REQUIRED_BUILD
        ):
            _vr_lang = _virtual_reader_browser_lang_param()
            url = f"http://127.0.0.1:{port}/?lang={_vr_lang}&_cb={int(time.time() * 1000)}"
            webbrowser.open(url)
            return True, ""
    proc = _virtual_reader_child_proc
    if proc is not None and proc.poll() is None:
        try:
            proc.terminate()
            proc.wait(timeout=2)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass
    _virtual_reader_child_proc = None
    return False, _virtual_reader_tr("msg.virtual_reader_server_timeout")


def run_virtual_reader_server_from_cli(argv: List[str]) -> int:
    debug_log = os.getenv("DAILY_LOGGER_READER_SERVER_LOG", "").strip()

    def _debug(message: str) -> None:
        if not debug_log:
            return
        try:
            Path(debug_log).write_text(message, encoding="utf-8")
        except OSError:
            pass

    paths = virtual_journal_reader_addon_paths()
    if paths is None:
        _debug("Virtual Journal Reader addon is missing.\n")
        print("Virtual Journal Reader addon is missing.", file=sys.stderr)
        return 1
    script_path, dist_dir = paths
    sys.modules.setdefault("daily_logger", sys.modules[__name__])
    os.environ["VIRTUAL_READER_DIST"] = str(dist_dir.resolve())
    old_argv = sys.argv[:]
    sys.argv = [str(script_path), *argv]
    try:
        runpy.run_path(str(script_path), run_name="__main__")
        return 0
    except SystemExit as exc:
        if exc.code is None:
            return 0
        if isinstance(exc.code, int):
            if exc.code != 0:
                _debug(f"Virtual Reader server exited with code {exc.code}.\n")
            return exc.code
        print(exc.code, file=sys.stderr)
        _debug(f"Virtual Reader server exited: {exc.code}\n")
        return 1
    except Exception:
        _debug(traceback.format_exc())
        return 1
    finally:
        sys.argv = old_argv


def open_virtual_reader_nav_action() -> Tuple[bool, str]:
    """Nav rail Virtual Reader: open reader in browser when addon exists; else open Journal.xlsx via file URI."""
    paths = virtual_journal_reader_addon_paths()
    if paths is not None:
        ok, err = open_virtual_journal_reader_in_browser()
        return ok, err
    try:
        jp = ensure_workbook(MODULES["J"])
        uri = jp.resolve().as_uri()
        webbrowser.open(uri)
        return True, ""
    except Exception as exc:
        return False, _virtual_reader_tr("msg.virtual_reader_open_journal_fail", err=str(exc))


def delete_journal_entry_at(sheet_name: str, row_index: int) -> bool:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)
    if sheet_name not in wb.sheetnames:
        return False
    ws = wb[sheet_name]
    if row_index < 2:
        return False
    ws.delete_rows(row_index, 1)

    has_remaining_data = False
    for candidate_row in range(2, ws.max_row + 1):
        values = [
            ws.cell(row=candidate_row, column=col).value
            for col in range(1, len(module.headers) + 1)
        ]
        if not is_row_empty(values):
            has_remaining_data = True
            break
    if not has_remaining_data:
        wb.remove(ws)

    rebuild_master_journal_from_daily_pages(wb, module)
    reorder_journal_sheets(wb)
    save_workbook_with_retry(wb, workbook_path)
    return True


def load_all_journal_entries() -> List[Tuple[datetime, str, str]]:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)
    entries: List[Tuple[datetime, str, str]] = []

    for sheet in wb.worksheets:
        if sheet.title == MASTER_JOURNAL_SHEET:
            continue
        try:
            sheet_date = datetime.strptime(sheet.title, "%Y-%m-%d")
        except ValueError:
            continue

        for row_index in range(2, sheet.max_row + 1):
            values = [
                sheet.cell(row=row_index, column=col).value
                for col in range(1, len(module.headers) + 1)
            ]
            if is_row_empty(values):
                continue
            date_value = "" if values[0] is None else str(values[0])
            time_value = "" if values[1] is None else str(values[1])
            journal_value = "" if values[2] is None else str(values[2])
            if journal_value.strip():
                entries.append((sheet_date, f"{date_value} {time_value}".strip(), journal_value))

    entries.sort(key=lambda item: item[0])
    return entries


def _journal_entry_calendar_day(entry: Tuple[datetime, str, str]) -> date:
    """Day used for recap filters: prefer the entry row date column, else the sheet tab date."""
    sheet_dt, when_value, _text = entry
    default_year = sheet_dt.year
    raw = (when_value or "").strip()
    if raw:
        first = raw.split(None, 1)[0]
        parsed = parse_flexible_date(first, default_year)
        if parsed is not None:
            return parsed.date()
    return sheet_dt.date()


def build_journal_context() -> str:
    return build_journal_context_for_range(None)


def build_journal_context_for_range(
    date_range: Optional[Tuple[datetime, datetime]]
) -> str:
    entries = load_all_journal_entries()
    if not entries:
        return "No journal entries available."
    if date_range is not None:
        start_date, end_date = date_range
        lo = start_date.date()
        hi = end_date.date()
        entries = [
            item
            for item in entries
            if lo <= _journal_entry_calendar_day(item) <= hi
        ]
        if not entries:
            return "No journal entries available in the selected date range."
    lines = []
    for _, when_value, text in entries:
        lines.append(f"- [{when_value}] {text}")
    return "\n".join(lines)


def build_journal_context_for_date_set(dates: Any) -> str:
    """Build journal context for sheet-days that match any calendar date in ``dates``."""
    if not dates:
        return "No dates selected."
    day_set: set[date] = set()
    for d in dates:
        if isinstance(d, datetime):
            day_set.add(d.date())
        elif isinstance(d, date):
            day_set.add(d)
    if not day_set:
        return "No dates selected."
    entries = load_all_journal_entries()
    if not entries:
        return "No journal entries available."
    filtered = [item for item in entries if _journal_entry_calendar_day(item) in day_set]
    if not filtered:
        return "No journal entries available for the selected day(s)."
    lines = [f"- [{when_value}] {text}" for _, when_value, text in filtered]
    return "\n".join(lines)


def parse_recap_date_range(raw_range: str, default_year: int) -> Optional[Tuple[datetime, datetime]]:
    cleaned = " ".join(raw_range.strip().split())
    if not cleaned:
        return None
    normalized = cleaned.replace(".", "/")
    tokens: List[str]
    if "-" in normalized:
        tokens = [part.strip() for part in normalized.split("-", 1)]
    else:
        parts = normalized.split()
        if len(parts) == 1:
            single = parse_flexible_date(parts[0], default_year)
            if single is None:
                return None
            return single, single
        if len(parts) != 2:
            return None
        tokens = parts
    if len(tokens) == 1:
        single = parse_flexible_date(tokens[0], default_year)
        if single is None:
            return None
        return single, single
    if len(tokens) != 2 or not tokens[0] or not tokens[1]:
        return None
    start = parse_flexible_date(tokens[0], default_year)
    end = parse_flexible_date(tokens[1], default_year)
    if start is None or end is None:
        return None
    if end < start:
        start, end = end, start
    return start, end


def list_journal_dates_in_range(date_range: Tuple[datetime, datetime]) -> List[str]:
    start_date, end_date = date_range
    lo = start_date.date()
    hi = end_date.date()
    matched_dates: set[str] = set()
    for entry in load_all_journal_entries():
        d = _journal_entry_calendar_day(entry)
        if lo <= d <= hi:
            matched_dates.add(d.strftime("%m/%d/%Y"))
    return sorted(
        matched_dates,
        key=lambda value: datetime.strptime(value, "%m/%d/%Y"),
    )


def load_recap_context_from_file(raw_path: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """Return (context_text, resolved_path, error_message)."""
    if not raw_path.strip():
        return None, None, "Missing file path."
    token = raw_path.strip().strip('"').strip("'")
    candidates = [
        Path(token),
        BASE_DIR / token,
        DATA_DIR / token,
    ]
    file_path: Optional[Path] = None
    for cand in candidates:
        if cand.exists() and cand.is_file():
            file_path = cand
            break
    if file_path is None:
        return None, None, f"File not found: {token}"
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xlsm", ".xls"):
        return None, None, "Recap file lookup only supports text-like files (not Excel)."
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except OSError as exc:
        return None, None, f"Could not read file: {exc}"
    if not text.strip():
        return None, None, "Selected file is empty."
    clipped = text
    if len(clipped) > 120000:
        clipped = clipped[:120000] + "\n\n[Truncated for recap]"
    header = f"Recap source file: {file_path.resolve()}\n"
    return header + clipped, str(file_path.resolve()), None


def resolve_recap_target(
    raw_arg: str, default_year: int
) -> Tuple[Optional[Tuple[datetime, datetime]], Optional[str], Optional[str], Optional[str]]:
    """Return (date_range, file_context, file_path, error)."""
    arg = raw_arg.strip()
    if not arg:
        return None, None, None, "Missing recap argument."
    recap_range = parse_recap_date_range(arg, default_year)
    if recap_range is not None:
        return recap_range, None, None, None
    file_context, file_path, file_error = load_recap_context_from_file(arg)
    if file_error is None:
        return None, file_context, file_path, None
    return None, None, None, (
        "Invalid recap target. Use a date range (e.g. 4/27 - 4/30) "
        "or a file path (e.g. notes.txt)."
    )


def get_openai_api_key() -> Optional[str]:
    env_key = os.getenv("OPENAI_API_KEY", "").strip()
    if env_key:
        return env_key
    if API_KEY_FILE.exists():
        try:
            file_key = API_KEY_FILE.read_text(encoding="utf-8").strip()
            if file_key:
                return file_key
        except OSError:
            return None
    return None


def save_openai_api_key(api_key: str) -> bool:
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        API_KEY_FILE.write_text(api_key.strip(), encoding="utf-8")
        return True
    except OSError:
        return False


def _normalize_desktop_pet_image_bytes(raw: bytes) -> bytes:
    try:
        from PIL import Image

        image = Image.open(io.BytesIO(raw)).convert("RGBA")
        width, height = image.size
        if width <= 0 or height <= 0:
            return raw
        alpha = image.getchannel("A")
        if alpha.getextrema() == (255, 255):
            pixels = image.load()
            corner_points = [
                (0, 0),
                (max(0, width - 1), 0),
                (0, max(0, height - 1)),
                (max(0, width - 1), max(0, height - 1)),
            ]
            corner_colors = [pixels[x, y][:3] for x, y in corner_points]
            base_color = tuple(sum(color[i] for color in corner_colors) // len(corner_colors) for i in range(3))

            def _near_background(rgb: Tuple[int, int, int]) -> bool:
                return sum((int(rgb[i]) - int(base_color[i])) ** 2 for i in range(3)) <= 38 * 38

            visited = set()
            stack = list(corner_points)
            while stack:
                x, y = stack.pop()
                if x < 0 or y < 0 or x >= width or y >= height or (x, y) in visited:
                    continue
                visited.add((x, y))
                r, g, b, a = pixels[x, y]
                if a <= 0 or not _near_background((r, g, b)):
                    continue
                pixels[x, y] = (r, g, b, 0)
                stack.extend(((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)))
        bbox = image.getchannel("A").getbbox()
        if bbox:
            pad = 24
            left = max(0, bbox[0] - pad)
            top = max(0, bbox[1] - pad)
            right = min(width, bbox[2] + pad)
            bottom = min(height, bbox[3] + pad)
            image = image.crop((left, top, right, bottom))
        output = io.BytesIO()
        image.save(output, format="PNG")
        return output.getvalue()
    except Exception:
        return raw


def _prepare_pet_reference_image(path: Path) -> Tuple[Optional[Tuple[str, bytes, str]], Optional[str]]:
    try:
        if not path.exists() or not path.is_file():
            return None, "Reference image was not found."
        if path.stat().st_size > 50 * 1024 * 1024:
            return None, "Reference image is larger than 50 MB."
        from PIL import Image

        with Image.open(path) as image:
            image.thumbnail((1536, 1536))
            if image.mode not in ("RGB", "RGBA"):
                image = image.convert("RGBA")
            output = io.BytesIO()
            image.save(output, format="PNG")
        return ("reference.png", output.getvalue(), "image/png"), None
    except Exception as exc:
        return None, f"Could not read the reference image: {exc}"


def _multipart_form_data(
    fields: Dict[str, str],
    files: Dict[str, Tuple[str, bytes, str]],
) -> Tuple[bytes, str]:
    boundary = f"----DailyLogger{secrets.token_hex(16)}"
    chunks: List[bytes] = []
    for name, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{name}"\r\n\r\n'.encode("utf-8"))
        chunks.append(str(value).encode("utf-8"))
        chunks.append(b"\r\n")
    for name, (filename, data, content_type) in files.items():
        safe_name = Path(filename).name or "image.png"
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(
            (
                f'Content-Disposition: form-data; name="{name}"; filename="{safe_name}"\r\n'
                f"Content-Type: {content_type or 'application/octet-stream'}\r\n\r\n"
            ).encode("utf-8")
        )
        chunks.append(data)
        chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


def generate_desktop_pet_image(
    level: int,
    name: str = "Companion",
    reference_image_path: Optional[Path] = None,
    output_dir: Optional[Path] = None,
    design_prompt: str = "",
) -> Tuple[Optional[Path], Optional[str]]:
    api_key = get_openai_api_key()
    if not api_key:
        return None, "OPENAI_API_KEY is not set. Add your token first, then generate the companion look."

    pet_level = max(1, int(level or 1))
    safe_name = re.sub(r"[^A-Za-z0-9 _-]+", "", str(name or "Companion")).strip() or "Companion"
    prompt = (
        "Create one original cute Q-style chibi desktop companion sprite named "
        f"{safe_name}. The default direction is a charming chibi anime boy or girl mascot, "
        "with the option to softly borrow animal-like pet/dog/fantasy companion energy. "
        "Do not resemble any existing copyrighted character. Transparent background, "
        "full body, front three-quarter view, soft rounded shapes, clean outline, readable at "
        "small desktop-icon size, no text, no logo, no UI frame, no black/solid background. "
        f"It is level {pet_level}; higher level means slightly more confident, brighter, and magical, "
        "but still simple and friendly."
    )
    notes = re.sub(r"\s+", " ", str(design_prompt or "")).strip()
    if notes:
        prompt += f" User editable design notes: {notes}"
    if reference_image_path is not None:
        reference_file, ref_err = _prepare_pet_reference_image(Path(reference_image_path))
        if ref_err or reference_file is None:
            return None, ref_err or "Could not prepare reference image."
        prompt = (
            prompt
            + " Use the uploaded image only as loose visual inspiration for hairstyle, outfit palette, "
            "pose language, or creature personality. Redesign it as an original cute chibi desktop "
            "companion sprite, not a direct copy."
        )
        payload, content_type = _multipart_form_data(
            {
                "model": OPENAI_IMAGE_MODEL,
                "prompt": prompt,
                "n": "1",
                "size": "1024x1024",
                "quality": "low",
                "background": "transparent",
                "output_format": "png",
            },
            {"image": reference_file},
        )
        req = request.Request(
            OPENAI_IMAGE_EDITS_URL,
            data=payload,
            method="POST",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": content_type,
            },
        )
    else:
        payload_data: Dict[str, object] = {
            "model": OPENAI_IMAGE_MODEL,
            "prompt": prompt,
            "n": 1,
            "size": "1024x1024",
            "quality": "low",
            "background": "transparent",
            "output_format": "png",
        }
        payload = json.dumps(payload_data).encode("utf-8")
        req = request.Request(
            OPENAI_IMAGE_GENERATIONS_URL,
            data=payload,
            method="POST",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
        )
    try:
        with request.urlopen(req, timeout=180) as response:
            body = response.read().decode("utf-8")
    except error.HTTPError as exc:
        try:
            details = exc.read().decode("utf-8")
        except Exception:
            details = str(exc)
        return None, f"Image generation API error ({exc.code}): {details}"
    except Exception as exc:
        return None, f"Could not contact image generation API: {exc}"

    try:
        parsed = json.loads(body)
        first = (parsed.get("data") or [{}])[0]
        if not isinstance(first, dict):
            return None, "Image generation returned an unexpected response."
        b64 = str(first.get("b64_json") or "").strip()
        if b64:
            raw = base64.b64decode(b64)
        else:
            url = str(first.get("url") or "").strip()
            if not url:
                return None, "Image generation did not return image data."
            with request.urlopen(url, timeout=90) as image_response:
                raw = image_response.read()
        if not raw:
            return None, "Image generation returned an empty image."
        raw = _normalize_desktop_pet_image_bytes(raw)
        target_dir = Path(output_dir) if output_dir is not None else DESKTOP_PET_ASSET_DIR
        target_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = target_dir / f"companion_{timestamp}_{secrets.token_hex(3)}.png"
        tmp_path = path.with_suffix(".tmp")
        tmp_path.write_bytes(raw)
        tmp_path.replace(path)
        return path, None
    except Exception as exc:
        return None, f"Could not save generated pet image: {exc}"


def delete_openai_api_key() -> bool:
    if not API_KEY_FILE.exists():
        return True
    try:
        API_KEY_FILE.unlink()
        return True
    except OSError:
        return False


def copy_text_to_clipboard(text: str) -> bool:
    # Prefer native Windows clipboard command for CLI reliability.
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", "Set-Clipboard -Value @'\n" + text + "\n'@"],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0
    except OSError:
        pass
    if tk is not None:
        try:
            root = tk.Tk()
            root.withdraw()
            root.clipboard_clear()
            root.clipboard_append(text)
            root.update()
            root.destroy()
            return True
        except Exception:
            return False
    return False


def ensure_openai_api_key_for_ai() -> bool:
    existing = get_openai_api_key()
    if existing:
        return True

    print("AI feature needs an OpenAI API key.")
    pasted = input("Paste your OpenAI API key (or press Enter to cancel): ").strip()
    if is_enter_equivalent(pasted):
        print("No API key entered. Returning to main menu.")
        return False
    if not save_openai_api_key(pasted):
        print("Could not save API key file. Check folder permissions and try again.")
        return False
    print("API key saved for future use.")
    return True


def load_preferences() -> Dict[str, str]:
    if not PREFS_FILE.exists():
        return {}
    try:
        raw = json.loads(PREFS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    result: Dict[str, str] = {}
    for key, value in raw.items():
        if isinstance(key, str) and isinstance(value, str):
            result[key] = value
    return result


def save_preferences(prefs: Dict[str, str]) -> bool:
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        PREFS_FILE.write_text(json.dumps(prefs, indent=2), encoding="utf-8")
        return True
    except OSError:
        return False


def normalize_release_tag(tag: str) -> str:
    text = (tag or "").strip()
    if text.lower().startswith("v"):
        text = text[1:]
    return text.strip()


def release_version_parts(tag: str) -> Tuple[int, ...]:
    normalized = normalize_release_tag(tag)
    parts: List[int] = []
    for raw in normalized.split("."):
        match = re.match(r"(\d+)", raw.strip())
        if not match:
            parts.append(0)
            continue
        try:
            parts.append(int(match.group(1)))
        except ValueError:
            parts.append(0)
    while parts and parts[-1] == 0:
        parts.pop()
    return tuple(parts or [0])


def release_tag_is_newer(latest_tag: str, current_version: str = APP_VERSION) -> bool:
    latest = list(release_version_parts(latest_tag))
    current = list(release_version_parts(current_version))
    max_len = max(len(latest), len(current))
    latest.extend([0] * (max_len - len(latest)))
    current.extend([0] * (max_len - len(current)))
    return tuple(latest) > tuple(current)


def update_check_enabled() -> bool:
    prefs = load_preferences()
    return prefs.get(UPDATE_CHECK_ENABLED_PREF_KEY, "true").strip().lower() != "false"


def save_update_check_enabled(enabled: bool) -> bool:
    prefs = load_preferences()
    prefs[UPDATE_CHECK_ENABLED_PREF_KEY] = "true" if enabled else "false"
    return save_preferences(prefs)


def today_update_check_key() -> str:
    return date.today().isoformat()


def release_notes_preview(body: str, limit: int = 1000) -> str:
    text = (body or "").strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)].rstrip() + "..."


def parse_github_release_info(raw: Dict[str, Any]) -> Dict[str, Any]:
    assets: Dict[str, str] = {}
    for asset in raw.get("assets", []):
        if not isinstance(asset, dict):
            continue
        name = str(asset.get("name") or "").strip()
        url = str(asset.get("browser_download_url") or "").strip()
        if name and url:
            assets[name] = url
    tag = str(raw.get("tag_name") or "").strip()
    html_url = str(raw.get("html_url") or "").strip() or APP_RELEASE_PAGE_URL
    return {
        "tag": tag,
        "version": normalize_release_tag(tag),
        "name": str(raw.get("name") or tag or "Daily Logger release").strip(),
        "body": str(raw.get("body") or "").strip(),
        "html_url": html_url,
        "published_at": str(raw.get("published_at") or "").strip(),
        "assets": assets,
    }


def fetch_latest_release_info(timeout_sec: float = 8.0) -> Tuple[Optional[Dict[str, Any]], str]:
    req = request.Request(
        APP_RELEASE_API_URL,
        headers={
            "Accept": "application/vnd.github+json",
            "User-Agent": f"DailyLogger/{APP_VERSION}",
        },
    )
    try:
        with request.urlopen(req, timeout=timeout_sec) as response:
            raw_text = response.read().decode("utf-8", errors="replace")
        parsed = json.loads(raw_text)
        if not isinstance(parsed, dict):
            return None, "GitHub returned an unexpected release response."
        info = parse_github_release_info(parsed)
        if not info.get("tag"):
            return None, "GitHub release response did not include a version tag."
        return info, ""
    except error.HTTPError as exc:
        return None, f"GitHub update check failed ({exc.code})."
    except (OSError, TimeoutError, json.JSONDecodeError) as exc:
        return None, f"Could not check for updates: {exc}"


LOCAL_TRANSCRIPTION_ADDON_STAGE_PREFIX = "local_transcription_pending"
LOCAL_TRANSCRIPTION_ADDON_VERSION = "helper-v5"
LOCAL_TRANSCRIPTION_HELPER_EXE_NAME = "DailyLoggerLocalTranscriber.exe"
LOCAL_TRANSCRIPTION_CURRENT_FILE = LOCAL_TRANSCRIPTION_ADDON_DIR / "current.json"
MEDIA_TOOLS_ADDON_MARKER = MEDIA_TOOLS_ADDON_DIR / "addon.json"
MEDIA_TOOLS_FFMPEG_EXE = MEDIA_TOOLS_ADDON_DIR / "ffmpeg.exe"
_LOCAL_TRANSCRIPTION_HELPER_HEALTH: Dict[str, Any] = {"path": "", "time": 0.0, "ok": False, "err": ""}


def directory_size_bytes(path: Path) -> int:
    try:
        if path.is_file():
            return int(path.stat().st_size)
        if not path.is_dir():
            return 0
    except OSError:
        return 0
    total = 0
    try:
        for child in path.rglob("*"):
            try:
                if child.is_file():
                    total += int(child.stat().st_size)
            except OSError:
                continue
    except OSError:
        return total
    return total


def format_size_short(num_bytes: int) -> str:
    value = float(max(0, int(num_bytes)))
    units = ("B", "KB", "MB", "GB")
    unit = units[0]
    for unit in units:
        if value < 1024 or unit == units[-1]:
            break
        value /= 1024
    if unit in ("B", "KB"):
        return f"{value:.0f} {unit}"
    if value >= 100:
        return f"{value:.0f} {unit}"
    return f"{value:.1f} {unit}"


def _local_helper_exe_in_dir(path: Path) -> Optional[Path]:
    candidate = path / LOCAL_TRANSCRIPTION_HELPER_EXE_NAME
    if candidate.is_file():
        return candidate
    try:
        matches = sorted(path.rglob(LOCAL_TRANSCRIPTION_HELPER_EXE_NAME), key=lambda p: p.stat().st_mtime, reverse=True)
    except OSError:
        return None
    return matches[0] if matches else None


def _read_local_transcription_current_data() -> Dict[str, str]:
    try:
        data = json.loads(LOCAL_TRANSCRIPTION_CURRENT_FILE.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            return {str(k): str(v) for k, v in data.items() if isinstance(k, str)}
    except Exception:
        pass
    return {}


def _read_local_transcription_current() -> str:
    data = _read_local_transcription_current_data()
    return str(data.get("runtime") or "").strip()


def _local_transcription_current_version() -> str:
    data = _read_local_transcription_current_data()
    return str(data.get("version") or "").strip()


def _local_transcription_helper_path() -> Optional[Path]:
    runtime_name = _read_local_transcription_current()
    if runtime_name:
        helper = _local_helper_exe_in_dir(LOCAL_TRANSCRIPTION_ADDON_DIR / runtime_name)
        if helper is not None:
            return helper
    try:
        runtimes = [
            p for p in LOCAL_TRANSCRIPTION_ADDON_DIR.iterdir()
            if p.is_dir() and p.name.startswith("runtime-")
        ]
        runtimes.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    except OSError:
        runtimes = []
    for runtime in runtimes:
        helper = _local_helper_exe_in_dir(runtime)
        if helper is not None:
            return helper
    helper = _local_helper_exe_in_dir(LOCAL_TRANSCRIPTION_ADDON_DIR)
    if helper is not None:
        return helper
    if not getattr(sys, "frozen", False):
        source_helper = BASE_DIR / "local_transcriber_helper.py"
        if source_helper.is_file():
            return source_helper
    return None


def _local_transcription_helper_command() -> Optional[List[str]]:
    helper = _local_transcription_helper_path()
    if helper is None:
        return None
    if helper.suffix.lower() == ".py":
        return [sys.executable, str(helper)]
    return [str(helper)]


def local_transcription_addon_is_installed() -> bool:
    return _local_transcription_helper_command() is not None


def local_transcription_runtime_available() -> bool:
    return local_transcription_addon_is_installed()


def _hidden_subprocess_kwargs() -> Dict[str, Any]:
    kwargs: Dict[str, Any] = {
        "stdin": subprocess.DEVNULL,
        "stdout": subprocess.PIPE,
        "stderr": subprocess.PIPE,
        "text": True,
        "encoding": "utf-8",
        "errors": "replace",
    }
    if os.name == "nt":
        kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    return kwargs


def _local_helper_environment() -> Dict[str, str]:
    env = os.environ.copy()
    ffmpeg = _find_ffmpeg_executable()
    if ffmpeg:
        env["DAILYLOGGER_FFMPEG"] = ffmpeg
    return env


def _run_local_transcriber_json(
    args: List[str],
    *,
    on_event: Optional[Callable[[Dict[str, Any]], None]] = None,
    timeout: Optional[float] = None,
) -> Tuple[bool, str, Dict[str, Any]]:
    command = _local_transcription_helper_command()
    if command is None:
        return False, "Local transcription add-on is not installed.", {}
    full_command = command + args + ["--json"]
    try:
        proc = subprocess.Popen(full_command, env=_local_helper_environment(), **_hidden_subprocess_kwargs())
    except OSError as exc:
        return False, f"Could not start local transcription helper: {exc}", {}

    last_event: Dict[str, Any] = {}
    error_message = ""
    started = time.monotonic()
    try:
        assert proc.stdout is not None
        for line in proc.stdout:
            if timeout is not None and time.monotonic() - started > timeout:
                proc.kill()
                return False, "Local transcription helper timed out.", last_event
            raw = line.strip()
            if not raw:
                continue
            try:
                event = json.loads(raw)
            except json.JSONDecodeError:
                continue
            last_event = event if isinstance(event, dict) else {}
            if callable(on_event) and last_event:
                try:
                    on_event(last_event)
                except Exception:
                    pass
            if last_event.get("event") == "error":
                error_message = str(last_event.get("message") or "Local transcription helper failed.")
        stderr = ""
        if proc.stderr is not None:
            try:
                stderr = proc.stderr.read().strip()
            except Exception:
                stderr = ""
        code = proc.wait(timeout=5)
    except Exception as exc:
        try:
            proc.kill()
        except Exception:
            pass
        return False, f"Local transcription helper failed: {exc}", last_event
    if code != 0:
        return False, error_message or stderr or f"Local transcription helper exited with code {code}.", last_event
    if error_message:
        return False, error_message, last_event
    return True, "", last_event


def ensure_local_transcription_runtime_loaded(force: bool = False) -> Tuple[bool, str]:
    helper = _local_transcription_helper_path()
    helper_key = str(helper.resolve()) if helper is not None else ""
    now = time.monotonic()
    if helper is not None and getattr(sys, "frozen", False):
        current_version = _local_transcription_current_version()
        if current_version and current_version != LOCAL_TRANSCRIPTION_ADDON_VERSION:
            _LOCAL_TRANSCRIPTION_HELPER_HEALTH.update({
                "path": helper_key,
                "time": now,
                "ok": False,
                "err": "Local transcription add-on needs an update.",
            })
            return False, "Local transcription add-on needs an update."
    if (
        not force
        and helper_key
        and _LOCAL_TRANSCRIPTION_HELPER_HEALTH.get("path") == helper_key
        and now - float(_LOCAL_TRANSCRIPTION_HELPER_HEALTH.get("time") or 0.0) < 8.0
    ):
        return bool(_LOCAL_TRANSCRIPTION_HELPER_HEALTH.get("ok")), str(_LOCAL_TRANSCRIPTION_HELPER_HEALTH.get("err") or "")
    if helper is None:
        _LOCAL_TRANSCRIPTION_HELPER_HEALTH.update({"path": "", "time": now, "ok": False, "err": "Local transcription add-on is not installed."})
        return False, "Local transcription add-on is not installed."
    ok, err, _event = _run_local_transcriber_json(["health"], timeout=30)
    _LOCAL_TRANSCRIPTION_HELPER_HEALTH.update({"path": helper_key, "time": now, "ok": ok, "err": err})
    return ok, err


def _safe_extract_zip(zip_path: Path, target_dir: Path) -> Tuple[bool, str]:
    try:
        with zipfile.ZipFile(zip_path, "r") as archive:
            for info in archive.infolist():
                name = info.filename.replace("\\", "/")
                if not name or name.startswith("/") or re.match(r"^[A-Za-z]:", name):
                    return False, "Add-on ZIP contains an unsafe path."
                parts = [part for part in PurePosixPath(name).parts if part not in ("", ".")]
                if any(part == ".." for part in parts):
                    return False, "Add-on ZIP contains an unsafe path."
            archive.extractall(target_dir)
        return True, ""
    except (OSError, zipfile.BadZipFile) as exc:
        return False, f"Could not extract add-on ZIP: {exc}"


def _addon_zip_candidate_paths(zip_name: str) -> List[Path]:
    candidates = [
        BASE_DIR / zip_name,
        BASE_DIR.parent / zip_name,
        BASE_DIR / "dist" / zip_name,
        BASE_DIR.parent / "dist" / zip_name,
        ADDON_DOWNLOAD_DIR / zip_name,
    ]
    try:
        candidates.append(Path(sys.executable).resolve().parent / zip_name)
        candidates.append(Path(sys.executable).resolve().parent.parent / zip_name)
    except Exception:
        pass
    seen: set[str] = set()
    unique: List[Path] = []
    for path in candidates:
        try:
            key = str(path.resolve()).casefold()
        except OSError:
            key = str(path).casefold()
        if key in seen:
            continue
        seen.add(key)
        unique.append(path)
    return unique


def _find_local_addon_zip(zip_name: str) -> Optional[Path]:
    for candidate in _addon_zip_candidate_paths(zip_name):
        try:
            if candidate.is_file():
                return candidate
        except OSError:
            continue
    return None


def _download_release_addon_zip(
    zip_name: str,
    progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[Optional[Path], str]:
    url = f"{ADDON_RELEASE_BASE_URL}/{zip_name}"
    try:
        ADDON_DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
        target = ADDON_DOWNLOAD_DIR / zip_name
        tmp = target.with_suffix(target.suffix + ".part")
        req = request.Request(url, headers={"User-Agent": "DailyLogger"})
        with request.urlopen(req, timeout=120) as response, tmp.open("wb") as out:
            try:
                total = int(response.headers.get("Content-Length") or "0")
            except (TypeError, ValueError):
                total = 0
            downloaded = 0
            last_percent = -1
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                out.write(chunk)
                downloaded += len(chunk)
                if progress is not None:
                    percent = int(downloaded * 100 / total) if total > 0 else 0
                    if percent != last_percent or total <= 0:
                        last_percent = percent
                        try:
                            progress(downloaded, total)
                        except Exception:
                            pass
        tmp.replace(target)
        if not zipfile.is_zipfile(target):
            try:
                target.unlink(missing_ok=True)
            except OSError:
                pass
            return None, f"Downloaded file is not a valid add-on ZIP:\n{url}"
        return target, ""
    except (OSError, error.URLError, TimeoutError) as exc:
        return None, f"Could not download add-on from GitHub:\n{url}\n\n{exc}"


def resolve_or_download_addon_zip(
    zip_name: str,
    progress: Optional[Callable[[int, int], None]] = None,
) -> Tuple[Optional[Path], str]:
    local_zip = _find_local_addon_zip(zip_name)
    if local_zip is not None:
        return local_zip, ""
    return _download_release_addon_zip(zip_name, progress=progress)


def _find_extracted_local_addon_root(extracted_dir: Path) -> Optional[Path]:
    try:
        matches = sorted(extracted_dir.rglob(LOCAL_TRANSCRIPTION_HELPER_EXE_NAME))
    except OSError:
        return None
    if not matches:
        return None
    return matches[0].parent


def finalize_pending_local_transcription_addon() -> Tuple[bool, str]:
    try:
        LOCAL_TRANSCRIPTION_ADDON_DIR.mkdir(parents=True, exist_ok=True)
        if _local_transcription_helper_path() is None:
            runtimes = [
                p for p in LOCAL_TRANSCRIPTION_ADDON_DIR.iterdir()
                if p.is_dir() and p.name.startswith("runtime-") and _local_helper_exe_in_dir(p) is not None
            ]
            runtimes.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            if runtimes:
                LOCAL_TRANSCRIPTION_CURRENT_FILE.write_text(
                    json.dumps({"runtime": runtimes[0].name, "version": LOCAL_TRANSCRIPTION_ADDON_VERSION}, indent=2),
                    encoding="utf-8",
                )
        return True, ""
    except OSError as exc:
        return (
            False,
            f"Could not finish local transcription add-on startup cleanup: {exc}",
        )


def _cleanup_staged_local_transcription_addons(keep: Optional[Path] = None) -> None:
    parent = LOCAL_TRANSCRIPTION_ADDON_DIR.parent
    try:
        keep_resolved = keep.resolve() if keep is not None else None
    except OSError:
        keep_resolved = keep
    try:
        staged_dirs = [
            p for p in parent.iterdir()
            if p.is_dir() and p.name.startswith(LOCAL_TRANSCRIPTION_ADDON_STAGE_PREFIX)
        ]
    except OSError:
        return
    for staged in staged_dirs:
        try:
            if keep_resolved is not None and staged.resolve() == keep_resolved:
                continue
        except OSError:
            pass
        try:
            shutil.rmtree(staged)
        except OSError:
            continue


def _cleanup_old_local_transcription_runtimes(keep: Optional[Path] = None) -> None:
    try:
        keep_resolved = keep.resolve() if keep is not None else None
    except OSError:
        keep_resolved = keep
    try:
        runtime_dirs = [
            p for p in LOCAL_TRANSCRIPTION_ADDON_DIR.iterdir()
            if p.is_dir() and p.name.startswith("runtime-")
        ]
    except OSError:
        return
    for runtime_dir in runtime_dirs:
        try:
            if keep_resolved is not None and runtime_dir.resolve() == keep_resolved:
                continue
        except OSError:
            pass
        try:
            shutil.rmtree(runtime_dir)
        except OSError:
            continue


def install_local_transcription_addon(zip_path: Path) -> Tuple[bool, str]:
    if not zip_path.is_file():
        return False, "Choose DailyLoggerLocalTranscriptionAddon.zip first."
    parent = LOCAL_TRANSCRIPTION_ADDON_DIR
    temp_dir = LOCAL_TRANSCRIPTION_ADDON_DIR.parent / f".local_transcription_install_{secrets.token_hex(4)}"
    runtime_name = f"runtime-{datetime.now().strftime('%Y%m%d%H%M%S')}-{secrets.token_hex(4)}"
    runtime_dir = LOCAL_TRANSCRIPTION_ADDON_DIR / runtime_name
    try:
        parent.mkdir(parents=True, exist_ok=True)
        temp_dir.mkdir(parents=True, exist_ok=False)
        ok, err_msg = _safe_extract_zip(zip_path, temp_dir)
        if not ok:
            return False, err_msg
        extracted_root = _find_extracted_local_addon_root(temp_dir)
        if extracted_root is None:
            return False, "That ZIP does not contain DailyLoggerLocalTranscriber.exe."
        shutil.move(str(extracted_root), str(runtime_dir))
        if _local_helper_exe_in_dir(runtime_dir) is None:
            return False, "That local transcription add-on is incomplete."
        LOCAL_TRANSCRIPTION_CURRENT_FILE.write_text(
            json.dumps(
                {
                    "runtime": runtime_name,
                    "version": LOCAL_TRANSCRIPTION_ADDON_VERSION,
                    "installed_at": datetime.now().isoformat(timespec="seconds"),
                },
                indent=2,
            ),
            encoding="utf-8",
        )
        _LOCAL_TRANSCRIPTION_HELPER_HEALTH.update({"path": "", "time": 0.0, "ok": False, "err": ""})
        _cleanup_staged_local_transcription_addons()
        _cleanup_old_local_transcription_runtimes(keep=runtime_dir)
        return True, ""
    except OSError as exc:
        return False, f"Could not install local transcription add-on: {exc}"
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except OSError:
            pass


def uninstall_local_transcription_addon() -> Tuple[bool, str]:
    _clear_local_transcription_model_cache()
    _LOCAL_TRANSCRIPTION_HELPER_HEALTH.update({"path": "", "time": 0.0, "ok": False, "err": ""})
    try:
        _cleanup_staged_local_transcription_addons()
        if LOCAL_TRANSCRIPTION_ADDON_DIR.exists():
            shutil.rmtree(LOCAL_TRANSCRIPTION_ADDON_DIR)
        return True, ""
    except OSError as exc:
        return (
            False,
            "Could not remove the local transcription add-on. Close and reopen Daily Logger, then try again. "
            f"Details: {exc}",
        )


def media_tools_addon_is_installed() -> bool:
    return MEDIA_TOOLS_FFMPEG_EXE.is_file()


def _find_extracted_media_tools_root(extracted_dir: Path) -> Optional[Path]:
    candidates = [extracted_dir, extracted_dir / "media_tools"]
    try:
        candidates.extend([p for p in extracted_dir.iterdir() if p.is_dir()])
    except OSError:
        pass
    for candidate in candidates:
        if (candidate / "ffmpeg.exe").is_file():
            return candidate
    return None


def install_media_tools_addon(zip_path: Path) -> Tuple[bool, str]:
    if not zip_path.is_file():
        return False, "Choose DailyLoggerMediaToolsAddon.zip first."
    parent = MEDIA_TOOLS_ADDON_DIR.parent
    temp_dir = parent / f".media_tools_install_{secrets.token_hex(4)}"
    try:
        parent.mkdir(parents=True, exist_ok=True)
        temp_dir.mkdir(parents=True, exist_ok=False)
        ok, err_msg = _safe_extract_zip(zip_path, temp_dir)
        if not ok:
            return False, err_msg
        extracted_root = _find_extracted_media_tools_root(temp_dir)
        if extracted_root is None:
            return False, "That ZIP does not look like a Daily Logger Media Tools add-on."
        if MEDIA_TOOLS_ADDON_DIR.exists():
            shutil.rmtree(MEDIA_TOOLS_ADDON_DIR)
        shutil.move(str(extracted_root), str(MEDIA_TOOLS_ADDON_DIR))
        return True, ""
    except OSError as exc:
        return False, f"Could not install Media Tools add-on: {exc}"
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except OSError:
            pass


def uninstall_media_tools_addon() -> Tuple[bool, str]:
    try:
        if MEDIA_TOOLS_ADDON_DIR.exists():
            shutil.rmtree(MEDIA_TOOLS_ADDON_DIR)
        return True, ""
    except OSError as exc:
        return (
            False,
            "Could not remove the Media Tools add-on. Close and reopen Daily Logger, then try again. "
            f"Details: {exc}",
        )


def normalize_transcription_model_choice(value: str) -> str:
    raw = (value or "").strip().lower()
    cloud_map = {m.lower(): m for m in TRANSCRIPTION_CLOUD_MODEL_NAMES}
    if raw in ("", "local", "local:"):
        return TRANSCRIPTION_DEFAULT_MODEL
        if raw in (TRANSCRIPTION_MODEL_CLOUD, "openai", "cloud:openai"):
            return TRANSCRIPTION_MODEL_CLOUD
    if raw in cloud_map:
        cloud_name = cloud_map[raw]
        if cloud_name == OPENAI_TRANSCRIPTION_MODEL:
            return TRANSCRIPTION_MODEL_CLOUD
        return f"cloud:{cloud_name}"
    if raw.startswith("cloud:"):
        cloud_name = raw.split(":", 1)[1].strip()
        if not cloud_name or cloud_name == TRANSCRIPTION_DEFAULT_CLOUD_MODEL.lower():
            return TRANSCRIPTION_MODEL_CLOUD
        if cloud_name in cloud_map:
            return f"cloud:{cloud_map[cloud_name]}"
        return TRANSCRIPTION_MODEL_CLOUD
    if raw.startswith("local:"):
        model_name = raw.split(":", 1)[1].strip()
    else:
        model_name = raw
    if model_name in TRANSCRIPTION_LOCAL_MODEL_NAMES:
        return f"local:{model_name}"
    return TRANSCRIPTION_DEFAULT_MODEL


def get_selected_transcription_model_choice() -> str:
    prefs = load_preferences()
    return normalize_transcription_model_choice(
        prefs.get(TRANSCRIPTION_MODEL_PREF_KEY, TRANSCRIPTION_DEFAULT_MODEL)
    )


def save_selected_transcription_model_choice(choice: str) -> bool:
    prefs = load_preferences()
    prefs[TRANSCRIPTION_MODEL_PREF_KEY] = normalize_transcription_model_choice(choice)
    return save_preferences(prefs)


def normalize_record_source_mode(value: str) -> str:
    raw = (value or "").strip().lower().replace("-", "_").replace(" ", "_")
    if raw in ("both", "record_both", "mic_computer", "mic_and_computer", "microphone_and_computer"):
        return RECORD_SOURCE_BOTH
    if raw in ("mic", "microphone", "microphone_only", "record_mic", "record_microphone"):
        return RECORD_SOURCE_MIC
    if raw in (
        "computer",
        "computer_audio",
        "computer_only",
        "system",
        "system_audio",
        "speaker",
        "speakers",
        "loopback",
        "record_computer",
    ):
        return RECORD_SOURCE_COMPUTER
    return RECORD_SOURCE_BOTH


def get_selected_record_source_mode() -> str:
    prefs = load_preferences()
    return normalize_record_source_mode(prefs.get(RECORD_SOURCE_PREF_KEY, RECORD_SOURCE_BOTH))


def save_selected_record_source_mode(mode: str) -> bool:
    prefs = load_preferences()
    prefs[RECORD_SOURCE_PREF_KEY] = normalize_record_source_mode(mode)
    return save_preferences(prefs)


def iphone_passive_receive_enabled() -> bool:
    prefs = load_preferences()
    return prefs.get(IPHONE_PASSIVE_RECEIVE_PREF_KEY, "true").strip().lower() != "false"


def save_iphone_passive_receive_enabled(enabled: bool) -> bool:
    prefs = load_preferences()
    prefs[IPHONE_PASSIVE_RECEIVE_PREF_KEY] = "true" if enabled else "false"
    return save_preferences(prefs)


def transcription_model_is_local(choice: str) -> bool:
    return normalize_transcription_model_choice(choice).startswith("local:")


def transcription_model_is_cloud(choice: str) -> bool:
    normalized = normalize_transcription_model_choice(choice)
    return normalized == TRANSCRIPTION_MODEL_CLOUD or normalized.startswith("cloud:")


def transcription_cloud_model_name(choice: str) -> str:
    normalized = normalize_transcription_model_choice(choice)
    if normalized.startswith("cloud:"):
        return normalized.split(":", 1)[1]
    return TRANSCRIPTION_DEFAULT_CLOUD_MODEL


def transcription_local_model_name(choice: str) -> str:
    normalized = normalize_transcription_model_choice(choice)
    if not normalized.startswith("local:"):
        return ""
    return normalized.split(":", 1)[1]


def local_transcription_model_path(model_name: str) -> Path:
    safe_name = model_name if model_name in TRANSCRIPTION_LOCAL_MODEL_NAMES else "small"
    return LOCAL_TRANSCRIPTION_MODEL_DIR / safe_name


def local_transcription_model_is_downloaded(model_name: str) -> bool:
    path = local_transcription_model_path(model_name)
    return (
        path.is_dir()
        and (path / "config.json").is_file()
        and (path / "model.bin").is_file()
        and (path / "tokenizer.json").is_file()
    )


def downloaded_local_transcription_model_names() -> List[str]:
    return [
        model_name
        for model_name in TRANSCRIPTION_LOCAL_MODEL_NAMES
        if local_transcription_model_is_downloaded(model_name)
    ]


def usable_local_transcription_model_names() -> List[str]:
    runtime_ok, _runtime_err = ensure_local_transcription_runtime_loaded()
    if not runtime_ok:
        return []
    return downloaded_local_transcription_model_names()


def first_usable_local_transcription_choice() -> str:
    names = usable_local_transcription_model_names()
    if not names:
        return ""
    return f"local:{names[0]}"


def transcription_cloud_is_ready() -> bool:
    return bool(get_openai_api_key())


def transcription_has_any_ready_model() -> bool:
    return transcription_cloud_is_ready() or bool(usable_local_transcription_model_names())


def download_local_transcription_model(
    model_name: str,
    *,
    progress: Optional[Callable[[int], None]] = None,
    status: Optional[Callable[[str], None]] = None,
) -> Tuple[bool, str]:
    if model_name not in TRANSCRIPTION_LOCAL_MODEL_NAMES:
        return False, f"Unsupported local transcription model: {model_name}"
    ok, err_msg = ensure_local_transcription_runtime_loaded()
    if not ok:
        return False, err_msg

    def _event(event: Dict[str, Any]) -> None:
        if event.get("event") in ("status", "progress"):
            message = str(event.get("message") or "").strip()
            if message and status is not None:
                try:
                    status(message)
                except Exception:
                    pass
            if progress is not None and "percent" in event:
                try:
                    progress(int(float(event.get("percent") or 0)))
                except Exception:
                    pass

    ok, err_msg, _last = _run_local_transcriber_json(
        [
            "download",
            "--model",
            model_name,
            "--models-dir",
            str(LOCAL_TRANSCRIPTION_MODEL_DIR),
        ],
        on_event=_event,
        timeout=None,
    )
    if not ok:
        return False, err_msg
    if not local_transcription_model_is_downloaded(model_name):
        return False, f"Downloaded model '{model_name}' is incomplete."
    return True, ""


def uninstall_local_transcription_model(model_name: str) -> Tuple[bool, str]:
    if model_name not in TRANSCRIPTION_LOCAL_MODEL_NAMES:
        return False, f"Unsupported local transcription model: {model_name}"
    try:
        root = LOCAL_TRANSCRIPTION_MODEL_DIR.resolve()
        target = local_transcription_model_path(model_name).resolve()
    except OSError as exc:
        return False, f"Could not locate local model folder: {exc}"
    try:
        target.relative_to(root)
    except ValueError:
        return False, "Refusing to remove a folder outside the Daily Logger model directory."
    try:
        if target.exists():
            shutil.rmtree(target)
        return True, ""
    except OSError as exc:
        return False, f"Could not uninstall local transcription model '{model_name}': {exc}"


def ensure_iphone_inbox_dir() -> Path:
    IPHONE_INBOX_DIR.mkdir(parents=True, exist_ok=True)
    return IPHONE_INBOX_DIR


def ensure_iphone_incoming_dir() -> Path:
    IPHONE_INCOMING_DIR.mkdir(parents=True, exist_ok=True)
    return IPHONE_INCOMING_DIR


def ensure_iphone_declined_dir() -> Path:
    IPHONE_DECLINED_DIR.mkdir(parents=True, exist_ok=True)
    return IPHONE_DECLINED_DIR


def get_or_create_iphone_import_token() -> str:
    prefs = load_preferences()
    token = prefs.get(IPHONE_IMPORT_TOKEN_PREF_KEY, "").strip()
    if len(token) >= 16 and re.fullmatch(r"[A-Za-z0-9._~-]+", token):
        return token
    token = secrets.token_urlsafe(18).rstrip("=")
    prefs[IPHONE_IMPORT_TOKEN_PREF_KEY] = token
    save_preferences(prefs)
    return token


def get_lan_ip_address() -> str:
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
            s.connect(("8.8.8.8", 80))
            ip = str(s.getsockname()[0])
            if ip and not ip.startswith("127."):
                return ip
    except OSError:
        pass
    try:
        host_ip = socket.gethostbyname(socket.gethostname())
        if host_ip and not host_ip.startswith("127."):
            return host_ip
    except OSError:
        pass
    return "127.0.0.1"


def get_computer_network_name_candidates() -> List[str]:
    names: List[str] = []
    seen: set[str] = set()
    for raw in (os.getenv("COMPUTERNAME", ""), socket.gethostname()):
        cleaned = re.sub(r"[^A-Za-z0-9-]+", "", str(raw or "").strip())
        key = cleaned.casefold()
        if cleaned and not cleaned.startswith("-") and key not in seen:
            names.append(cleaned)
            seen.add(key)
    if names:
        local_name = f"{names[0]}.local"
        if local_name.casefold() not in seen:
            names.append(local_name)
    return names[:2]


def build_iphone_upload_url(host: str, port: int, token: str) -> str:
    return f"http://{host}:{port}/upload?token={token}"


def build_iphone_mobile_page_url(host: str, port: int, token: str) -> str:
    return f"http://{host}:{port}/iphone?token={token}"


def build_iphone_receiver_urls(port: int, token: str) -> Dict[str, str]:
    ip = get_lan_ip_address()
    shortcut_url = build_iphone_upload_url(ip, port, token)
    urls: Dict[str, str] = {
        "wifi": shortcut_url,
        "shortcut": shortcut_url,
        "mobile": build_iphone_mobile_page_url(ip, port, token),
    }
    return urls


def _filename_from_content_disposition(value: str) -> str:
    if not value:
        return ""
    match = re.search(r"filename\*=UTF-8''([^;]+)", value, flags=re.IGNORECASE)
    if match:
        return unquote(match.group(1).strip().strip('"'))
    match = re.search(r'filename="([^"]+)"', value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r"filename=([^;]+)", value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip().strip('"')
    return ""


def _safe_iphone_upload_suffix(filename: str, content_type: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix in TRANSCRIPTION_MEDIA_SUFFIXES:
        return suffix
    ctype = (content_type or "").split(";", 1)[0].strip().lower()
    suffix = IPHONE_IMPORT_CONTENT_TYPE_SUFFIXES.get(ctype, "")
    if suffix in TRANSCRIPTION_MEDIA_SUFFIXES:
        return suffix
    return ""


def _infer_iphone_upload_suffix_from_file(path: Path, filename: str, content_type: str) -> str:
    suffix = _safe_iphone_upload_suffix(filename, content_type)
    if suffix:
        return suffix
    try:
        header = path.read_bytes()[:512]
    except OSError:
        header = b""
    if header.startswith(b"RIFF") and header[8:12] == b"WAVE":
        return ".wav"
    if header.startswith(b"ID3") or header[:2] in (b"\xff\xfb", b"\xff\xf3", b"\xff\xf2"):
        return ".mp3"
    if header.startswith(b"OggS"):
        return ".ogg"
    if header.startswith(b"fLaC"):
        return ".flac"
    if header.startswith(b"caff"):
        return ".caf"
    if len(header) >= 2 and header[0] == 0xFF and (header[1] & 0xF0) == 0xF0:
        return ".aac"
    if header.startswith(b"\x1a\x45\xdf\xa3"):
        return ".webm"
    if len(header) >= 12 and header[4:8] == b"ftyp":
        brands = header[8:80]
        if any(brand in brands for brand in (b"M4A", b"M4B", b"M4P")):
            return ".m4a"
        if b"qt  " in brands:
            return ".mov"
        ctype = (content_type or "").split(";", 1)[0].strip().lower()
        if ctype.startswith("audio/"):
            return ".m4a"
        if ctype.startswith("video/"):
            return ".mov"
        return ".mp4"
    return ""


def _sanitize_iphone_upload_name_with_suffix(filename: str, suffix: str) -> Tuple[str, str]:
    suffix = (suffix or "").lower()
    if suffix not in TRANSCRIPTION_MEDIA_SUFFIXES:
        return "", ""
    raw_stem = Path(filename).stem if filename else ""
    stem = re.sub(r"[^A-Za-z0-9._ -]+", "_", raw_stem).strip(" ._-")
    if not stem:
        stem = "iphone_video" if suffix in TRANSCRIPTION_VIDEO_SUFFIXES else "iphone_audio"
    if len(stem) > 80:
        stem = stem[:80].rstrip(" ._-") or "iphone_media"
    return stem, suffix


def _sanitize_iphone_upload_name(filename: str, content_type: str) -> Tuple[str, str]:
    return _sanitize_iphone_upload_name_with_suffix(
        filename,
        _safe_iphone_upload_suffix(filename, content_type),
    )


def _unique_iphone_inbox_path_for_suffix(
    filename: str,
    suffix: str,
    *,
    folder: Optional[Path] = None,
) -> Tuple[Optional[Path], str]:
    stem, suffix = _sanitize_iphone_upload_name_with_suffix(filename, suffix)
    if not suffix:
        allowed = ", ".join(sorted(TRANSCRIPTION_MEDIA_SUFFIXES))
        return None, f"Unsupported upload type. Use: {allowed}"
    inbox = folder or ensure_iphone_inbox_dir()
    inbox.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    for attempt in range(200):
        nonce = secrets.token_hex(3)
        extra = f"_{attempt}" if attempt else ""
        candidate = inbox / f"{stamp}_{stem}_{nonce}{extra}{suffix}"
        if not candidate.exists() and not candidate.with_suffix(candidate.suffix + ".part").exists():
            return candidate, ""
    return None, "Could not create a unique iPhone Inbox filename."


def _unique_iphone_inbox_path(
    filename: str,
    content_type: str,
    *,
    folder: Optional[Path] = None,
) -> Tuple[Optional[Path], str]:
    stem, suffix = _sanitize_iphone_upload_name(filename, content_type)
    if not suffix:
        allowed = ", ".join(sorted(TRANSCRIPTION_MEDIA_SUFFIXES))
        return None, f"Unsupported upload type. Use: {allowed}"
    return _unique_iphone_inbox_path_for_suffix(filename, suffix, folder=folder)


def _unique_existing_file_dest(folder: Path, src: Path) -> Optional[Path]:
    try:
        folder.mkdir(parents=True, exist_ok=True)
    except OSError:
        return None
    dest = folder / src.name
    if not dest.exists():
        return dest
    stem = dest.stem
    suffix = dest.suffix
    for attempt in range(1, 500):
        alt = folder / f"{stem}_{attempt}{suffix}"
        if not alt.exists():
            return alt
    return None


def apply_iphone_last_modified(path: Path, raw_value: str) -> None:
    try:
        last_modified = float(raw_value)
        if last_modified > 10_000_000_000:
            last_modified /= 1000.0
        if 946684800 <= last_modified <= 4102444800:
            os.utime(path, (last_modified, last_modified))
    except (OSError, TypeError, ValueError):
        pass


def list_pending_iphone_inbox_files() -> List[Path]:
    try:
        inbox = ensure_iphone_inbox_dir()
    except OSError:
        return []
    items: List[Path] = []
    try:
        for path in inbox.iterdir():
            if (
                path.is_file()
                and not path.name.lower().endswith(".part")
                and is_transcription_media_file(path)
            ):
                items.append(path)
    except OSError:
        return []
    return sorted(items, key=lambda p: (p.stat().st_mtime if p.exists() else float("inf"), p.name.lower()))


def list_incoming_iphone_files() -> List[Path]:
    try:
        incoming = ensure_iphone_incoming_dir()
    except OSError:
        return []
    items: List[Path] = []
    try:
        for path in incoming.iterdir():
            if (
                path.is_file()
                and not path.name.lower().endswith(".part")
                and is_transcription_media_file(path)
            ):
                items.append(path)
    except OSError:
        return []
    return sorted(items, key=lambda p: (p.stat().st_mtime if p.exists() else float("inf"), p.name.lower()))


def mark_iphone_inbox_files_processed(paths: List[Path]) -> None:
    try:
        processed_dir = ensure_iphone_inbox_dir() / "Processed"
        processed_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        return
    for src in paths:
        try:
            if not src.is_file() or src.parent.resolve() != IPHONE_INBOX_DIR.resolve():
                continue
            dest = _unique_existing_file_dest(processed_dir, src)
            if dest is None:
                continue
            shutil.move(str(src), str(dest))
        except OSError:
            continue


def move_iphone_file_to_inbox(src: Path) -> Optional[Path]:
    try:
        if not src.is_file() or not is_transcription_media_file(src):
            return None
        dest = _unique_existing_file_dest(ensure_iphone_inbox_dir(), src)
        if dest is None:
            return None
        shutil.move(str(src), str(dest))
        return dest
    except OSError:
        return None


def decline_iphone_file(src: Path) -> bool:
    try:
        if not src.is_file():
            return False
        dest = _unique_existing_file_dest(ensure_iphone_declined_dir(), src)
        if dest is None:
            return False
        shutil.move(str(src), str(dest))
        return True
    except OSError:
        return False


def _is_pref_true(value: str) -> bool:
    return value.strip().lower() == "true"


def ensure_backup_folder() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)


def _list_backup_zip_files() -> List[Path]:
    ensure_backup_folder()
    return sorted(
        [path for path in BACKUP_DIR.glob("*.zip") if path.is_file()],
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def run_backup_now() -> Optional[Path]:
    with _backup_lock:
        ensure_backup_folder()
        items_to_backup = [
            path
            for path in DATA_DIR.iterdir()
            if path.name.lower() != BACKUP_DIR.name.lower()
        ]
        if not items_to_backup:
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        zip_path = BACKUP_DIR / f"backup_{timestamp}.zip"
        temp_zip_path = BACKUP_DIR / f".backup_{timestamp}.tmp"

        try:
            with zipfile.ZipFile(
                temp_zip_path,
                mode="w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=BACKUP_COMPRESSION_LEVEL,
            ) as archive:
                for item in items_to_backup:
                    if item.is_file():
                        archive.write(item, arcname=item.name)
                        continue
                    if item.is_dir():
                        for nested in item.rglob("*"):
                            if nested.is_file():
                                archive.write(nested, arcname=str(nested.relative_to(DATA_DIR)))
            temp_zip_path.replace(zip_path)
        except Exception:
            try:
                temp_zip_path.unlink()
            except OSError:
                pass
            raise
        return zip_path


def trim_backups_if_limited(prefs: Dict[str, str]) -> None:
    if not _is_pref_true(prefs.get("backup_limited", "false")):
        return
    backups = _list_backup_zip_files()
    if len(backups) <= 3:
        return
    oldest_backup = backups[-1]
    try:
        oldest_backup.unlink()
        print(f"Backup limited mode: removed oldest backup {oldest_backup.name}")
    except OSError:
        print(f"Backup limited mode: could not remove {oldest_backup.name}")


def evict_oldest_backup_if_limited_full(prefs: Dict[str, str]) -> None:
    if not _is_pref_true(prefs.get("backup_limited", "false")):
        return
    backups = _list_backup_zip_files()
    if len(backups) < 3:
        return
    oldest_backup = backups[-1]
    try:
        oldest_backup.unlink()
        print(f"Backup limited mode: removed oldest backup {oldest_backup.name} before new backup")
    except OSError:
        print(f"Backup limited mode: could not remove {oldest_backup.name} before new backup")


def maybe_run_daily_auto_backup() -> None:
    prefs = load_preferences()
    backup_enabled = prefs.get("backup_enabled", "true")
    if not _is_pref_true(backup_enabled):
        return

    ensure_backup_folder()
    today = datetime.now().strftime("%Y-%m-%d")
    last_program_run_date = prefs.get("last_program_run_date", "").strip()
    if last_program_run_date != today:
        try:
            evict_oldest_backup_if_limited_full(prefs)
            backup_path = run_backup_now()
        except Exception as exc:
            print(f"Auto backup failed: {exc}")
        else:
            if backup_path is None:
                print("Auto backup skipped: nothing in daily_logs to back up.")
            else:
                print(f"Auto backup created: {backup_path.name}")
                trim_backups_if_limited(prefs)
                prefs["last_backup_date"] = today

    prefs["backup_enabled"] = "true" if _is_pref_true(backup_enabled) else "false"
    prefs["last_program_run_date"] = today
    if not save_preferences(prefs):
        print("Warning: could not save backup preferences.")


def start_daily_auto_backup_in_background(delay_sec: float = AUTO_BACKUP_START_DELAY_SEC) -> Optional[threading.Thread]:
    prefs = load_preferences()
    if not _is_pref_true(prefs.get("backup_enabled", "true")):
        return None
    today = datetime.now().strftime("%Y-%m-%d")
    if prefs.get("last_program_run_date", "").strip() == today:
        return None

    def _worker() -> None:
        if delay_sec > 0:
            time.sleep(delay_sec)
        maybe_run_daily_auto_backup()

    thread = threading.Thread(target=_worker, name="DailyLoggerAutoBackup", daemon=True)
    thread.start()
    return thread


def prompt_for_app_name() -> str:
    entered = input(
        "What would you like to name this app? (Press Enter for default name): "
    ).strip()
    if is_enter_equivalent(entered):
        return "Daily Logger"
    return entered


def get_or_create_app_name() -> str:
    prefs = load_preferences()
    app_name = prefs.get("app_name", "").strip()
    if app_name:
        return app_name
    app_name = prompt_for_app_name()
    prefs["app_name"] = app_name
    if not save_preferences(prefs):
        print("Warning: could not save app name preference.")
    return app_name


def rename_app_name() -> str:
    app_name = prompt_for_app_name()
    prefs = load_preferences()
    prefs["app_name"] = app_name
    if save_preferences(prefs):
        print(f'App renamed to "{app_name}".')
    else:
        print("Could not save new app name preference.")
    return app_name


def rename_app_name_to(new_name: str) -> str:
    app_name = new_name.strip() or "Daily Logger"
    prefs = load_preferences()
    prefs["app_name"] = app_name
    if save_preferences(prefs):
        print(f'App renamed to "{app_name}".')
    else:
        print("Could not save new app name preference.")
    return app_name


def get_startup_folder() -> Optional[Path]:
    appdata = os.getenv("APPDATA", "").strip()
    if not appdata:
        return None
    return Path(appdata) / "Microsoft" / "Windows" / "Start Menu" / "Programs" / "Startup"


def get_startup_shortcut_path() -> Optional[Path]:
    startup_dir = get_startup_folder()
    if startup_dir is None:
        return None
    return startup_dir / STARTUP_SHORTCUT_NAME


def create_startup_shortcut() -> bool:
    shortcut_path = get_startup_shortcut_path()
    if shortcut_path is None:
        return False
    try:
        shortcut_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError:
        return False
    target_path = str((BASE_DIR / "launch_daily_logger.bat").resolve())
    ps_script = (
        "$WshShell = New-Object -ComObject WScript.Shell; "
        f"$Shortcut = $WshShell.CreateShortcut('{str(shortcut_path)}'); "
        f"$Shortcut.TargetPath = '{target_path}'; "
        f"$Shortcut.WorkingDirectory = '{str(BASE_DIR)}'; "
        "$Shortcut.Save();"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0 and shortcut_path.exists()
    except OSError:
        return False


def remove_startup_shortcut() -> bool:
    shortcut_path = get_startup_shortcut_path()
    if shortcut_path is None:
        return False
    if not shortcut_path.exists():
        return True
    try:
        shortcut_path.unlink()
        return True
    except OSError:
        return False


def is_startup_enabled() -> bool:
    shortcut_path = get_startup_shortcut_path()
    return bool(shortcut_path and shortcut_path.exists())


def open_current_directory_in_explorer() -> bool:
    return open_path_with_default_app(USER_DATA_ROOT)


def open_path_with_default_app(path: Path) -> bool:
    try:
        if sys.platform == "win32":
            os.startfile(str(path))  # type: ignore[attr-defined]
            return True
        return False
    except OSError:
        return False


def _resolve_path_for_compare(path: Path) -> Path:
    try:
        return path.resolve()
    except OSError:
        return path.absolute()


def _path_is_relative_to(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def _current_app_folder() -> Path:
    if getattr(sys, "frozen", False):
        return _resolve_path_for_compare(Path(sys.executable).parent)
    return _resolve_path_for_compare(BASE_DIR)


def _path_is_current_app_file(path: Path) -> bool:
    resolved = _resolve_path_for_compare(path)
    current_app = _current_app_folder()
    return resolved == current_app or _path_is_relative_to(resolved, current_app)


def _remove_path_quietly(path: Path, *, protect_current_app: bool = True) -> bool:
    if protect_current_app and _path_is_current_app_file(path):
        return False
    try:
        if path.is_dir():
            shutil.rmtree(path)
            return not path.exists()
        if path.exists():
            path.unlink()
            return not path.exists()
        return True
    except OSError:
        return False


def _remove_empty_dir_quietly(path: Path) -> bool:
    if _path_is_current_app_file(path):
        return False
    try:
        path.rmdir()
        return True
    except OSError:
        return False


def _remove_daily_logger_start_menu_shortcuts() -> int:
    removed = 0
    programs = get_start_menu_programs_dir()
    if programs is None:
        return removed
    folder = programs / "Daily Logger"
    for name in (
        "Daily Logger BAT Launcher.lnk",
        "Daily Logger Journal Excel.lnk",
        "Virtual Journal Reader.lnk",
    ):
        shortcut = folder / name
        if shortcut.exists() and _remove_path_quietly(shortcut, protect_current_app=False):
            removed += 1
    _remove_empty_dir_quietly(folder)
    return removed


def _cleanup_download_leftovers() -> int:
    removed = 0
    artifact_names = (
        APP_PORTABLE_ZIP_NAME,
        LOCAL_TRANSCRIPTION_ADDON_ZIP_NAME,
        MEDIA_TOOLS_ADDON_ZIP_NAME,
    )
    candidate_dirs = [
        ADDON_DOWNLOAD_DIR,
        USER_DATA_ROOT,
        BASE_DIR,
        BASE_DIR.parent,
    ]
    seen: set[Path] = set()
    for directory in candidate_dirs:
        for name in artifact_names:
            candidate = directory / name
            resolved = _resolve_path_for_compare(candidate)
            if resolved in seen:
                continue
            seen.add(resolved)
            if candidate.exists() and _remove_path_quietly(candidate, protect_current_app=False):
                removed += 1
    return removed


def _safe_to_schedule_app_folder_delete(app_folder: Path, exe_path: Path) -> Tuple[bool, str]:
    folder = _resolve_path_for_compare(app_folder)
    exe = _resolve_path_for_compare(exe_path)
    if not _path_is_relative_to(exe, folder):
        return False, "the app EXE is not inside the app folder"
    if not exe.is_file():
        return False, "the app EXE could not be found"
    try:
        if folder == Path(folder.anchor):
            return False, "the app folder path is unsafe"
    except OSError:
        return False, "the app folder path is unsafe"
    if str(folder).strip() in ("", "\\", "/"):
        return False, "the app folder path is unsafe"
    try:
        if folder.parent == folder:
            return False, "the app folder parent is unsafe"
    except OSError:
        return False, "the app folder parent is unsafe"
    return True, ""


def _safe_empty_parent_after_app_delete(app_folder: Path) -> Optional[Path]:
    folder = _resolve_path_for_compare(app_folder)
    parent = _resolve_path_for_compare(folder.parent)
    try:
        if parent == folder or parent == Path(parent.anchor):
            return None
    except OSError:
        return None
    parent_name = parent.name.casefold()
    folder_name = folder.name.casefold()
    if folder_name != "dailylogger":
        return None
    if not (
        parent_name == "dailyloggerportable"
        or parent_name.startswith("dailyloggerportable")
        or parent_name in {"daily logger portable", "daily logger"}
    ):
        return None
    return parent


def _schedule_windows_app_folder_delete(app_folder: Path, exe_path: Path) -> Tuple[bool, str]:
    ok, reason = _safe_to_schedule_app_folder_delete(app_folder, exe_path)
    if not ok:
        return False, reason
    script_path = Path(tempfile.gettempdir()) / f"daily_logger_uninstall_{int(time.time())}_{secrets.token_hex(3)}.cmd"
    app_dir = str(_resolve_path_for_compare(app_folder))
    parent_dir = ""
    safe_parent = _safe_empty_parent_after_app_delete(app_folder)
    if safe_parent is not None:
        parent_dir = str(safe_parent)
    pid = os.getpid()
    script = (
        "@echo off\n"
        "setlocal\n"
        f'set "APP_DIR={app_dir}"\n'
        f'set "APP_PARENT={parent_dir}"\n'
        f"set \"APP_PID={pid}\"\n"
        ":wait_for_daily_logger\n"
        'tasklist /FI "PID eq %APP_PID%" 2>nul | find "%APP_PID%" >nul\n'
        "if not errorlevel 1 (\n"
        "  timeout /t 1 /nobreak >nul\n"
        "  goto wait_for_daily_logger\n"
        ")\n"
        'cd /d "%TEMP%" >nul 2>&1\n'
        'rmdir /s /q "%APP_DIR%" >nul 2>&1\n'
        'if not "%APP_PARENT%"=="" rmdir "%APP_PARENT%" >nul 2>&1\n'
        'del /f /q "%~f0" >nul 2>&1\n'
    )
    try:
        script_path.write_text(script, encoding="utf-8")
        subprocess.Popen(
            ["cmd", "/c", str(script_path)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            stdin=subprocess.DEVNULL,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0)
            | getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
        )
        return True, ""
    except OSError as exc:
        return False, str(exc)


def run_clean_uninstall() -> None:
    removed_items = 0
    if remove_startup_shortcut():
        removed_items += 1
    removed_items += _remove_daily_logger_start_menu_shortcuts()

    cleanup_targets = [
        (DATA_DIR, False),
        (SETTINGS_DIR, False),
        (USER_DATA_ROOT / "addons", False),
        (USER_DATA_ROOT / "models", False),
        (ADDON_DOWNLOAD_DIR, False),
        (LEGACY_DATA_DIR, False),
        (LEGACY_SETTINGS_DIR, False),
        (USER_DATA_ROOT / "_internal", True),
    ]
    for target, protect_current_app in cleanup_targets:
        if target.exists() and _remove_path_quietly(target, protect_current_app=protect_current_app):
            removed_items += 1
    removed_items += _cleanup_download_leftovers()

    if USER_DATA_ROOT.exists() and _remove_path_quietly(USER_DATA_ROOT, protect_current_app=True):
        removed_items += 1
    else:
        _remove_empty_dir_quietly(USER_DATA_ROOT)

    scheduled_app_delete = False
    schedule_err = ""
    if getattr(sys, "frozen", False):
        scheduled_app_delete, schedule_err = _schedule_windows_app_folder_delete(
            _current_app_folder(),
            Path(sys.executable),
        )

    print(
        "Uninstall cleanup complete. Removed Daily Logger user data, add-ons, models, "
        "downloads, and shortcuts."
    )
    if getattr(sys, "frozen", False):
        if scheduled_app_delete:
            print(f"Current Daily Logger app folder will be removed after exit: {_current_app_folder()}")
        else:
            print(f"Could not schedule current app folder removal: {schedule_err}")
            print(f"Current Daily Logger app files were kept: {_current_app_folder()}")
    else:
        print("Source/dev mode detected; repository files were kept.")


def get_start_menu_programs_dir() -> Optional[Path]:
    appdata = os.getenv("APPDATA", "").strip()
    if not appdata:
        return None
    return Path(appdata) / "Microsoft" / "Windows" / "Start Menu" / "Programs"


def create_start_menu_search_shortcut(
    shortcut_path: Path,
    target_path: Path,
    working_directory: Path,
    description: str,
) -> bool:
    """Create a .lnk in Start Menu Programs so Windows Search can surface the target."""
    try:
        shortcut_path.parent.mkdir(parents=True, exist_ok=True)
    except OSError:
        return False
    target = str(target_path.resolve()).replace("'", "''")
    work_dir = str(working_directory.resolve()).replace("'", "''")
    desc = description.replace("'", "''")
    lnk = str(shortcut_path).replace("'", "''")
    ps_script = (
        "$WshShell = New-Object -ComObject WScript.Shell; "
        f"$Shortcut = $WshShell.CreateShortcut('{lnk}'); "
        f"$Shortcut.TargetPath = '{target}'; "
        f"$Shortcut.WorkingDirectory = '{work_dir}'; "
        f"$Shortcut.Description = '{desc}'; "
        "$Shortcut.Save();"
    )
    try:
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
            capture_output=True,
            text=True,
            check=False,
        )
        return result.returncode == 0 and shortcut_path.exists()
    except OSError:
        return False


def sb_create_bat_search_shortcut() -> bool:
    programs = get_start_menu_programs_dir()
    if programs is None:
        return False
    folder = programs / "Daily Logger"
    shortcut_path = folder / "Daily Logger BAT Launcher.lnk"
    bat_path = BASE_DIR / "launch_daily_logger.bat"
    if not bat_path.exists():
        print(f"Missing launcher file: {bat_path}")
        return False
    ok = create_start_menu_search_shortcut(
        shortcut_path,
        bat_path,
        BASE_DIR,
        "Daily Logger batch launcher - search: Daily Logger, BAT, batch, logger",
    )
    if ok:
        print(f"Search shortcut created: {shortcut_path}")
        print("Try Windows search for: Daily Logger, BAT, or batch.")
    return ok


def sb_create_journal_search_shortcut() -> bool:
    programs = get_start_menu_programs_dir()
    if programs is None:
        return False
    journal_path = ensure_workbook(MODULES["J"])
    folder = programs / "Daily Logger"
    shortcut_path = folder / "Daily Logger Journal Excel.lnk"
    ok = create_start_menu_search_shortcut(
        shortcut_path,
        journal_path,
        journal_path.parent,
        "Daily Logger journal workbook - search: Journal, Excel, Daily Logger, xlsx",
    )
    if ok:
        print(f"Search shortcut created: {shortcut_path}")
        print("Try Windows search for: Daily Logger Journal, Excel, or Journal.")
    return ok


def sb_create_reader_search_shortcut() -> bool:
    programs = get_start_menu_programs_dir()
    if programs is None:
        return False
    bat_path = BASE_DIR / "launch_journal_reader.bat"
    if not bat_path.exists():
        print(f"Missing launcher file: {bat_path}")
        return False
    folder = programs / "Daily Logger"
    shortcut_path = folder / "Virtual Journal Reader.lnk"
    ok = create_start_menu_search_shortcut(
        shortcut_path,
        bat_path,
        BASE_DIR,
        "Virtual Journal Reader - search: Virtual Reader, Journal Reader, Daily Logger",
    )
    if ok:
        print(f"Search shortcut created: {shortcut_path}")
        print("Try Windows search for: Virtual Journal Reader.")
    return ok


def load_wifi_warn_list() -> List[str]:
    if not WIFI_WARN_FILE.exists():
        return []
    try:
        parsed = json.loads(WIFI_WARN_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(parsed, list):
        return []
    result: List[str] = []
    seen = set()
    for item in parsed:
        if not isinstance(item, str):
            continue
        cleaned = item.strip()
        key = cleaned.lower()
        if not cleaned or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


def save_wifi_warn_list(names: List[str]) -> bool:
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        WIFI_WARN_FILE.write_text(json.dumps(names, indent=2), encoding="utf-8")
        return True
    except OSError:
        return False


def add_wifi_warn_name(name: str) -> bool:
    cleaned = name.strip()
    if not cleaned:
        return False
    existing = load_wifi_warn_list()
    existing_lower = {item.lower() for item in existing}
    if cleaned.lower() in existing_lower:
        return True
    existing.append(cleaned)
    return save_wifi_warn_list(existing)


def get_current_wifi_name() -> Optional[str]:
    try:
        result = subprocess.run(
            ["netsh", "wlan", "show", "interfaces"],
            capture_output=True,
            text=True,
            check=False,
        )
    except OSError:
        return None
    if result.returncode != 0:
        return None
    for line in result.stdout.splitlines():
        stripped = line.strip()
        if not stripped.startswith("SSID"):
            continue
        if stripped.startswith("SSID BSSID"):
            continue
        if ":" not in stripped:
            continue
        value = stripped.split(":", 1)[1].strip()
        if value:
            return value
    return None


def maybe_warn_for_current_wifi() -> None:
    warned_names = load_wifi_warn_list()
    if not warned_names:
        return
    current_wifi = get_current_wifi_name()
    if not current_wifi:
        return
    warned_lower = {name.lower() for name in warned_names}
    if current_wifi.lower() in warned_lower:
        print(red_text(f'Warning: you are on "{current_wifi}" connection, it might not work.'))


def load_journal_window_draft() -> Optional[Dict[str, object]]:
    if not JOURNAL_WINDOW_DRAFT_FILE.exists():
        return None
    try:
        parsed = json.loads(JOURNAL_WINDOW_DRAFT_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    if not isinstance(parsed, dict):
        return None
    return parsed


def save_journal_window_draft(draft: Dict[str, object]) -> bool:
    try:
        SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        JOURNAL_WINDOW_DRAFT_FILE.write_text(json.dumps(draft, indent=2), encoding="utf-8")
        return True
    except OSError:
        return False


def clear_journal_window_draft() -> None:
    try:
        if JOURNAL_WINDOW_DRAFT_FILE.exists():
            JOURNAL_WINDOW_DRAFT_FILE.unlink()
    except OSError:
        pass


def normalize_window_time_input(raw: str) -> Optional[str]:
    cleaned = raw.strip()
    if not cleaned:
        return datetime.now().strftime("%I:%M%p").lstrip("0")
    if cleaned.lower() in ("n/a", "na"):
        return "N/A"
    if cleaned.lower() == "rn":
        return datetime.now().strftime("%I:%M%p").lstrip("0")
    normalized = cleaned.upper().replace(" ", "")
    try:
        parsed = datetime.strptime(normalized, "%I:%M%p")
        return parsed.strftime("%I:%M%p").lstrip("0")
    except ValueError:
        return None


def _read_wav_mono_int16(path: Path) -> Tuple[Optional[Any], int, Optional[str]]:
    """Load 16-bit PCM WAV as mono int16 ndarray. Returns (samples, sample_rate, error)."""
    try:
        import numpy as np
    except Exception as exc:
        return None, 0, str(exc)
    try:
        with wave.open(str(path), "rb") as wf:
            ch = wf.getnchannels()
            sw = wf.getsampwidth()
            rate = wf.getframerate() or 16000
            nframes = wf.getnframes()
            raw = wf.readframes(nframes)
    except Exception as exc:
        return None, 0, str(exc)
    if sw != 2:
        return None, 0, "Whisper preprocessing needs 16-bit PCM WAV."
    data = np.frombuffer(raw, dtype=np.int16)
    if ch == 1:
        mono = data
    elif ch >= 2:
        flat = data.reshape(-1, ch).astype(np.float32)
        mono = np.mean(flat, axis=1).astype(np.int16)
    else:
        return None, 0, "Invalid WAV channel count."
    return mono, int(rate), None


def _rms_per_frame_int16(samples: Any, frame: int) -> Any:
    import numpy as np

    n = (int(samples.shape[0]) // frame) * frame
    if n <= 0:
        return np.array([], dtype=np.float64)
    blocks = samples[:n].reshape(-1, frame).astype(np.float64)
    return np.sqrt(np.mean(blocks * blocks, axis=1))


def _adaptive_whisper_silence_rms(rms: Any) -> float:
    """Estimate a per-recording silence threshold so noisy rooms do not look like speech."""
    try:
        import numpy as np

        if rms is None or int(rms.size) < 1:
            return float(WHISPER_PRE_SILENCE_RMS)
        noise_floor = float(np.percentile(rms, WHISPER_PRE_NOISE_PERCENTILE))
        if not np.isfinite(noise_floor):
            return float(WHISPER_PRE_SILENCE_RMS)
        return max(float(WHISPER_PRE_SILENCE_RMS), noise_floor * WHISPER_PRE_NOISE_MULTIPLIER)
    except Exception:
        return float(WHISPER_PRE_SILENCE_RMS)


def preprocess_wav_for_whisper(samples: Any, sample_rate: int) -> Tuple[Any, Optional[str]]:
    """Trim edge silence and shorten long internal silences. Returns (mono int16 ndarray, error)."""
    try:
        import numpy as np
    except Exception as exc:
        return None, str(exc)
    if samples is None or int(samples.shape[0]) < 1:
        return None, "Empty audio."
    rate = max(1, int(sample_rate))
    frame = max(int(rate * (WHISPER_PRE_FRAME_MS / 1000.0)), 1)
    rms = _rms_per_frame_int16(samples, frame)
    if rms.size < 1:
        return samples, None
    thr = _adaptive_whisper_silence_rms(rms)
    voiced = rms > thr
    if not bool(np.any(voiced)):
        return samples[:0], "No speech detected (audio is mostly silence)."
    first = int(np.argmax(voiced))
    last = int(rms.shape[0] - 1 - np.argmax(voiced[::-1]))
    pad = int(rate * (WHISPER_PRE_EDGE_PAD_MS / 1000.0))
    min_sp = int(rate * (WHISPER_PRE_MIN_SPEECH_MS / 1000.0))
    start = max(0, first * frame - pad)
    end = min(int(samples.shape[0]), (last + 1) * frame + pad)
    if end - start < min_sp:
        start = max(0, first * frame)
        end = min(int(samples.shape[0]), (last + 1) * frame)
    trimmed = samples[start:end].copy()
    if int(trimmed.shape[0]) < min_sp:
        return trimmed[:0], "No speech detected (audio is mostly silence)."

    rms2 = _rms_per_frame_int16(trimmed, frame)
    if rms2.size < 1:
        return trimmed, None
    v2 = rms2 > thr
    max_gap = int((rate * WHISPER_PRE_MAX_INTERNAL_SILENCE_SEC) / frame)
    keep_gap = max(1, int((rate * WHISPER_PRE_KEEP_INTERNAL_SILENCE_SEC) / frame))
    keep_samples = keep_gap * frame

    out_parts: List[Any] = []
    f = 0
    nfr = int(v2.shape[0])
    while f < nfr:
        while f < nfr and not bool(v2[f]):
            f += 1
        if f >= nfr:
            break
        t = f
        while t < nfr and bool(v2[t]):
            t += 1
        out_parts.append(trimmed[f * frame : t * frame])
        if t >= nfr:
            break
        u = t
        while u < nfr and not bool(v2[u]):
            u += 1
        silence_frames = u - t
        if silence_frames > max_gap:
            out_parts.append(np.zeros(keep_samples, dtype=np.int16))
        else:
            out_parts.append(trimmed[t * frame : u * frame])
        f = u

    if not out_parts:
        return trimmed, None
    merged = np.concatenate(out_parts, axis=0)
    n_sample_full = (int(trimmed.shape[0]) // frame) * frame
    if n_sample_full < int(trimmed.shape[0]):
        merged = np.concatenate([merged, trimmed[n_sample_full:]], axis=0)
    if int(merged.shape[0]) < 1:
        return merged, "No speech detected after preprocessing."
    return merged, None


def prepare_wav_path_for_whisper(source: Path) -> Tuple[Path, Optional[str], Optional[Path]]:
    """Pick WAV bytes to upload: trimmed/collapsed copy when possible.

    Returns (upload_path, fatal_error_string_or_none, temp_path_to_delete_or_none).
    """
    mono, rate, err = _read_wav_mono_int16(source)
    if err is not None or mono is None:
        return source, None, None
    processed, perr = preprocess_wav_for_whisper(mono, rate)
    if perr is not None:
        return source, perr, None
    try:
        import numpy as np
    except Exception:
        return source, None, None
    if processed is None or int(processed.shape[0]) < 1:
        return source, "No speech detected (audio is mostly silence).", None
    if int(processed.shape[0]) == int(mono.shape[0]) and bool(np.array_equal(processed, mono)):
        return source, None, None
    fd, tmp_name = tempfile.mkstemp(suffix=".wav", prefix="whisper_pre_")
    os.close(fd)
    tmp = Path(tmp_name)
    werr = write_mono_int16_wav(tmp, processed, rate)
    if werr is not None:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        return source, None, None
    return tmp, None, tmp


def _media_size_mb(num_bytes: int) -> str:
    return f"{num_bytes / 1024 / 1024:.1f} MB"


def transcription_content_type_for_path(path: Path) -> str:
    return TRANSCRIPTION_CONTENT_TYPES.get(path.suffix.lower(), "application/octet-stream")


def is_transcription_media_file(path: Path) -> bool:
    return path.suffix.lower() in TRANSCRIPTION_MEDIA_SUFFIXES


def is_transcription_video_file(path: Path) -> bool:
    return path.suffix.lower() in TRANSCRIPTION_VIDEO_SUFFIXES


def transcription_path_needs_media_tools(path: Path) -> bool:
    suffix = path.suffix.lower()
    if suffix in TRANSCRIPTION_VIDEO_SUFFIXES or suffix in TRANSCRIPTION_FORCE_CONVERT_SUFFIXES:
        return True
    if suffix not in TRANSCRIPTION_MEDIA_SUFFIXES or suffix == ".wav":
        return False
    try:
        return int(path.stat().st_size) >= TRANSCRIPTION_DIRECT_UPLOAD_MAX_BYTES
    except OSError:
        return False


def any_transcription_path_needs_media_tools(paths: List[Path]) -> bool:
    return any(transcription_path_needs_media_tools(path) for path in paths)


def _find_ffmpeg_executable() -> Optional[str]:
    candidates: List[Optional[str]] = [str(MEDIA_TOOLS_FFMPEG_EXE)]
    if not getattr(sys, "frozen", False):
        try:
            import imageio_ffmpeg

            candidates.append(str(imageio_ffmpeg.get_ffmpeg_exe()))
        except Exception:
            pass
    candidates.extend([
        shutil.which("ffmpeg"),
        str(BASE_DIR / "ffmpeg.exe"),
        str(BASE_DIR / "_internal" / "ffmpeg.exe"),
    ])
    try:
        candidates.append(str(Path(sys.executable).with_name("ffmpeg.exe")))
    except Exception:
        pass
    for candidate in candidates:
        if not candidate:
            continue
        try:
            p = Path(candidate)
            if p.is_file():
                return str(p)
        except OSError:
            continue
    return None


def _temp_media_path(suffix: str, prefix: str) -> Path:
    fd, tmp_name = tempfile.mkstemp(suffix=suffix, prefix=prefix)
    os.close(fd)
    return Path(tmp_name)


def _temp_media_dir(prefix: str) -> Path:
    return Path(tempfile.mkdtemp(prefix=prefix))


def _cleanup_transcription_temp(temp_path: Optional[Path]) -> None:
    if temp_path is None:
        return
    try:
        if temp_path.is_dir():
            shutil.rmtree(temp_path, ignore_errors=True)
        else:
            temp_path.unlink(missing_ok=True)
    except OSError:
        pass


def _run_ffmpeg_extract(source: Path, target: Path, *, copy_audio: bool = False) -> Tuple[bool, str]:
    ffmpeg = _find_ffmpeg_executable()
    if not ffmpeg:
        return False, "Media Tools add-on is not installed. Open Download Manager and install Media Tools first."
    codec_args = (
        ["-c:a", "copy"]
        if copy_audio
        else ["-ac", "1", "-c:a", "aac", "-b:a", TRANSCRIPTION_CONVERTED_AUDIO_BITRATE]
    )
    cmd = [
        ffmpeg,
        "-y",
        "-hide_banner",
        "-loglevel",
        "error",
        "-i",
        str(source),
        "-vn",
        *codec_args,
        str(target),
    ]
    try:
        result = run_hidden_subprocess(cmd, capture_output=True, text=True, timeout=300)
    except Exception as exc:
        return False, str(exc)
    if result.returncode != 0:
        details = (result.stderr or result.stdout or "").strip()
        return False, details or f"ffmpeg exited with code {result.returncode}."
    try:
        if target.is_file() and target.stat().st_size > 0:
            return True, ""
    except OSError:
        pass
    return False, "ffmpeg did not create an audio file."


def _run_ffmpeg_segment_audio(source: Path, target_pattern: Path) -> Tuple[bool, str]:
    ffmpeg = _find_ffmpeg_executable()
    if not ffmpeg:
        return False, "Media Tools add-on is not installed. Open Download Manager and install Media Tools first."
    cmd = [
        ffmpeg,
        "-y",
        "-hide_banner",
        "-loglevel",
        "error",
        "-i",
        str(source),
        "-vn",
        "-map",
        "0:a:0?",
        "-ac",
        "1",
        "-c:a",
        "aac",
        "-b:a",
        TRANSCRIPTION_CONVERTED_AUDIO_BITRATE,
        "-f",
        "segment",
        "-segment_time",
        str(TRANSCRIPTION_AUDIO_CHUNK_SEC),
        "-reset_timestamps",
        "1",
        str(target_pattern),
    ]
    try:
        result = run_hidden_subprocess(cmd, capture_output=True, text=True, timeout=900)
    except Exception as exc:
        return False, str(exc)
    if result.returncode != 0:
        details = (result.stderr or result.stdout or "").strip()
        return False, details or f"ffmpeg exited with code {result.returncode}."
    return True, ""


def _looks_like_no_audio_media_error(text: str) -> bool:
    low = (text or "").casefold()
    markers = (
        "stream map",
        "matches no streams",
        "does not contain any stream",
        "output file does not contain",
        "no audio",
        "audio:0",
    )
    return any(marker in low for marker in markers)


def _extract_media_audio_for_transcription(source: Path) -> Tuple[List[Path], Optional[str], Optional[Path]]:
    temp_dir = _temp_media_dir("transcribe_audio_")
    pattern = temp_dir / "part_%03d.m4a"
    ok, err = _run_ffmpeg_segment_audio(source, pattern)
    if not ok:
        _cleanup_transcription_temp(temp_dir)
        if _looks_like_no_audio_media_error(err):
            return [], f"No audio track found in {source.name}.", None
        return [], f"Could not extract audio from that media file: {err}", None
    parts = sorted(
        (p for p in temp_dir.glob("part_*.m4a") if p.is_file()),
        key=lambda p: p.name.lower(),
    )
    parts = [p for p in parts if p.stat().st_size > 0]
    if not parts:
        _cleanup_transcription_temp(temp_dir)
        return [], f"No audio track found in {source.name}.", None
    oversized = [p for p in parts if p.stat().st_size >= TRANSCRIPTION_DIRECT_UPLOAD_MAX_BYTES]
    if oversized:
        largest = max(int(p.stat().st_size) for p in oversized)
        _cleanup_transcription_temp(temp_dir)
        return (
            [],
            "The selected media is still too large after extracting audio "
            f"({_media_size_mb(largest)} in one chunk). Use a shorter clip or compress it first.",
            None,
        )
    return parts, None, temp_dir


def _copy_mov_as_mp4_for_transcription(source: Path) -> Tuple[Path, Optional[str], Optional[Path]]:
    tmp = _temp_media_path(".mp4", "transcribe_iphone_")
    try:
        shutil.copyfile(source, tmp)
    except OSError as exc:
        try:
            tmp.unlink(missing_ok=True)
        except OSError:
            pass
        return source, f"Could not prepare iPhone video for transcription: {exc}", None
    return tmp, None, tmp


def prepare_paths_for_transcription(source: Path) -> Tuple[List[Path], Optional[str], Optional[Path]]:
    """Return ordered upload files for WAV, audio, and video files."""
    suffix = source.suffix.lower()
    if suffix == ".wav":
        upload, err, temp_path = prepare_wav_path_for_whisper(source)
        if err is not None:
            return [], err, None
        return [upload], None, temp_path
    if not is_transcription_media_file(source):
        allowed = ", ".join(sorted(TRANSCRIPTION_MEDIA_SUFFIXES))
        return [], f"Unsupported transcription file type '{suffix or '(none)'}'. Use: {allowed}", None
    try:
        source_size = int(source.stat().st_size)
    except OSError as exc:
        return [], f"Could not read audio file: {exc}", None
    if source_size <= 0:
        return [], "Empty audio.", None

    if is_transcription_video_file(source):
        return _extract_media_audio_for_transcription(source)
    if suffix in TRANSCRIPTION_FORCE_CONVERT_SUFFIXES:
        return _extract_media_audio_for_transcription(source)
    if source_size >= TRANSCRIPTION_DIRECT_UPLOAD_MAX_BYTES:
        return _extract_media_audio_for_transcription(source)
    return [source], None, None


def prepare_path_for_transcription(source: Path) -> Tuple[Path, Optional[str], Optional[Path]]:
    """Return (upload_path, error, temp_path) for WAV, audio, and supported video files."""
    uploads, err, temp_path = prepare_paths_for_transcription(source)
    if err is not None:
        return source, err, temp_path
    if not uploads:
        return source, "Whisper returned empty text.", temp_path
    return uploads[0], None, temp_path


def _default_whisper_prompt(language: Optional[str]) -> str:
    lang = (language or "").strip().lower()
    if lang == "en":
        lang_hint = "The speech is English."
    elif lang == "zh":
        lang_hint = "The speech is Mandarin Chinese. Use Simplified Chinese characters."
    else:
        lang_hint = (
            "The speech is English, Mandarin Chinese, or a mix of those two. "
            "Use Simplified Chinese characters for Chinese speech."
        )
    return (
        f"{lang_hint} Transcribe verbatim in the spoken language. Do not translate. "
        "Do not output any other language. Do not add subtitles, timestamps, filler, "
        "or repeated phrases. Common short test phrases may include: can you hear me, "
        "test, \u6d4b\u8bd5, and counting numbers. If the audio is silent or unclear, "
        "return empty text."
    )


def _whisper_prompt(language: Optional[str], prompt: Optional[str]) -> str:
    base = _default_whisper_prompt(language)
    extra = (prompt or "").strip()
    if extra:
        base = f"{base}\nContext words and names: {extra}"
    return base[:WHISPER_TRANSCRIBE_PROMPT_CHAR_LIMIT]


def _transcript_repeat_key(text: str) -> str:
    return re.sub(r"[\W_]+", "", (text or "").casefold(), flags=re.UNICODE)


def _strip_repeated_phrase_prefix(text: str, previous: str) -> str:
    prev = previous.strip()
    cur = text.strip()
    if not prev or len(_transcript_repeat_key(prev)) < 2:
        return text
    if not cur.casefold().startswith(prev.casefold()):
        return text
    rest = cur[len(prev) :]
    if not rest or (not rest[:1].isspace() and rest[:1] not in ",\uFF0C\u3001;\uFF1B"):
        return text
    leading = text[: len(text) - len(text.lstrip())]
    return leading + rest.lstrip(" ,\uFF0C\u3001;\uFF1B")


def _collapse_repeated_transcript_clauses(text: str) -> str:
    """Collapse comma-separated repeat spam inside one sentence."""
    sentence_re = re.compile(r"([^.!?\n\r\u3002\uff01\uff1f]+)([.!?\u3002\uff01\uff1f]*)(\s*)", re.UNICODE)
    clause_sep_re = re.compile(r"([,\uFF0C\u3001;\uFF1B]\s*)", re.UNICODE)
    pieces: List[str] = []
    pos = 0
    for sentence_match in sentence_re.finditer(text):
        if sentence_match.start() > pos:
            pieces.append(text[pos : sentence_match.start()])
        body, end, space = sentence_match.groups()
        tokens = clause_sep_re.split(body)
        out: List[str] = []
        last_key = ""
        last_phrase = ""
        last_clause_start = 0
        i = 0
        while i < len(tokens):
            phrase = tokens[i]
            sep = tokens[i + 1] if i + 1 < len(tokens) else ""
            adjusted = _strip_repeated_phrase_prefix(phrase, last_phrase)
            prefix_stripped = adjusted != phrase
            key = _transcript_repeat_key(adjusted)
            if prefix_stripped and key:
                out = out[:last_clause_start]
                last_key = ""
                last_phrase = ""
            if key and key == last_key:
                i += 2
                continue
            clause_start = len(out)
            out.append(adjusted)
            if sep:
                out.append(sep)
            if key:
                last_clause_start = clause_start
                last_key = key
                last_phrase = adjusted
            i += 2
        collapsed = re.sub(r"\s*[,\uFF0C\u3001;\uFF1B]\s*$", "", "".join(out)).strip()
        pieces.append(collapsed + end + space)
        pos = sentence_match.end()
    if pos < len(text):
        pieces.append(text[pos:])
    return "".join(pieces).strip()


def _collapse_repeated_transcript_sentences(text: str) -> str:
    """Collapse consecutive duplicate sentence-like phrases from Whisper hallucinations."""
    if not text.strip():
        return ""
    pieces: List[str] = []
    last_key = ""
    repeat_count = 0
    pos = 0
    sentence_re = re.compile(r"([^.!?\n\r\u3002\uff01\uff1f]+[.!?\u3002\uff01\uff1f]*)(\s*)", re.UNICODE)
    for match in sentence_re.finditer(text):
        if match.start() > pos:
            pieces.append(text[pos : match.start()])
        sentence = match.group(1)
        space = match.group(2)
        key = _transcript_repeat_key(sentence)
        if key and key == last_key:
            repeat_count += 1
            if repeat_count <= WHISPER_REPEAT_SENTENCE_KEEP:
                pieces.append(sentence + space)
        else:
            last_key = key
            repeat_count = 1
            pieces.append(sentence + space)
        pos = match.end()
    if pos < len(text):
        pieces.append(text[pos:])
    return "".join(pieces).strip()


def _is_cjk_letter(ch: str) -> bool:
    return (
        "\u3400" <= ch <= "\u4dbf"
        or "\u4e00" <= ch <= "\u9fff"
        or "\uf900" <= ch <= "\ufaff"
        or "\U00020000" <= ch <= "\U0002ebef"
    )


def _is_latin_letter(ch: str) -> bool:
    return "LATIN" in unicodedata.name(ch, "")


def _cjk_count(text: str) -> int:
    return sum(1 for ch in text or "" if _is_cjk_letter(ch))


def _latin_count(text: str) -> int:
    return sum(1 for ch in text or "" if _is_latin_letter(ch))


def _unsupported_transcript_script_ratio(text: str) -> Tuple[int, float]:
    letters = 0
    unsupported = 0
    for ch in text:
        if not unicodedata.category(ch).startswith("L"):
            continue
        letters += 1
        if not (_is_latin_letter(ch) or _is_cjk_letter(ch)):
            unsupported += 1
    if letters <= 0:
        return 0, 0.0
    return letters, unsupported / float(letters)


JOURNAL_PUNCTUATION_TRANSLATION = str.maketrans(
    {
        "\uff0c": ",",
        "\u3002": ".",
        "\u3001": ",",
        "\uff1b": ";",
        "\uff1a": ":",
        "\uff1f": "?",
        "\uff01": "!",
        "\uff08": "(",
        "\uff09": ")",
        "\u3010": "[",
        "\u3011": "]",
        "\u201c": '"',
        "\u201d": '"',
        "\u2018": "'",
        "\u2019": "'",
    }
)


TRADITIONAL_CHINESE_TRANSLATION = str.maketrans(
    {
        "\u807d": "\u542c",
        "\u898b": "\u89c1",
        "\u55ce": "\u5417",
        "\u8207": "\u4e0e",
        "\u9019": "\u8fd9",
        "\u500b": "\u4e2a",
        "\u6703": "\u4f1a",
        "\u8eca": "\u8f66",
        "\u9304": "\u5f55",
        "\u8a18": "\u8bb0",
        "\u9ede": "\u70b9",
        "\u958b": "\u5f00",
        "\u95dc": "\u5173",
        "\u9580": "\u95e8",
        "\u88e1": "\u91cc",
        "\u9084": "\u8fd8",
        "\u8b93": "\u8ba9",
        "\u8aaa": "\u8bf4",
        "\u8a71": "\u8bdd",
        "\u5f8c": "\u540e",
        "\u6642": "\u65f6",
        "\u9593": "\u95f4",
        "\u7121": "\u65e0",
        "\u767c": "\u53d1",
        "\u8655": "\u5904",
        "\u5099": "\u5907",
        "\u6a94": "\u6863",
        "\u8f49": "\u8f6c",
        "\u5beb": "\u5199",
        "\u8072": "\u58f0",
        "\u8996": "\u89c6",
        "\u8a0a": "\u8baf",
        "\u9ad4": "\u4f53",
        "\u7d71": "\u7edf",
        "\u6a5f": "\u673a",
        "\u4e26": "\u5e76",
        "\u96d9": "\u53cc",
        "\u9801": "\u9875",
        "\u5716": "\u56fe",
        "\u5831": "\u62a5",
        "\u8acb": "\u8bf7",
        "\u8f38": "\u8f93",
        "\u9078": "\u9009",
        "\u64c7": "\u62e9",
        "\u555f": "\u542f",
        "\u52d5": "\u52a8",
        "\u61c9": "\u5e94",
        "\u8a72": "\u8be5",
        "\u984c": "\u9898",
        "\u8abf": "\u8c03",
        "\u8a66": "\u8bd5",
        "\u932f": "\u9519",
        "\u8aa4": "\u8bef",
        "\u96f2": "\u4e91",
        "\u5fa9": "\u590d",
        "\u88fd": "\u5236",
        "\u9ebc": "\u4e48",
        "\u70ba": "\u4e3a",
        "\u5f9e": "\u4ece",
        "\u5c0d": "\u5bf9",
        "\u8b1b": "\u8bb2",
        "\u73fe": "\u73b0",
        "\u5be6": "\u5b9e",
        "\u9a57": "\u9a8c",
        "\u865f": "\u53f7",
        "\u78bc": "\u7801",
        "\u8a08": "\u8ba1",
        "\u5283": "\u5212",
        "\u9577": "\u957f",
        "\u986f": "\u663e",
        "\u7c21": "\u7b80",
        "\u6f22": "\u6c49",
        "\u8a9e": "\u8bed",
        "\u8b80": "\u8bfb",
        "\u66f8": "\u4e66",
        "\u5132": "\u50a8",
        "\u8a2d": "\u8bbe",
        "\u96fb": "\u7535",
        "\u8166": "\u8111",
        "\u7db2": "\u7f51",
        "\u7d61": "\u7edc",
        "\u9023": "\u8fde",
        "\u6e2c": "\u6d4b",
        "\u6578": "\u6570",
        "\u64da": "\u636e",
        "\u5eab": "\u5e93",
        "\u96e2": "\u79bb",
        "\u7dda": "\u7ebf",
        "\u8cc7": "\u8d44",
        "\u54e1": "\u5458",
        "\u5b78": "\u5b66",
        "\u7fd2": "\u4e60",
        "\u5be9": "\u5ba1",
        "\u7522": "\u4ea7",
        "\u696d": "\u4e1a",
        "\u52d9": "\u52a1",
        "\u9810": "\u9884",
        "\u89bd": "\u89c8",
        "\u58d3": "\u538b",
        "\u7e2e": "\u7f29",
        "\u91cb": "\u91ca",
        "\u7a2e": "\u79cd",
        "\u985e": "\u7c7b",
        "\u5c0e": "\u5bfc",
        "\u6a19": "\u6807",
        "\u7c64": "\u7b7e",
        "\u7c3d": "\u7b7e",
        "\u7e6a": "\u7ed8",
        "\u756b": "\u753b",
        "\u984f": "\u989c",
        "\u64ca": "\u51fb",
        "\u522a": "\u5220",
        "\u6aa2": "\u68c0",
        "\u7576": "\u5f53",
        "\u7e8c": "\u7eed",
        "\u65b7": "\u65ad",
        "\u554f": "\u95ee",
    }
)


def simplify_chinese_text(text: str) -> str:
    return (text or "").translate(TRADITIONAL_CHINESE_TRANSLATION)


def normalize_common_short_chinese_transcription_misses(text: str) -> str:
    fixed = simplify_chinese_text(text or "")
    if not fixed:
        return ""
    fixed = re.sub(
        r"(?i)\bwei\s+wei\s+wei\b[,\s]*(?=\u542c\u5f97\u89c1\u5417)",
        "\u5582\u5582\u5582,",
        fixed,
    )
    fixed = re.sub(
        r"(?i)\bwei\s+(?:lu\s*)?wei\b[,\s]*(?=\u542c\u5f97\u89c1\u5417)",
        "\u5582\u5582\u5582,",
        fixed,
    )
    return fixed


def normalize_journal_text_punctuation(text: str) -> str:
    """Use baseline-friendly ASCII punctuation in mixed Chinese/English journal text."""
    if not text:
        return ""
    fixed = normalize_common_short_chinese_transcription_misses(text).translate(JOURNAL_PUNCTUATION_TRANSLATION)
    fixed = re.sub(r"\s+([,.;:?!])", r"\1", fixed)
    return fixed


def clean_whisper_transcript(text: str, language: Optional[str]) -> str:
    """Normalize Whisper text and reject obvious non-English/non-Chinese hallucinations."""
    cleaned = re.sub(r"\s+", " ", (text or "").strip())
    if not cleaned:
        return ""
    cleaned = _collapse_repeated_transcript_clauses(cleaned)
    cleaned = _collapse_repeated_transcript_sentences(cleaned)
    cleaned = normalize_journal_text_punctuation(cleaned)
    letters, unsupported_ratio = _unsupported_transcript_script_ratio(cleaned)
    if (
        letters >= WHISPER_UNSUPPORTED_SCRIPT_MIN_LETTERS
        and unsupported_ratio > WHISPER_UNSUPPORTED_SCRIPT_RATIO
    ):
        return (
            "Whisper transcription rejected: the result looked like a language outside "
            "English/Chinese. Try again, or choose English/Chinese explicitly."
        )
    return cleaned


def _is_whisper_rejection_message(text: str) -> bool:
    return (text or "").strip().startswith("Whisper transcription rejected:")


def _looks_like_repeat_noise(text: str) -> bool:
    compact = " ".join((text or "").split()).strip()
    if len(compact) < 16:
        return False
    letters_only = [ch.casefold() for ch in compact if unicodedata.category(ch).startswith("L")]
    if len(letters_only) >= 8 and len(set(letters_only)) <= 2:
        return True
    words = re.findall(r"[A-Za-z\u3400-\u9fff]+", compact.casefold())
    if len(words) >= 8:
        return len(set(words)) <= max(2, len(words) // 5)
    return False


def _looks_like_possible_chinese_translation(text: str) -> bool:
    cleaned = (text or "").strip().casefold()
    if not cleaned or _cjk_count(cleaned) > 0:
        return False
    latin = _latin_count(cleaned)
    if latin <= 0:
        return True
    if latin > 240:
        return False
    markers = (
        "can you hear me",
        "do you hear me",
        "are you there",
        "hello hello",
        "hello, hello",
        "testing",
        "test test",
        "i can hear",
        "can hear you",
        "what should",
        "how should",
        "what do we do",
        "how do we",
        "let me try",
        "start recording",
        "continue",
    )
    if any(marker in cleaned for marker in markers):
        return True
    short_words = re.findall(r"[a-z]+", cleaned)
    if 0 < len(short_words) <= 5 and any(word in {"hello", "testing", "test"} for word in short_words):
        return True
    return False


def _should_use_cloud_chinese_retry(first_text: str, fallback_text: str) -> bool:
    first = normalize_journal_text_punctuation(first_text or "").strip()
    fallback = normalize_journal_text_punctuation(fallback_text or "").strip()
    if not _looks_like_possible_chinese_translation(first):
        return False
    cjk = _cjk_count(fallback)
    if cjk < 2:
        return False
    if _looks_like_repeat_noise(fallback):
        return False
    fallback_latin = _latin_count(fallback)
    first_latin = max(1, _latin_count(first))
    if cjk >= 4 and fallback_latin <= max(24, cjk * 3):
        return True
    return cjk >= max(2, first_latin // 8)


def _should_try_cloud_chinese_retry(language: Optional[str], first_text: str) -> bool:
    return (
        language is None
        and not _is_likely_api_error_message_global(first_text)
        and _looks_like_possible_chinese_translation(first_text)
    )


def _transcribe_audio_openai_single_with_retry(
    file_path: Path,
    language: Optional[str],
    *,
    model_name: Optional[str] = None,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    content_type: Optional[str] = None,
    progress: Optional[Callable[[int], None]] = None,
) -> str:
    result = _transcribe_audio_openai_single(
        file_path,
        language,
        model_name=model_name,
        prompt=prompt,
        temperature=temperature,
        content_type=content_type,
        progress=progress,
    )
    if not _should_try_cloud_chinese_retry(language, result):
        return result

    retry_result = _transcribe_audio_openai_single(
        file_path,
        "zh",
        model_name=model_name,
        prompt=prompt,
        temperature=temperature,
        content_type=content_type,
        progress=None,
    )
    if (
        not _is_likely_api_error_message_global(retry_result)
        and _should_use_cloud_chinese_retry(result, retry_result)
    ):
        return retry_result
    return result




def _transcribe_audio_openai_single(
    file_path: Path,
    language: Optional[str],
    *,
    model_name: Optional[str] = None,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    content_type: Optional[str] = None,
    progress: Optional[Callable[[int], None]] = None,
) -> str:
    """Single-request Whisper upload. Caller handles retries/fallback strategy."""

    def _pg(p: int) -> None:
        if progress is not None:
            try:
                progress(min(100, max(0, int(p))))
            except Exception:
                pass

    _pg(10)
    boundary = uuid.uuid4().hex.encode("ascii")
    crlf = b"\r\n"
    body_chunks: List[bytes] = []

    def add_field(name: str, value: str) -> None:
        body_chunks.append(b"--" + boundary + crlf)
        body_chunks.append(
            f'Content-Disposition: form-data; name="{name}"'.encode("utf-8") + crlf + crlf
        )
        body_chunks.append(value.encode("utf-8") + crlf)

    api_key = get_openai_api_key()
    if not api_key:
        return "OPENAI_API_KEY is not set. Use TOKEN ADD in the main menu or set the environment variable."

    _pg(14)
    selected_cloud_model = (model_name or TRANSCRIPTION_DEFAULT_CLOUD_MODEL).strip() or TRANSCRIPTION_DEFAULT_CLOUD_MODEL
    add_field("model", selected_cloud_model)
    if language:
        add_field("language", language)
    is_diarize_model = selected_cloud_model == "gpt-4o-transcribe-diarize"
    if is_diarize_model:
        add_field("response_format", "diarized_json")
        add_field("chunking_strategy", "auto")
    else:
        prompt_text = _whisper_prompt(language, prompt)
        if prompt_text:
            add_field("prompt", prompt_text)
        add_field("temperature", str(temperature))

    _pg(22)
    filename = file_path.name
    try:
        audio_bytes = file_path.read_bytes()
    except OSError as exc:
        return f"Could not read audio file: {exc}"

    _pg(30)
    body_chunks.append(b"--" + boundary + crlf)
    body_chunks.append(
        f'Content-Disposition: form-data; name="file"; filename="{filename}"'.encode("utf-8")
        + crlf
    )
    upload_content_type = content_type or transcription_content_type_for_path(file_path)
    body_chunks.append(f"Content-Type: {upload_content_type}".encode("ascii", "ignore") + crlf + crlf)
    body_chunks.append(audio_bytes + crlf)
    body_chunks.append(b"--" + boundary + b"--" + crlf)
    body = b"".join(body_chunks)

    _pg(38)
    req = request.Request(
        OPENAI_TRANSCRIPTION_URL,
        data=body,
        method="POST",
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": f"multipart/form-data; boundary={boundary.decode('ascii')}",
        },
    )

    try:
        with request.urlopen(req, timeout=120) as response:
            raw_bytes = response.read()
        _pg(76)
        raw = raw_bytes.decode("utf-8")
    except error.HTTPError as exc:
        _pg(50)
        try:
            details = exc.read().decode("utf-8")
        except Exception:
            details = str(exc)
        result = f"Whisper API error ({exc.code}): {details}"
    except Exception as exc:
        _pg(50)
        result = f"Whisper request failed: {exc}"
    else:
        try:
            parsed = json.loads(raw)
            text = parsed.get("text")
            if is_diarize_model and isinstance(parsed.get("segments"), list):
                segment_texts: List[str] = []
                for segment in parsed.get("segments", []):
                    if not isinstance(segment, dict):
                        continue
                    speaker = str(segment.get("speaker") or "").strip()
                    part = clean_whisper_transcript(str(segment.get("text") or ""), language)
                    if not part:
                        continue
                    segment_texts.append(f"{speaker}: {part}" if speaker else part)
                result = normalize_journal_text_punctuation(" ".join(segment_texts).strip())
                _pg(94)
            elif isinstance(text, str):
                result = clean_whisper_transcript(text, language)
                _pg(94)
            else:
                result = "Whisper returned an unexpected response format."
                _pg(88)
        except json.JSONDecodeError:
            result = "Whisper returned invalid JSON."
            _pg(88)

    return result


def _whisper_context_too_long_error(text: str) -> bool:
    needle = text.strip().lower()
    if not needle:
        return False
    markers = (
        "maximum context length",
        "context length exceeded",
        "prompt is too long",
        "too many tokens",
        "reduce the length",
        "request too large",
        "payload too large",
        "content size limit",
        "maximum content size",
        "26214400",
        "413",
        "entity too large",
    )
    return any(m in needle for m in markers)


def _is_likely_api_error_message_global(text: str) -> bool:
    """Module-level variant used by non-UI helpers (UI also defines its own)."""
    t = (text or "").strip()
    if not t:
        return False
    prefixes = (
        "OPENAI_API_KEY",
        "ChatGPT API error",
        "Failed to contact ChatGPT",
        "ChatGPT returned",
        "No response received",
        "Whisper API error",
        "Whisper request failed",
        "Whisper returned",
        "Whisper transcription rejected",
        "Local transcription failed",
        "Local transcription add-on",
        "Local transcription model",
        "Could not load local transcription add-on",
        "Could not load local transcription model",
        "Media Tools add-on",
        "Unsupported transcription file type",
        "That media file is too large",
        "That iPhone video is too large",
        "The selected media is still too large",
        "Could not prepare iPhone video",
        "Could not extract audio",
        "Could not read audio file",
        "No audio track found",
        "Recording needs optional packages",
        "No speech detected",
        "Empty audio.",
    )
    return any(t.startswith(p) for p in prefixes)


def whisper_chunk_duration_sec(sample_rate: int) -> int:
    """Seconds of mono int16 audio per chunk so WAV uploads stay under Whisper size limits."""
    rate = max(1, int(sample_rate))
    pcm_bps = 2 * rate
    max_sec_budget = int(WHISPER_SAFE_CHUNK_PCM_BYTES // pcm_bps)
    return max(45, min(WHISPER_TRANSCRIBE_CHUNK_SEC, max_sec_budget))


def _transcribe_audio_openai_chunked(
    file_path: Path,
    language: Optional[str],
    *,
    model_name: Optional[str] = None,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    on_part: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[int], None]] = None,
) -> str:
    """Fallback for oversized uploads/context: split WAV and merge partial transcripts."""

    def _pg(p: int) -> None:
        if progress is not None:
            try:
                progress(min(100, max(0, int(p))))
            except Exception:
                pass

    _pg(12)
    mono, rate, read_err = _read_wav_mono_int16(file_path)
    if read_err is not None or mono is None:
        return _transcribe_audio_openai_single_with_retry(
            file_path,
            language,
            model_name=model_name,
            prompt=prompt,
            temperature=temperature,
            progress=progress,
        )
    chunk_sec = whisper_chunk_duration_sec(int(rate))
    chunk_samples = max(int(rate * chunk_sec), 1)
    if int(mono.shape[0]) <= chunk_samples:
        return _transcribe_audio_openai_single_with_retry(
            file_path,
            language,
            model_name=model_name,
            prompt=prompt,
            temperature=temperature,
            progress=progress,
        )

    transcripts: List[str] = []
    sample_count = int(mono.shape[0])
    chunk_starts = list(range(0, sample_count, chunk_samples))
    n_chunks = max(len(chunk_starts), 1)
    _pg(18)
    for ci, start in enumerate(chunk_starts):
        end = min(start + chunk_samples, sample_count)
        part = mono[start:end]
        if int(part.shape[0]) < 1:
            continue
        fd, tmp_name = tempfile.mkstemp(suffix=".wav", prefix="whisper_chunk_")
        os.close(fd)
        tmp = Path(tmp_name)
        try:
            werr = write_mono_int16_wav(tmp, part, rate)
            if werr is not None:
                return f"Could not write chunked audio: {werr}"
            chunk_result = _transcribe_audio_openai_single_with_retry(
                tmp,
                language,
                model_name=model_name,
                prompt=prompt,
                temperature=temperature,
                progress=None,
            ).strip()
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass
        if _is_likely_api_error_message_global(chunk_result):
            return chunk_result
        if chunk_result:
            if transcripts and _transcript_repeat_key(chunk_result) == _transcript_repeat_key(transcripts[-1]):
                _pg(22 + int(72 * (ci + 1) / n_chunks))
                continue
            transcripts.append(chunk_result)
            if on_part is not None:
                try:
                    on_part(chunk_result)
                except Exception:
                    pass
        _pg(22 + int(72 * (ci + 1) / n_chunks))
    merged = clean_whisper_transcript(" ".join(t for t in transcripts if t.strip()).strip(), language)
    _pg(97)
    if merged:
        return merged
    return "Whisper returned empty text."


def _transcribe_prepared_upload_openai(
    upload_path: Path,
    language: Optional[str],
    *,
    model_name: Optional[str] = None,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    on_part: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[int], None]] = None,
) -> str:
    upload_is_wav = upload_path.suffix.lower() == ".wav"
    upload_content_type = transcription_content_type_for_path(upload_path)
    upl_sz = 0
    try:
        upl_sz = int(upload_path.stat().st_size)
    except OSError:
        pass
    if upload_is_wav and upl_sz >= WHISPER_SKIP_SINGLE_FILE_BYTES:
        return _transcribe_audio_openai_chunked(
            upload_path,
            language,
            model_name=model_name,
            prompt=prompt,
            temperature=temperature,
            on_part=on_part,
            progress=progress,
        )
    if not upload_is_wav and upl_sz >= TRANSCRIPTION_DIRECT_UPLOAD_MAX_BYTES:
        return (
            "That media file is too large for one transcription upload "
            f"({_media_size_mb(upl_sz)}). Use a shorter clip or compress it first."
        )
    result = _transcribe_audio_openai_single_with_retry(
        upload_path,
        language,
        model_name=model_name,
        prompt=prompt,
        temperature=temperature,
        content_type=upload_content_type,
        progress=progress,
    )
    if upload_is_wav and _whisper_context_too_long_error(result):
        return _transcribe_audio_openai_chunked(
            upload_path,
            language,
            model_name=model_name,
            prompt=prompt,
            temperature=temperature,
            on_part=on_part,
            progress=progress,
        )
    if on_part is not None and not _is_likely_api_error_message_global(result):
        final_text = normalize_journal_text_punctuation(result.strip())
        if final_text:
            try:
                on_part(final_text)
            except Exception:
                pass
    return result


_LOCAL_TRANSCRIPTION_MODEL_CACHE: Dict[str, Any] = {}
_LOCAL_TRANSCRIPTION_MODEL_CACHE_LOCK = threading.Lock()


def _get_local_transcription_model(model_name: str) -> Tuple[Optional[Any], Optional[str]]:
    return None, "Local transcription models are loaded by the helper add-on."


def _clear_local_transcription_model_cache(model_name: Optional[str] = None) -> None:
    with _LOCAL_TRANSCRIPTION_MODEL_CACHE_LOCK:
        if not model_name:
            _LOCAL_TRANSCRIPTION_MODEL_CACHE.clear()
            return
        prefix = f"{model_name}:"
        for key in list(_LOCAL_TRANSCRIPTION_MODEL_CACHE.keys()):
            if key.startswith(prefix):
                _LOCAL_TRANSCRIPTION_MODEL_CACHE.pop(key, None)


def _transcribe_prepared_upload_local(
    upload_path: Path,
    model_name: str,
    language: Optional[str],
    *,
    prompt: Optional[str] = None,
    on_part: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[int], None]] = None,
) -> str:
    def _pg(p: int) -> None:
        if progress is not None:
            try:
                progress(min(100, max(0, int(p))))
            except Exception:
                pass

    _pg(8)
    runtime_ok, runtime_err = ensure_local_transcription_runtime_loaded()
    if not runtime_ok:
        return runtime_err
    transcripts: List[str] = []

    def _event(event: Dict[str, Any]) -> None:
        name = str(event.get("event") or "")
        if "percent" in event:
            try:
                _pg(int(float(event.get("percent") or 0)))
            except Exception:
                pass
        if name == "segment":
            text = clean_whisper_transcript(str(event.get("text") or ""), language).strip()
            if not text or _is_whisper_rejection_message(text):
                return
            if transcripts and _transcript_repeat_key(text) == _transcript_repeat_key(transcripts[-1]):
                return
            transcripts.append(text)
            if on_part is not None:
                try:
                    on_part(text)
                except Exception:
                    pass

    helper_prompt = _whisper_prompt(language, prompt) if language or (prompt or "").strip() else ""
    helper_args = [
        "transcribe",
        "--model",
        model_name,
        "--input",
        str(upload_path),
        "--models-dir",
        str(LOCAL_TRANSCRIPTION_MODEL_DIR),
        "--language",
        language or "auto",
        "--prompt",
        helper_prompt,
        "--compute-type",
        TRANSCRIPTION_LOCAL_COMPUTE_TYPE,
        "--cpu-threads",
        str(TRANSCRIPTION_LOCAL_CPU_THREADS),
    ]
    ok, err_msg, last = _run_local_transcriber_json(
        helper_args,
        on_event=_event,
        timeout=None,
    )
    if not ok:
        return err_msg
    final_text = str(last.get("text") or "").strip() if last.get("event") == "complete" else ""
    if final_text:
        merged = clean_whisper_transcript(final_text, language)
        if _is_whisper_rejection_message(merged) and transcripts:
            merged = clean_whisper_transcript(" ".join(t for t in transcripts if t.strip()).strip(), language)
    else:
        merged = clean_whisper_transcript(" ".join(t for t in transcripts if t.strip()).strip(), language)
    _pg(98)
    return merged or "Whisper returned empty text."


def transcribe_audio_with_model(
    file_path: Path,
    language: Optional[str],
    model_choice: str,
    *,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    on_part: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[int], None]] = None,
    status: Optional[Callable[[str], None]] = None,
) -> str:
    choice = normalize_transcription_model_choice(model_choice)
    if transcription_model_is_cloud(choice):
        return transcribe_audio_openai(
            file_path,
            language,
            model_name=transcription_cloud_model_name(choice),
            prompt=prompt,
            temperature=temperature,
            on_part=on_part,
            progress=progress,
            status=status,
        )

    model_name = transcription_local_model_name(choice)
    last_pct = [0]

    def _p(pct: int) -> None:
        v = min(100, max(0, int(pct)))
        if v < last_pct[0]:
            v = last_pct[0]
        else:
            last_pct[0] = v
        if progress is not None:
            try:
                progress(v)
            except Exception:
                pass

    def _status(text: str) -> None:
        if status is not None:
            try:
                status(text)
            except Exception:
                pass

    _p(2)
    _status("Preparing audio for local transcription...")
    if file_path.suffix.lower() == ".wav":
        upload_paths, temp_upload = [file_path], None
    else:
        upload_paths, prep_err, temp_upload = prepare_paths_for_transcription(file_path)
        if prep_err is not None:
            return prep_err
    if not upload_paths:
        return "Whisper returned empty text."
    _p(6)
    try:
        transcripts: List[str] = []
        n_uploads = max(len(upload_paths), 1)
        if n_uploads > 1:
            _status(f"Prepared {n_uploads} audio parts. Transcribing part 1/{n_uploads}...")
        for idx, upload_path in enumerate(upload_paths):
            def _part_progress(pct: int, _idx: int = idx) -> None:
                base = 6 + int(92 * (_idx / n_uploads))
                span = max(1, int(92 / n_uploads))
                _p(base + int(span * min(100, max(0, int(pct))) / 100))

            _status(f"Transcribing audio part {idx + 1}/{n_uploads}...")
            result = _transcribe_prepared_upload_local(
                upload_path,
                model_name,
                language,
                prompt=prompt,
                on_part=on_part,
                progress=_part_progress,
            )
            if _is_likely_api_error_message_global(result):
                return result
            if result.strip():
                transcripts.append(result.strip())
            _status(f"Finished audio part {idx + 1}/{n_uploads}.")
        _p(99)
        merged = clean_whisper_transcript(" ".join(transcripts).strip(), language)
        _status("Transcription complete.")
        return merged or "Whisper returned empty text."
    finally:
        if temp_upload is not None:
            _cleanup_transcription_temp(temp_upload)


def transcribe_audio_openai(
    file_path: Path,
    language: Optional[str],
    *,
    model_name: Optional[str] = None,
    prompt: Optional[str] = None,
    temperature: float = 0.0,
    on_part: Optional[Callable[[str], None]] = None,
    progress: Optional[Callable[[int], None]] = None,
    status: Optional[Callable[[str], None]] = None,
) -> str:
    """Send local audio to Whisper with fallback for long context/uploads."""
    last_pct = [0]

    def _p(pct: int) -> None:
        v = min(100, max(0, int(pct)))
        if v < last_pct[0]:
            v = last_pct[0]
        else:
            last_pct[0] = v
        if progress is not None:
            try:
                progress(v)
            except Exception:
                pass

    def _status(text: str) -> None:
        if status is not None:
            try:
                status(text)
            except Exception:
                pass

    _p(2)
    _status("Preparing audio for transcription...")
    upload_paths, prep_err, temp_upload = prepare_paths_for_transcription(file_path)
    if prep_err is not None:
        return prep_err
    if not upload_paths:
        return "Whisper returned empty text."
    _p(6)
    try:
        transcripts: List[str] = []
        n_uploads = max(len(upload_paths), 1)
        if n_uploads > 1:
            _status(f"Prepared {n_uploads} audio parts. Transcribing part 1/{n_uploads}...")
        for idx, upload_path in enumerate(upload_paths):
            def _part_progress(pct: int, _idx: int = idx) -> None:
                base = 6 + int(92 * (_idx / n_uploads))
                span = max(1, int(92 / n_uploads))
                _p(base + int(span * min(100, max(0, int(pct))) / 100))

            _status(f"Transcribing audio part {idx + 1}/{n_uploads}...")
            result = _transcribe_prepared_upload_openai(
                upload_path,
                language,
                model_name=model_name,
                prompt=prompt,
                temperature=temperature,
                on_part=on_part,
                progress=_part_progress,
            )
            if _is_likely_api_error_message_global(result):
                return result
            if result.strip():
                transcripts.append(result.strip())
            _status(f"Finished audio part {idx + 1}/{n_uploads}.")
        _p(99)
        merged = clean_whisper_transcript(" ".join(transcripts).strip(), language)
        _status("Transcription complete.")
        return merged or "Whisper returned empty text."
    finally:
        if temp_upload is not None:
            _cleanup_transcription_temp(temp_upload)


def archive_journal_recording(wav_path: Path) -> Optional[Path]:
    """Copy a session WAV into RECORDING_DIR.

    Files are named rcdYYYYMMDD.wav, then rcdYYYYMMDD1.wav, rcdYYYYMMDD2.wav, ... for the same day.
    """
    try:
        RECORDING_DIR.mkdir(parents=True, exist_ok=True)
        day = datetime.now().strftime("%Y%m%d")
        base = f"rcd{day}"
        for n in range(0, 10000):
            name = f"{base}.wav" if n == 0 else f"{base}{n}.wav"
            dest = RECORDING_DIR / name
            if dest.exists():
                continue
            shutil.copy2(wav_path, dest)
            return dest.resolve()
    except OSError:
        return None


def latest_archived_journal_wav() -> Optional[Path]:
    """Newest journal clip in ``RECORDING_DIR`` (``rcd*.wav`` by modification time), or ``None``."""
    try:
        if not RECORDING_DIR.is_dir():
            return None
        best_mtime: float = -1.0
        best_path: Optional[Path] = None
        for p in RECORDING_DIR.iterdir():
            if not p.is_file() or p.suffix.lower() != ".wav":
                continue
            if not p.stem.lower().startswith("rcd"):
                continue
            try:
                mtime = float(p.stat().st_mtime)
            except OSError:
                continue
            if mtime > best_mtime:
                best_mtime = mtime
                best_path = p.resolve()
        return best_path
    except OSError:
        return None


def wav_mono_duration_seconds(path: Path) -> float:
    """Return duration in seconds for a readable mono WAV, or 0.0 on error."""
    try:
        with wave.open(str(path), "rb") as wf:
            rate = wf.getframerate() or 16000
            return wf.getnframes() / float(rate)
    except Exception:
        return 0.0


def estimate_whisper_cost_usd(wav_path: Path) -> Tuple[float, float]:
    """Return (duration_sec, approximate_usd) using WHISPER_USD_PER_MIN."""
    dur = wav_mono_duration_seconds(wav_path)
    usd = (dur / 60.0) * WHISPER_USD_PER_MIN if dur > 0 else 0.0
    return dur, usd


def bind_hover_tooltip(widget: Any, text_callable: Callable[[], str]) -> None:
    """Show a tooltip only for this widget; place inside its toplevel, hugging edges when clipped."""
    if tk is None:
        return
    tip: Dict[str, Optional[Any]] = {"w": None}

    def hide(_evt: Optional[Any] = None) -> None:
        tw = tip["w"]
        if tw is not None:
            try:
                tw.destroy()
            except tk.TclError:
                pass
            tip["w"] = None

    def show(evt: Any) -> None:
        hide()
        msg = (text_callable() or "").strip()
        if not msg:
            return
        tw = tk.Toplevel(widget)
        tip["w"] = tw
        tw.wm_overrideredirect(True)
        tw.wm_attributes("-topmost", True)
        lbl = tk.Label(
            tw,
            text=msg,
            justify="left",
            background="#ffffe0",
            foreground="#000000",
            relief="solid",
            borderwidth=1,
            font=("Segoe UI", 9),
            wraplength=TOOLTIP_WRAP_PX,
        )
        lbl.pack(ipadx=4, ipady=2)
        m = 8
        try:
            top = widget.winfo_toplevel()
            top.update_idletasks()
            win_x = int(top.winfo_rootx())
            win_y = int(top.winfo_rooty())
            win_w = max(int(top.winfo_width()), 160)
            win_h = max(int(top.winfo_height()), 120)
        except tk.TclError:
            win_x, win_y = 0, 0
            win_w, win_h = 800, 600
        win_r = win_x + win_w
        win_b = win_y + win_h
        cx = int(evt.x_root)
        cy = int(evt.y_root)
        pref_x = cx + 12
        pref_y = cy + 12
        space_right = max(0, win_r - m - pref_x)
        space_left = max(0, pref_x - win_x - m)
        if space_right >= space_left:
            wrap = max(100, min(TOOLTIP_WRAP_PX_MAX, space_right - 8))
        else:
            wrap = max(100, min(TOOLTIP_WRAP_PX_MAX, space_left - 8))
        max_inner = max(100, win_w - 2 * m - 16)
        lbl.config(wraplength=min(wrap, max_inner, TOOLTIP_WRAP_PX_MAX))
        tw.update_idletasks()
        tip_w = int(tw.winfo_reqwidth())
        tip_h = int(tw.winfo_reqheight())
        if tip_w > win_w - 2 * m:
            lbl.config(wraplength=max_inner)
            tw.update_idletasks()
            tip_w = int(tw.winfo_reqwidth())
            tip_h = int(tw.winfo_reqheight())
        x = pref_x
        if x + tip_w > win_r - m:
            x = win_r - m - tip_w
        if x < win_x + m:
            x = win_x + m
        y = pref_y
        if y + tip_h > win_b - m:
            y = win_b - m - tip_h
        if y < win_y + m:
            y = win_y + m
        x = max(win_x + m, min(x, win_r - tip_w - m))
        y = max(win_y + m, min(y, win_b - tip_h - m))
        tw.wm_geometry(f"+{x}+{y}")

    widget.bind("<Enter>", show, add="+")
    widget.bind("<Leave>", hide, add="+")
    widget.bind("<ButtonPress>", hide, add="+")


def bind_button_hover_if_enabled(
    widget: Any,
    get_rest_style: Callable[[], Tuple[str, str, str, str, str]],
    hover_bg: Union[str, Callable[[], str]],
    hover_fg: Union[str, Callable[[], str]],
) -> None:
    """Apply hover colors on <Enter> only when state is normal; <Leave> restores idle look.

    get_rest_style returns (state, bg, fg, activebackground, activeforeground) for the
    non-hover appearance; state should match widget.cget('state') logic for that moment.
    hover_bg / hover_fg may be callables (e.g. lambda: theme.hover_primary) so themes can
    change without rebinding.
    """
    if tk is None:
        return

    def _hover_color(spec: Union[str, Callable[[], str]]) -> str:
        return spec() if callable(spec) else spec

    def on_leave(_evt: Optional[Any] = None) -> None:
        try:
            st, bg, fg, abg, afg = get_rest_style()
        except tk.TclError:
            return
        kw: Dict[str, Any] = {
            "bg": bg,
            "fg": fg,
            "activebackground": abg,
            "activeforeground": afg,
        }
        if str(st) == "disabled":
            kw["disabledforeground"] = fg
        try:
            widget.config(**kw)
        except tk.TclError:
            pass

    def on_enter(_evt: Optional[Any] = None) -> None:
        try:
            st, _b, _f, _ab, _af = get_rest_style()
        except tk.TclError:
            return
        if str(st) != "normal":
            return
        hb = _hover_color(hover_bg)
        hf = _hover_color(hover_fg)
        try:
            widget.config(
                bg=hb,
                fg=hf,
                activebackground=hb,
                activeforeground=hf,
            )
        except tk.TclError:
            pass

    widget.bind("<Enter>", on_enter, add="+")
    widget.bind("<Leave>", on_leave, add="+")


def write_mono_int16_wav(path: Path, samples: object, sample_rate: int) -> Optional[str]:
    """Write mono int16 PCM to WAV. Returns error string or None on success."""
    try:
        import numpy as np
    except Exception as exc:
        return str(exc)
    if not isinstance(samples, np.ndarray):
        return "Internal error: audio must be a numpy array."
    arr = np.atleast_1d(samples.squeeze())
    if arr.dtype != np.int16:
        arr = arr.astype(np.int16)
    if arr.size == 0:
        return "Empty audio buffer."
    try:
        with wave.open(str(path), "wb") as wf:
            wf.setnchannels(1)
            wf.setsampwidth(2)
            wf.setframerate(sample_rate)
            wf.writeframes(arr.tobytes())
    except OSError as exc:
        return str(exc)
    return None


def _put_latest_audio_block(audio_queue: "queue.Queue[Any]", block: Any) -> None:
    try:
        audio_queue.put_nowait(block)
        return
    except queue.Full:
        pass
    try:
        audio_queue.get_nowait()
    except queue.Empty:
        pass
    try:
        audio_queue.put_nowait(block)
    except queue.Full:
        pass


def _drain_latest_audio_block(audio_queue: "queue.Queue[Any]") -> Optional[Any]:
    latest: Optional[Any] = None
    while True:
        try:
            latest = audio_queue.get_nowait()
        except queue.Empty:
            return latest


def record_sources_session_wav(
    output_path: Path,
    stop_event: threading.Event,
    *,
    source_enabled_events: Dict[str, threading.Event],
    sample_rate: int = 16000,
    chunk_interval_sec: float = LIVE_STT_CHUNK_INTERVAL_SEC,
    on_audio_chunk: Optional[Callable[[Path], None]] = None,
    on_pcm_block: Optional[Callable[[Any], None]] = None,
    pause_event: Optional[threading.Event] = None,
    on_source_error: Optional[Callable[[str, str], None]] = None,
) -> Optional[str]:
    """Record selected sources until stop_event and save one mixed mono WAV.

    Optional on_audio_chunk(path): periodic temp WAV paths for live STT (legacy).
    Optional on_pcm_block(block): each mixed block as int16 numpy array (mono); runs in the record thread.
    Optional pause_event: while set, input is still read (to avoid device overrun) but not written to the
    output buffer and on_pcm_block is not called so metering/waveform can stay frozen until resumed.
    """
    try:
        import numpy as np
        import sounddevice as sd
    except Exception as exc:
        return (
            "Recording needs optional packages. Install with:\n"
            f"  {sys.executable} -m pip install sounddevice numpy\n"
            f"Details: {exc}"
        )

    frames: List[object] = []
    last_flushed_samples = 0
    next_chunk_at = time.monotonic() + chunk_interval_sec
    block_samples = (
        WAVEFORM_INPUT_BLOCK_SAMPLES if on_pcm_block is not None else 4096
    )
    block_interval = max(0.02, block_samples / max(1, sample_rate))
    source_stop = threading.Event()
    source_queues: Dict[str, "queue.Queue[Any]"] = {
        RECORD_SOURCE_MIC: queue.Queue(maxsize=12),
        RECORD_SOURCE_COMPUTER: queue.Queue(maxsize=12),
    }
    source_threads: Dict[str, threading.Thread] = {}
    source_errors: Dict[str, str] = {}
    source_lock = threading.Lock()

    def _source_enabled(source: str) -> bool:
        evt = source_enabled_events.get(source)
        return evt is not None and evt.is_set()

    def _set_source_failed(source: str, err: str) -> None:
        detail = (err or "").strip() or "Unknown error"
        with source_lock:
            if source in source_errors:
                return
            source_errors[source] = detail
        evt = source_enabled_events.get(source)
        if evt is not None:
            evt.clear()
        if on_source_error is not None:
            try:
                on_source_error(source, detail)
            except Exception:
                pass

    def _normalize_input_block(data: Any) -> Any:
        arr = np.asarray(data)
        if arr.size == 0:
            return np.zeros((0, 1), dtype=np.int16)
        if arr.dtype != np.int16:
            arr = arr.astype(np.int16)
        if arr.ndim == 1:
            arr = arr.reshape(-1, 1)
        elif arr.ndim > 1 and arr.shape[1] > 1:
            arr = np.mean(arr.astype(np.float64), axis=1).astype(np.int16).reshape(-1, 1)
        return arr.reshape(-1, 1)

    def _soundcard_float_block_to_int16(data: Any) -> Any:
        arr = np.asarray(data, dtype=np.float32)
        if arr.size == 0:
            return np.zeros((0, 1), dtype=np.int16)
        if arr.ndim == 1:
            mono = arr
        elif arr.shape[1] == 1:
            mono = arr[:, 0]
        else:
            mono = np.mean(arr, axis=1)
        mono = np.nan_to_num(mono, nan=0.0, posinf=0.0, neginf=0.0)
        mono = np.clip(mono, -1.0, 1.0)
        return (mono * 32767.0).astype(np.int16).reshape(-1, 1)

    def _mic_capture_loop() -> None:
        try:
            with sd.InputStream(
                samplerate=sample_rate,
                channels=1,
                dtype=np.int16,
                blocksize=block_samples,
            ) as stream:
                while not source_stop.is_set():
                    data, _overflowed = stream.read(block_samples)
                    block = _normalize_input_block(data)
                    if block.size:
                        _put_latest_audio_block(source_queues[RECORD_SOURCE_MIC], block.copy())
        except Exception as exc:
            _set_source_failed(RECORD_SOURCE_MIC, str(exc))

    def _computer_capture_loop() -> None:
        try:
            import soundcard as sc
        except Exception as exc:
            _set_source_failed(
                RECORD_SOURCE_COMPUTER,
                "Install soundcard for Windows computer-audio recording. "
                f"Details: {exc}",
            )
            return
        try:
            speaker = sc.default_speaker()
            loopback = sc.get_microphone(id=str(speaker.name), include_loopback=True)
            with loopback.recorder(
                samplerate=sample_rate,
                channels=1,
                blocksize=block_samples,
            ) as recorder:
                while not source_stop.is_set():
                    data = recorder.record(numframes=block_samples)
                    block = _soundcard_float_block_to_int16(data)
                    if block.size:
                        _put_latest_audio_block(
                            source_queues[RECORD_SOURCE_COMPUTER],
                            block.copy(),
                        )
        except Exception as exc:
            _set_source_failed(RECORD_SOURCE_COMPUTER, str(exc))

    def _ensure_source_thread(source: str) -> None:
        if source in source_threads or source in source_errors:
            return
        target = _mic_capture_loop if source == RECORD_SOURCE_MIC else _computer_capture_loop
        thread = threading.Thread(target=target, daemon=True)
        source_threads[source] = thread
        thread.start()

    def _enabled_source_names() -> List[str]:
        return [
            source
            for source in (RECORD_SOURCE_MIC, RECORD_SOURCE_COMPUTER)
            if _source_enabled(source)
        ]

    try:
        while not stop_event.is_set():
            active_sources = _enabled_source_names()
            if not active_sources:
                if source_errors:
                    break
                time.sleep(0.05)
                continue
            for source in active_sources:
                _ensure_source_thread(source)

            blocks: List[Any] = []
            for source in active_sources:
                block = _drain_latest_audio_block(source_queues[source])
                if block is None:
                    try:
                        block = source_queues[source].get(timeout=block_interval * 0.75)
                    except queue.Empty:
                        block = None
                if block is not None:
                    normalized = _normalize_input_block(block)
                    if normalized.size:
                        blocks.append(normalized)
            if not blocks:
                time.sleep(block_interval * 0.25)
                continue

            max_len = max(int(block.shape[0]) for block in blocks)
            mixed = np.zeros(max_len, dtype=np.float64)
            for block in blocks:
                flat = block.reshape(-1).astype(np.float64)
                if flat.shape[0] < max_len:
                    padded = np.zeros(max_len, dtype=np.float64)
                    padded[: flat.shape[0]] = flat
                    flat = padded
                mixed += flat
            if len(blocks) > 1:
                mixed /= float(len(blocks))
            mixed_block = np.clip(mixed, -32768, 32767).astype(np.int16).reshape(-1, 1)

            if pause_event is not None and pause_event.is_set():
                continue
            frames.append(mixed_block.copy())
            if on_pcm_block is not None:
                try:
                    on_pcm_block(mixed_block.copy())
                except Exception:
                    pass
            if on_audio_chunk is not None:
                now = time.monotonic()
                if now >= next_chunk_at and frames:
                    big = np.concatenate(frames, axis=0)
                    delta = big[last_flushed_samples:]
                    if delta.size >= LIVE_STT_MIN_CHUNK_SAMPLES:
                        fd, tmp_name = tempfile.mkstemp(suffix=".wav", prefix="stt_chunk_")
                        os.close(fd)
                        chunk_path = Path(tmp_name)
                        werr = write_mono_int16_wav(chunk_path, delta, sample_rate)
                        if werr is None:
                            on_audio_chunk(chunk_path)
                            last_flushed_samples = int(big.shape[0])
                        else:
                            try:
                                chunk_path.unlink(missing_ok=True)
                            except OSError:
                                pass
                    next_chunk_at = now + chunk_interval_sec
    except Exception as exc:
        return str(exc)
    finally:
        source_stop.set()
        for thread in list(source_threads.values()):
            if thread.is_alive():
                thread.join(timeout=1.0)

    if not frames:
        if source_errors:
            return "; ".join(f"{source}: {err}" for source, err in source_errors.items())
        return "No audio captured."

    audio = np.concatenate(frames, axis=0)
    werr = write_mono_int16_wav(output_path, audio, sample_rate)
    if werr is not None:
        return werr
    return None


def record_microphone_session_wav(
    output_path: Path,
    stop_event: threading.Event,
    *,
    sample_rate: int = 16000,
    chunk_interval_sec: float = LIVE_STT_CHUNK_INTERVAL_SEC,
    on_audio_chunk: Optional[Callable[[Path], None]] = None,
    on_pcm_block: Optional[Callable[[Any], None]] = None,
    pause_event: Optional[threading.Event] = None,
) -> Optional[str]:
    mic_event = threading.Event()
    mic_event.set()
    return record_sources_session_wav(
        output_path,
        stop_event,
        source_enabled_events={RECORD_SOURCE_MIC: mic_event},
        sample_rate=sample_rate,
        chunk_interval_sec=chunk_interval_sec,
        on_audio_chunk=on_audio_chunk,
        on_pcm_block=on_pcm_block,
        pause_event=pause_event,
    )


def generate_journal_report_from_sources(
    journal_text: str,
    speech_transcript: str,
    *,
    progress: Optional[Callable[[str], None]] = None,
) -> str:
    system_message = (
        "You produce clear, professional summaries of daily work notes. "
        "Highlight key activities, decisions, blockers, and suggested follow-ups. "
        "Use short sections with bullets where appropriate."
    )
    if progress is not None:
        try:
            progress("Preparing report")
        except Exception:
            pass
    user_content = (
        "### Journal text\n"
        + (normalize_journal_text_punctuation(journal_text.strip()) or "(empty)")
        + "\n\n### Speech-to-text transcript\n"
        + (normalize_journal_text_punctuation(speech_transcript.strip()) or "(none)")
    )
    messages: List[Dict[str, object]] = [
        {"role": "system", "content": system_message},
        {"role": "user", "content": user_content},
    ]
    return normalize_journal_text_punctuation(chat_completion(
        messages,
        model=OPENAI_THINKING_MODEL,
        reasoning_effort="high",
        progress=progress,
        timeout_sec=180,
        attempts=4,
    ))


def open_journal_window_editor(
    draft_data: Optional[Dict[str, object]] = None,
    *,
    start_auto_backup: bool = False,
) -> bool:
    if tk is None or messagebox is None:
        print("Window mode is not available on this Python setup.")
        return False

    now = datetime.now()
    default_date = now.strftime("%m/%d/%Y")
    default_time = now.strftime("%I:%M%p").lstrip("0")
    draft_text = ""
    draft_speech = ""
    draft_report = ""
    draft_date = default_date
    draft_time = default_time
    edit_target_sheet = ""
    edit_target_row = 0
    if draft_data:
        draft_text = normalize_journal_text_punctuation(str(draft_data.get("text", "") or ""))
        draft_speech = normalize_journal_text_punctuation(
            str(draft_data.get("speech_transcript", "") or "")
        )
        draft_report = normalize_journal_text_punctuation(str(draft_data.get("ai_report", "") or ""))
        draft_date = str(draft_data.get("date", default_date) or default_date)
        draft_time = str(draft_data.get("time", default_time) or default_time)
        edit_target_sheet = str(draft_data.get("edit_target_sheet", "") or "")
        try:
            edit_target_row = int(draft_data.get("edit_target_row", 0) or 0)
        except (TypeError, ValueError):
            edit_target_row = 0

    root = tk.Tk()
    journal_cleanup_callbacks: List[Callable[[], None]] = []
    uninstall_shutdown_requested = {"v": False}

    def destroy_journal_window() -> None:
        while journal_cleanup_callbacks:
            cb = journal_cleanup_callbacks.pop()
            try:
                cb()
            except Exception:
                pass
        shutdown_virtual_reader_child_server()
        root.destroy()

    root_prefs = load_preferences()
    fullscreen_state = {
        "enabled": str(root_prefs.get(JOURNAL_PREF_FULLSCREEN_KEY, "false")).strip().lower()
        == "true"
    }
    ui_lang_holder: List[str] = [
        normalize_ui_language(str(root_prefs.get(UI_LANGUAGE_PREF_KEY, "en")))
    ]

    def tr(key: str, **kwargs: object) -> str:
        return ui_translate(ui_lang_holder[0], key, **kwargs)

    window_app_name = root_prefs.get("app_name", "Daily Logger").strip() or "Daily Logger"
    root.title(window_app_name)
    root.geometry("1360x720")
    root.minsize(1020, 620)
    try:
        root.attributes("-fullscreen", bool(fullscreen_state["enabled"]))
    except tk.TclError:
        fullscreen_state["enabled"] = False
    theme_holder: List[JournalWindowThemeSpec] = [load_journal_window_theme_spec()]

    def th() -> JournalWindowThemeSpec:
        return theme_holder[0]


    t_init = th()
    root.configure(bg=t_init.surface)
    startup_total_steps = 6
    startup_progress = {"value": 0}
    startup_overlay = tk.Frame(root, bg=t_init.surface, bd=0, highlightthickness=0)
    startup_overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
    startup_box = tk.Frame(startup_overlay, bg=t_init.surface, bd=0, highlightthickness=0)
    startup_box.place(relx=0.5, rely=0.5, anchor="center")
    splash_w = 460
    splash_title = tk.Label(
        startup_box,
        text=tr("splash.title", app=window_app_name),
        bg=t_init.surface,
        fg=t_init.text,
        font=("Segoe UI", 11, "bold"),
        anchor="w",
    )
    splash_title.pack(fill="x", padx=16, pady=(14, 6))
    splash_detail = tk.Label(
        startup_box,
        text=tr("splash.detail.theme"),
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
    )
    splash_detail.pack(fill="x", padx=16, pady=(0, 10))
    startup_bar: Any
    startup_canvas: Optional[Any] = None
    startup_fill: Optional[Any] = None
    if ttk is not None:
        startup_bar = ttk.Progressbar(
            startup_box,
            orient="horizontal",
            mode="determinate",
            maximum=float(startup_total_steps),
            length=splash_w - 32,
        )
        startup_bar.pack(padx=16, pady=(0, 14))
    else:
        startup_canvas = tk.Canvas(
            startup_box,
            width=splash_w - 32,
            height=16,
            bg=t_init.field,
            highlightthickness=1,
            highlightbackground=t_init.border,
        )
        startup_canvas.pack(padx=16, pady=(0, 14))
        startup_fill = startup_canvas.create_rectangle(0, 0, 0, 16, fill=t_init.accent, width=0)
        startup_bar = None

    def _startup_step(detail_key: str) -> None:
        startup_progress["value"] = min(startup_total_steps, startup_progress["value"] + 1)
        splash_detail.config(text=tr(detail_key))
        if startup_bar is not None:
            startup_bar["value"] = float(startup_progress["value"])
        elif startup_canvas is not None and startup_fill is not None:
            bar_w = int((splash_w - 32) * (startup_progress["value"] / startup_total_steps))
            startup_canvas.coords(startup_fill, 0, 0, bar_w, 16)
        startup_overlay.lift()
        root.update_idletasks()

    _startup_step("splash.detail.theme")
    _jw_style: Any = None
    if ttk is not None:
        _jw_style = ttk.Style(root)
        try:
            _jw_style.theme_use("clam")
        except tk.TclError:
            pass
        _jw_style.configure("Journal.TCombobox", **t_init.ttk_combobox_kwargs())
        if t_init.is_dark:
            _jw_style.map(
                "Journal.TCombobox",
                fieldbackground=[
                    ("readonly", t_init.field),
                    ("disabled", t_init.btn_disabled),
                ],
                selectbackground=[("readonly", t_init.accent)],
                selectforeground=[("readonly", "white")],
            )
        else:
            _jw_style.map(
                "Journal.TCombobox",
                fieldbackground=[
                    ("readonly", t_init.field),
                    ("disabled", t_init.btn_disabled),
                ],
            )
    # Bring the journal window to front so it does not hide behind the console.
    root.lift()
    root.attributes("-topmost", True)
    root.after(250, lambda: root.attributes("-topmost", False))
    root.focus_force()
    # Mutable so console commands (like editprev) can toggle edit mode in-place.
    is_edit_mode = {"v": bool(edit_target_sheet and edit_target_row > 0)}

    shell = tk.Frame(root, bg=t_init.surface, bd=0, highlightthickness=0)
    shell.pack(fill="both", expand=True)
    shell.grid_rowconfigure(0, weight=1)
    shell.grid_columnconfigure(1, weight=1)
    shell.grid_columnconfigure(0, minsize=170)

    nav_rail = tk.Frame(shell, bg=t_init.panel, width=170, bd=0, highlightthickness=0)
    nav_rail.grid(row=0, column=0, sticky="nsw")
    nav_rail.grid_rowconfigure(98, weight=1)
    nav_rail.grid_rowconfigure(100, weight=0)
    nav_rail.grid_columnconfigure(0, weight=1)
    nav_rail.grid_propagate(False)

    content_host = tk.Frame(shell, bg=t_init.surface, bd=0, highlightthickness=0)
    content_host.grid(row=0, column=1, sticky="nsew")
    content_host.grid_rowconfigure(0, weight=1)
    content_host.grid_rowconfigure(1, weight=0)
    content_host.grid_columnconfigure(0, weight=1)
    console_input_holder: Dict[str, Any] = {"row": None}
    console_output_holder: Dict[str, Any] = {"widget": None}
    console_session_updates: List[str] = []
    console_session_last: Dict[str, str] = {}

    journal_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    ai_recap_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    chatbot_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    desktop_pet_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    console_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    settings_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    module_library_page = tk.Frame(content_host, bg=t_init.surface, bd=0, highlightthickness=0)
    for _p in (
        journal_page,
        ai_recap_page,
        chatbot_page,
        desktop_pet_page,
        console_page,
        settings_page,
        module_library_page,
    ):
        _p.grid(row=0, column=0, sticky="nsew")
    # Ensure first paint shows Journal instead of last-created stacked page.
    journal_page.tkraise()

    nav_collapsed = {"value": False}
    nav_animating = {"value": False}
    nav_full_width = 170
    nav_restore_page = {"key": "journal"}
    nav_title = tk.Label(
        nav_rail,
        text=window_app_name,
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 10, "bold"),
    )
    nav_title.grid(row=0, column=0, sticky="w", padx=(12, 0), pady=(14, 10))

    nav_buttons: Dict[str, Any] = {}
    nav_module_widgets: Dict[str, Dict[str, Any]] = {}
    nav_module_order_state: Dict[str, List[str]] = {"ids": []}
    nav_edit_state: Dict[str, Any] = {
        "active": False,
        "press_module": "",
        "press_after": None,
        "long_hold": False,
        "dragging": False,
        "drag_module": "",
        "placeholder_index": None,
        "hover_replace_module": "",
        "library_drag_module": "",
    }
    active_page = {"key": "journal"}
    active_page_frame: Dict[str, Any] = {"frame": None}
    page_leave_reset_handlers: Dict[str, Callable[[], None]] = {}

    def _layout_console_row(frame: Any) -> None:
        frame.update_idletasks()
        fw = frame.winfo_width()
        reveal_width = nav_summon_btn.winfo_width() if nav_collapsed["value"] else 0
        left_margin = 20 + (reveal_width + 8 if nav_collapsed["value"] else 0)
        right_margin = 20
        if save_entry_btn_holder.get("btn") is not None and frame is journal_page:
            save_entry_btn.update_idletasks()
            save_w = max(save_entry_btn.winfo_width(), save_entry_btn.winfo_reqwidth())
            right_margin = max(right_margin, save_w + 36)
            button_row.place(
                in_=frame,
                relx=1.0,
                rely=1.0,
                x=-20,
                y=-12,
                anchor="se",
            )
            button_row.lift()
        row_w = max(280, fw - left_margin - right_margin)
        console_row = console_input_holder.get("row")
        if console_row is not None:
            console_row.place(
                in_=frame,
                x=left_margin,
                rely=1.0,
                y=-12,
                anchor="sw",
                width=row_w,
            )
            console_row.lift()
        if save_entry_btn_holder.get("btn") is not None and frame is journal_page:
            button_row.lift()

    def show_page(page_key: str) -> None:
        if page_key == "desktop_pet" and not COMPANION_FEATURE_ENABLED:
            page_key = "journal"
        page_map = {
            "journal": journal_page,
            "ai_recap": ai_recap_page,
            "chatbot": chatbot_page,
            "console": console_page,
            "settings": settings_page,
            "module_library": module_library_page,
        }
        prev_key = active_page["key"]
        if page_key == "console":
            _clear_console_hint()
        frame = page_map.get(page_key, journal_page)
        frame.tkraise()
        active_page["key"] = page_key
        if prev_key != page_key:
            reset_fn = page_leave_reset_handlers.get(prev_key)
            if reset_fn is not None:
                reset_fn()
        active_page_frame["frame"] = frame
        console_row = console_input_holder.get("row")
        if console_row is not None:
            _layout_console_row(frame)
        try:
            _refresh_sidebar_module_styles()
        except NameError:
            pass
        if COMPANION_FEATURE_ENABLED:
            try:
                _draw_journal_pet()
            except NameError:
                pass

    page_toggle_buttons: List[Any] = []
    nav_summon_btn = tk.Button(
        content_host,
        text="\u25b6",
        bg=t_init.toolbar_btn_config()[0],
        fg=t_init.toolbar_btn_config()[1],
        activebackground=t_init.toolbar_btn_config()[2],
        activeforeground=t_init.toolbar_btn_config()[3],
        relief="flat",
        font=("Segoe UI", 10, "bold"),
        padx=2,
        pady=12,
        cursor="hand2",
        bd=0,
        highlightthickness=0,
        width=1,
    )

    def _place_page_toggle(btn: Any) -> None:
        if nav_animating["value"]:
            return
        if nav_collapsed["value"]:
            btn.place_forget()
        else:
            nav_rail.update_idletasks()
            ph = nav_rail.winfo_height()
            y = max(56, (ph // 2) - 14)
            # Place on the right seam of the Pages rail.
            x = max(0, nav_rail.winfo_width() - 12)
            btn.place(x=x, y=y)

    def _place_nav_summon() -> None:
        if nav_animating["value"]:
            return
        content_host.update_idletasks()
        jh = content_host.winfo_height()
        y = max(56, (jh // 2) - 14)
        nav_summon_btn.place(x=0, y=y)
        nav_summon_btn.lift()

    def _register_page_toggle(parent: Any) -> Any:
        if page_toggle_buttons:
            return page_toggle_buttons[0]
        btn = tk.Button(
            nav_rail,
            text="\u25c0",
            bg=t_init.toolbar_btn_config()[0],
            fg=t_init.toolbar_btn_config()[1],
            activebackground=t_init.toolbar_btn_config()[2],
            activeforeground=t_init.toolbar_btn_config()[3],
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=2,
            pady=12,
            cursor="hand2",
            bd=0,
            highlightthickness=0,
            width=1,
        )
        btn.config(command=lambda b=btn: set_nav_visible(False))
        bind_button_hover_if_enabled(
            btn,
            lambda: (
                "normal",
                th().toolbar_btn_config()[0],
                th().toolbar_btn_config()[1],
                th().toolbar_btn_config()[2],
                th().toolbar_btn_config()[3],
            ),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        page_toggle_buttons.append(btn)
        _place_page_toggle(btn)
        nav_rail.bind(
            "<Configure>",
            lambda _e, b=btn: _place_page_toggle(b) if not nav_animating["value"] else None,
            add="+",
        )
        return btn

    def set_nav_visible(visible: bool) -> None:
        if nav_animating["value"]:
            return
        target_collapsed = not visible
        if nav_collapsed["value"] == target_collapsed:
            return

        nav_animating["value"] = True
        nav_collapsed["value"] = not visible

        def _animate_width(start: int, target: int, done: Callable[[], None]) -> None:
            duration_ms = 220.0
            t0 = time.perf_counter()

            def _tick() -> None:
                elapsed = (time.perf_counter() - t0) * 1000.0
                p = min(1.0, elapsed / duration_ms)
                eased = 1.0 - ((1.0 - p) ** 3)
                nxt = int(round(start + (target - start) * eased))
                shell.grid_columnconfigure(0, minsize=nxt)
                nav_rail.config(width=nxt)
                if p >= 1.0:
                    done()
                else:
                    root.after(16, _tick)

            _tick()

        if visible:
            nav_rail.grid()
            nav_summon_btn.place_forget()
            shell.grid_columnconfigure(0, minsize=0)
            nav_rail.config(width=0)

            def _on_expand_done() -> None:
                nav_animating["value"] = False
                for btn in page_toggle_buttons:
                    _place_page_toggle(btn)
                restore_key = nav_restore_page.get("key", "journal")
                if restore_key in (
                    "journal",
                    "ai_recap",
                    "chatbot",
                    "console",
                    "settings",
                    "module_library",
                ):
                    show_page(restore_key)

            _animate_width(0, nav_full_width, _on_expand_done)
        else:
            for btn in page_toggle_buttons:
                btn.place_forget()
            nav_restore_page["key"] = active_page["key"]

            def _on_collapse_done() -> None:
                nav_animating["value"] = False
                shell.grid_columnconfigure(0, minsize=0)
                nav_rail.config(width=0)
                nav_rail.grid_remove()
                _place_nav_summon()

            _animate_width(nav_full_width, 0, _on_collapse_done)

    top = tk.Frame(journal_page, bg=t_init.panel, bd=0, highlightthickness=0)
    top.pack(fill="x", padx=t_init.pad_outer, pady=t_init.pad_top_y)
    top.grid_columnconfigure(5, weight=1)
    _register_page_toggle(journal_page)
    date_lbl = tk.Label(
        top,
        text="Date (mm/dd/yyyy):",
        bg=t_init.panel,
        fg=t_init.muted,
        font=t_init.date_label_font,
    )
    date_lbl.grid(row=0, column=0, sticky="w", padx=(12, 0), pady=12)
    date_entry: object
    if DateEntry is not None:
        date_entry = DateEntry(
            top,
            width=14,
            date_pattern="mm/dd/yyyy",
            state="normal",  # Keep typing enabled while allowing popup calendar selection.
            background=t_init.field,
            foreground=t_init.text,
            borderwidth=1,
        )
        date_entry.grid(row=0, column=1, padx=(8, 20), pady=12, sticky="w")
        try:
            date_entry.set_date(draft_date)
        except Exception:
            date_entry.delete(0, "end")
            date_entry.insert(0, draft_date)
    else:
        date_entry = tk.Entry(
            top,
            width=16,
            bg=t_init.field,
            fg=t_init.text,
            insertbackground=t_init.text,
            relief="flat",
            highlightthickness=1,
            highlightbackground=t_init.border,
            highlightcolor=t_init.accent,
            font=("Segoe UI", 10),
        )
        date_entry.grid(row=0, column=1, padx=(8, 20), pady=12, sticky="w")
        date_entry.insert(0, draft_date)
    time_lbl = tk.Label(
        top,
        text="Time (hh:mmAM/PM or rn):",
        bg=t_init.panel,
        fg=t_init.muted,
        font=t_init.date_label_font,
    )
    time_lbl.grid(row=0, column=2, sticky="w", pady=12)
    time_entry = tk.Entry(
        top,
        width=16,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
        font=("Segoe UI", 10),
    )
    time_entry.grid(row=0, column=3, padx=(8, 0), pady=12)
    time_entry.insert(0, draft_time)
    def update_date_time_to_now() -> None:
        current_now = datetime.now()
        date_entry.delete(0, "end")
        date_entry.insert(0, current_now.strftime("%m/%d/%Y"))
        time_entry.delete(0, "end")
        time_entry.insert(0, current_now.strftime("%I:%M%p").lstrip("0"))
        save_draft()
    _ut_bg, _ut_fg, _ut_abg, _ut_afg = t_init.toolbar_btn_config()
    update_time_btn = tk.Button(
        top,
        text="Update Time",
        command=update_date_time_to_now,
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
    )
    update_time_btn.grid(row=0, column=4, padx=(12, 12), sticky="w")
    bind_button_hover_if_enabled(
        update_time_btn,
        lambda: th().toolbar_bind_rest(),
        lambda: th().toolbar_hover()[0],
        lambda: th().toolbar_hover()[1],
    )
    find_row = tk.Frame(journal_page, bg=t_init.panel, bd=0, highlightthickness=0)
    find_row.pack(fill="x", padx=t_init.pad_outer, pady=(0, 6))
    find_row.grid_columnconfigure(8, weight=1)
    find_lbl = tk.Label(
        find_row,
        text="Find:",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 9, "bold"),
    )
    find_lbl.grid(row=0, column=0, sticky="w", padx=(12, 6), pady=8)
    find_scope_var = tk.StringVar(value="all")
    find_scope_all_rb = tk.Radiobutton(
        find_row,
        text="All",
        value="all",
        variable=find_scope_var,
        bg=t_init.panel,
        fg=t_init.muted,
        activebackground=t_init.panel,
        activeforeground=t_init.text,
        selectcolor=t_init.panel,
        font=("Segoe UI", 9),
        highlightthickness=0,
        bd=0,
        padx=4,
    )
    find_scope_all_rb.grid(row=0, column=1, sticky="w", padx=(2, 4), pady=8)
    find_scope_one_rb = tk.Radiobutton(
        find_row,
        text="Current box",
        value="one",
        variable=find_scope_var,
        bg=t_init.panel,
        fg=t_init.muted,
        activebackground=t_init.panel,
        activeforeground=t_init.text,
        selectcolor=t_init.panel,
        font=("Segoe UI", 9),
        highlightthickness=0,
        bd=0,
        padx=4,
    )
    find_scope_one_rb.grid(row=0, column=2, sticky="w", padx=(0, 8), pady=8)
    find_var = tk.StringVar(value="")
    find_entry = tk.Entry(
        find_row,
        textvariable=find_var,
        width=28,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
        font=("Segoe UI", 10),
    )
    find_entry.grid(row=0, column=3, sticky="w", pady=8)
    find_case_var = tk.BooleanVar(value=False)
    find_case_chk = tk.Checkbutton(
        find_row,
        text="Case",
        variable=find_case_var,
        bg=t_init.panel,
        fg=t_init.muted,
        activebackground=t_init.panel,
        activeforeground=t_init.text,
        selectcolor=t_init.panel,
        font=("Segoe UI", 9),
        highlightthickness=0,
        bd=0,
        padx=4,
    )
    find_case_chk.grid(row=0, column=4, sticky="w", padx=(8, 0), pady=8)
    find_word_var = tk.BooleanVar(value=False)
    find_word_chk = tk.Checkbutton(
        find_row,
        text="Word",
        variable=find_word_var,
        bg=t_init.panel,
        fg=t_init.muted,
        activebackground=t_init.panel,
        activeforeground=t_init.text,
        selectcolor=t_init.panel,
        font=("Segoe UI", 9),
        highlightthickness=0,
        bd=0,
        padx=4,
    )
    find_word_chk.grid(row=0, column=5, sticky="w", padx=(4, 0), pady=8)
    bind_hover_tooltip(
        find_scope_all_rb,
        lambda: tr("tip.find_all"),
    )
    bind_hover_tooltip(
        find_scope_one_rb,
        lambda: tr("tip.find_one"),
    )
    bind_hover_tooltip(
        find_case_chk,
        lambda: tr("tip.find_case"),
    )
    bind_hover_tooltip(
        find_word_chk,
        lambda: tr("tip.find_word"),
    )
    find_status = tk.Label(
        find_row,
        text="",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 9),
    )
    find_status.grid(row=0, column=6, sticky="w", padx=(8, 0), pady=8)
    find_prev_btn = tk.Button(
        find_row,
        text="Prev",
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=4,
        cursor="hand2",
    )
    find_prev_btn.grid(row=0, column=7, sticky="e", padx=(10, 6), pady=8)
    find_next_btn = tk.Button(
        find_row,
        text="Next",
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=4,
        cursor="hand2",
    )
    find_next_btn.grid(row=0, column=8, sticky="e", padx=6, pady=8)
    find_close_btn = tk.Button(
        find_row,
        text="Close",
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=4,
        cursor="hand2",
    )
    find_close_btn.grid(row=0, column=9, sticky="e", padx=(6, 12), pady=8)
    find_row.pack_forget()

    center = tk.Frame(journal_page, bg=t_init.surface)
    center.pack(
        fill="both",
        expand=True,
        padx=t_init.pad_outer,
        pady=(0, t_init.pad_center_y + JOURNAL_WINDOW_CONSOLE_RESERVE_BOTTOM),
    )
    center.grid_columnconfigure(0, weight=2)
    center.grid_columnconfigure(1, weight=2)
    center.grid_rowconfigure(0, weight=1)

    left_col = tk.Frame(center, bg=t_init.surface)
    left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
    left_col.grid_columnconfigure(0, weight=1)
    left_col.grid_rowconfigure(1, weight=1)
    journal_title_lbl = tk.Label(
        left_col,
        text="Journal Text",
        bg=t_init.surface,
        fg=t_init.muted,
        font=t_init.section_label_font,
    )
    journal_title_lbl.grid(row=0, column=0, sticky="w", pady=(0, 6))
    editor_frame = tk.Frame(left_col, bg=t_init.panel, bd=0, highlightthickness=0)
    editor_frame.grid(row=1, column=0, sticky="nsew")
    editor_frame.grid_rowconfigure(0, weight=1)
    editor_frame.grid_columnconfigure(0, weight=1)
    text_box = tk.Text(
        editor_frame,
        wrap="word",
        height=12,
        undo=True,
        autoseparators=True,
        maxundo=-1,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        padx=12,
        pady=12,
        font=(JOURNAL_TEXT_FONT_FAMILY, 11),
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    scroll_bar = JournalSlimScrollbar(
        editor_frame,
        command=text_box.yview,
        theme=t_init,
        width=JOURNAL_TEXT_SCROLLBAR_WIDTH,
    )
    text_box.configure(yscrollcommand=scroll_bar.set)
    text_box.grid(row=0, column=0, sticky="nsew", padx=(12, 0), pady=12)
    scroll_bar.grid(row=0, column=1, sticky="nsew", padx=(0, 12), pady=12)
    text_box.insert("1.0", draft_text)
    text_box.focus_set()
    root.after(50, text_box.focus_set)

    right_col = tk.Frame(center, bg=t_init.surface)
    right_col.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
    right_col.grid_rowconfigure(0, weight=1)
    right_col.grid_rowconfigure(1, weight=1)
    right_col.grid_columnconfigure(0, weight=1)

    stt_outer = tk.Frame(right_col, bg=t_init.surface)
    stt_outer.grid(row=0, column=0, sticky="nsew", pady=(0, 6))
    stt_outer.grid_columnconfigure(0, weight=1)
    stt_header = tk.Frame(stt_outer, bg=t_init.surface)
    stt_header.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    stt_header.grid_columnconfigure(1, weight=1)
    stt_title_lbl = tk.Label(
        stt_header,
        text="Speech to text",
        bg=t_init.surface,
        fg=t_init.muted,
        font=t_init.section_label_font,
    )
    stt_title_lbl.grid(row=0, column=0, sticky="w")
    stt_saved_path_var = tk.StringVar(value="")
    stt_saved_path_entry = tk.Entry(
        stt_header,
        textvariable=stt_saved_path_var,
        state="readonly",
        readonlybackground=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 8),
        relief="flat",
        borderwidth=0,
        highlightthickness=0,
        highlightbackground=t_init.surface,
        insertwidth=0,
        justify="right",
        takefocus=1,
        selectbackground=t_init.accent,
        selectforeground="white",
        cursor="xterm",
    )
    stt_saved_path_entry.grid(row=0, column=1, sticky="ew", padx=(14, 8))
    stt_saved_path_full_text: Dict[str, str] = {"text": ""}

    def _set_stt_saved_path_display(text: str) -> None:
        raw = str(text or "")
        stt_saved_path_full_text["text"] = raw
        display = raw
        match = re.search(r"([A-Za-z]:\\[^\n]+)$", raw.strip())
        if match:
            path_text = match.group(1).strip()
            path_name = Path(path_text).name
            if path_name:
                display = raw[: match.start(1)] + path_name
        stt_saved_path_entry.config(state="normal")
        stt_saved_path_var.set(display)
        stt_saved_path_entry.config(state="readonly")

    bind_hover_tooltip(stt_saved_path_entry, lambda: stt_saved_path_full_text["text"])

    def open_journal_recording_folder() -> None:
        try:
            RECORDING_DIR.mkdir(parents=True, exist_ok=True)
        except OSError:
            pass
        open_path_with_default_app(RECORDING_DIR)

    open_recording_btn = tk.Button(
        stt_header,
        text="Open",
        command=open_journal_recording_folder,
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 8, "bold"),
        padx=6,
        pady=2,
        cursor="hand2",
    )
    open_recording_btn.grid(row=0, column=2, sticky="e")

    def open_recording_tooltip_text() -> str:
        return tr("tip.open_recordings")

    bind_hover_tooltip(open_recording_btn, open_recording_tooltip_text)
    bind_button_hover_if_enabled(
        open_recording_btn,
        lambda: th().toolbar_bind_rest(),
        lambda: th().toolbar_hover()[0],
        lambda: th().toolbar_hover()[1],
    )

    stt_top = tk.Frame(stt_outer, bg=t_init.panel, bd=0, highlightthickness=0)
    stt_top.grid(row=1, column=0, sticky="ew", pady=(0, 6))
    stt_top.grid_columnconfigure(4, weight=1)
    lang_var = tk.StringVar(value="Auto")

    stt_status = tk.Label(
        stt_top,
        text="",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
        wraplength=420,
    )

    if ttk is not None:
        lang_combo = ttk.Combobox(
            stt_top,
            textvariable=lang_var,
            values=("Auto", "English", "\u7b80\u4f53\u4e2d\u6587"),
            state="readonly",
            width=11,
            style="Journal.TCombobox",
        )
    else:
        lang_combo = tk.OptionMenu(stt_top, lang_var, "Auto", "English", "\u7b80\u4f53\u4e2d\u6587")
        lang_combo.config(bg=t_init.panel, fg=t_init.text, highlightthickness=0)

    stt_frame = tk.Frame(stt_outer, bg=t_init.panel, bd=0, highlightthickness=0)
    stt_frame.grid(row=2, column=0, sticky="nsew")
    stt_frame.grid_rowconfigure(0, weight=0)
    stt_frame.grid_rowconfigure(1, weight=1)
    stt_frame.grid_columnconfigure(0, weight=1)
    stt_frame.grid_columnconfigure(1, minsize=JOURNAL_TEXT_SCROLLBAR_WIDTH)
    stt_frame.grid_columnconfigure(2, minsize=JOURNAL_SIDE_ACTION_GRID_MINSIZE)
    stt_outer.grid_rowconfigure(2, weight=1)

    wave_canvas = tk.Canvas(
        stt_frame,
        height=52,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    wave_canvas.grid(row=0, column=0, columnspan=3, sticky="ew", padx=10, pady=(10, 4))

    stt_box = tk.Text(
        stt_frame,
        wrap="word",
        height=8,
        undo=True,
        autoseparators=True,
        maxundo=-1,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        padx=10,
        pady=10,
        font=(JOURNAL_TEXT_FONT_FAMILY, 10),
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    stt_scroll = JournalSlimScrollbar(
        stt_frame,
        command=stt_box.yview,
        theme=t_init,
        width=JOURNAL_TEXT_SCROLLBAR_WIDTH,
    )
    stt_box.configure(yscrollcommand=stt_scroll.set)
    stt_box.grid(row=1, column=0, sticky="nsew", padx=(10, 0), pady=(4, 10))
    stt_scroll.grid(row=1, column=1, sticky="nsew", padx=(0, 2), pady=(4, 10))
    transcribe_hover = tk.Frame(stt_frame, bg=t_init.panel)
    transcribe_hover.grid(row=1, column=2, sticky="nsew", padx=(2, 10), pady=(4, 10))
    _tid = t_init.transcribe_idle_disabled_config()
    transcription_model_combo_map: Dict[str, str] = {}
    transcription_model_selector_is_combo = False

    def _make_transcription_model_selector(parent: Any, variable: Any) -> Any:
        return tk.Button(
            parent,
            text="▼",
            state="normal",
            width=3,
            bg=_tid[0],
            fg=_tid[1],
            activebackground=_tid[2],
            activeforeground=_tid[3],
            disabledforeground=_tid[4],
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=0,
            pady=8,
            cursor="hand2",
        )

    transcribe_btn_row = tk.Frame(transcribe_hover, bg=t_init.panel)
    transcribe_btn_row.pack(fill="x")
    transcribe_btn_row.grid_columnconfigure(0, weight=1)
    transcribe_btn = tk.Button(
        transcribe_btn_row,
        text="Transcribe",
        state="disabled",
        width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
        bg=_tid[0],
        fg=_tid[1],
        activebackground=_tid[2],
        activeforeground=_tid[3],
        disabledforeground=_tid[4],
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=8,
        cursor="hand2",
    )
    transcribe_btn.grid(row=0, column=0, sticky="ew")
    transcribe_model_var = tk.StringVar(value="Model")
    transcribe_model_btn = _make_transcription_model_selector(
        transcribe_btn_row,
        transcribe_model_var,
    )
    transcribe_model_btn.grid(row=0, column=1, sticky="ew", padx=(2, 0))
    transcribe_file_btn_row = tk.Frame(transcribe_hover, bg=t_init.panel)
    transcribe_file_btn_row.pack(fill="x", pady=(6, 0))
    transcribe_file_btn_row.grid_columnconfigure(0, weight=1)
    transcribe_file_btn = tk.Button(
        transcribe_file_btn_row,
        text="Transcribe File",
        state="disabled",
        width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
        bg=_tid[0],
        fg=_tid[1],
        activebackground=_tid[2],
        activeforeground=_tid[3],
        disabledforeground=_tid[4],
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=8,
        cursor="hand2",
    )
    transcribe_file_btn.grid(row=0, column=0, sticky="ew")
    transcribe_file_model_var = tk.StringVar(value="Model")
    transcribe_file_model_btn = _make_transcription_model_selector(
        transcribe_file_btn_row,
        transcribe_file_model_var,
    )
    transcribe_file_model_btn.grid(row=0, column=1, sticky="ew", padx=(2, 0))
    receive_iphone_btn_row = tk.Frame(transcribe_hover, bg=t_init.panel)
    receive_iphone_btn_row.pack(fill="x", pady=(6, 0))
    receive_iphone_btn_row.grid_columnconfigure(0, weight=1)
    receive_iphone_btn = tk.Button(
        receive_iphone_btn_row,
        text="Receive from iPhone",
        state="normal",
        width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
        bg=_tid[0],
        fg=_tid[1],
        activebackground=_tid[2],
        activeforeground=_tid[3],
        disabledforeground=_tid[4],
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=8,
        cursor="hand2",
    )
    receive_iphone_btn.grid(row=0, column=0, columnspan=2, sticky="ew")
    stt_box.insert("1.0", draft_speech)

    report_outer = tk.Frame(right_col, bg=t_init.surface)
    report_outer.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
    report_outer.grid_rowconfigure(1, weight=1)
    report_outer.grid_columnconfigure(0, weight=1)
    report_header = tk.Frame(report_outer, bg=t_init.surface)
    report_header.grid(row=0, column=0, sticky="ew", pady=(0, 4))
    report_header.grid_columnconfigure(1, weight=1)
    report_title_lbl = tk.Label(
        report_header,
        text="AI report",
        bg=t_init.surface,
        fg=t_init.muted,
        font=t_init.section_label_font,
    )
    report_title_lbl.grid(row=0, column=0, sticky="w")
    report_status = tk.Label(
        report_header,
        text="",
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="e",
    )
    report_status.grid(row=0, column=1, sticky="e", padx=(8, 0))

    report_frame = tk.Frame(report_outer, bg=t_init.panel, bd=0, highlightthickness=0)
    report_frame.grid(row=1, column=0, sticky="nsew")
    report_frame.grid_rowconfigure(0, weight=1)
    report_frame.grid_columnconfigure(0, weight=1)
    report_frame.grid_columnconfigure(1, minsize=JOURNAL_TEXT_SCROLLBAR_WIDTH)
    report_frame.grid_columnconfigure(2, minsize=JOURNAL_SIDE_ACTION_GRID_MINSIZE)
    report_box = tk.Text(
        report_frame,
        wrap="word",
        height=8,
        undo=True,
        autoseparators=True,
        maxundo=-1,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        padx=10,
        pady=10,
        font=(JOURNAL_TEXT_FONT_FAMILY, 10),
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    report_scroll = JournalSlimScrollbar(
        report_frame,
        command=report_box.yview,
        theme=t_init,
        width=JOURNAL_TEXT_SCROLLBAR_WIDTH,
    )
    report_box.configure(yscrollcommand=report_scroll.set)
    report_box.grid(row=0, column=0, sticky="nsew", padx=(10, 0), pady=(4, 10))
    report_scroll.grid(row=0, column=1, sticky="nsew", padx=(0, 2), pady=(4, 10))
    gen_report_hover = tk.Frame(report_frame, bg=t_init.panel)
    gen_report_hover.grid(row=0, column=2, sticky="nsew", padx=(2, 10), pady=(4, 10))
    _, _gn, _gf, _gab, _gaf = t_init.gen_bind_rest()
    gen_report_btn_row = tk.Frame(gen_report_hover, bg=t_init.panel)
    gen_report_btn_row.pack(fill="x")
    gen_report_btn_row.grid_columnconfigure(0, weight=1)
    gen_report_btn_row.grid_columnconfigure(1, minsize=max(1, transcribe_model_btn.winfo_reqwidth()))
    gen_button = tk.Button(
        gen_report_btn_row,
        text="Generate report",
        width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
        bg=_gn,
        fg=_gf,
        activebackground=_gab,
        activeforeground=_gaf,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=8,
        cursor="hand2",
    )
    gen_button.grid(row=0, column=0, columnspan=2, sticky="ew")
    report_box.insert("1.0", draft_report)

    journal_pet_transparent = "#00ff7f"
    journal_pet_window = tk.Toplevel(root)
    journal_pet_window.overrideredirect(True)
    journal_pet_window.transient(root)
    try:
        journal_pet_window.attributes("-transparentcolor", journal_pet_transparent)
    except tk.TclError:
        pass
    journal_pet_window.withdraw()
    journal_pet_canvas = tk.Canvas(
        journal_pet_window,
        width=104,
        height=58,
        bg=journal_pet_transparent,
        bd=0,
        highlightthickness=0,
        cursor="hand2",
    )
    journal_pet_canvas.pack(fill="both", expand=True)

    _register_page_toggle(desktop_pet_page)
    desktop_pet_page.grid_rowconfigure(0, weight=1)
    desktop_pet_page.grid_columnconfigure(0, weight=1)
    pet_wrap = tk.Frame(desktop_pet_page, bg=t_init.surface)
    pet_wrap.grid(row=0, column=0, sticky="nsew", padx=28, pady=28)
    pet_wrap.grid_columnconfigure(0, weight=2)
    pet_wrap.grid_columnconfigure(1, weight=1)
    pet_wrap.grid_rowconfigure(2, weight=1)
    pet_title_lbl = tk.Label(
        pet_wrap,
        text=tr("pet.title"),
        bg=t_init.surface,
        fg=t_init.text,
        font=("Segoe UI", 18, "bold"),
        anchor="w",
    )
    pet_title_lbl.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))
    pet_body_lbl = tk.Label(
        pet_wrap,
        text=tr("pet.body"),
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 10),
        anchor="w",
        justify="left",
        wraplength=760,
    )
    pet_body_lbl.grid(row=1, column=0, columnspan=2, sticky="new", pady=(0, 16))
    pet_stage = tk.Canvas(
        pet_wrap,
        height=360,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    pet_stage.grid(row=2, column=0, sticky="nsew", padx=(0, 16), pady=(0, 0))
    pet_info_holder = tk.Frame(pet_wrap, bg=t_init.panel, highlightthickness=1, highlightbackground=t_init.border)
    pet_info_holder.grid(row=2, column=1, sticky="nsew")
    pet_info_holder.grid_rowconfigure(0, weight=1)
    pet_info_holder.grid_columnconfigure(0, weight=1)
    pet_info_canvas = tk.Canvas(
        pet_info_holder,
        bg=t_init.panel,
        bd=0,
        highlightthickness=0,
        yscrollincrement=16,
    )
    pet_info_scroll = tk.Scrollbar(
        pet_info_holder,
        command=pet_info_canvas.yview,
        bg=t_init.panel,
        troughcolor=t_init.field,
        activebackground=t_init.accent,
        bd=0,
        highlightthickness=0,
        width=11,
    )
    pet_info_canvas.configure(yscrollcommand=pet_info_scroll.set)
    pet_info_canvas.grid(row=0, column=0, sticky="nsew")
    pet_info_scroll.grid(row=0, column=1, sticky="ns")
    pet_info = tk.Frame(pet_info_canvas, bg=t_init.panel)
    pet_info_window = pet_info_canvas.create_window((0, 0), window=pet_info, anchor="nw")
    pet_info.grid_columnconfigure(0, weight=1)

    def _sync_pet_info_scroll(_evt: Optional[Any] = None) -> None:
        try:
            pet_info_canvas.configure(scrollregion=pet_info_canvas.bbox("all"))
            pet_info_canvas.itemconfigure(pet_info_window, width=pet_info_canvas.winfo_width())
        except tk.TclError:
            pass

    def _pet_info_mousewheel(evt: Any) -> str:
        try:
            pet_info_canvas.yview_scroll(int(-1 * (evt.delta / 120)), "units")
        except tk.TclError:
            pass
        return "break"

    pet_info.bind("<Configure>", _sync_pet_info_scroll, add="+")
    pet_info_canvas.bind("<Configure>", _sync_pet_info_scroll, add="+")
    pet_info_canvas.bind("<MouseWheel>", _pet_info_mousewheel, add="+")
    pet_info.bind("<MouseWheel>", _pet_info_mousewheel, add="+")
    pet_level_lbl = tk.Label(
        pet_info,
        text="",
        bg=t_init.panel,
        fg=t_init.text,
        font=("Segoe UI", 16, "bold"),
        anchor="w",
    )
    pet_level_lbl.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 4))
    pet_xp_lbl = tk.Label(
        pet_info,
        text="",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 10),
        anchor="w",
        justify="left",
    )
    pet_xp_lbl.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 10))
    pet_mood_lbl = tk.Label(
        pet_info,
        text="",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 10),
        anchor="w",
        justify="left",
        wraplength=260,
    )
    pet_mood_lbl.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 14))
    pet_name_frame = tk.Frame(pet_info, bg=t_init.panel, bd=0, highlightthickness=0)
    pet_name_frame.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 10))
    pet_name_frame.grid_columnconfigure(1, weight=1)
    pet_name_lbl = tk.Label(
        pet_name_frame,
        text=tr("pet.name_label"),
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 9, "bold"),
        anchor="w",
    )
    pet_name_lbl.grid(row=0, column=0, sticky="w", padx=(0, 8))
    pet_name_var = tk.StringVar(value="")
    pet_name_entry = tk.Entry(
        pet_name_frame,
        textvariable=pet_name_var,
        bg=t_init.field,
        fg=t_init.text,
        disabledbackground=t_init.btn_disabled,
        disabledforeground=t_init.disabled_fg,
        insertbackground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9),
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    pet_name_entry.grid(row=0, column=1, sticky="ew", padx=(0, 6))
    pet_name_save_btn = tk.Button(
        pet_name_frame,
        text=tr("pet.name_save"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=4,
        cursor="hand2",
    )
    pet_name_save_btn.grid(row=0, column=2, sticky="ns")
    pet_feed_btn = tk.Button(
        pet_info,
        text=tr("pet.feed_now"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    pet_feed_btn.grid(row=4, column=0, sticky="ew", padx=16, pady=(0, 8))
    pet_art_btn = tk.Button(
        pet_info,
        text=tr("pet.generate_look"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    pet_art_btn.grid(row=5, column=0, sticky="ew", padx=16, pady=(0, 8))
    pet_art_ref_btn = tk.Button(
        pet_info,
        text=tr("pet.generate_from_image"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    pet_art_ref_btn.grid(row=6, column=0, sticky="ew", padx=16, pady=(0, 8))
    pet_desktop_btn = tk.Button(
        pet_info,
        text=tr("pet.desktop_mode"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    pet_desktop_btn.grid(row=7, column=0, sticky="ew", padx=16, pady=(0, 8))
    pet_journal_btn = tk.Button(
        pet_info,
        text=tr("pet.journal_hide"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    pet_journal_btn.grid(row=8, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_profile_frame = tk.Frame(pet_info, bg=t_init.panel, bd=0, highlightthickness=0)
    pet_profile_frame.grid(row=9, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_profile_frame.grid_columnconfigure(0, weight=1)
    pet_profile_frame.grid_columnconfigure(1, weight=1)
    pet_export_btn = tk.Button(
        pet_profile_frame,
        text=tr("pet.export_profile"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
        cursor="hand2",
    )
    pet_export_btn.grid(row=0, column=0, sticky="ew", padx=(0, 4), pady=(0, 6))
    pet_import_btn = tk.Button(
        pet_profile_frame,
        text=tr("pet.import_profile"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
        cursor="hand2",
    )
    pet_import_btn.grid(row=0, column=1, sticky="ew", padx=(4, 0), pady=(0, 6))
    pet_folder_btn = tk.Button(
        pet_profile_frame,
        text=tr("pet.open_folder"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
        cursor="hand2",
    )
    pet_folder_btn.grid(row=1, column=0, columnspan=2, sticky="ew")
    pet_delete_btn = tk.Button(
        pet_profile_frame,
        text=tr("pet.delete_profile"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
        cursor="hand2",
    )
    pet_delete_btn.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(6, 0))
    pet_modules_lbl = tk.Label(
        pet_info,
        text=tr("pet.modules_title"),
        bg=t_init.panel,
        fg=t_init.text,
        font=("Segoe UI", 11, "bold"),
        anchor="w",
    )
    pet_modules_lbl.grid(row=10, column=0, sticky="ew", padx=16, pady=(0, 8))
    pet_bubble_module = tk.Frame(
        pet_info,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
    )
    pet_bubble_module.grid(row=11, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_bubble_module.grid_columnconfigure(0, weight=1)
    pet_bubble_title_lbl = tk.Label(
        pet_bubble_module,
        text=tr("pet.module_bubble_title"),
        bg=t_init.field,
        fg=t_init.text,
        font=("Segoe UI", 10, "bold"),
        anchor="w",
    )
    pet_bubble_title_lbl.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 2))
    pet_bubble_status_lbl = tk.Label(
        pet_bubble_module,
        text="",
        bg=t_init.field,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
        wraplength=230,
    )
    pet_bubble_status_lbl.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
    pet_bubble_btn = tk.Button(
        pet_bubble_module,
        text="",
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    pet_bubble_btn.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))
    pet_memory_module = tk.Frame(
        pet_info,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
    )
    pet_memory_module.grid(row=12, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_memory_module.grid_columnconfigure(0, weight=1)
    pet_memory_title_lbl = tk.Label(
        pet_memory_module,
        text=tr("pet.module_memory_title"),
        bg=t_init.field,
        fg=t_init.text,
        font=("Segoe UI", 10, "bold"),
        anchor="w",
    )
    pet_memory_title_lbl.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 2))
    pet_memory_status_lbl = tk.Label(
        pet_memory_module,
        text="",
        bg=t_init.field,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
        wraplength=230,
    )
    pet_memory_status_lbl.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
    pet_memory_btn = tk.Button(
        pet_memory_module,
        text="",
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    pet_memory_btn.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))
    pet_tts_module = tk.Frame(
        pet_info,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
    )
    pet_tts_module.grid(row=13, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_tts_module.grid_columnconfigure(0, weight=1)
    pet_tts_title_lbl = tk.Label(
        pet_tts_module,
        text=tr("pet.module_tts_title"),
        bg=t_init.field,
        fg=t_init.text,
        font=("Segoe UI", 10, "bold"),
        anchor="w",
    )
    pet_tts_title_lbl.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 2))
    pet_tts_status_lbl = tk.Label(
        pet_tts_module,
        text="",
        bg=t_init.field,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
        wraplength=230,
    )
    pet_tts_status_lbl.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
    pet_tts_btn = tk.Button(
        pet_tts_module,
        text="",
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    pet_tts_btn.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 10))
    pet_chat_module = tk.Frame(
        pet_info,
        bg=t_init.field,
        highlightthickness=1,
        highlightbackground=t_init.border,
    )
    pet_chat_module.grid(row=14, column=0, sticky="ew", padx=16, pady=(0, 16))
    pet_chat_module.grid_columnconfigure(0, weight=1)
    pet_chat_title_lbl = tk.Label(
        pet_chat_module,
        text=tr("pet.module_chat_title"),
        bg=t_init.field,
        fg=t_init.text,
        font=("Segoe UI", 10, "bold"),
        anchor="w",
    )
    pet_chat_title_lbl.grid(row=0, column=0, sticky="ew", padx=10, pady=(8, 2))
    pet_chat_status_lbl = tk.Label(
        pet_chat_module,
        text="",
        bg=t_init.field,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
        wraplength=230,
    )
    pet_chat_status_lbl.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
    pet_chat_btn = tk.Button(
        pet_chat_module,
        text="",
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    pet_chat_btn.grid(row=2, column=0, sticky="ew", padx=10, pady=(0, 8))
    pet_chat_entry_row = tk.Frame(pet_chat_module, bg=t_init.field)
    pet_chat_entry_row.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))
    pet_chat_entry_row.grid_columnconfigure(0, weight=1)
    pet_chat_entry = tk.Entry(
        pet_chat_entry_row,
        bg=t_init.panel,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9),
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    pet_chat_entry.grid(row=0, column=0, sticky="ew", padx=(0, 6))
    pet_chat_ask_btn = tk.Button(
        pet_chat_entry_row,
        text=tr("pet.chat_ask"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=4,
        cursor="hand2",
    )
    pet_chat_ask_btn.grid(row=0, column=1, sticky="ns")

    def _bind_pet_info_mousewheel_recursive(widget: Any) -> None:
        try:
            widget.bind("<MouseWheel>", _pet_info_mousewheel, add="+")
            for child in widget.winfo_children():
                _bind_pet_info_mousewheel_recursive(child)
        except tk.TclError:
            pass

    root.after_idle(lambda: _bind_pet_info_mousewheel_recursive(pet_info_holder))

    def _safe_companion_profile_name(name: str) -> str:
        clean = re.sub(r"[^A-Za-z0-9 _.-]+", "", str(name or "Companion")).strip(" ._-")
        return clean[:48] or "Companion"

    def _default_companion_state(*, active: bool = False) -> Dict[str, Any]:
        return {
            "active": active,
            "name": "Companion",
            "profile_id": "",
            "created_at": "",
            "xp": 0,
            "fed_units": 0,
            "desktop_enabled": False,
            "journal_visible": True,
            "desktop_x": 0,
            "desktop_y": 0,
            "journal_x": 0,
            "journal_y": 0,
            "stage_x": 0,
            "stage_y": 0,
            "image_path": "",
            "modules": {"bubble": False, "memory": False, "tts": False, "chat": False},
            "downloads": {"bubble": False, "memory": False, "tts": False, "chat": False},
            "memory": {},
            "last_gain": 0,
        }

    def _read_json_file(path: Path) -> Dict[str, Any]:
        try:
            parsed = json.loads(path.read_text(encoding="utf-8"))
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}

    def _write_json_atomic(path: Path, data: Dict[str, Any]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(path)

    def _normalize_companion_state(data: Dict[str, Any], *, active: bool = True) -> Dict[str, Any]:
        state = _default_companion_state(active=active)
        state["name"] = str(data.get("name") or "Companion").strip() or "Companion"
        state["profile_id"] = str(data.get("profile_id") or "").strip()
        state["created_at"] = str(data.get("created_at") or "").strip()
        for key in ("xp", "fed_units", "desktop_x", "desktop_y", "journal_x", "journal_y", "stage_x", "stage_y"):
            try:
                state[key] = max(0, int(data.get(key, 0) or 0))
            except (TypeError, ValueError):
                state[key] = 0
        state["desktop_enabled"] = bool(data.get("desktop_enabled", False))
        state["journal_visible"] = bool(data.get("journal_visible", True))
        image_path = str(data.get("image_path") or "").strip()
        state["image_path"] = image_path if image_path and Path(image_path).exists() else ""
        modules_raw = data.get("modules") if isinstance(data.get("modules"), dict) else {}
        downloads_raw = data.get("downloads") if isinstance(data.get("downloads"), dict) else {}
        modules: Dict[str, bool] = {}
        downloads: Dict[str, bool] = {}
        for module_id in ("bubble", "memory", "tts", "chat"):
            modules[module_id] = bool(modules_raw.get(module_id, False)) if isinstance(modules_raw, dict) else False
            downloads[module_id] = (
                bool(downloads_raw.get(module_id, False) or modules[module_id])
                if isinstance(downloads_raw, dict)
                else modules[module_id]
            )
        state["modules"] = modules
        state["downloads"] = downloads
        memory_raw = data.get("memory") if isinstance(data.get("memory"), dict) else {}
        memory: Dict[str, int] = {}
        if isinstance(memory_raw, dict):
            for key, value in memory_raw.items():
                term = str(key or "").strip()[:32]
                if not term:
                    continue
                try:
                    count = max(1, int(value))
                except (TypeError, ValueError):
                    continue
                memory[term] = min(999, count)
        state["memory"] = memory
        return state

    def _profile_dir_has_companion(profile_dir: Path) -> bool:
        if not profile_dir.exists() or not profile_dir.is_dir():
            return False
        if (profile_dir / "profile.json").exists():
            return True
        return any(profile_dir.glob("*.png"))

    def _load_companion_state_from_dir(profile_dir: Path) -> Optional[Dict[str, Any]]:
        if not _profile_dir_has_companion(profile_dir):
            return None
        profile = _read_json_file(profile_dir / "profile.json")
        stats = _read_json_file(profile_dir / "stats.json")
        memory = _read_json_file(profile_dir / "memory.json")
        data: Dict[str, Any] = {}
        data.update(stats)
        data.update(profile)
        data["memory"] = memory
        image_file = str(profile.get("image_file") or "").strip()
        image_path = profile_dir / image_file if image_file else Path()
        if not image_file or not image_path.exists():
            images = sorted(profile_dir.glob("*.png"))
            image_path = images[0] if images else Path()
        if image_path and image_path.exists():
            data["image_path"] = str(image_path)
        state = _normalize_companion_state(data, active=True)
        state["profile_dir"] = str(profile_dir)
        return state

    def _write_companion_state_to_dir(profile_dir: Path, state: Dict[str, Any]) -> None:
        profile_dir.mkdir(parents=True, exist_ok=True)
        image_file = ""
        image_path = Path(str(state.get("image_path") or ""))
        if image_path.exists():
            target = profile_dir / "companion.png"
            try:
                if image_path.resolve() != target.resolve():
                    shutil.copy2(image_path, target)
                image_file = target.name
                state["image_path"] = str(target)
            except Exception:
                image_file = image_path.name
        now = datetime.now().isoformat(timespec="seconds")
        profile = {
            "profile_version": 1,
            "profile_id": str(state.get("profile_id") or secrets.token_hex(8)),
            "name": str(state.get("name") or "Companion"),
            "image_file": image_file,
            "created_at": str(state.get("created_at") or now),
            "updated_at": now,
        }
        stats = {
            "xp": int(state.get("xp", 0) or 0),
            "fed_units": int(state.get("fed_units", 0) or 0),
            "desktop_enabled": bool(state.get("desktop_enabled", False)),
            "journal_visible": bool(state.get("journal_visible", True)),
            "desktop_x": int(state.get("desktop_x", 0) or 0),
            "desktop_y": int(state.get("desktop_y", 0) or 0),
            "journal_x": int(state.get("journal_x", 0) or 0),
            "journal_y": int(state.get("journal_y", 0) or 0),
            "stage_x": int(state.get("stage_x", 0) or 0),
            "stage_y": int(state.get("stage_y", 0) or 0),
            "modules": dict(state.get("modules") or {}),
            "downloads": dict(state.get("downloads") or {}),
        }
        _write_json_atomic(profile_dir / "profile.json", profile)
        _write_json_atomic(profile_dir / "stats.json", stats)
        _write_json_atomic(profile_dir / "memory.json", dict(state.get("memory") or {}))
        state["profile_id"] = profile["profile_id"]
        state["created_at"] = profile["created_at"]
        state["profile_dir"] = str(profile_dir)

    def _load_legacy_companion_state_from_preferences() -> Optional[Dict[str, Any]]:
        raw = load_preferences().get(DESKTOP_PET_STATE_PREF_KEY, "")
        if not raw:
            return None
        try:
            parsed = json.loads(str(raw))
            if not isinstance(parsed, dict):
                return None
        except Exception:
            return None
        state = _normalize_companion_state(parsed, active=True)
        if not any(
            [
                state.get("image_path"),
                int(state.get("xp", 0) or 0),
                int(state.get("fed_units", 0) or 0),
                bool(state.get("memory")),
                any(dict(state.get("modules") or {}).values()),
            ]
        ):
            return None
        return state

    def _load_desktop_pet_state() -> Dict[str, Any]:
        COMPANION_ROOT_DIR.mkdir(parents=True, exist_ok=True)
        current_state = _load_companion_state_from_dir(COMPANION_CURRENT_DIR)
        if current_state is not None:
            return current_state
        legacy_state = _load_legacy_companion_state_from_preferences()
        if legacy_state is not None:
            _write_companion_state_to_dir(COMPANION_CURRENT_DIR, legacy_state)
            return legacy_state
        return _default_companion_state(active=False)

    pet_state = (
        _load_desktop_pet_state()
        if COMPANION_FEATURE_ENABLED
        else _default_companion_state(active=False)
    )
    pet_anim_state = {"tick": 0, "x": 42.0, "direction": 1}
    pet_art_busy = {"v": False}
    pet_chat_busy = {"v": False}
    pet_bubble_state = {"text": "", "until": 0.0, "last_auto": 0.0}
    pet_tts_state: Dict[str, Any] = {"proc": None, "last": 0.0, "error": ""}
    pet_image_cache: Dict[str, Any] = {}
    journal_pet_state: Dict[str, Any] = {
        "dragging": False,
        "drag_dx": 0,
        "drag_dy": 0,
        "x": 0.58,
        "y": 0.74,
        "direction": 1,
        "tick": 0,
        "handoff": False,
        "hitbox": (0.0, 0.0, 0.0, 0.0),
    }
    pet_stage_drag_state: Dict[str, Any] = {
        "dragging": False,
        "active": False,
        "dx": 0,
        "dy": 0,
        "x": 42.0,
        "y": 0.0,
        "handoff": False,
        "hitbox": (0.0, 0.0, 0.0, 0.0),
    }
    pet_overlay_state: Dict[str, Any] = {
        "window": None,
        "canvas": None,
        "dragging": False,
        "drag_dx": 0,
        "drag_dy": 0,
        "x": int(pet_state.get("desktop_x", 0) or 0),
        "y": int(pet_state.get("desktop_y", 0) or 0),
        "direction": 1,
        "falling": False,
        "vy": 0.0,
        "tick": 0,
    }

    def _save_desktop_pet_state() -> None:
        if not bool(pet_state.get("active", False)):
            prefs = load_preferences()
            prefs[DESKTOP_PET_STATE_PREF_KEY] = json.dumps({"active": False})
            save_preferences(prefs)
            return
        _write_companion_state_to_dir(COMPANION_CURRENT_DIR, pet_state)
        prefs = load_preferences()
        prefs[DESKTOP_PET_STATE_PREF_KEY] = json.dumps(
            {
                "active": True,
                "name": str(pet_state.get("name") or "Companion"),
                "xp": int(pet_state.get("xp", 0) or 0),
                "fed_units": int(pet_state.get("fed_units", 0) or 0),
                "desktop_enabled": bool(pet_state.get("desktop_enabled", False)),
                "journal_visible": bool(pet_state.get("journal_visible", True)),
                "desktop_x": int(pet_state.get("desktop_x", 0) or 0),
                "desktop_y": int(pet_state.get("desktop_y", 0) or 0),
                "journal_x": int(pet_state.get("journal_x", 0) or 0),
                "journal_y": int(pet_state.get("journal_y", 0) or 0),
                "stage_x": int(pet_state.get("stage_x", 0) or 0),
                "stage_y": int(pet_state.get("stage_y", 0) or 0),
                "image_path": str(pet_state.get("image_path") or ""),
                "modules": dict(pet_state.get("modules") or {}),
                "downloads": dict(pet_state.get("downloads") or {}),
                "memory": dict(pet_state.get("memory") or {}),
            }
        )
        save_preferences(prefs)

    def _archive_current_companion(*, copy_only: bool = False) -> Optional[Path]:
        if not _profile_dir_has_companion(COMPANION_CURRENT_DIR):
            return None
        loaded = _load_companion_state_from_dir(COMPANION_CURRENT_DIR) or pet_state
        base_name = _safe_companion_profile_name(str(loaded.get("name") or "Companion"))
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target = COMPANION_ROOT_DIR / base_name
        if target.exists():
            target = COMPANION_ROOT_DIR / f"{base_name}_{stamp}"
        COMPANION_ROOT_DIR.mkdir(parents=True, exist_ok=True)
        if copy_only:
            shutil.copytree(COMPANION_CURRENT_DIR, target)
        else:
            shutil.move(str(COMPANION_CURRENT_DIR), str(target))
        return target

    def _set_no_current_companion() -> None:
        pet_state.clear()
        pet_state.update(_default_companion_state(active=False))
        pet_image_cache.clear()
        _save_desktop_pet_state()
        _refresh_desktop_pet_ui()

    def _replace_current_companion_from_dir(source_dir: Path) -> Tuple[bool, str]:
        source_dir = Path(source_dir)
        if not _profile_dir_has_companion(source_dir):
            return False, tr("pet.import_invalid")
        if source_dir.resolve() == COMPANION_CURRENT_DIR.resolve():
            return True, tr("pet.import_already_current")
        try:
            _archive_current_companion(copy_only=False)
            if COMPANION_CURRENT_DIR.exists():
                shutil.rmtree(COMPANION_CURRENT_DIR)
            shutil.copytree(source_dir, COMPANION_CURRENT_DIR)
            loaded = _load_companion_state_from_dir(COMPANION_CURRENT_DIR)
            if loaded is None:
                return False, tr("pet.import_invalid")
            pet_state.clear()
            pet_state.update(loaded)
            pet_image_cache.clear()
            _save_desktop_pet_state()
            _refresh_desktop_pet_ui()
            return True, tr("pet.imported")
        except Exception as exc:
            return False, tr("pet.import_failed").format(error=str(exc))

    def _export_current_companion_profile() -> None:
        if not bool(pet_state.get("active", False)):
            messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
            return
        try:
            _save_desktop_pet_state()
            target = _archive_current_companion(copy_only=True)
            if target is None:
                messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
                return
            messagebox.showinfo(tr("pet.title"), tr("pet.exported").format(path=str(target)))
            _append_console_session_update(f"Companion exported: {target}", key=f"companion:export:{int(time.time())}")
        except Exception as exc:
            messagebox.showerror(tr("pet.title"), tr("pet.export_failed").format(error=str(exc))[:4000])

    def _import_companion_profile() -> None:
        if filedialog is None:
            messagebox.showerror(tr("pet.title"), tr("pet.import_unavailable"))
            return
        COMPANION_ROOT_DIR.mkdir(parents=True, exist_ok=True)
        selected = filedialog.askdirectory(
            title=tr("pet.import_picker_title"),
            initialdir=str(COMPANION_ROOT_DIR),
            mustexist=True,
        )
        if not selected:
            return
        ok, msg = _replace_current_companion_from_dir(Path(selected))
        if ok:
            messagebox.showinfo(tr("pet.title"), msg)
            _append_console_session_update(
                f"Companion imported: {selected}",
                key=f"companion:import:{int(time.time())}",
            )
        else:
            messagebox.showerror(tr("pet.title"), msg[:4000])

    def _open_companion_folder() -> None:
        try:
            COMPANION_ROOT_DIR.mkdir(parents=True, exist_ok=True)
            if os.name == "nt":
                os.startfile(str(COMPANION_ROOT_DIR))  # type: ignore[attr-defined]
            else:
                webbrowser.open(COMPANION_ROOT_DIR.as_uri())
        except Exception as exc:
            messagebox.showerror(tr("pet.title"), tr("pet.folder_open_failed").format(error=str(exc))[:4000])

    def _companion_is_fresh_default() -> bool:
        return (
            bool(pet_state.get("active", False))
            and str(pet_state.get("name") or "Companion").strip() == "Companion"
            and int(pet_state.get("xp", 0) or 0) <= 0
            and int(pet_state.get("fed_units", 0) or 0) <= 0
        )

    def _save_companion_name() -> None:
        if not bool(pet_state.get("active", False)):
            messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
            return
        name = re.sub(r"\s+", " ", pet_name_var.get()).strip() or "Companion"
        pet_state["name"] = name[:48]
        _save_desktop_pet_state()
        _refresh_desktop_pet_ui()
        pet_mood_lbl.config(text=tr("pet.name_saved"))
        _append_console_session_update(f"Companion renamed: {pet_state['name']}", key=f"companion:rename:{int(time.time())}")

    def _delete_current_companion() -> None:
        try:
            _destroy_pet_overlay()
            try:
                journal_pet_window.withdraw()
            except tk.TclError:
                pass
            if COMPANION_CURRENT_DIR.exists():
                shutil.rmtree(COMPANION_CURRENT_DIR)
            _set_no_current_companion()
            _append_console_session_update("Current companion deleted.", key=f"companion:delete:{int(time.time())}")
        except Exception as exc:
            messagebox.showerror(tr("pet.title"), tr("pet.delete_failed").format(error=str(exc))[:4000])

    def _desktop_pet_level() -> Tuple[int, int, int]:
        xp = max(0, int(pet_state.get("xp", 0) or 0))
        level = (xp // 100) + 1
        return level, xp % 100, 100

    def _pet_modules() -> Dict[str, bool]:
        modules = pet_state.get("modules")
        if not isinstance(modules, dict):
            modules = {}
            pet_state["modules"] = modules
        return modules  # type: ignore[return-value]

    def _pet_downloads() -> Dict[str, bool]:
        downloads = pet_state.get("downloads")
        if not isinstance(downloads, dict):
            downloads = {}
            pet_state["downloads"] = downloads
        return downloads  # type: ignore[return-value]

    def _pet_bubble_installed() -> bool:
        return bool(_pet_modules().get("bubble", False)) and _pet_module_downloaded("bubble")

    def _pet_memory_unlocked() -> bool:
        level, _, _ = _desktop_pet_level()
        return level >= DESKTOP_PET_MEMORY_UNLOCK_LEVEL

    def _pet_memory_installed() -> bool:
        return bool(_pet_modules().get("memory", False)) and _pet_memory_unlocked() and _pet_module_downloaded("memory")

    def _pet_tts_unlocked() -> bool:
        level, _, _ = _desktop_pet_level()
        return level >= DESKTOP_PET_TTS_UNLOCK_LEVEL

    def _pet_tts_installed() -> bool:
        return bool(_pet_modules().get("tts", False)) and _pet_tts_unlocked() and _pet_module_downloaded("tts")

    def _pet_chat_unlocked() -> bool:
        level, _, _ = _desktop_pet_level()
        return level >= DESKTOP_PET_CHAT_UNLOCK_LEVEL

    def _pet_chat_installed() -> bool:
        return bool(_pet_modules().get("chat", False)) and _pet_chat_unlocked() and _pet_module_downloaded("chat")

    def _pet_reactions_enabled() -> bool:
        return _pet_bubble_installed() or _pet_tts_installed() or _pet_chat_installed()

    def _pet_module_unlocked(module_id: str) -> bool:
        if module_id == "bubble":
            return True
        if module_id == "memory":
            return _pet_memory_unlocked()
        if module_id == "tts":
            return _pet_tts_unlocked()
        if module_id == "chat":
            return _pet_chat_unlocked()
        return False

    def _pet_module_downloaded(module_id: str) -> bool:
        modules = _pet_modules()
        downloads = _pet_downloads()
        if bool(modules.get(module_id, False)):
            downloads[module_id] = True
            return True
        manifest = DESKTOP_PET_MODULE_DIR / module_id / "module.json"
        if manifest.exists():
            downloads[module_id] = True
            return True
        return bool(downloads.get(module_id, False))

    def _pet_module_title(module_id: str) -> str:
        if module_id == "bubble":
            return tr("pet.module_bubble_title")
        if module_id == "memory":
            return tr("pet.module_memory_title")
        if module_id == "tts":
            return tr("pet.module_tts_title")
        if module_id == "chat":
            return tr("pet.module_chat_title")
        return module_id

    def _download_pet_module(module_id: str) -> None:
        if module_id not in {"bubble", "memory", "tts", "chat"}:
            return
        if not _pet_module_unlocked(module_id):
            if module_id == "tts":
                msg = tr("pet.module_tts_locked").format(level=DESKTOP_PET_TTS_UNLOCK_LEVEL)
            elif module_id == "chat":
                msg = tr("pet.module_chat_locked").format(level=DESKTOP_PET_CHAT_UNLOCK_LEVEL)
            elif module_id == "memory":
                msg = tr("pet.module_memory_locked").format(level=DESKTOP_PET_MEMORY_UNLOCK_LEVEL)
            else:
                msg = tr("pet.module_locked")
            _set_pet_bubble(msg, seconds=7, force=True)
            return
        try:
            module_dir = DESKTOP_PET_MODULE_DIR / module_id
            module_dir.mkdir(parents=True, exist_ok=True)
            manifest = {
                "id": module_id,
                "title": _pet_module_title(module_id),
                "version": 1,
                "downloaded_at": datetime.now().isoformat(timespec="seconds"),
            }
            manifest_path = module_dir / "module.json"
            tmp_path = module_dir / "module.tmp"
            tmp_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
            tmp_path.replace(manifest_path)
            _pet_downloads()[module_id] = True
            _save_desktop_pet_state()
            _append_console_session_update(
                f"Companion module downloaded: {module_id}",
                key=f"pet:module:{module_id}:downloaded",
            )
            _refresh_pet_module_ui()
            _set_pet_bubble(
                tr("pet.module_downloaded_say").format(module=_pet_module_title(module_id)),
                seconds=6,
                force=True,
            )
        except Exception as exc:
            messagebox.showerror(
                tr("pet.title"),
                tr("pet.module_download_failed").format(error=str(exc))[:4000],
            )

    def _speak_pet_text(text: str, *, force: bool = False) -> Optional[str]:
        if not force and not _pet_tts_installed():
            return None
        if os.name != "nt":
            return tr("pet.tts_windows_only")
        clean = re.sub(r"\s+", " ", text or "").strip()
        if not clean:
            return None
        now_ts = time.time()
        if not force and now_ts - float(pet_tts_state.get("last", 0.0) or 0.0) < 3.0:
            return None
        pet_tts_state["last"] = now_ts
        clean = clean[:220]
        try:
            old_proc = pet_tts_state.get("proc")
            if isinstance(old_proc, subprocess.Popen) and old_proc.poll() is None:
                old_proc.terminate()
        except Exception:
            pass
        try:
            text_b64 = base64.b64encode(clean.encode("utf-8")).decode("ascii")
            script = (
                "Add-Type -AssemblyName System.Speech;"
                "$s = New-Object System.Speech.Synthesis.SpeechSynthesizer;"
                "$s.Volume = 70;"
                "$s.Rate = 1;"
                f"$text = [Text.Encoding]::UTF8.GetString([Convert]::FromBase64String('{text_b64}'));"
                "$s.Speak($text);"
                "$s.Dispose();"
            )
            encoded = base64.b64encode(script.encode("utf-16le")).decode("ascii")
            kwargs: Dict[str, Any] = {
                "stdin": subprocess.DEVNULL,
                "stdout": subprocess.DEVNULL,
                "stderr": subprocess.DEVNULL,
            }
            if os.name == "nt":
                kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0)
            pet_tts_state["proc"] = subprocess.Popen(
                [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-EncodedCommand",
                    encoded,
                ],
                **kwargs,
            )
            pet_tts_state["error"] = ""
            return None
        except Exception as exc:
            err = tr("pet.tts_failed").format(error=str(exc))
            pet_tts_state["error"] = err
            return err

    def _active_pet_bubble_text() -> str:
        if not _pet_reactions_enabled():
            return ""
        if time.time() > float(pet_bubble_state.get("until", 0.0) or 0.0):
            return ""
        return str(pet_bubble_state.get("text") or "").strip()

    def _set_pet_bubble(text: str, *, seconds: float = 6.0, force: bool = False) -> None:
        if not force and not _pet_reactions_enabled():
            return
        clean = re.sub(r"\s+", " ", text or "").strip()
        if not clean:
            return
        pet_bubble_state["text"] = clean[:120]
        pet_bubble_state["until"] = time.time() + max(1.0, float(seconds))
        tts_err = _speak_pet_text(clean)
        if tts_err:
            _append_console_session_update(tts_err, key=f"pet:tts:error:{int(time.time())}")
        try:
            _draw_desktop_pet()
            _draw_journal_pet()
            _draw_pet_overlay()
        except Exception:
            pass

    def _top_pet_memory_terms(limit: int = 3) -> List[str]:
        memory = pet_state.get("memory")
        if not isinstance(memory, dict):
            return []
        ranked = sorted(
            ((str(k), int(v)) for k, v in memory.items() if str(k).strip()),
            key=lambda item: (-item[1], item[0].lower()),
        )
        return [term for term, _count in ranked[: max(0, int(limit))]]

    def _update_pet_memory_from_entry(*parts: str) -> None:
        combined = " ".join(str(part or "") for part in parts)
        if not combined.strip():
            return
        common_words = {
            "about",
            "after",
            "again",
            "also",
            "because",
            "before",
            "could",
            "daily",
            "entry",
            "from",
            "have",
            "journal",
            "just",
            "like",
            "more",
            "need",
            "notes",
            "that",
            "then",
            "there",
            "they",
            "this",
            "today",
            "want",
            "when",
            "with",
            "would",
            "you",
        }
        found: Dict[str, int] = {}
        for token in re.findall(r"[A-Za-z][A-Za-z0-9_+-]{3,}|[\u4e00-\u9fff]{2,8}", combined):
            term = token.strip()
            if not term:
                continue
            key = term.lower() if term.isascii() else term
            if key in common_words:
                continue
            found[key] = found.get(key, 0) + 1
        if not found:
            return
        memory = pet_state.get("memory")
        if not isinstance(memory, dict):
            memory = {}
            pet_state["memory"] = memory
        for term, count in found.items():
            try:
                memory[term] = min(999, int(memory.get(term, 0) or 0) + count)
            except Exception:
                memory[term] = count
        trimmed = sorted(
            ((str(k), int(v)) for k, v in memory.items() if str(k).strip()),
            key=lambda item: (-item[1], item[0].lower()),
        )[:18]
        pet_state["memory"] = {term: count for term, count in trimmed}
        _save_desktop_pet_state()

    def _pet_memory_hint() -> str:
        if not _pet_memory_installed():
            return ""
        terms = _top_pet_memory_terms(2)
        if not terms:
            return ""
        return tr("pet.memory_hint").format(top=", ".join(terms))

    def _pet_growth_profile(level: Optional[int] = None) -> Dict[str, str]:
        if level is None:
            level, _, _ = _desktop_pet_level()
        pet_level = max(1, int(level or 1))
        if pet_level >= 10:
            return {
                "key": "wise",
                "label": tr("pet.stage_wise"),
                "chat": (
                    "You are now a wise journal companion. You may synthesize patterns from memory "
                    "and offer one gentle, practical next step when helpful."
                ),
                "reaction": tr("pet.stage_wise_reaction"),
            }
        if pet_level >= 7:
            return {
                "key": "insightful",
                "label": tr("pet.stage_insightful"),
                "chat": (
                    "You are becoming insightful. Connect the user's question to current context or "
                    "memory when it is genuinely relevant, and keep the tone warm."
                ),
                "reaction": tr("pet.stage_insightful_reaction"),
            }
        if pet_level >= 4:
            return {
                "key": "curious",
                "label": tr("pet.stage_curious"),
                "chat": (
                    "You are curious and supportive. Ask at most one small clarifying question only "
                    "when it would help the user think."
                ),
                "reaction": tr("pet.stage_curious_reaction"),
            }
        if pet_level >= 2:
            return {
                "key": "learning",
                "label": tr("pet.stage_learning"),
                "chat": (
                    "You are still learning. Keep replies simple, cheerful, and concrete. "
                    "Do not over-analyze."
                ),
                "reaction": tr("pet.stage_learning_reaction"),
            }
        return {
            "key": "hatchling",
            "label": tr("pet.stage_hatchling"),
            "chat": (
                "You are a tiny new companion. Reply very simply and cutely, like a supportive pet "
                "that is just learning the user's world."
            ),
            "reaction": tr("pet.stage_hatchling_reaction"),
        }

    def _pet_chat_context() -> str:
        def _clip(label: str, value: str, limit: int) -> str:
            clean = re.sub(r"\s+", " ", value or "").strip()
            if not clean:
                return ""
            if len(clean) > limit:
                clean = clean[:limit].rstrip() + "..."
            return f"{label}: {clean}"

        level, _, _ = _desktop_pet_level()
        profile = _pet_growth_profile(level)
        memory_terms = _top_pet_memory_terms(8) if _pet_memory_installed() else []
        parts = [
            f"Pet level: {level}, XP: {current_xp}/{needed_xp}",
            f"Pet intelligence stage: {profile['label']}",
            (
                f"Local memory topics: {', '.join(memory_terms) if memory_terms else '(none yet)'}"
                if _pet_memory_installed()
                else "Local memory topics: (Memory module not installed)"
            ),
        ]
        for label, widget, limit in (
            ("Current journal draft", text_box, 900),
            ("Current speech-to-text draft", stt_box, 650),
            ("Current AI report draft", report_box, 650),
        ):
            try:
                item = _clip(label, widget.get("1.0", "end-1c"), limit)
            except tk.TclError:
                item = ""
            if item:
                parts.append(item)
        return "\n".join(parts)

    def _refresh_pet_chat_controls() -> None:
        t = th()
        ready = bool(pet_state.get("active", False)) and _pet_chat_installed() and not bool(pet_chat_busy.get("v", False))
        state = "normal" if ready else "disabled"
        pet_chat_entry.config(
            state=state,
            bg=t.panel if ready else t.btn_disabled,
            fg=t.text if ready else t.disabled_fg,
            disabledforeground=t.disabled_fg,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        pet_chat_ask_btn.config(
            text=tr("pet.chat_thinking") if pet_chat_busy.get("v") else tr("pet.chat_ask"),
            state=state,
            bg=t.btn_secondary if ready else t.btn_disabled,
            fg=t.text if ready else t.disabled_fg,
            disabledforeground=t.disabled_fg,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2" if ready else "arrow",
        )

    def _ask_pet_chat() -> None:
        if bool(pet_chat_busy.get("v", False)):
            return
        if not _pet_chat_installed():
            if not _pet_chat_unlocked():
                _set_pet_bubble(
                    tr("pet.module_chat_locked").format(level=DESKTOP_PET_CHAT_UNLOCK_LEVEL),
                    seconds=7,
                    force=True,
                )
            return
        question = re.sub(r"\s+", " ", pet_chat_entry.get()).strip()
        if not question:
            messagebox.showinfo(tr("pet.title"), tr("pet.chat_empty"))
            return
        if not get_openai_api_key():
            messagebox.showerror(tr("pet.title"), tr("pet.chat_no_api_key"))
            return

        pet_chat_busy["v"] = True
        _refresh_pet_module_ui()
        _append_console_session_update("Companion chat started.", key=f"pet:chat:start:{int(time.time())}")

        context = _pet_chat_context()
        ui_language = "Chinese" if ui_lang_holder[0] == "zh" else "English"
        profile = _pet_growth_profile()
        messages: List[Dict[str, object]] = [
            {
                "role": "system",
                "content": (
                    "You are the user's small Q-style companion inside Daily Logger. "
                    "Answer warmly in no more than two short sentences. "
                    "Use the user's language when possible. Do not invent private facts. "
                    "Local memory topics and current draft text are hints, not complete truth. "
                    f"Current intelligence stage: {profile['label']}. {profile['chat']}"
                ),
            },
            {
                "role": "user",
                "content": (
                    f"Daily Logger UI language: {ui_language}\n"
                    f"{context}\n\n"
                    f"User asks their companion: {question}"
                ),
            },
        ]

        def _work() -> None:
            try:
                answer = chat_completion(messages, model=OPENAI_MODEL, timeout_sec=45, attempts=2)
            except Exception as exc:
                answer = f"ERROR: {exc}"
            root.after(0, lambda result=answer: _done(result))

        def _done(result: str) -> None:
            pet_chat_busy["v"] = False
            _refresh_pet_module_ui()
            if _is_likely_api_error_message_global(result) or result.startswith("ERROR:"):
                error_text = result[6:].strip() if result.startswith("ERROR:") else result
                msg = tr("pet.chat_failed").format(error=error_text)
                pet_mood_lbl.config(text=msg[:240])
                _append_console_session_update(
                    f"Companion chat failed: {error_text}",
                    key=f"pet:chat:fail:{int(time.time())}",
                )
                messagebox.showerror(tr("pet.title"), msg[:4000])
                return
            clean = re.sub(r"\s+", " ", result or "").strip()
            if len(clean) > 240:
                clean = clean[:240].rstrip() + "..."
            pet_chat_entry.delete(0, "end")
            _set_pet_bubble(clean, seconds=10, force=True)
            pet_mood_lbl.config(text=clean)
            _append_console_session_update(
                "Companion chat answered.",
                key=f"pet:chat:done:{int(time.time())}",
            )

        threading.Thread(target=_work, daemon=True).start()

    def _draw_pet_bubble(canvas: Any, x: float, y: float, text: str, *, width_chars: int = 24) -> None:
        clean = re.sub(r"\s+", " ", text or "").strip()
        if not clean:
            return
        lines = textwrap.wrap(clean, width=max(8, int(width_chars)))[:4]
        if not lines:
            return
        max_line = max(len(line) for line in lines)
        box_w = max(86, min(240, max_line * 7 + 22))
        box_h = 18 + (len(lines) * 16)
        x1 = float(x)
        y1 = float(y)
        x2 = x1 + box_w
        y2 = y1 + box_h
        canvas.create_rectangle(x1, y1, x2, y2, fill="#fff8d6", outline="#f2c94c", width=2)
        canvas.create_polygon(
            x1 + 20,
            y2,
            x1 + 34,
            y2,
            x1 + 24,
            y2 + 12,
            fill="#fff8d6",
            outline="#f2c94c",
        )
        for idx, line in enumerate(lines):
            canvas.create_text(
                x1 + 11,
                y1 + 12 + (idx * 16),
                text=line,
                fill="#2b2730",
                font=("Segoe UI", 8, "bold"),
                anchor="w",
            )

    def _refresh_pet_module_ui() -> None:
        t = th()

        def _button_style(btn: Any, enabled: bool) -> None:
            btn.config(
                state=("normal" if enabled else "disabled"),
                bg=(t.btn_secondary if enabled else t.btn_disabled),
                fg=(t.text if enabled else t.disabled_fg),
                disabledforeground=t.disabled_fg,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
                cursor=("hand2" if enabled else "arrow"),
            )

        if not bool(pet_state.get("active", False)):
            pet_bubble_title_lbl.config(text=tr("pet.module_bubble_title"))
            pet_memory_title_lbl.config(text=tr("pet.module_memory_title"))
            pet_tts_title_lbl.config(text=tr("pet.module_tts_title"))
            pet_chat_title_lbl.config(text=tr("pet.module_chat_title"))
            for status_lbl in (
                pet_bubble_status_lbl,
                pet_memory_status_lbl,
                pet_tts_status_lbl,
                pet_chat_status_lbl,
            ):
                status_lbl.config(text=tr("pet.no_current"))
            for btn in (pet_bubble_btn, pet_memory_btn, pet_tts_btn, pet_chat_btn):
                btn.config(text=tr("pet.module_download"))
                _button_style(btn, False)
            _refresh_pet_chat_controls()
            return

        bubble_installed = _pet_bubble_installed()
        bubble_downloaded = _pet_module_downloaded("bubble")
        pet_bubble_title_lbl.config(text=tr("pet.module_bubble_title"))
        if bubble_installed:
            pet_bubble_status_lbl.config(text=tr("pet.module_bubble_installed"))
            pet_bubble_btn.config(text=tr("pet.module_disable"))
        elif bubble_downloaded:
            pet_bubble_status_lbl.config(text=tr("pet.module_downloaded"))
            pet_bubble_btn.config(text=tr("pet.module_install"))
        else:
            pet_bubble_status_lbl.config(text=tr("pet.module_bubble_available"))
            pet_bubble_btn.config(text=tr("pet.module_download"))
        _button_style(pet_bubble_btn, True)

        memory_unlocked = _pet_memory_unlocked()
        memory_downloaded = _pet_module_downloaded("memory")
        memory_enabled = _pet_memory_installed()
        pet_memory_title_lbl.config(text=tr("pet.module_memory_title"))
        if memory_enabled:
            pet_memory_status_lbl.config(text=tr("pet.module_memory_installed"))
            pet_memory_btn.config(text=tr("pet.module_disable"))
        elif not memory_unlocked:
            pet_memory_status_lbl.config(
                text=tr("pet.module_memory_locked").format(level=DESKTOP_PET_MEMORY_UNLOCK_LEVEL)
            )
            pet_memory_btn.config(text=tr("pet.module_download"))
        elif memory_downloaded:
            pet_memory_status_lbl.config(text=tr("pet.module_downloaded"))
            pet_memory_btn.config(text=tr("pet.module_install"))
        else:
            pet_memory_status_lbl.config(text=tr("pet.module_memory_available"))
            pet_memory_btn.config(text=tr("pet.module_download"))
        _button_style(pet_memory_btn, memory_unlocked)

        tts_unlocked = _pet_tts_unlocked()
        tts_downloaded = _pet_module_downloaded("tts")
        tts_enabled = _pet_tts_installed()
        pet_tts_title_lbl.config(text=tr("pet.module_tts_title"))
        if tts_enabled:
            pet_tts_status_lbl.config(text=tr("pet.module_tts_installed"))
            pet_tts_btn.config(text=tr("pet.module_disable"))
        elif not tts_unlocked:
            pet_tts_status_lbl.config(
                text=tr("pet.module_tts_locked").format(level=DESKTOP_PET_TTS_UNLOCK_LEVEL)
            )
            pet_tts_btn.config(text=tr("pet.module_download"))
        elif tts_downloaded:
            pet_tts_status_lbl.config(text=tr("pet.module_downloaded"))
            pet_tts_btn.config(text=tr("pet.module_install"))
        else:
            pet_tts_status_lbl.config(text=tr("pet.module_tts_available"))
            pet_tts_btn.config(text=tr("pet.module_download"))
        _button_style(pet_tts_btn, tts_unlocked)

        chat_unlocked = _pet_chat_unlocked()
        chat_downloaded = _pet_module_downloaded("chat")
        chat_enabled = _pet_chat_installed()
        pet_chat_title_lbl.config(text=tr("pet.module_chat_title"))
        if chat_enabled:
            pet_chat_status_lbl.config(
                text=tr("pet.chat_thinking")
                if pet_chat_busy.get("v")
                else tr("pet.module_chat_installed")
            )
            pet_chat_btn.config(text=tr("pet.module_disable"))
        elif not chat_unlocked:
            pet_chat_status_lbl.config(
                text=tr("pet.module_chat_locked").format(level=DESKTOP_PET_CHAT_UNLOCK_LEVEL)
            )
            pet_chat_btn.config(text=tr("pet.module_download"))
        elif chat_downloaded:
            pet_chat_status_lbl.config(text=tr("pet.module_downloaded"))
            pet_chat_btn.config(text=tr("pet.module_install"))
        else:
            pet_chat_status_lbl.config(text=tr("pet.module_chat_available"))
            pet_chat_btn.config(text=tr("pet.module_download"))
        _button_style(pet_chat_btn, chat_unlocked and not bool(pet_chat_busy.get("v", False)))
        _refresh_pet_chat_controls()

    def _toggle_pet_bubble_module() -> None:
        if not _pet_module_downloaded("bubble"):
            _download_pet_module("bubble")
            return
        modules = _pet_modules()
        now_installed = not bool(modules.get("bubble", False))
        modules["bubble"] = now_installed
        _save_desktop_pet_state()
        _refresh_pet_module_ui()
        if now_installed:
            _set_pet_bubble(tr("pet.bubble_installed_say"), seconds=7, force=True)
        else:
            pet_bubble_state["until"] = 0.0
            _draw_desktop_pet()
            _draw_pet_overlay()

    def _toggle_pet_memory_module() -> None:
        if not _pet_memory_unlocked():
            _set_pet_bubble(
                tr("pet.module_memory_locked").format(level=DESKTOP_PET_MEMORY_UNLOCK_LEVEL),
                seconds=7,
                force=True,
            )
            return
        if not _pet_module_downloaded("memory"):
            _download_pet_module("memory")
            return
        modules = _pet_modules()
        now_installed = not bool(modules.get("memory", False))
        modules["memory"] = now_installed
        _save_desktop_pet_state()
        _refresh_pet_module_ui()
        if now_installed:
            _update_pet_memory_from_entry(
                text_box.get("1.0", "end-1c"),
                stt_box.get("1.0", "end-1c"),
                report_box.get("1.0", "end-1c"),
            )
            phrase = _pet_memory_hint() or tr("pet.memory_installed_say")
            _set_pet_bubble(phrase, seconds=8, force=True)

    def _toggle_pet_tts_module() -> None:
        if not _pet_tts_unlocked():
            _set_pet_bubble(
                tr("pet.module_tts_locked").format(level=DESKTOP_PET_TTS_UNLOCK_LEVEL),
                seconds=7,
                force=True,
            )
            return
        if not _pet_module_downloaded("tts"):
            _download_pet_module("tts")
            return
        modules = _pet_modules()
        now_installed = not bool(modules.get("tts", False))
        modules["tts"] = now_installed
        _save_desktop_pet_state()
        _refresh_pet_module_ui()
        if now_installed:
            phrase = tr("pet.tts_installed_say")
            _set_pet_bubble(phrase, seconds=7, force=True)
        else:
            try:
                old_proc = pet_tts_state.get("proc")
                if isinstance(old_proc, subprocess.Popen) and old_proc.poll() is None:
                    old_proc.terminate()
            except Exception:
                pass

    def _toggle_pet_chat_module() -> None:
        if bool(pet_chat_busy.get("v", False)):
            return
        if not _pet_chat_unlocked():
            _set_pet_bubble(
                tr("pet.module_chat_locked").format(level=DESKTOP_PET_CHAT_UNLOCK_LEVEL),
                seconds=7,
                force=True,
            )
            return
        if not _pet_module_downloaded("chat"):
            _download_pet_module("chat")
            return
        modules = _pet_modules()
        now_installed = not bool(modules.get("chat", False))
        modules["chat"] = now_installed
        _save_desktop_pet_state()
        _refresh_pet_module_ui()
        if now_installed:
            _set_pet_bubble(tr("pet.chat_installed_say"), seconds=8, force=True)

    def _current_pet_feed_units() -> int:
        text = " ".join(
            (
                text_box.get("1.0", "end-1c"),
                stt_box.get("1.0", "end-1c"),
                report_box.get("1.0", "end-1c"),
            )
        )
        compact = re.sub(r"\s+", "", text)
        return len(compact)

    pet_session_seen_units = {"v": _current_pet_feed_units()}

    def _add_desktop_pet_xp(amount: int, *, save: bool = True) -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        gain = max(0, int(amount))
        if gain <= 0:
            return
        old_level, _, _ = _desktop_pet_level()
        pet_state["xp"] = max(0, int(pet_state.get("xp", 0) or 0)) + gain
        pet_state["last_gain"] = gain
        if save:
            _save_desktop_pet_state()
        new_level, _, _ = _desktop_pet_level()
        if new_level > old_level:
            profile = _pet_growth_profile(new_level)
            _append_console_session_update(
                f"Companion reached level {new_level}: {profile['label']}.",
                key=f"pet:level:{new_level}",
            )
            _set_pet_bubble(
                tr("pet.bubble_level").format(level=new_level, stage=profile["label"]),
                seconds=8,
            )
        _refresh_desktop_pet_ui()

    def _feed_desktop_pet_from_current_entry(*, force: bool = False) -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        if not bool(pet_state.get("active", False)):
            return
        units = _current_pet_feed_units()
        delta = units - int(pet_session_seen_units.get("v", 0) or 0)
        if force:
            delta = max(delta, min(300, max(0, units // 3)))
        if delta < 60 and not force:
            return
        if delta <= 0:
            return
        pet_session_seen_units["v"] = max(units, int(pet_session_seen_units.get("v", 0) or 0))
        pet_state["fed_units"] = int(pet_state.get("fed_units", 0) or 0) + delta
        _add_desktop_pet_xp(max(1, delta // 18))
        if _pet_reactions_enabled() and time.time() - float(pet_bubble_state.get("last_auto", 0.0) or 0.0) > 18:
            pet_bubble_state["last_auto"] = time.time()
            _set_pet_bubble(tr("pet.bubble_writing"), seconds=5)

    def _feed_desktop_pet_on_save(journal_text: str, speech_text: str, report_text: str) -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        if not bool(pet_state.get("active", False)):
            return
        total_units = len(re.sub(r"\s+", "", f"{journal_text} {speech_text} {report_text}"))
        if total_units <= 0:
            return
        remaining_units = max(0, total_units - int(pet_session_seen_units.get("v", 0) or 0))
        if remaining_units:
            pet_session_seen_units["v"] = total_units
            pet_state["fed_units"] = int(pet_state.get("fed_units", 0) or 0) + remaining_units
        bonus = min(80, 8 + (total_units // 120) + (remaining_units // 24))
        _add_desktop_pet_xp(bonus)
        if _pet_memory_installed():
            _update_pet_memory_from_entry(journal_text, speech_text, report_text)
        memory_hint = _pet_memory_hint()
        _set_pet_bubble(
            memory_hint if memory_hint else tr("pet.bubble_saved"),
            seconds=8,
        )

    def _load_pet_photo(cache_name: str, max_size: Tuple[int, int]) -> Optional[Any]:
        image_path = str(pet_state.get("image_path") or "").strip()
        if not image_path:
            return None
        path = Path(image_path)
        if not path.exists():
            pet_state["image_path"] = ""
            _save_desktop_pet_state()
            return None
        try:
            mtime = path.stat().st_mtime
            key = f"{cache_name}:{path}:{mtime}:{max_size[0]}x{max_size[1]}"
            cached = pet_image_cache.get(key)
            if cached is not None:
                return cached
            from PIL import Image, ImageTk

            normalized = _normalize_desktop_pet_image_bytes(path.read_bytes())
            image = Image.open(io.BytesIO(normalized)).convert("RGBA")
            try:
                resample = Image.Resampling.LANCZOS
            except AttributeError:
                resample = Image.LANCZOS
            image.thumbnail(max_size, resample)
            photo = ImageTk.PhotoImage(image)
            stale = [k for k in pet_image_cache if k.startswith(f"{cache_name}:")]
            for old_key in stale:
                pet_image_cache.pop(old_key, None)
            pet_image_cache[key] = photo
            return photo
        except Exception:
            return None

    def _draw_desktop_pet() -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        t = th()
        pet_stage.delete("all")
        width = max(280, int(pet_stage.winfo_width() or 520))
        height = max(260, int(pet_stage.winfo_height() or 360))
        pet_stage.configure(bg=t.field, highlightbackground=t.border, highlightcolor=t.accent)
        if not bool(pet_state.get("active", False)):
            pet_stage.create_text(
                width / 2,
                height / 2,
                text=tr("pet.no_current_stage"),
                fill=t.muted,
                font=("Segoe UI", 12, "bold"),
                justify="center",
                width=max(220, width - 80),
            )
            pet_stage_drag_state["hitbox"] = (0.0, 0.0, 0.0, 0.0)
            return
        tick = int(pet_anim_state.get("tick", 0) or 0)
        if pet_stage_drag_state.get("dragging"):
            x = max(20.0, min(float(width - 120), float(pet_stage_drag_state.get("x", 42.0) or 42.0)))
        else:
            saved_x = int(pet_state.get("stage_x", 0) or 0)
            x = max(44.0, min(float(width - 118), float(saved_x or (width / 2 - 52))))
        pet_anim_state.update({"tick": tick + 1, "x": x, "direction": 1})
        level, current_xp, needed_xp = _desktop_pet_level()
        breath = 4 if (tick // 12) % 2 == 0 else 0
        blink = (tick % 80) in (0, 1, 2)
        y = (
            max(20.0, min(float(height - 132), float(pet_stage_drag_state.get("y", 0.0) or 0.0)))
            if pet_stage_drag_state.get("dragging")
            else max(20.0, min(float(height - 132), float(int(pet_state.get("stage_y", 0) or 0) or (height - 132))))
        )
        pet_stage_drag_state["hitbox"] = (x - 28, y, x + 132, y + 132)
        photo = _load_pet_photo("stage", (150 + breath, 150 - min(2, breath)))
        if photo is not None:
            pet_stage.create_oval(x - 28, y + 82, x + 132, y + 126, fill=t.panel, outline="")
            pet_stage.create_image(x + 52, y + 68, image=photo)
            bubble_text = _active_pet_bubble_text()
            if bubble_text:
                bubble_x = max(18, min(width - 258, x + 124))
                _draw_pet_bubble(pet_stage, bubble_x, max(54, y - 20), bubble_text, width_chars=28)
            return
        body = "#8FE6D6" if level < 5 else "#B8A2FF"
        belly = "#F7F5FF"
        accent = "#FFB86B" if level < 8 else "#FFE066"
        outline = t.accent
        pet_stage.create_oval(x - 28, y + 34, x + 112, y + 104, fill=t.panel, outline="")
        pet_stage.create_oval(x + 4 - breath, y + 28, x + 104 + breath, y + 124, fill=body, outline=outline, width=2)
        pet_stage.create_oval(x + 26, y + 60, x + 82, y + 112, fill=belly, outline="")
        pet_stage.create_oval(x + 6, y + 8, x + 100, y + 88, fill=body, outline=outline, width=2)
        pet_stage.create_oval(x + 2, y + 2, x + 34, y + 42, fill=body, outline=outline, width=2)
        pet_stage.create_oval(x + 72, y + 2, x + 104, y + 42, fill=body, outline=outline, width=2)
        if blink:
            pet_stage.create_line(x + 29, y + 44, x + 39, y + 44, fill="#10131C", width=2)
            pet_stage.create_line(x + 66, y + 44, x + 76, y + 44, fill="#10131C", width=2)
        else:
            pet_stage.create_oval(x + 29, y + 38, x + 39, y + 50, fill="#10131C", outline="")
            pet_stage.create_oval(x + 66, y + 38, x + 76, y + 50, fill="#10131C", outline="")
            pet_stage.create_oval(x + 32, y + 40, x + 35, y + 43, fill="white", outline="")
            pet_stage.create_oval(x + 69, y + 40, x + 72, y + 43, fill="white", outline="")
        pet_stage.create_arc(x + 42, y + 49, x + 64, y + 66, start=200, extent=140, style="arc", outline="#10131C", width=2)
        pet_stage.create_oval(x + 16, y + 52, x + 30, y + 64, fill="#FF8FB3", outline="")
        pet_stage.create_oval(x + 74, y + 52, x + 88, y + 64, fill="#FF8FB3", outline="")
        pet_stage.create_line(x + 88, y + 74, x + 126, y + 54 + (4 if (tick // 8) % 2 else -2), fill=accent, width=8, smooth=True)
        bubble_text = _active_pet_bubble_text()
        if bubble_text:
            bubble_x = max(18, min(width - 258, x + 126))
            _draw_pet_bubble(pet_stage, bubble_x, max(54, y - 20), bubble_text, width_chars=28)

    def _journal_pet_visible_now() -> bool:
        if not COMPANION_FEATURE_ENABLED:
            return False
        try:
            if root.state() in ("withdrawn", "iconic"):
                return False
        except tk.TclError:
            return False
        return active_page["key"] == "journal" and bool(pet_state.get("active", False)) and bool(pet_state.get("journal_visible", True)) and not bool(
            pet_state.get("desktop_enabled", False)
        )

    def _position_journal_pet_canvas(*, animate: bool = True) -> None:
        if not _journal_pet_visible_now():
            try:
                journal_pet_window.withdraw()
            except tk.TclError:
                pass
            return
        try:
            page_w = max(160, int(journal_page.winfo_width() or root.winfo_width() or 800))
            page_h = max(120, int(journal_page.winfo_height() or root.winfo_height() or 600))
            x = int(pet_state.get("journal_x", 0) or 0)
            y = int(pet_state.get("journal_y", 0) or 0)
            if x <= 0 and y <= 0:
                try:
                    x = max(0, int(restore_draft_btn.winfo_rootx() - journal_page.winfo_rootx()) - 112)
                    y = max(0, int(restore_draft_btn.winfo_rooty() - journal_page.winfo_rooty()) + 2)
                except tk.TclError:
                    x = max(0, page_w - 230)
                    y = 22
            x = max(0, min(page_w - 108, x))
            y = max(0, min(page_h - 62, y))
            pet_state["journal_x"] = x
            pet_state["journal_y"] = y
            screen_x = int(journal_page.winfo_rootx()) + x
            screen_y = int(journal_page.winfo_rooty()) + y
            journal_pet_window.geometry(f"104x58+{screen_x}+{screen_y}")
            journal_pet_window.deiconify()
            journal_pet_window.lift(root)
        except tk.TclError:
            return

    def _draw_journal_pet() -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        if not _journal_pet_visible_now():
            try:
                journal_pet_canvas.delete("all")
                journal_pet_window.withdraw()
            except tk.TclError:
                pass
            return
        try:
            t = th()
            _position_journal_pet_canvas(animate=not bool(journal_pet_state.get("dragging")))
            journal_pet_canvas.configure(bg=journal_pet_transparent)
            journal_pet_canvas.delete("all")
            tick = int(journal_pet_state.get("tick", 0) or 0)
            journal_pet_state["tick"] = tick + 1
            level, _current_xp, _needed_xp = _desktop_pet_level()
            breath = 2 if (tick // 12) % 2 == 0 else 0
            blink = (tick % 70) in (0, 1, 2)
            shadow = t.field
            photo = _load_pet_photo("journal", (56 + breath, 56 - min(1, breath)))
            if photo is not None:
                journal_pet_canvas.create_oval(24, 43, 82, 55, fill=shadow, outline="")
                journal_pet_canvas.create_image(53, 28, image=photo)
                hitbox = (18.0, 2.0, 88.0, 56.0)
            else:
                body = "#8FE6D6" if level < 5 else "#B8A2FF"
                belly = "#F7F5FF"
                accent = "#FFB86B" if level < 8 else "#FFE066"
                outline = t.accent
                x = 24
                y = 2
                journal_pet_canvas.create_oval(x - 1, y + 42, x + 62, y + 56, fill=shadow, outline="")
                journal_pet_canvas.create_oval(x + 10 - breath, y + 30, x + 50 + breath, y + 56, fill=body, outline=outline, width=2)
                journal_pet_canvas.create_oval(x + 21, y + 39, x + 39, y + 55, fill=belly, outline="")
                journal_pet_canvas.create_oval(x + 8, y + 9, x + 52, y + 43, fill=body, outline=outline, width=2)
                journal_pet_canvas.create_oval(x + 3, y + 4, x + 18, y + 22, fill=body, outline=outline, width=2)
                journal_pet_canvas.create_oval(x + 42, y + 4, x + 57, y + 22, fill=body, outline=outline, width=2)
                if blink:
                    journal_pet_canvas.create_line(x + 22, y + 24, x + 28, y + 24, fill="#10131C", width=2)
                    journal_pet_canvas.create_line(x + 36, y + 24, x + 42, y + 24, fill="#10131C", width=2)
                else:
                    journal_pet_canvas.create_oval(x + 22, y + 21, x + 27, y + 27, fill="#10131C", outline="")
                    journal_pet_canvas.create_oval(x + 37, y + 21, x + 42, y + 27, fill="#10131C", outline="")
                    journal_pet_canvas.create_oval(x + 24, y + 22, x + 25, y + 23, fill="white", outline="")
                    journal_pet_canvas.create_oval(x + 39, y + 22, x + 40, y + 23, fill="white", outline="")
                journal_pet_canvas.create_arc(x + 29, y + 28, x + 36, y + 35, start=200, extent=140, style="arc", outline="#10131C", width=1)
                journal_pet_canvas.create_line(x + 48, y + 38, x + 67, y + 29 + (2 if (tick // 8) % 2 else -1), fill=accent, width=4, smooth=True)
                hitbox = (18.0, 1.0, 92.0, 58.0)
            journal_pet_state["hitbox"] = hitbox
        except tk.TclError:
            pass

    def _refresh_journal_pet_button() -> None:
        try:
            pet_journal_btn.config(
                text=tr("pet.journal_hide")
                if bool(pet_state.get("journal_visible", True))
                else tr("pet.journal_show")
            )
        except tk.TclError:
            pass

    def _refresh_desktop_pet_ui() -> None:
        if not COMPANION_FEATURE_ENABLED:
            return
        active = bool(pet_state.get("active", False))
        if active and not _profile_dir_has_companion(COMPANION_CURRENT_DIR):
            pet_state.clear()
            pet_state.update(_default_companion_state(active=False))
            pet_image_cache.clear()
            _save_desktop_pet_state()
            active = False
        level, current_xp, needed_xp = _desktop_pet_level()
        profile = _pet_growth_profile(level)
        if active:
            pet_level_lbl.config(text=tr("pet.level").format(level=level))
            try:
                if root.focus_get() is not pet_name_entry and pet_name_var.get() != str(pet_state.get("name") or "Companion"):
                    pet_name_var.set(str(pet_state.get("name") or "Companion"))
            except tk.TclError:
                pass
            pet_xp_lbl.config(
                text=tr("pet.xp").format(
                    current=current_xp,
                    needed=needed_xp,
                    total=int(pet_state.get("fed_units", 0) or 0),
                    stage=profile["label"],
                )
            )
            last_gain = int(pet_state.get("last_gain", 0) or 0)
            pet_mood_lbl.config(
                text=(
                    tr("pet.mood_gain").format(xp=last_gain)
                    if last_gain > 0
                    else (_pet_memory_hint() or profile["reaction"] or tr("pet.mood_idle"))
                )
            )
        else:
            pet_level_lbl.config(text=tr("pet.title"))
            try:
                if root.focus_get() is not pet_name_entry:
                    pet_name_var.set("")
            except tk.TclError:
                pass
            pet_xp_lbl.config(text=tr("pet.no_current"))
            pet_mood_lbl.config(text=tr("pet.no_current"))
        t = th()
        for btn in (pet_feed_btn, pet_desktop_btn, pet_journal_btn, pet_export_btn, pet_delete_btn, pet_name_save_btn):
            btn.config(
                state=("normal" if active else "disabled"),
                bg=("#8B2F3A" if btn is pet_delete_btn and active else (t.btn_secondary if active else t.btn_disabled)),
                fg=("white" if btn is pet_delete_btn and active else (t.text if active else t.disabled_fg)),
                disabledforeground=t.disabled_fg,
                activebackground=("#A43B48" if btn is pet_delete_btn and active else t.secondary_hover),
                activeforeground=("white" if btn is pet_delete_btn and active else t.text),
                cursor=("hand2" if active else "arrow"),
            )
        pet_name_entry.config(
            state=("normal" if active else "disabled"),
            bg=(t.field if active else t.btn_disabled),
            fg=(t.text if active else t.disabled_fg),
            disabledbackground=t.btn_disabled,
            disabledforeground=t.disabled_fg,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        pet_desktop_btn.config(
            text=tr("pet.desktop_disable")
            if active and bool(pet_state.get("desktop_enabled", False))
            else tr("pet.desktop_enable")
        )
        _refresh_pet_module_ui()
        _refresh_journal_pet_button()
        _draw_desktop_pet()
        _draw_journal_pet()

    def _refresh_pet_art_button() -> None:
        t = th()
        if bool(pet_art_busy.get("v", False)):
            for btn in (pet_art_btn, pet_art_ref_btn):
                btn.config(
                    text=tr("pet.generating") if btn is pet_art_btn else tr("pet.generate_from_image"),
                    state="disabled",
                    bg=t.btn_disabled,
                    fg=t.disabled_fg,
                    disabledforeground=t.disabled_fg,
                    cursor="arrow",
                )
        else:
            for btn, label in (
                (pet_art_btn, tr("pet.generate_look")),
                (pet_art_ref_btn, tr("pet.generate_from_image")),
            ):
                btn.config(
                    text=label,
                    state="normal",
                    bg=t.btn_secondary,
                    fg=t.text,
                    activebackground=t.secondary_hover,
                    activeforeground=t.text,
                    cursor="hand2",
                )

    def _ask_companion_design_prompt() -> Optional[str]:
        t = th()
        result: Dict[str, Optional[str]] = {"value": None}
        prompt_win = tk.Toplevel(root)
        prompt_win.title(tr("pet.prompt_title"))
        prompt_win.configure(bg=t.surface)
        prompt_win.transient(root)
        prompt_win.grab_set()
        prompt_win.geometry("560x360")
        prompt_win.minsize(420, 280)
        prompt_win.grid_columnconfigure(0, weight=1)
        prompt_win.grid_rowconfigure(1, weight=1)
        tk.Label(
            prompt_win,
            text=tr("pet.prompt_body"),
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 10),
            anchor="w",
            justify="left",
            wraplength=500,
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 10))
        prompt_text = tk.Text(
            prompt_win,
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            relief="flat",
            font=("Segoe UI", 10),
            wrap="word",
            undo=True,
            highlightthickness=1,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        prompt_text.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 12))
        prompt_text.insert("1.0", tr("pet.prompt_default"))
        button_bar = tk.Frame(prompt_win, bg=t.surface)
        button_bar.grid(row=2, column=0, sticky="e", padx=18, pady=(0, 18))

        def _close_with_value(value: Optional[str]) -> None:
            result["value"] = value
            try:
                prompt_win.grab_release()
            except tk.TclError:
                pass
            prompt_win.destroy()

        cancel_btn = tk.Button(
            button_bar,
            text=tr("common.cancel"),
            command=lambda: _close_with_value(None),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=6,
            cursor="hand2",
        )
        cancel_btn.pack(side="left", padx=(0, 8))
        generate_btn = tk.Button(
            button_bar,
            text=tr("pet.prompt_generate"),
            command=lambda: _close_with_value(prompt_text.get("1.0", "end-1c").strip()),
            bg=t.accent,
            fg="white",
            activebackground=t.hover_primary,
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=6,
            cursor="hand2",
        )
        generate_btn.pack(side="left")
        prompt_win.protocol("WM_DELETE_WINDOW", lambda: _close_with_value(None))
        try:
            root.update_idletasks()
            x = root.winfo_rootx() + max(40, (root.winfo_width() - 560) // 2)
            y = root.winfo_rooty() + max(40, (root.winfo_height() - 360) // 2)
            prompt_win.geometry(f"560x360+{x}+{y}")
        except tk.TclError:
            pass
        prompt_text.focus_set()
        root.wait_window(prompt_win)
        return result["value"]

    def _generate_desktop_pet_look(reference_image: Optional[Path] = None) -> None:
        if bool(pet_art_busy.get("v", False)):
            return
        if not get_openai_api_key():
            messagebox.showerror(tr("pet.title"), tr("pet.no_api_key"))
            return
        design_prompt = ""
        if reference_image is None:
            chosen_prompt = _ask_companion_design_prompt()
            if chosen_prompt is None:
                return
            design_prompt = chosen_prompt
        elif not messagebox.askyesno(tr("pet.generate_confirm_title"), tr("pet.generate_confirm")):
            return
        level, _, _ = _desktop_pet_level()
        pet_art_busy["v"] = True
        _refresh_pet_art_button()
        pet_mood_lbl.config(text=tr("pet.generating"))
        _append_console_session_update("Generating companion look...", key=f"pet:generate:{int(time.time())}")

        def _work() -> None:
            try:
                if COMPANION_INCOMING_DIR.exists():
                    shutil.rmtree(COMPANION_INCOMING_DIR)
                COMPANION_INCOMING_DIR.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            path, err = generate_desktop_pet_image(
                level,
                str(pet_state.get("name") or "Companion"),
                reference_image_path=reference_image,
                output_dir=COMPANION_INCOMING_DIR,
                design_prompt=design_prompt,
            )
            root.after(0, lambda p=path, e=err: _done(p, e))

        def _done(path: Optional[Path], err: Optional[str]) -> None:
            pet_art_busy["v"] = False
            _refresh_pet_art_button()
            if err or path is None:
                pet_mood_lbl.config(text=tr("pet.generate_failed"))
                _append_console_session_update(
                    f"Companion look generation failed: {err or 'No image returned.'}",
                    key=f"pet:generate:fail:{int(time.time())}",
                )
                messagebox.showerror(tr("pet.title"), (err or tr("pet.generate_failed"))[:4000])
                return
            try:
                if _companion_is_fresh_default():
                    if COMPANION_CURRENT_DIR.exists():
                        shutil.rmtree(COMPANION_CURRENT_DIR)
                else:
                    _archive_current_companion(copy_only=False)
                if COMPANION_CURRENT_DIR.exists():
                    shutil.rmtree(COMPANION_CURRENT_DIR)
                COMPANION_CURRENT_DIR.mkdir(parents=True, exist_ok=True)
                target_image = COMPANION_CURRENT_DIR / "companion.png"
                shutil.move(str(path), str(target_image))
            except Exception as exc:
                pet_mood_lbl.config(text=tr("pet.generate_failed"))
                messagebox.showerror(tr("pet.title"), tr("pet.profile_replace_failed").format(error=str(exc))[:4000])
                return
            pet_state.clear()
            pet_state.update(_default_companion_state(active=True))
            pet_state["name"] = "Companion"
            pet_state["image_path"] = str(target_image)
            pet_state["created_at"] = datetime.now().isoformat(timespec="seconds")
            pet_image_cache.clear()
            _save_desktop_pet_state()
            _append_console_session_update(
                f"Companion generated: {target_image.name}",
                key=f"companion:generate:done:{int(time.time())}",
            )
            _refresh_desktop_pet_ui()
            pet_mood_lbl.config(text=tr("pet.generated"))
            _draw_pet_overlay()

        threading.Thread(target=_work, daemon=True).start()

    def _generate_desktop_pet_look_from_image() -> None:
        if filedialog is None:
            messagebox.showerror(tr("pet.title"), tr("pet.reference_picker_unavailable"))
            return
        selected = filedialog.askopenfilename(
            title=tr("pet.reference_picker_title"),
            filetypes=[
                (tr("pet.reference_filetype_images"), "*.png *.jpg *.jpeg *.webp *.bmp"),
                (tr("pet.reference_filetype_all"), "*.*"),
            ],
        )
        if not selected:
            return
        _generate_desktop_pet_look(Path(selected))

    def _screen_bottom_y() -> int:
        try:
            return max(0, int(root.winfo_screenheight()) - 152)
        except tk.TclError:
            return 0

    def _clamp_overlay_position(x: int, y: int) -> Tuple[int, int]:
        try:
            sw = max(180, int(root.winfo_screenwidth()))
            sh = max(180, int(root.winfo_screenheight()))
        except tk.TclError:
            sw, sh = 1280, 720
        return (max(0, min(sw - 156, int(x))), max(0, min(sh - 148, int(y))))

    def _stage_pet_pointer_inside(event: Any) -> bool:
        try:
            width = int(pet_stage.winfo_width())
            height = int(pet_stage.winfo_height())
        except tk.TclError:
            return False
        x = int(getattr(event, "x", 0) or 0)
        y = int(getattr(event, "y", 0) or 0)
        return -24 <= x <= width + 24 and -24 <= y <= height + 24

    def _handoff_stage_pet_to_desktop(event: Any) -> None:
        if pet_stage_drag_state.get("handoff"):
            return
        pet_stage_drag_state["handoff"] = True
        try:
            root_x = int(getattr(event, "x_root", 0) or 0)
            root_y = int(getattr(event, "y_root", 0) or 0)
        except Exception:
            root_x, root_y = 40, _screen_bottom_y()
        x, y = _clamp_overlay_position(root_x - 74, root_y - 72)
        pet_overlay_state["x"] = x
        pet_overlay_state["y"] = y
        pet_overlay_state["falling"] = False
        pet_overlay_state["vy"] = 0.0
        pet_state["desktop_enabled"] = True
        pet_state["desktop_x"] = x
        pet_state["desktop_y"] = y
        _save_desktop_pet_state()
        _show_pet_overlay(drop=False)
        _set_pet_bubble(tr("pet.drag_desktop_say"), seconds=5, force=True)
        pet_stage_drag_state.update({"dragging": False, "active": False})
        pet_stage.configure(cursor="")
        _append_console_session_update(
            "Companion dragged out to desktop idle.",
            key=f"pet:dragout:{int(time.time())}",
        )
        _refresh_desktop_pet_ui()

    def _journal_pet_pointer_inside_page(event: Any) -> bool:
        try:
            root_x = int(getattr(event, "x_root", 0) or 0)
            root_y = int(getattr(event, "y_root", 0) or 0)
            page_x = int(journal_page.winfo_rootx())
            page_y = int(journal_page.winfo_rooty())
            page_w = int(journal_page.winfo_width())
            page_h = int(journal_page.winfo_height())
        except tk.TclError:
            return False
        return page_x - 24 <= root_x <= page_x + page_w + 24 and page_y - 24 <= root_y <= page_y + page_h + 24

    def _handoff_journal_pet_to_desktop(event: Any) -> None:
        if journal_pet_state.get("handoff"):
            return
        journal_pet_state["handoff"] = True
        try:
            root_x = int(getattr(event, "x_root", 0) or 0)
            root_y = int(getattr(event, "y_root", 0) or 0)
        except Exception:
            root_x, root_y = 40, _screen_bottom_y()
        x, y = _clamp_overlay_position(root_x - 74, root_y - 72)
        pet_overlay_state["x"] = x
        pet_overlay_state["y"] = y
        pet_overlay_state["falling"] = False
        pet_overlay_state["vy"] = 0.0
        pet_state["desktop_enabled"] = True
        pet_state["desktop_x"] = x
        pet_state["desktop_y"] = y
        _save_desktop_pet_state()
        _show_pet_overlay(drop=False)
        _set_pet_bubble(tr("pet.drag_desktop_say"), seconds=5, force=True)
        journal_pet_state.update({"dragging": False, "handoff": False})
        try:
            journal_pet_window.withdraw()
            journal_pet_canvas.configure(cursor="hand2")
        except tk.TclError:
            pass
        _append_console_session_update(
            "Companion dragged from journal to desktop idle.",
            key=f"pet:journal-dragout:{int(time.time())}",
        )
        _refresh_desktop_pet_ui()

    def _toggle_journal_pet_visible() -> None:
        if not bool(pet_state.get("active", False)):
            messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
            return
        pet_state["journal_visible"] = not bool(pet_state.get("journal_visible", True))
        _save_desktop_pet_state()
        if bool(pet_state.get("journal_visible", True)):
            _set_pet_bubble(tr("pet.journal_show_say"), seconds=4, force=True)
        else:
            try:
                journal_pet_window.withdraw()
            except tk.TclError:
                pass
        _refresh_desktop_pet_ui()

    def _on_journal_pet_press(event: Any) -> str:
        if not _journal_pet_visible_now():
            return ""
        x = float(getattr(event, "x", 0) or 0)
        y = float(getattr(event, "y", 0) or 0)
        x1, y1, x2, y2 = journal_pet_state.get("hitbox", (0.0, 0.0, 0.0, 0.0))
        if not (x1 <= x <= x2 and y1 <= y <= y2):
            return ""
        journal_pet_state["dragging"] = True
        journal_pet_state["handoff"] = False
        journal_pet_state["drag_dx"] = int(x)
        journal_pet_state["drag_dy"] = int(y)
        try:
            journal_pet_canvas.configure(cursor="fleur")
        except tk.TclError:
            pass
        _draw_journal_pet()
        return "break"

    def _on_journal_pet_motion(event: Any) -> str:
        if not journal_pet_state.get("dragging"):
            return ""
        try:
            page_x = int(journal_page.winfo_rootx())
            page_y = int(journal_page.winfo_rooty())
            page_w = max(120, int(journal_page.winfo_width()))
            page_h = max(90, int(journal_page.winfo_height()))
            new_x = int(getattr(event, "x_root", 0) or 0) - page_x - int(journal_pet_state.get("drag_dx", 0) or 0)
            new_y = int(getattr(event, "y_root", 0) or 0) - page_y - int(journal_pet_state.get("drag_dy", 0) or 0)
            pet_state["journal_x"] = max(0, min(page_w - 108, new_x))
            pet_state["journal_y"] = max(0, min(page_h - 62, new_y))
            _position_journal_pet_canvas()
        except tk.TclError:
            pass
        _draw_journal_pet()
        return "break"

    def _on_journal_pet_release(_event: Optional[Any] = None) -> str:
        if not journal_pet_state.get("dragging"):
            return ""
        journal_pet_state.update({"dragging": False, "handoff": False})
        try:
            journal_pet_canvas.configure(cursor="hand2")
        except tk.TclError:
            pass
        _save_desktop_pet_state()
        _draw_journal_pet()
        return "break"

    def _on_pet_stage_press(event: Any) -> str:
        x = float(getattr(event, "x", 0) or 0)
        y = float(getattr(event, "y", 0) or 0)
        x1, y1, x2, y2 = pet_stage_drag_state.get("hitbox", (0.0, 0.0, 0.0, 0.0))
        if not (x1 <= x <= x2 and y1 <= y <= y2):
            return ""
        pet_stage_drag_state["dragging"] = True
        pet_stage_drag_state["active"] = True
        pet_stage_drag_state["handoff"] = False
        pet_stage_drag_state["x"] = float(pet_anim_state.get("x", 42.0) or 42.0)
        pet_stage_drag_state["y"] = max(0.0, y1)
        pet_stage_drag_state["dx"] = int(x - float(pet_stage_drag_state["x"]))
        pet_stage_drag_state["dy"] = int(y - float(pet_stage_drag_state["y"]))
        pet_stage.configure(cursor="hand2")
        _draw_desktop_pet()
        return "break"

    def _on_pet_stage_motion(event: Any) -> str:
        if not pet_stage_drag_state.get("dragging"):
            return ""
        if not _stage_pet_pointer_inside(event):
            _handoff_stage_pet_to_desktop(event)
            return "break"
        pet_stage_drag_state["x"] = float(getattr(event, "x", 0) or 0) - int(pet_stage_drag_state.get("dx", 0) or 0)
        pet_stage_drag_state["y"] = float(getattr(event, "y", 0) or 0) - int(pet_stage_drag_state.get("dy", 0) or 0)
        _draw_desktop_pet()
        return "break"

    def _on_pet_stage_release(event: Any) -> str:
        if not pet_stage_drag_state.get("dragging"):
            return ""
        if not pet_stage_drag_state.get("handoff"):
            x = max(42.0, float(pet_stage_drag_state.get("x", 42.0) or 42.0))
            y = max(20.0, float(pet_stage_drag_state.get("y", 0.0) or 0.0))
            pet_anim_state["x"] = x
            pet_state["stage_x"] = int(x)
            pet_state["stage_y"] = int(y)
            _save_desktop_pet_state()
        pet_stage_drag_state.update({"dragging": False, "active": False, "handoff": False})
        pet_stage.configure(cursor="")
        _draw_desktop_pet()
        return "break"

    def _draw_pet_overlay() -> None:
        canvas = pet_overlay_state.get("canvas")
        if canvas is None:
            return
        try:
            t = th()
            canvas.delete("all")
            if not bool(pet_state.get("active", False)):
                return
            tick = int(pet_overlay_state.get("tick", 0) or 0)
            level, _current_xp, _needed_xp = _desktop_pet_level()
            bounce = 4 if (tick // 9) % 2 == 0 else 0
            bubble_text = _active_pet_bubble_text()
            photo = _load_pet_photo("overlay", (94, 94) if bubble_text else (118, 118))
            if photo is not None:
                canvas.create_oval(14, 96, 134, 132, fill="#2A2D44", outline="")
                if bubble_text:
                    _draw_pet_bubble(canvas, 8, 4, bubble_text, width_chars=16)
                canvas.create_image(74, (88 if bubble_text else 64) - bounce, image=photo)
                return
            body = "#8FE6D6" if level < 5 else "#B8A2FF"
            belly = "#F7F5FF"
            accent = "#FFB86B" if level < 8 else "#FFE066"
            outline = t.accent
            x = 26
            y = (32 if bubble_text else 12) - bounce
            if bubble_text:
                _draw_pet_bubble(canvas, 8, 4, bubble_text, width_chars=16)
            canvas.create_oval(x - 10, y + 58, x + 110, y + 118, fill="#2A2D44", outline="")
            canvas.create_oval(x + 8, y + 42, x + 92, y + 124, fill=body, outline=outline, width=2)
            canvas.create_oval(x + 30, y + 72, x + 72, y + 116, fill=belly, outline="")
            canvas.create_oval(x + 8, y + 18, x + 92, y + 88, fill=body, outline=outline, width=2)
            canvas.create_oval(x + 4, y + 10, x + 32, y + 46, fill=body, outline=outline, width=2)
            canvas.create_oval(x + 68, y + 10, x + 96, y + 46, fill=body, outline=outline, width=2)
            canvas.create_oval(x + 30, y + 48, x + 39, y + 58, fill="#10131C", outline="")
            canvas.create_oval(x + 62, y + 48, x + 71, y + 58, fill="#10131C", outline="")
            canvas.create_oval(x + 33, y + 50, x + 36, y + 53, fill="white", outline="")
            canvas.create_oval(x + 65, y + 50, x + 68, y + 53, fill="white", outline="")
            canvas.create_arc(x + 43, y + 58, x + 59, y + 72, start=200, extent=140, style="arc", outline="#10131C", width=2)
            canvas.create_oval(x + 18, y + 61, x + 30, y + 72, fill="#FF8FB3", outline="")
            canvas.create_oval(x + 72, y + 61, x + 84, y + 72, fill="#FF8FB3", outline="")
            canvas.create_line(x + 80, y + 82, x + 118, y + 64 - bounce, fill=accent, width=7, smooth=True)
        except tk.TclError:
            pass

    def _save_overlay_position() -> None:
        x, y = _clamp_overlay_position(
            int(pet_overlay_state.get("x", 0) or 0),
            int(pet_overlay_state.get("y", 0) or 0),
        )
        pet_overlay_state["x"] = x
        pet_overlay_state["y"] = y
        pet_state["desktop_x"] = x
        pet_state["desktop_y"] = y
        _save_desktop_pet_state()

    def _restore_journal_from_pet(_event: Optional[Any] = None) -> str:
        try:
            root.deiconify()
            root.lift()
            root.focus_force()
            show_page("journal")
        except tk.TclError:
            pass
        return "break"

    def _destroy_pet_overlay() -> None:
        win = pet_overlay_state.get("window")
        pet_overlay_state.update({"window": None, "canvas": None, "dragging": False, "falling": False})
        if win is not None:
            try:
                win.destroy()
            except tk.TclError:
                pass

    def _destroy_journal_pet_window() -> None:
        try:
            journal_pet_window.destroy()
        except tk.TclError:
            pass

    def _create_pet_overlay() -> None:
        if not bool(pet_state.get("active", False)):
            return
        win = pet_overlay_state.get("window")
        try:
            if win is not None and bool(win.winfo_exists()):
                return
        except tk.TclError:
            pet_overlay_state["window"] = None
        transparent = "#00ff7f"
        win = tk.Toplevel(root)
        win.overrideredirect(True)
        try:
            win.attributes("-topmost", True)
            win.attributes("-transparentcolor", transparent)
        except tk.TclError:
            pass
        canvas = tk.Canvas(win, width=148, height=144, bg=transparent, bd=0, highlightthickness=0, cursor="hand2")
        canvas.pack(fill="both", expand=True)
        pet_overlay_state["window"] = win
        pet_overlay_state["canvas"] = canvas

        def _press(event: Any) -> str:
            pet_overlay_state["dragging"] = True
            pet_overlay_state["falling"] = False
            pet_overlay_state["drag_dx"] = int(getattr(event, "x", 0) or 0)
            pet_overlay_state["drag_dy"] = int(getattr(event, "y", 0) or 0)
            return "break"

        def _motion(event: Any) -> str:
            if not pet_overlay_state.get("dragging"):
                return "break"
            x = int(getattr(event, "x_root", 0) or 0) - int(pet_overlay_state.get("drag_dx", 0) or 0)
            y = int(getattr(event, "y_root", 0) or 0) - int(pet_overlay_state.get("drag_dy", 0) or 0)
            x, y = _clamp_overlay_position(x, y)
            pet_overlay_state["x"] = x
            pet_overlay_state["y"] = y
            try:
                win.geometry(f"148x144+{x}+{y}")
            except tk.TclError:
                pass
            return "break"

        def _release(_event: Optional[Any] = None) -> str:
            pet_overlay_state["dragging"] = False
            _save_overlay_position()
            return "break"

        def _right_click(_event: Optional[Any] = None) -> str:
            _set_desktop_pet_enabled(False)
            return "break"

        canvas.bind("<ButtonPress-1>", _press, add="+")
        canvas.bind("<B1-Motion>", _motion, add="+")
        canvas.bind("<ButtonRelease-1>", _release, add="+")
        canvas.bind("<Double-Button-1>", _restore_journal_from_pet, add="+")
        canvas.bind("<Button-3>", _right_click, add="+")
        _draw_pet_overlay()

    def _show_pet_overlay(*, drop: bool = False) -> None:
        if not bool(pet_state.get("active", False)):
            _destroy_pet_overlay()
            return
        _create_pet_overlay()
        win = pet_overlay_state.get("window")
        if win is None:
            return
        x = int(pet_overlay_state.get("x", 0) or 0)
        y = int(pet_overlay_state.get("y", 0) or 0)
        if x <= 0 and y <= 0:
            try:
                x = int(root.winfo_rootx() + root.winfo_width() - 190)
                y = _screen_bottom_y()
            except tk.TclError:
                x, y = 40, _screen_bottom_y()
        if drop:
            try:
                y = int(root.winfo_rooty() + root.winfo_height() - 110)
            except tk.TclError:
                y = max(0, _screen_bottom_y() - 160)
            pet_overlay_state["falling"] = True
            pet_overlay_state["vy"] = 0.0
        x, y = _clamp_overlay_position(x, y)
        pet_overlay_state["x"] = x
        pet_overlay_state["y"] = y
        try:
            win.geometry(f"148x144+{x}+{y}")
            win.deiconify()
            win.lift()
        except tk.TclError:
            pass

    def _tick_pet_overlay() -> None:
        try:
            win = pet_overlay_state.get("window")
            if win is not None and bool(win.winfo_exists()):
                pet_overlay_state["tick"] = int(pet_overlay_state.get("tick", 0) or 0) + 1
                x = int(pet_overlay_state.get("x", 0) or 0)
                y = int(pet_overlay_state.get("y", 0) or 0)
                if not pet_overlay_state.get("dragging"):
                    if pet_overlay_state.get("falling"):
                        target_y = _screen_bottom_y()
                        vy = float(pet_overlay_state.get("vy", 0.0) or 0.0) + 1.8
                        y = min(target_y, int(y + vy))
                        pet_overlay_state["vy"] = vy
                        if y >= target_y:
                            pet_overlay_state["falling"] = False
                            _save_overlay_position()
                    x, y = _clamp_overlay_position(x, y)
                    pet_overlay_state["x"] = x
                    pet_overlay_state["y"] = y
                    win.geometry(f"148x144+{x}+{y}")
                _draw_pet_overlay()
            root.after(90, _tick_pet_overlay)
        except tk.TclError:
            return

    def _set_desktop_pet_enabled(enabled: bool, *, drop: bool = False) -> None:
        if enabled and not bool(pet_state.get("active", False)):
            messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
            _destroy_pet_overlay()
            return
        was_backgrounded = False
        try:
            was_backgrounded = root.state() == "withdrawn"
        except tk.TclError:
            pass
        pet_state["desktop_enabled"] = bool(enabled)
        if enabled:
            _show_pet_overlay(drop=drop)
        else:
            _destroy_pet_overlay()
        _save_desktop_pet_state()
        if not enabled and was_backgrounded:
            try:
                destroy_journal_window()
            except tk.TclError:
                pass
            return
        _refresh_desktop_pet_ui()

    def _toggle_desktop_pet_mode() -> None:
        if not bool(pet_state.get("active", False)):
            messagebox.showinfo(tr("pet.title"), tr("pet.no_current"))
            return
        _set_desktop_pet_enabled(not bool(pet_state.get("desktop_enabled", False)), drop=True)

    def _background_journal_with_pet() -> None:
        if not bool(pet_state.get("active", False)):
            destroy_journal_window()
            return
        _set_desktop_pet_enabled(True, drop=True)
        try:
            root.withdraw()
        except tk.TclError:
            pass

    def _finish_or_background_journal_window() -> None:
        if bool(pet_state.get("desktop_enabled", False)):
            _background_journal_with_pet()
            return
        destroy_journal_window()

    def _tick_desktop_pet() -> None:
        try:
            _draw_desktop_pet()
            _draw_journal_pet()
            root.after(90, _tick_desktop_pet)
        except tk.TclError:
            return

    if COMPANION_FEATURE_ENABLED:
        pet_feed_btn.config(command=lambda: _feed_desktop_pet_from_current_entry(force=True))
        pet_art_btn.config(command=_generate_desktop_pet_look)
        pet_art_ref_btn.config(command=_generate_desktop_pet_look_from_image)
        pet_name_save_btn.config(command=_save_companion_name)
        pet_name_entry.bind("<Return>", lambda _event: (_save_companion_name(), "break")[1])
        pet_desktop_btn.config(command=_toggle_desktop_pet_mode, state="normal", cursor="hand2")
        pet_journal_btn.config(command=_toggle_journal_pet_visible, state="normal", cursor="hand2")
        pet_export_btn.config(command=_export_current_companion_profile, state="normal", cursor="hand2")
        pet_import_btn.config(command=_import_companion_profile, state="normal", cursor="hand2")
        pet_folder_btn.config(command=_open_companion_folder, state="normal", cursor="hand2")
        pet_delete_btn.config(command=_delete_current_companion, state="normal", cursor="hand2")
        pet_bubble_btn.config(command=_toggle_pet_bubble_module)
        pet_memory_btn.config(command=_toggle_pet_memory_module)
        pet_tts_btn.config(command=_toggle_pet_tts_module)
        pet_chat_btn.config(command=_toggle_pet_chat_module)
        pet_chat_ask_btn.config(command=_ask_pet_chat)
        pet_chat_entry.bind("<Return>", lambda _event: (_ask_pet_chat(), "break")[1])
        pet_stage.bind("<ButtonPress-1>", _on_pet_stage_press, add="+")
        pet_stage.bind("<B1-Motion>", _on_pet_stage_motion, add="+")
        pet_stage.bind("<ButtonRelease-1>", _on_pet_stage_release, add="+")
        journal_pet_canvas.bind("<ButtonPress-1>", _on_journal_pet_press, add="+")
        journal_pet_canvas.bind("<B1-Motion>", _on_journal_pet_motion, add="+")
        journal_pet_canvas.bind("<ButtonRelease-1>", _on_journal_pet_release, add="+")
        journal_pet_canvas.bind("<Button-3>", lambda _event: (_toggle_journal_pet_visible(), "break")[1], add="+")
        journal_page.bind("<Configure>", lambda _event: root.after_idle(_draw_journal_pet), add="+")
        _refresh_desktop_pet_ui()
        _refresh_pet_art_button()
        _refresh_pet_module_ui()
        root.after(120, _tick_desktop_pet)
        root.after(120, _tick_pet_overlay)
        journal_cleanup_callbacks.append(_destroy_pet_overlay)
        journal_cleanup_callbacks.append(_destroy_journal_pet_window)
        if bool(pet_state.get("desktop_enabled", False)):
            root.after(500, lambda: _show_pet_overlay(drop=False))
    else:
        journal_cleanup_callbacks.append(_destroy_journal_pet_window)

    def _sync_journal_side_action_columns() -> None:
        try:
            root.update_idletasks()
            action_width = max(
                JOURNAL_SIDE_ACTION_GRID_MINSIZE,
                int(transcribe_hover.winfo_reqwidth()),
                int(transcribe_hover.winfo_width()),
                int(transcribe_btn_row.winfo_reqwidth()),
                int(transcribe_file_btn_row.winfo_reqwidth()),
                int(receive_iphone_btn_row.winfo_reqwidth()),
                int(gen_report_hover.winfo_reqwidth()),
                int(gen_report_hover.winfo_width()),
                int(gen_report_btn_row.winfo_reqwidth()),
                int(gen_button.winfo_reqwidth()),
            )
            stt_frame.grid_columnconfigure(1, minsize=JOURNAL_TEXT_SCROLLBAR_WIDTH)
            report_frame.grid_columnconfigure(1, minsize=JOURNAL_TEXT_SCROLLBAR_WIDTH)
            stt_frame.grid_columnconfigure(2, minsize=action_width)
            report_frame.grid_columnconfigure(2, minsize=action_width)
            transcribe_hover.configure(width=action_width)
            gen_report_hover.configure(width=action_width)
            root.update_idletasks()
        except tk.TclError:
            pass

    root.after_idle(_sync_journal_side_action_columns)
    root.after(250, _sync_journal_side_action_columns)
    root.after(1000, _sync_journal_side_action_columns)

    placeholder_frames: List[Any] = []
    placeholder_title_labels: List[Any] = []
    placeholder_body_labels: List[Any] = []

    api_key_prompt_hooks: Dict[str, Callable[[], None]] = {}
    transcription_models_manager_hooks: Dict[str, Callable[[], None]] = {}

    def build_ai_recap_and_chatbot_pages() -> None:
        _register_page_toggle(ai_recap_page)
        _register_page_toggle(chatbot_page)

        t0 = t_init
        _tb, _tf, _tab, _taf = t0.toolbar_btn_config()

        # --- Shared: append styled lines to a read-only transcript ---
        def _append_transcript(box: Any, role: str, body: str) -> None:
            box.config(state="normal")
            if role == "user":
                box.insert("end", tr("chat.you") + "\n", ("t_meta",))
                box.insert("end", (body or "").strip() + "\n\n", ("t_user",))
            else:
                box.insert("end", tr("chat.assistant") + "\n", ("t_meta",))
                box.insert("end", (body or "").strip() + "\n\n", ("t_bot",))
            box.config(state="disabled")
            box.see("end")

        # ========== AI Recap ==========
        recap_wrap = tk.Frame(ai_recap_page, bg=t0.surface)
        recap_wrap.pack(
            fill="both",
            expand=True,
            padx=t0.pad_outer,
            pady=(0, t0.pad_center_y + JOURNAL_WINDOW_CONSOLE_RESERVE_BOTTOM),
        )
        recap_wrap.grid_columnconfigure(0, weight=1)
        recap_wrap.grid_rowconfigure(3, weight=1)

        recap_title = tk.Label(
            recap_wrap,
            text=tr("recap.title"),
            bg=t0.surface,
            fg=t0.text,
            font=("Segoe UI", 15, "bold"),
            anchor="w",
        )
        recap_title.grid(row=0, column=0, sticky="w", pady=(0, 8))

        recap_top = tk.Frame(recap_wrap, bg=t0.panel, highlightthickness=1, highlightbackground=t0.border)
        recap_top.grid(row=1, column=0, sticky="ew", pady=(0, 8))
        recap_top.grid_columnconfigure(4, weight=1)

        recap_thinking_var = tk.BooleanVar(value=True)
        recap_thinking_chk = tk.Checkbutton(
            recap_top,
            text=tr("recap.thinking"),
            variable=recap_thinking_var,
            bg=t0.panel,
            fg=t0.muted,
            activebackground=t0.panel,
            activeforeground=t0.text,
            selectcolor=t0.field,
            font=("Segoe UI", 9),
        )
        recap_thinking_chk.grid(row=0, column=0, padx=(10, 8), pady=8, sticky="w")
        bind_hover_tooltip(recap_thinking_chk, lambda: tr("tip.thinking_model"))

        recap_from_fr = tk.Frame(recap_top, bg=t0.panel)
        recap_from_fr.grid(row=0, column=1, padx=(0, 10), pady=8, sticky="w")
        recap_from_lbl = tk.Label(
            recap_from_fr,
            text=tr("recap.from"),
            bg=t0.panel,
            fg=t0.muted,
            font=t0.date_label_font,
        )
        recap_from_lbl.pack(side="left", padx=(0, 8))
        recap_from_de: Any = None
        recap_to_de: Any = None
        _today = datetime.now().date()
        if DateEntry is not None:
            recap_from_de = DateEntry(
                recap_from_fr,
                width=14,
                date_pattern="mm/dd/yyyy",
                state="normal",
                background=t0.field,
                foreground=t0.text,
                borderwidth=1,
            )
            recap_from_de.pack(side="left")
            try:
                recap_from_de.set_date(_today)
            except Exception:
                pass
        else:
            tk.Label(
                recap_from_fr,
                text=tr("recap.install_dates"),
                bg=t0.panel,
                fg=t0.muted,
                font=("Segoe UI", 9),
            ).pack(side="left")

        recap_to_var = tk.BooleanVar(value=False)
        recap_all_journal_var = tk.BooleanVar(value=True)
        recap_to_wrap = tk.Frame(recap_top, bg=t0.panel)
        recap_to_wrap.grid(row=0, column=2, padx=(0, 8), pady=8, sticky="w")
        recap_to_chk = tk.Checkbutton(
            recap_to_wrap,
            text=tr("recap.to_chk"),
            variable=recap_to_var,
            bg=t0.panel,
            fg=t0.muted,
            activebackground=t0.panel,
            activeforeground=t0.text,
            selectcolor=t0.field,
            font=("Segoe UI", 9),
        )
        recap_to_chk.pack(side="left")
        recap_all_journal_chk = tk.Checkbutton(
            recap_top,
            text=tr("recap.all_journal"),
            variable=recap_all_journal_var,
            bg=t0.panel,
            fg=t0.muted,
            activebackground=t0.panel,
            activeforeground=t0.text,
            selectcolor=t0.field,
            font=("Segoe UI", 9),
        )
        recap_all_journal_chk.grid(row=1, column=0, columnspan=8, sticky="w", padx=(10, 0), pady=(0, 6))
        bind_hover_tooltip(recap_all_journal_chk, lambda: tr("tip.recap_all_journal"))
        recap_through_fr = tk.Frame(recap_top, bg=t0.panel)
        recap_through_lbl = tk.Label(
            recap_through_fr,
            text=tr("recap.through"),
            bg=t0.panel,
            fg=t0.muted,
            font=t0.date_label_font,
        )
        recap_through_lbl.pack(side="left", padx=(0, 8))
        if DateEntry is not None:
            recap_to_de = DateEntry(
                recap_through_fr,
                width=14,
                date_pattern="mm/dd/yyyy",
                state="normal",
                background=t0.field,
                foreground=t0.text,
                borderwidth=1,
            )
            recap_to_de.pack(side="left")
            try:
                recap_to_de.set_date(_today)
            except Exception:
                pass
        else:
            tk.Label(
                recap_through_fr,
                text=tr("recap.through_placeholder"),
                bg=t0.panel,
                fg=t0.muted,
                font=("Segoe UI", 9),
            ).pack(side="left")

        recap_cal_row = tk.Frame(recap_wrap, bg=t0.surface)
        recap_cal_row.grid_columnconfigure(1, weight=1)
        recap_selected_dates: set = set()
        recap_calendar: Any = None
        if Calendar is not None:
            recap_calendar = Calendar(
                recap_cal_row,
                selectmode="day",
                showweeknumbers=False,
                background=t0.field,
                foreground=t0.text,
                headersbackground=t0.panel,
                headersforeground=t0.text,
                weekendbackground=t0.field,
                weekendforeground=t0.muted,
                normalbackground=t0.field,
                normalforeground=t0.text,
                othermonthbackground=t0.field,
                othermonthforeground=t0.muted,
                selectbackground=t0.accent,
                selectforeground="white",
                bordercolor=t0.border,
                font=("Segoe UI", 9),
            )
            recap_calendar.grid(row=0, column=0, sticky="nw", padx=(0, 12), pady=(0, 4))
        else:
            tk.Label(
                recap_cal_row,
                text=tr("recap.install_calendar"),
                bg=t0.surface,
                fg=t0.muted,
                font=("Segoe UI", 9),
                wraplength=400,
                justify="left",
            ).grid(row=0, column=0, sticky="w", pady=(0, 4))

        recap_sel_lbl = tk.Label(
            recap_cal_row,
            text=tr("recap.selected.none"),
            bg=t0.surface,
            fg=t0.muted,
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
        )
        recap_sel_lbl.grid(row=0, column=1, sticky="nw", pady=(0, 4))

        recap_cal_marks_tag = "recap_sel"

        def recap_refresh_cal_marks() -> None:
            if recap_calendar is None:
                return
            try:
                recap_calendar.calevent_remove("all")
            except Exception:
                pass
            t = th()
            for d in recap_selected_dates:
                try:
                    recap_calendar.calevent_create(d, "", recap_cal_marks_tag)
                except Exception:
                    pass
            try:
                recap_calendar.tag_config(recap_cal_marks_tag, background=t.accent, foreground="white")
            except Exception:
                pass

        def recap_update_sel_label() -> None:
            if recap_all_journal_var.get():
                recap_sel_lbl.config(text=tr("recap.all_journal_active"))
                return
            if not recap_selected_dates:
                recap_sel_lbl.config(text=tr("recap.selected.none"))
                return
            ordered = sorted(recap_selected_dates)
            parts = [x.strftime("%m/%d/%Y") for x in ordered]
            recap_sel_lbl.config(text=tr("recap.selected.prefix") + ", ".join(parts))

        def recap_sync_to_checkbox() -> None:
            if recap_all_journal_var.get():
                recap_to_chk.config(state="disabled")
                return
            if recap_to_var.get():
                return
            if len(recap_selected_dates) > 1:
                recap_to_chk.config(state="disabled")
            else:
                recap_to_chk.config(state="normal")

        def recap_to_tooltip() -> str:
            if str(recap_to_chk.cget("state")) == "disabled":
                return tr("tip.recap_to_disabled")
            return tr("tip.recap_to")

        bind_hover_tooltip(recap_to_wrap, recap_to_tooltip)

        def on_recap_calendar_toggle(_evt: Optional[Any] = None) -> None:
            if recap_all_journal_var.get() or recap_to_var.get() or recap_calendar is None:
                return
            try:
                picked = recap_calendar.selection_get()
            except Exception:
                return
            if picked in recap_selected_dates:
                recap_selected_dates.remove(picked)
            else:
                recap_selected_dates.add(picked)
            recap_refresh_cal_marks()
            recap_update_sel_label()
            recap_sync_to_checkbox()

        if recap_calendar is not None:
            recap_calendar.bind("<<CalendarSelected>>", on_recap_calendar_toggle)

        recap_session: Dict[str, Any] = {"messages": [], "bootstrapped": False, "busy": False}
        recap_pending_images: List[Path] = []
        recap_pending_files: List[Path] = []

        def recap_refresh_date_controls() -> None:
            busy = bool(recap_session.get("busy"))
            if recap_all_journal_var.get():
                try:
                    recap_from_fr.grid_remove()
                except tk.TclError:
                    pass
                try:
                    recap_to_wrap.grid_remove()
                except tk.TclError:
                    pass
                try:
                    recap_through_fr.grid_remove()
                except tk.TclError:
                    pass
                try:
                    recap_cal_row.grid_remove()
                except tk.TclError:
                    pass
                if recap_from_de is not None:
                    try:
                        recap_from_de.config(state="disabled")
                    except tk.TclError:
                        pass
                if recap_to_de is not None:
                    try:
                        recap_to_de.config(state="disabled")
                    except tk.TclError:
                        pass
                recap_to_chk.config(state="disabled")
                if recap_calendar is not None:
                    try:
                        recap_calendar.config(state="disabled")
                    except tk.TclError:
                        pass
                recap_all_journal_chk.config(state=("disabled" if busy else "normal"))
                recap_update_sel_label()
                return
            try:
                recap_from_fr.grid(row=0, column=1, padx=(0, 10), pady=8, sticky="w")
            except tk.TclError:
                pass
            try:
                recap_to_wrap.grid(row=0, column=2, padx=(0, 8), pady=8, sticky="w")
            except tk.TclError:
                pass
            if recap_from_de is not None:
                try:
                    recap_from_de.config(state=("disabled" if busy else "normal"))
                except tk.TclError:
                    pass
            if recap_to_de is not None:
                try:
                    recap_to_de.config(state=("disabled" if busy else "normal"))
                except tk.TclError:
                    pass
            if recap_calendar is not None:
                try:
                    recap_calendar.config(state=("disabled" if busy else "normal"))
                except tk.TclError:
                    pass
            recap_all_journal_chk.config(state=("disabled" if busy else "normal"))
            if not recap_to_var.get():
                try:
                    recap_cal_row.grid(row=2, column=0, sticky="ew", pady=(0, 8))
                except tk.TclError:
                    pass
            else:
                try:
                    recap_cal_row.grid_remove()
                except tk.TclError:
                    pass
            recap_sync_to_checkbox()
            recap_update_sel_label()

        def on_recap_all_journal_toggle(*_a: Any) -> None:
            if recap_all_journal_var.get():
                recap_to_var.set(False)
            recap_refresh_date_controls()

        recap_all_journal_var.trace_add("write", on_recap_all_journal_toggle)

        def on_recap_to_mode(*_a: Any) -> None:
            if recap_to_var.get():
                if not recap_all_journal_var.get():
                    recap_through_fr.grid(row=0, column=3, padx=(0, 8), pady=8, sticky="w")
                    recap_cal_row.grid_remove()
                if len(recap_selected_dates) == 1 and DateEntry is not None:
                    only = next(iter(recap_selected_dates))
                    if recap_from_de is not None and recap_to_de is not None:
                        try:
                            recap_from_de.set_date(only)
                            recap_to_de.set_date(only)
                        except Exception:
                            pass
            else:
                recap_through_fr.grid_remove()
                if not recap_all_journal_var.get():
                    recap_cal_row.grid(row=2, column=0, sticky="ew", pady=(0, 8))
                recap_refresh_cal_marks()
                recap_update_sel_label()
                recap_sync_to_checkbox()
            recap_refresh_date_controls()

        recap_to_var.trace_add("write", lambda *_: on_recap_to_mode())
        recap_through_fr.grid_remove()
        recap_refresh_date_controls()

        recap_mid = tk.Frame(recap_wrap, bg=t0.surface)
        recap_mid.grid(row=3, column=0, sticky="nsew", pady=(0, 8))
        recap_mid.grid_rowconfigure(0, weight=1)
        recap_mid.grid_columnconfigure(0, weight=1)

        recap_transcript = tk.Text(
            recap_mid,
            wrap="word",
            state="disabled",
            font=("Segoe UI", 10),
            bg=t0.field,
            fg=t0.text,
            insertbackground=t0.text,
            relief="flat",
            padx=12,
            pady=12,
            highlightthickness=1,
            highlightbackground=t0.border,
            highlightcolor=t0.accent,
        )
        recap_ts = tk.Scrollbar(
            recap_mid,
            command=recap_transcript.yview,
            bg=t0.panel,
            troughcolor=t0.field,
            activebackground=t0.accent,
            bd=0,
            highlightthickness=0,
            width=11,
        )
        recap_transcript.configure(yscrollcommand=recap_ts.set)
        recap_transcript.grid(row=0, column=0, sticky="nsew")
        recap_ts.grid(row=0, column=1, sticky="ns")
        recap_transcript.tag_configure("t_meta", foreground=t0.muted, font=("Segoe UI", 9, "bold"))
        recap_transcript.tag_configure("t_user", foreground=t0.text, font=("Segoe UI", 10))
        recap_transcript.tag_configure("t_bot", foreground=t0.text, font=("Segoe UI", 10))

        recap_attach_row = tk.Frame(recap_wrap, bg=t0.surface)
        recap_attach_row.grid(row=4, column=0, sticky="ew", pady=(0, 4))
        recap_pending_lbl = tk.Label(
            recap_attach_row,
            text=tr("recap.attachments", what=tr("recap.attachments_none")),
            bg=t0.surface,
            fg=t0.muted,
            font=("Segoe UI", 9),
            anchor="w",
        )
        recap_pending_lbl.pack(side="left", fill="x", expand=True)

        def recap_refresh_pending_lbl() -> None:
            bits = []
            if recap_pending_images:
                bits.append(tr("recap.n_images", n=len(recap_pending_images)))
            if recap_pending_files:
                bits.append(tr("recap.n_files", n=len(recap_pending_files)))
            what = ", ".join(bits) if bits else tr("recap.attachments_none")
            recap_pending_lbl.config(text=tr("recap.attachments", what=what))

        def recap_pick_image() -> None:
            p = filedialog.askopenfilename(
                title="Attach image",
                filetypes=[
                    ("Images", "*.png *.jpg *.jpeg *.gif *.webp"),
                    ("All files", "*.*"),
                ],
            )
            if p:
                recap_pending_images.append(Path(p))
                recap_refresh_pending_lbl()

        def recap_pick_file() -> None:
            p = filedialog.askopenfilename(title="Attach file", filetypes=[("Text / data", "*.*")])
            if p:
                recap_pending_files.append(Path(p))
                recap_refresh_pending_lbl()

        recap_img_btn = tk.Button(
            recap_attach_row,
            text=tr("recap.image"),
            command=recap_pick_image,
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=4,
            cursor="hand2",
        )
        recap_img_btn.pack(side="right", padx=(6, 0))
        recap_file_btn = tk.Button(
            recap_attach_row,
            text=tr("recap.file"),
            command=recap_pick_file,
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=4,
            cursor="hand2",
        )
        recap_file_btn.pack(side="right", padx=(6, 0))

        recap_bottom = tk.Frame(recap_wrap, bg=t0.panel, highlightthickness=1, highlightbackground=t0.border)
        recap_bottom.grid(row=5, column=0, sticky="ew", pady=(0, 0))
        recap_bottom.grid_columnconfigure(0, weight=1)

        recap_input = tk.Text(
            recap_bottom,
            height=3,
            wrap="word",
            font=("Segoe UI", 10),
            bg=t0.field,
            fg=t0.text,
            insertbackground=t0.text,
            relief="flat",
            padx=10,
            pady=8,
            highlightthickness=0,
        )
        recap_input.grid(row=0, column=0, sticky="ew", padx=(8, 4), pady=8)

        recap_btn_fr = tk.Frame(recap_bottom, bg=t0.panel)
        recap_btn_fr.grid(row=0, column=1, sticky="ns", padx=(4, 8), pady=8)

        recap_send_btn = tk.Button(
            recap_btn_fr,
            text=tr("recap.send"),
            bg=t0.accent,
            fg="white",
            activebackground=t0.hover_primary,
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=8,
            cursor="hand2",
        )
        recap_send_btn.pack(fill="x", pady=(0, 6))
        recap_new_btn = tk.Button(
            recap_btn_fr,
            text=tr("recap.new_chat"),
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=6,
            cursor="hand2",
        )
        recap_new_btn.pack(fill="x")

        def _recap_send_rest_style() -> Tuple[str, str, str, str, str]:
            t = th()
            if str(recap_send_btn.cget("state")) != "normal":
                ds, gb, gf, dab, daf = t.gen_bind_disabled()
                return ds, gb, gf, dab, daf
            return ("normal", t.accent, "white", t.hover_primary, "white")

        bind_button_hover_if_enabled(
            recap_img_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            recap_file_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            recap_new_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            recap_send_btn,
            _recap_send_rest_style,
            lambda: th().hover_primary,
            lambda: "white",
        )

        _AI_SEND_SPIN = ("-", "/", "|", "\\")
        recap_send_spin: Dict[str, Any] = {"after_id": None, "i": 0}

        def _stop_recap_send_spinner() -> None:
            aid = recap_send_spin.get("after_id")
            if aid is not None:
                try:
                    root.after_cancel(aid)
                except (tk.TclError, ValueError):
                    pass
                recap_send_spin["after_id"] = None

        def _start_recap_send_spinner() -> None:
            _stop_recap_send_spinner()

            def _tick() -> None:
                if not recap_session.get("busy"):
                    recap_send_spin["after_id"] = None
                    return
                try:
                    i = recap_send_spin["i"] % len(_AI_SEND_SPIN)
                    recap_send_btn.config(text=tr("ai.send_busy_prefix") + _AI_SEND_SPIN[i])
                    recap_send_spin["i"] = recap_send_spin["i"] + 1
                    recap_send_spin["after_id"] = root.after(130, _tick)
                except tk.TclError:
                    recap_send_spin["after_id"] = None

            recap_send_spin["i"] = 0
            _tick()

        def recap_set_sending(sending: bool) -> None:
            recap_session["busy"] = sending
            st = "disabled" if sending else "normal"
            recap_send_btn.config(state=st)
            recap_new_btn.config(state=st)
            recap_img_btn.config(state=st)
            recap_file_btn.config(state=st)
            recap_input.config(state=st)
            recap_thinking_chk.config(state=st)
            if sending:
                _start_recap_send_spinner()
            else:
                _stop_recap_send_spinner()
                try:
                    recap_send_btn.config(text=tr("recap.send"))
                except tk.TclError:
                    pass
            recap_refresh_date_controls()

        def reset_recap_session(*_a: Any) -> None:
            recap_session["messages"].clear()
            recap_session["bootstrapped"] = False
            recap_session["busy"] = False
            recap_all_journal_var.set(True)
            recap_pending_images.clear()
            recap_pending_files.clear()
            recap_refresh_pending_lbl()
            recap_transcript.config(state="normal")
            recap_transcript.delete("1.0", "end")
            recap_transcript.config(state="disabled")
            recap_input.delete("1.0", "end")
            recap_selected_dates.clear()
            recap_update_sel_label()
            recap_refresh_cal_marks()
            recap_set_sending(False)

        def reset_recap_on_page_leave() -> None:
            reset_recap_session()
            recap_to_var.set(False)
            try:
                td = datetime.now().date()
                if recap_from_de is not None:
                    recap_from_de.set_date(td)
                if recap_to_de is not None:
                    recap_to_de.set_date(td)
            except Exception:
                pass
            recap_sync_to_checkbox()
            on_recap_to_mode()

        page_leave_reset_handlers["ai_recap"] = reset_recap_on_page_leave

        def recap_new_chat() -> None:
            if recap_session.get("busy"):
                return
            reset_recap_session()

        recap_new_btn.config(command=recap_new_chat)

        def recap_build_context() -> Optional[str]:
            if recap_all_journal_var.get():
                return build_journal_context()
            if recap_to_var.get():
                if DateEntry is None or recap_from_de is None or recap_to_de is None:
                    messagebox.showerror(tr("msg.ai_recap"), tr("recap.err.tkcal_range"))
                    return None
                try:
                    d0 = recap_from_de.get_date()
                    d1 = recap_to_de.get_date()
                except Exception as exc:
                    messagebox.showerror(tr("msg.ai_recap"), tr("recap.err.read_range", err=str(exc)))
                    return None
                start = datetime.combine(d0, datetime.min.time())
                end = datetime.combine(d1, datetime.min.time())
                if end < start:
                    start, end = end, start
                return build_journal_context_for_range((start, end))
            if recap_selected_dates:
                return build_journal_context_for_date_set(recap_selected_dates)
            if DateEntry is None or recap_from_de is None:
                messagebox.showerror(tr("msg.ai_recap"), tr("recap.err.from_tkcal"))
                return None
            try:
                only = recap_from_de.get_date()
            except Exception as exc:
                messagebox.showerror(tr("msg.ai_recap"), tr("recap.err.read_from", err=str(exc)))
                return None
            return build_journal_context_for_date_set({only})

        def recap_send() -> None:
            if recap_session["busy"]:
                return
            if not get_openai_api_key():
                go_settings = messagebox.askyesno(
                    tr("msg.no_api_key_use_ai_title"),
                    tr("msg.no_api_key_use_ai_body"),
                )
                if go_settings:
                    goto_tok = api_key_prompt_hooks.get("goto_token")
                    if callable(goto_tok):
                        goto_tok()
                return
            text = recap_input.get("1.0", "end-1c").strip()
            if not text and not recap_pending_images and not recap_pending_files:
                return
            ctx = recap_build_context()
            if ctx is None:
                return
            model_name = OPENAI_THINKING_MODEL if recap_thinking_var.get() else OPENAI_MODEL
            effort = "high" if recap_thinking_var.get() else None
            imgs = list(recap_pending_images)
            files = list(recap_pending_files)
            recap_set_sending(True)

            def kickoff() -> None:
                try:
                    if not recap_session["bootstrapped"]:
                        recap_session["messages"] = [
                            {
                                "role": "system",
                                "content": (
                                    "You answer questions only using the user's journal context. "
                                    "If the answer is not in the journal, say you do not know based on the journal."
                                ),
                            },
                            {"role": "system", "content": f"Journal context:\n{ctx}"},
                        ]
                        recap_session["bootstrapped"] = True
                    user_msg = build_user_message_with_attachments(text, imgs, files)
                    recap_session["messages"].append(user_msg)
                    answer = chat_completion(
                        recap_session["messages"],
                        model=model_name,
                        reasoning_effort=effort,
                    )

                    def done() -> None:
                        recap_set_sending(False)
                        if _is_likely_api_error_message(answer):
                            messagebox.showerror(tr("msg.ai_recap"), answer[:4000])
                            if recap_session["messages"] and recap_session["messages"][-1].get("role") == "user":
                                recap_session["messages"].pop()
                            return
                        recap_session["messages"].append({"role": "assistant", "content": answer})
                        _append_transcript(recap_transcript, "user", text or tr("chat.attachment_only"))
                        _append_transcript(recap_transcript, "assistant", answer)
                        recap_input.delete("1.0", "end")
                        recap_pending_images.clear()
                        recap_pending_files.clear()
                        recap_refresh_pending_lbl()

                    root.after(0, done)
                except Exception as exc:

                    def fail() -> None:
                        recap_set_sending(False)
                        messagebox.showerror(tr("msg.ai_recap"), str(exc))
                        if recap_session["messages"] and recap_session["messages"][-1].get("role") == "user":
                            recap_session["messages"].pop()

                    root.after(0, fail)

            threading.Thread(target=kickoff, daemon=True).start()

        recap_send_btn.config(command=recap_send)

        def recap_on_enter_key(event: Any) -> Optional[str]:
            # Text widget: <Return> alone may not consume the key; use KeyPress-Return / KP_Enter.
            if (getattr(event, "state", 0) or 0) & 0x0001:
                return None
            if (getattr(event, "state", 0) or 0) & 0x0004:
                return None
            recap_send()
            return "break"

        recap_input.bind("<KeyPress-Return>", recap_on_enter_key, add="+")
        recap_input.bind("<KeyPress-KP_Enter>", recap_on_enter_key, add="+")

        # ========== Chatbot ==========
        cb_wrap = tk.Frame(chatbot_page, bg=t0.surface)
        cb_wrap.pack(
            fill="both",
            expand=True,
            padx=t0.pad_outer,
            pady=(0, t0.pad_center_y + JOURNAL_WINDOW_CONSOLE_RESERVE_BOTTOM),
        )
        cb_wrap.grid_columnconfigure(0, weight=1)
        cb_wrap.grid_rowconfigure(1, weight=1)

        cb_title = tk.Label(
            cb_wrap,
            text=tr("chatbot.title"),
            bg=t0.surface,
            fg=t0.text,
            font=("Segoe UI", 15, "bold"),
            anchor="w",
        )
        cb_title.grid(row=0, column=0, sticky="w", pady=(0, 8))

        cb_top = tk.Frame(cb_wrap, bg=t0.panel, highlightthickness=1, highlightbackground=t0.border)
        cb_top.grid(row=1, column=0, sticky="nsew", pady=(0, 8))
        cb_top.grid_rowconfigure(0, weight=1)
        cb_top.grid_columnconfigure(0, weight=1)

        cb_transcript = tk.Text(
            cb_top,
            wrap="word",
            state="disabled",
            font=("Segoe UI", 10),
            bg=t0.field,
            fg=t0.text,
            insertbackground=t0.text,
            relief="flat",
            padx=12,
            pady=12,
            highlightthickness=1,
            highlightbackground=t0.border,
            highlightcolor=t0.accent,
        )
        cb_ts = tk.Scrollbar(
            cb_top,
            command=cb_transcript.yview,
            bg=t0.panel,
            troughcolor=t0.field,
            activebackground=t0.accent,
            bd=0,
            highlightthickness=0,
            width=11,
        )
        cb_transcript.configure(yscrollcommand=cb_ts.set)
        cb_transcript.grid(row=0, column=0, sticky="nsew")
        cb_ts.grid(row=0, column=1, sticky="ns")
        cb_transcript.tag_configure("t_meta", foreground=t0.muted, font=("Segoe UI", 9, "bold"))
        cb_transcript.tag_configure("t_user", foreground=t0.text, font=("Segoe UI", 10))
        cb_transcript.tag_configure("t_bot", foreground=t0.text, font=("Segoe UI", 10))

        cb_attach = tk.Frame(cb_wrap, bg=t0.surface)
        cb_attach.grid(row=2, column=0, sticky="ew", pady=(0, 4))
        cb_thinking_var = tk.BooleanVar(value=True)
        cb_thinking_chk = tk.Checkbutton(
            cb_attach,
            text=tr("chatbot.thinking"),
            variable=cb_thinking_var,
            bg=t0.surface,
            fg=t0.muted,
            activebackground=t0.surface,
            activeforeground=t0.text,
            selectcolor=t0.field,
            font=("Segoe UI", 9),
        )
        cb_thinking_chk.pack(side="left")
        bind_hover_tooltip(cb_thinking_chk, lambda: tr("tip.thinking_model"))
        cb_pending_lbl = tk.Label(
            cb_attach,
            text=tr("recap.attachments", what=tr("recap.attachments_none")),
            bg=t0.surface,
            fg=t0.muted,
            font=("Segoe UI", 9),
            anchor="w",
        )
        cb_pending_lbl.pack(side="left", fill="x", expand=True, padx=(12, 0))

        cb_session: Dict[str, Any] = {
            "messages": [{"role": "system", "content": "You are a helpful assistant."}],
            "busy": False,
        }
        cb_pending_images: List[Path] = []
        cb_pending_files: List[Path] = []

        def cb_refresh_pending() -> None:
            bits = []
            if cb_pending_images:
                bits.append(tr("recap.n_images", n=len(cb_pending_images)))
            if cb_pending_files:
                bits.append(tr("recap.n_files", n=len(cb_pending_files)))
            what = ", ".join(bits) if bits else tr("recap.attachments_none")
            cb_pending_lbl.config(text=tr("recap.attachments", what=what))

        def cb_pick_image() -> None:
            p = filedialog.askopenfilename(
                title="Attach image",
                filetypes=[("Images", "*.png *.jpg *.jpeg *.gif *.webp"), ("All files", "*.*")],
            )
            if p:
                cb_pending_images.append(Path(p))
                cb_refresh_pending()

        def cb_pick_file() -> None:
            p = filedialog.askopenfilename(title="Attach file", filetypes=[("All files", "*.*")])
            if p:
                cb_pending_files.append(Path(p))
                cb_refresh_pending()

        cb_img_btn = tk.Button(
            cb_attach,
            text=tr("recap.image"),
            command=cb_pick_image,
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=4,
            cursor="hand2",
        )
        cb_img_btn.pack(side="right", padx=(6, 0))
        cb_file_btn = tk.Button(
            cb_attach,
            text=tr("recap.file"),
            command=cb_pick_file,
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=4,
            cursor="hand2",
        )
        cb_file_btn.pack(side="right", padx=(6, 0))

        cb_bottom = tk.Frame(cb_wrap, bg=t0.panel, highlightthickness=1, highlightbackground=t0.border)
        cb_bottom.grid(row=3, column=0, sticky="ew")
        cb_bottom.grid_columnconfigure(0, weight=1)

        cb_input = tk.Text(
            cb_bottom,
            height=3,
            wrap="word",
            font=("Segoe UI", 10),
            bg=t0.field,
            fg=t0.text,
            insertbackground=t0.text,
            relief="flat",
            padx=10,
            pady=8,
            highlightthickness=0,
        )
        cb_input.grid(row=0, column=0, sticky="ew", padx=(8, 4), pady=8)

        cb_btn_fr = tk.Frame(cb_bottom, bg=t0.panel)
        cb_btn_fr.grid(row=0, column=1, sticky="ns", padx=(4, 8), pady=8)

        cb_send_btn = tk.Button(
            cb_btn_fr,
            text=tr("recap.send"),
            bg=t0.accent,
            fg="white",
            activebackground=t0.hover_primary,
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=8,
            cursor="hand2",
        )
        cb_send_btn.pack(fill="x", pady=(0, 6))
        cb_new_btn = tk.Button(
            cb_btn_fr,
            text=tr("recap.new_chat"),
            bg=_tb,
            fg=_tf,
            activebackground=_tab,
            activeforeground=_taf,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=6,
            cursor="hand2",
        )
        cb_new_btn.pack(fill="x")

        def _cb_send_rest_style() -> Tuple[str, str, str, str, str]:
            t = th()
            if str(cb_send_btn.cget("state")) != "normal":
                ds, gb, gf, dab, daf = t.gen_bind_disabled()
                return ds, gb, gf, dab, daf
            return ("normal", t.accent, "white", t.hover_primary, "white")

        bind_button_hover_if_enabled(
            cb_img_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            cb_file_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            cb_new_btn,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )
        bind_button_hover_if_enabled(
            cb_send_btn,
            _cb_send_rest_style,
            lambda: th().hover_primary,
            lambda: "white",
        )

        cb_send_spin: Dict[str, Any] = {"after_id": None, "i": 0}

        def _stop_cb_send_spinner() -> None:
            aid = cb_send_spin.get("after_id")
            if aid is not None:
                try:
                    root.after_cancel(aid)
                except (tk.TclError, ValueError):
                    pass
                cb_send_spin["after_id"] = None

        def _start_cb_send_spinner() -> None:
            _stop_cb_send_spinner()

            def _tick() -> None:
                if not cb_session.get("busy"):
                    cb_send_spin["after_id"] = None
                    return
                try:
                    i = cb_send_spin["i"] % len(_AI_SEND_SPIN)
                    cb_send_btn.config(text=tr("ai.send_busy_prefix") + _AI_SEND_SPIN[i])
                    cb_send_spin["i"] = cb_send_spin["i"] + 1
                    cb_send_spin["after_id"] = root.after(130, _tick)
                except tk.TclError:
                    cb_send_spin["after_id"] = None

            cb_send_spin["i"] = 0
            _tick()

        def cb_set_sending(sending: bool) -> None:
            cb_session["busy"] = sending
            st = "disabled" if sending else "normal"
            cb_send_btn.config(state=st)
            cb_new_btn.config(state=st)
            cb_img_btn.config(state=st)
            cb_file_btn.config(state=st)
            cb_input.config(state=st)
            cb_thinking_chk.config(state=st)
            if sending:
                _start_cb_send_spinner()
            else:
                _stop_cb_send_spinner()
                try:
                    cb_send_btn.config(text=tr("recap.send"))
                except tk.TclError:
                    pass

        def reset_chatbot_session(*_a: Any) -> None:
            cb_session["messages"] = [{"role": "system", "content": "You are a helpful assistant."}]
            cb_session["busy"] = False
            cb_pending_images.clear()
            cb_pending_files.clear()
            cb_refresh_pending()
            cb_transcript.config(state="normal")
            cb_transcript.delete("1.0", "end")
            cb_transcript.config(state="disabled")
            cb_input.delete("1.0", "end")
            cb_set_sending(False)

        page_leave_reset_handlers["chatbot"] = reset_chatbot_session

        def cb_new_chat() -> None:
            if cb_session.get("busy"):
                return
            reset_chatbot_session()

        cb_new_btn.config(command=cb_new_chat)

        def cb_send() -> None:
            if cb_session["busy"]:
                return
            if not get_openai_api_key():
                if messagebox.askyesno(
                    tr("msg.no_api_key_use_ai_title"),
                    tr("msg.no_api_key_use_ai_body"),
                ):
                    goto_tok = api_key_prompt_hooks.get("goto_token")
                    if callable(goto_tok):
                        goto_tok()
                return
            text = cb_input.get("1.0", "end-1c").strip()
            if not text and not cb_pending_images and not cb_pending_files:
                return
            model_name = OPENAI_THINKING_MODEL if cb_thinking_var.get() else OPENAI_MODEL
            effort = "high" if cb_thinking_var.get() else None
            imgs = list(cb_pending_images)
            files = list(cb_pending_files)
            cb_set_sending(True)

            def kickoff() -> None:
                try:
                    user_msg = build_user_message_with_attachments(text, imgs, files)
                    cb_session["messages"].append(user_msg)
                    answer = chat_completion(
                        cb_session["messages"],
                        model=model_name,
                        reasoning_effort=effort,
                    )

                    def done() -> None:
                        cb_set_sending(False)
                        if _is_likely_api_error_message(answer):
                            messagebox.showerror(tr("msg.chatbot"), answer[:4000])
                            if cb_session["messages"] and cb_session["messages"][-1].get("role") == "user":
                                cb_session["messages"].pop()
                            return
                        cb_session["messages"].append({"role": "assistant", "content": answer})
                        _append_transcript(cb_transcript, "user", text or tr("chat.attachment_only"))
                        _append_transcript(cb_transcript, "assistant", answer)
                        cb_input.delete("1.0", "end")
                        cb_pending_images.clear()
                        cb_pending_files.clear()
                        cb_refresh_pending()

                    root.after(0, done)
                except Exception as exc:

                    def fail() -> None:
                        cb_set_sending(False)
                        messagebox.showerror(tr("msg.chatbot"), str(exc))
                        if cb_session["messages"] and cb_session["messages"][-1].get("role") == "user":
                            cb_session["messages"].pop()

                    root.after(0, fail)

            threading.Thread(target=kickoff, daemon=True).start()

        cb_send_btn.config(command=cb_send)

        def cb_on_enter_key(event: Any) -> Optional[str]:
            if (getattr(event, "state", 0) or 0) & 0x0001:
                return None
            if (getattr(event, "state", 0) or 0) & 0x0004:
                return None
            cb_send()
            return "break"

        cb_input.bind("<KeyPress-Return>", cb_on_enter_key, add="+")
        cb_input.bind("<KeyPress-KP_Enter>", cb_on_enter_key, add="+")

        def apply_ai_recap_chatbot_theme() -> None:
            t = th()
            tb, tf, tab, taf = t.toolbar_btn_config()
            for fr in (
                recap_wrap,
                recap_title,
                recap_top,
                recap_thinking_chk,
                recap_all_journal_chk,
                recap_from_fr,
                recap_to_wrap,
                recap_to_chk,
                recap_through_fr,
                recap_cal_row,
                recap_sel_lbl,
                recap_mid,
                recap_attach_row,
                recap_pending_lbl,
                recap_bottom,
                recap_btn_fr,
                cb_wrap,
                cb_title,
                cb_attach,
                cb_pending_lbl,
                cb_bottom,
                cb_btn_fr,
            ):
                try:
                    fr.configure(bg=t.surface)
                except Exception:
                    try:
                        fr.configure(bg=t.panel)
                    except Exception:
                        pass
            recap_title.configure(bg=t.surface, fg=t.text)
            recap_top.configure(bg=t.panel, highlightbackground=t.border)
            recap_thinking_chk.configure(
                bg=t.panel,
                fg=t.muted,
                activebackground=t.panel,
                activeforeground=t.text,
                selectcolor=t.field,
            )
            recap_all_journal_chk.configure(
                bg=t.panel,
                fg=t.muted,
                activebackground=t.panel,
                activeforeground=t.text,
                selectcolor=t.field,
            )
            recap_to_wrap.configure(bg=t.panel)
            recap_to_chk.configure(
                bg=t.panel,
                fg=t.muted,
                activebackground=t.panel,
                activeforeground=t.text,
                selectcolor=t.field,
            )
            recap_from_fr.configure(bg=t.panel)
            for _w in recap_from_fr.winfo_children():
                if isinstance(_w, tk.Label):
                    _w.configure(bg=t.panel, fg=t.muted)
            recap_through_fr.configure(bg=t.panel)
            for _w in recap_through_fr.winfo_children():
                if isinstance(_w, tk.Label):
                    _w.configure(bg=t.panel, fg=t.muted)
            recap_cal_row.configure(bg=t.surface)
            recap_sel_lbl.configure(bg=t.surface, fg=t.muted)
            recap_mid.configure(bg=t.surface)
            recap_transcript.config(
                bg=t.field,
                fg=t.text,
                insertbackground=t.text,
                highlightbackground=t.border,
                highlightcolor=t.accent,
            )
            recap_ts.config(bg=t.panel, troughcolor=t.field, activebackground=t.accent)
            recap_transcript.tag_configure("t_meta", foreground=t.muted)
            recap_transcript.tag_configure("t_user", foreground=t.text)
            recap_transcript.tag_configure("t_bot", foreground=t.text)
            recap_attach_row.configure(bg=t.surface)
            recap_pending_lbl.configure(bg=t.surface, fg=t.muted)
            recap_bottom.configure(bg=t.panel, highlightbackground=t.border)
            recap_input.config(bg=t.field, fg=t.text, insertbackground=t.text)
            recap_btn_fr.configure(bg=t.panel)
            recap_send_btn.configure(
                bg=t.accent,
                fg="white",
                activebackground=t.hover_primary,
                activeforeground="white",
            )
            recap_new_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)
            recap_img_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)
            recap_file_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)
            if recap_calendar is not None:
                try:
                    recap_calendar.config(
                        background=t.field,
                        foreground=t.text,
                        headersbackground=t.panel,
                        headersforeground=t.text,
                        weekendbackground=t.field,
                        weekendforeground=t.muted,
                        normalbackground=t.field,
                        normalforeground=t.text,
                        othermonthbackground=t.field,
                        othermonthforeground=t.muted,
                        selectbackground=t.accent,
                        selectforeground="white",
                        bordercolor=t.border,
                    )
                except Exception:
                    pass
                recap_refresh_cal_marks()
            if DateEntry is not None and recap_from_de is not None:
                try:
                    recap_from_de.config(background=t.field, foreground=t.text)
                except tk.TclError:
                    pass
            if DateEntry is not None and recap_to_de is not None:
                try:
                    recap_to_de.config(background=t.field, foreground=t.text)
                except tk.TclError:
                    pass
            cb_wrap.configure(bg=t.surface)
            cb_title.configure(bg=t.surface, fg=t.text)
            cb_top.configure(bg=t.panel, highlightbackground=t.border)
            cb_transcript.config(
                bg=t.field,
                fg=t.text,
                insertbackground=t.text,
                highlightbackground=t.border,
                highlightcolor=t.accent,
            )
            cb_ts.config(bg=t.panel, troughcolor=t.field, activebackground=t.accent)
            cb_transcript.tag_configure("t_meta", foreground=t.muted)
            cb_transcript.tag_configure("t_user", foreground=t.text)
            cb_transcript.tag_configure("t_bot", foreground=t.text)
            cb_attach.configure(bg=t.surface)
            cb_thinking_chk.configure(
                bg=t.surface,
                fg=t.muted,
                activebackground=t.surface,
                activeforeground=t.text,
                selectcolor=t.field,
            )
            cb_pending_lbl.configure(bg=t.surface, fg=t.muted)
            cb_bottom.configure(bg=t.panel, highlightbackground=t.border)
            cb_input.config(bg=t.field, fg=t.text, insertbackground=t.text)
            cb_btn_fr.configure(bg=t.panel)
            cb_send_btn.configure(
                bg=t.accent,
                fg="white",
                activebackground=t.hover_primary,
                activeforeground="white",
            )
            cb_new_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)
            cb_img_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)
            cb_file_btn.configure(bg=tb, fg=tf, activebackground=tab, activeforeground=taf)

        def refresh_recap_chat_i18n() -> None:
            try:
                _has_key = bool(get_openai_api_key())
                recap_title.config(
                    text=tr("recap.title" if _has_key else "recap.title_no_key")
                )
                recap_thinking_chk.config(text=tr("recap.thinking"))
                recap_all_journal_chk.config(text=tr("recap.all_journal"))
                recap_from_lbl.config(text=tr("recap.from"))
                recap_to_chk.config(text=tr("recap.to_chk"))
                recap_through_lbl.config(text=tr("recap.through"))
                recap_img_btn.config(text=tr("recap.image"))
                recap_file_btn.config(text=tr("recap.file"))
                if not recap_session.get("busy"):
                    recap_send_btn.config(text=tr("recap.send"))
                recap_new_btn.config(text=tr("recap.new_chat"))
                recap_update_sel_label()
                recap_refresh_pending_lbl()
                cb_title.config(
                    text=tr("chatbot.title" if _has_key else "chatbot.title_no_key")
                )
                cb_thinking_chk.config(text=tr("chatbot.thinking"))
                cb_img_btn.config(text=tr("recap.image"))
                cb_file_btn.config(text=tr("recap.file"))
                if not cb_session.get("busy"):
                    cb_send_btn.config(text=tr("recap.send"))
                cb_new_btn.config(text=tr("recap.new_chat"))
                cb_refresh_pending()
            except tk.TclError:
                pass

        build_ai_recap_and_chatbot_pages._i18n = refresh_recap_chat_i18n  # type: ignore[attr-defined]
        build_ai_recap_and_chatbot_pages._apply_theme = apply_ai_recap_chatbot_theme  # type: ignore[attr-defined]

    build_ai_recap_and_chatbot_pages()
    settings_wrap = tk.Frame(settings_page, bg=t_init.surface)
    settings_wrap.pack(fill="both", expand=True, padx=20, pady=20)
    _register_page_toggle(settings_page)
    settings_title = tk.Label(
        settings_wrap,
        text=tr("settings.title"),
        bg=t_init.surface,
        fg=t_init.text,
        font=("Segoe UI", 16, "bold"),
        anchor="w",
    )
    settings_title.pack(anchor="w", pady=(0, 12))
    settings_status_var = tk.StringVar(value="")
    settings_status_lbl = tk.Label(
        settings_wrap,
        textvariable=settings_status_var,
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
        justify="left",
    )
    settings_status_lbl.pack(fill="x", pady=(0, 10))

    settings_rows: List[Any] = []
    settings_labels: List[Any] = []
    settings_label_keys: List[Tuple[Any, str]] = []

    def _bind_settings_tooltip(widget: Any, label_key: str) -> None:
        tip_key = f"tip.{label_key}"
        bind_hover_tooltip(widget, lambda _tip_key=tip_key: tr(_tip_key))

    def _make_settings_row(label_key: str) -> Tuple[Any, Any]:
        row = tk.Frame(settings_wrap, bg=t_init.surface)
        row.pack(fill="x", pady=(0, 10))
        lbl = tk.Label(
            row,
            text=tr(label_key),
            bg=t_init.surface,
            fg=t_init.muted,
            font=("Segoe UI", 10, "bold"),
            width=18,
            anchor="w",
        )
        lbl.pack(side="left")
        settings_rows.append(row)
        settings_labels.append(lbl)
        settings_label_keys.append((lbl, label_key))
        _bind_settings_tooltip(row, label_key)
        _bind_settings_tooltip(lbl, label_key)
        return row, lbl

    settings_prefs = load_preferences()
    settings_app_name = {"value": settings_prefs.get("app_name", "Daily Logger") or "Daily Logger"}
    console_hint_state: Dict[str, Any] = {"text": "", "apply": None, "reset_after_id": None}

    def _console_update_timestamp() -> str:
        return datetime.now().strftime("%I:%M:%S%p").lstrip("0")

    def _set_console_temp_message(msg: str, *, timeout_ms: int = 10000) -> None:
        text = (msg or "").strip()
        if not text:
            return
        console_hint_state["text"] = text
        _id = console_hint_state.get("reset_after_id")
        if _id is not None:
            try:
                root.after_cancel(_id)
            except Exception:
                pass
        console_hint_state["reset_after_id"] = root.after(timeout_ms, _clear_console_hint)
        apply_hint = console_hint_state.get("apply")
        if callable(apply_hint):
            apply_hint()

    def _append_console_session_update(msg: str, *, key: str = "") -> None:
        text = (msg or "").strip()
        if not text:
            return
        dedupe_key = key or text
        if console_session_last.get(dedupe_key) == text:
            return
        console_session_last[dedupe_key] = text
        line = f"[{_console_update_timestamp()}] {text}"
        console_session_updates.append(line)
        if len(console_session_updates) > 1000:
            del console_session_updates[: len(console_session_updates) - 1000]
        widget = console_output_holder.get("widget")
        if widget is None:
            return
        try:
            widget.config(state="normal")
            widget.insert("end", line + "\n")
            widget.see("end")
            widget.config(state="disabled")
        except tk.TclError:
            pass

    def _console_update_is_noisy_progress(text: str) -> bool:
        lowered = (text or "").casefold()
        if lowered.startswith("receiving ") and " / " in lowered:
            return True
        if lowered.startswith("transcribing ") and re.search(r"\(\d+%\)$", text or ""):
            return True
        return False

    def _publish_console_update(msg: str, *, key: str = "", temp: bool = True, log: Optional[bool] = None) -> None:
        text = (msg or "").strip()
        if not text:
            return
        if temp:
            _set_console_temp_message(text)
        should_log = (not _console_update_is_noisy_progress(text)) if log is None else bool(log)
        if should_log:
            _append_console_session_update(text, key=key)

    def _clear_console_hint() -> None:
        console_hint_state["text"] = ""
        _id = console_hint_state.get("reset_after_id")
        if _id is not None:
            try:
                root.after_cancel(_id)
            except Exception:
                pass
            console_hint_state["reset_after_id"] = None
        apply_hint = console_hint_state.get("apply")
        if callable(apply_hint):
            apply_hint()

    def _set_settings_status(msg: str) -> None:
        settings_status_var.set("")
        _publish_console_update(msg, key=f"settings:{msg}")

    lang_row, _ = _make_settings_row("settings.language")
    ui_lang_var = tk.StringVar(
        value=tr("settings.lang.chinese")
        if ui_lang_holder[0] == "zh"
        else tr("settings.lang.english")
    )
    lang_ui_combo: Any = None
    if ttk is not None:
        lang_ui_combo = ttk.Combobox(
            lang_row,
            textvariable=ui_lang_var,
            values=(tr("settings.lang.english"), tr("settings.lang.chinese")),
            state="readonly",
            width=14,
            style="Journal.TCombobox",
        )
        lang_ui_combo.pack(side="left", fill="x", expand=True, padx=(0, 8))
    else:
        lang_ui_combo = tk.OptionMenu(
            lang_row,
            ui_lang_var,
            tr("settings.lang.english"),
            tr("settings.lang.chinese"),
        )
        lang_ui_combo.pack(side="left", fill="x", expand=True, padx=(0, 8))
    _bind_settings_tooltip(lang_ui_combo, "settings.language")

    def _on_ui_language_selected(_evt: object | None = None) -> None:
        raw = ui_lang_var.get().strip()
        new_lang = "zh" if raw == tr("settings.lang.chinese") else "en"
        if new_lang == ui_lang_holder[0]:
            return
        prefs = load_preferences()
        prefs[UI_LANGUAGE_PREF_KEY] = new_lang
        save_preferences(prefs)
        ui_lang_holder[0] = new_lang
        apply_journal_window_colors()

    rename_row, _ = _make_settings_row("settings.rename")
    rename_entry = tk.Entry(
        rename_row,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
        font=("Segoe UI", 10),
    )
    rename_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
    rename_entry.insert(0, settings_app_name["value"])
    rename_btn = tk.Button(
        rename_row,
        text=tr("settings.rename_btn"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    rename_btn.pack(side="left")
    _bind_settings_tooltip(rename_entry, "settings.rename")
    _bind_settings_tooltip(rename_btn, "settings.rename")

    startup_row, _ = _make_settings_row("settings.startup")
    startup_state = {"enabled": is_startup_enabled()}
    startup_toggle_btn = tk.Button(
        startup_row,
        text=tr("settings.on") if startup_state["enabled"] else tr("settings.off"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=7,
    )
    startup_toggle_btn.pack(side="left")
    _bind_settings_tooltip(startup_toggle_btn, "settings.startup")

    iphone_receive_row, _ = _make_settings_row("settings.iphone_receive")
    iphone_receive_state = {"enabled": iphone_passive_receive_enabled()}
    iphone_receive_toggle_btn = tk.Button(
        iphone_receive_row,
        text=tr("settings.on") if iphone_receive_state["enabled"] else tr("settings.off"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=7,
    )
    iphone_receive_toggle_btn.pack(side="left")
    _bind_settings_tooltip(iphone_receive_toggle_btn, "settings.iphone_receive")

    transcription_models_row, _ = _make_settings_row("settings.transcription_models")
    transcription_models_btn = tk.Button(
        transcription_models_row,
        text=tr("settings.manage"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=9,
    )
    transcription_models_btn.pack(side="left")
    _bind_settings_tooltip(transcription_models_btn, "settings.transcription_models")

    updates_row, _ = _make_settings_row("settings.updates")
    updates_state = {"enabled": update_check_enabled(), "busy": False}
    updates_toggle_btn = tk.Button(
        updates_row,
        text=tr("settings.on") if updates_state["enabled"] else tr("settings.off"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=7,
    )
    updates_toggle_btn.pack(side="left", padx=(0, 8))
    _bind_settings_tooltip(updates_toggle_btn, "settings.updates")
    updates_check_btn = tk.Button(
        updates_row,
        text=tr("settings.check_now"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=12,
    )
    updates_check_btn.pack(side="left")
    _bind_settings_tooltip(updates_check_btn, "settings.updates")
    updates_status_var = tk.StringVar(value="")
    updates_status_lbl = tk.Label(
        updates_row,
        textvariable=updates_status_var,
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
    )
    updates_status_lbl.pack(side="left", fill="x", expand=True, padx=(10, 0))

    def _refresh_iphone_settings_toggle_btn() -> None:
        iphone_receive_state["enabled"] = iphone_passive_receive_enabled()
        try:
            iphone_receive_toggle_btn.config(
                text=tr("settings.on") if iphone_receive_state["enabled"] else tr("settings.off")
            )
        except tk.TclError:
            pass

    globals()["_daily_logger_refresh_iphone_settings_toggle"] = _refresh_iphone_settings_toggle_btn

    theme_row, _ = _make_settings_row("settings.theme")
    settings_theme_btn = tk.Button(
        theme_row,
        text=tr(journal_window_theme_current_label_key(t_init.id)),
        command=lambda: toggle_journal_window_theme(),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
    )
    settings_theme_btn.pack(side="left")
    _bind_settings_tooltip(settings_theme_btn, "settings.theme")

    display_row, _ = _make_settings_row("settings.display")
    display_fullscreen_btn = tk.Button(
        display_row,
        text=tr("settings.on") if fullscreen_state["enabled"] else tr("settings.off"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=7,
    )
    display_fullscreen_btn.pack(side="left", padx=(0, 8))
    _bind_settings_tooltip(display_fullscreen_btn, "settings.display")
    display_f11_lbl = tk.Label(
        display_row,
        text=tr("settings.display_f11"),
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 9),
        anchor="w",
    )
    display_f11_lbl.pack(side="left", fill="x", expand=True)
    _bind_settings_tooltip(display_f11_lbl, "settings.display")

    def _backup_mode_btn_label(mode_val: str) -> str:
        return tr(
            {"On": "backup.on", "Off": "backup.off", "Limited": "backup.limited"}.get(
                mode_val, "backup.off"
            )
        )

    backup_row, _ = _make_settings_row("settings.backup")
    backup_mode = {"value": "On"}
    if _is_pref_true(settings_prefs.get("backup_limited", "false")):
        backup_mode["value"] = "Limited"
    elif not _is_pref_true(settings_prefs.get("backup_enabled", "true")):
        backup_mode["value"] = "Off"
    backup_mode_btn = tk.Button(
        backup_row,
        text=_backup_mode_btn_label(backup_mode["value"]),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=9,
    )
    backup_mode_btn.pack(side="left", padx=(0, 8))
    backup_manual_btn = tk.Button(
        backup_row,
        text=tr("settings.manual"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
        width=9,
    )
    backup_manual_btn.pack(side="left")
    bind_hover_tooltip(
        backup_mode_btn,
        lambda: tr("tip.backup_mode"),
    )
    bind_hover_tooltip(
        backup_manual_btn,
        lambda: tr("tip.backup_manual"),
    )

    token_row, _ = _make_settings_row("settings.token")
    token_saved = {"value": get_openai_api_key() or ""}
    token_entry = tk.Entry(
        token_row,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
        font=("Consolas", 10),
    )
    token_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
    if token_saved["value"]:
        token_entry.insert(0, "*" * max(32, len(token_saved["value"])))
    token_save_btn = tk.Button(
        token_row,
        text=tr("settings.save"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
        width=7,
    )
    token_save_btn.pack(side="left", padx=(0, 8))
    token_copy_btn = tk.Button(
        token_row,
        text=tr("settings.copy"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
        width=7,
    )
    token_copy_btn.pack(side="left")
    _bind_settings_tooltip(token_entry, "settings.token")
    _bind_settings_tooltip(token_save_btn, "settings.token")
    _bind_settings_tooltip(token_copy_btn, "settings.token")

    start_menu_row, _ = _make_settings_row("settings.start_menu")
    start_menu_app_btn = tk.Button(
        start_menu_row,
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        text=tr("settings.start_menu_app"),
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
    )
    start_menu_app_btn.pack(side="left", padx=(0, 8))
    start_menu_journal_btn = tk.Button(
        start_menu_row,
        text=tr("settings.start_menu_journal"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
    )
    start_menu_journal_btn.pack(side="left", padx=(0, 8))
    start_menu_reader_btn = tk.Button(
        start_menu_row,
        text=tr("settings.start_menu_reader"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=6,
        cursor="hand2",
    )
    start_menu_reader_btn.pack(side="left")
    bind_hover_tooltip(
        start_menu_app_btn,
        lambda: tr("tip.start_menu_app"),
    )
    bind_hover_tooltip(
        start_menu_journal_btn,
        lambda: tr("tip.start_menu_journal"),
    )
    bind_hover_tooltip(
        start_menu_reader_btn,
        lambda: tr("tip.start_menu_reader"),
    )

    def _refresh_token_entry_mask() -> None:
        token_entry.delete(0, "end")
        if token_saved["value"]:
            token_entry.insert(0, "*" * max(32, len(token_saved["value"])))

    def _is_token_mask(value: str) -> bool:
        return bool(value) and all(ch == "*" for ch in value)

    def _on_rename_apply() -> None:
        new_name = rename_entry.get().strip() or "Daily Logger"
        updated = rename_app_name_to(new_name)
        settings_app_name["value"] = updated
        console_app_name["value"] = updated
        root.title(updated)
        nav_title.config(text=updated)
        rename_entry.delete(0, "end")
        rename_entry.insert(0, updated)
        _set_settings_status(tr("status.rename_ok", name=updated))

    def _on_toggle_startup() -> None:
        should_enable = not startup_state["enabled"]
        ok = create_startup_shortcut() if should_enable else remove_startup_shortcut()
        if not ok:
            _set_settings_status(tr("status.startup_fail"))
            return
        startup_state["enabled"] = should_enable
        startup_toggle_btn.config(
            text=tr("settings.on") if should_enable else tr("settings.off")
        )
        prefs = load_preferences()
        prefs["startup_enabled"] = "true" if should_enable else "false"
        save_preferences(prefs)
        _set_settings_status(
            tr("status.startup_on") if should_enable else tr("status.startup_off")
        )

    def _on_toggle_iphone_receive() -> None:
        should_enable = not iphone_receive_state["enabled"]
        if not save_iphone_passive_receive_enabled(should_enable):
            _set_settings_status(tr("status.iphone_receive_save_fail"))
            return
        iphone_receive_state["enabled"] = should_enable
        iphone_receive_toggle_btn.config(
            text=tr("settings.on") if should_enable else tr("settings.off")
        )
        if should_enable:
            start_iphone_receiver(show_setup=False, passive=True)
            _set_settings_status(tr("status.iphone_receive_on"))
        else:
            stop_iphone_receiver()
            _set_settings_status(tr("status.iphone_receive_off"))

    def _on_open_transcription_models() -> None:
        opener = transcription_models_manager_hooks.get("open")
        if callable(opener):
            opener()
            return
        _set_settings_status(tr("status.transcription_models_not_ready"))

    def _release_download_url(info: Dict[str, Any], asset_name: str) -> str:
        assets = info.get("assets")
        if isinstance(assets, dict):
            url = str(assets.get(asset_name) or "").strip()
            if url:
                return url
        return f"{ADDON_RELEASE_BASE_URL}/{asset_name}"

    def _release_asset_names(info: Dict[str, Any]) -> List[str]:
        assets = info.get("assets")
        if not isinstance(assets, dict):
            return []
        return [name for name in UPDATE_RELEASE_ASSET_NAMES if name in assets]

    def _installed_addon_release_names(info: Dict[str, Any]) -> List[str]:
        assets = info.get("assets")
        if not isinstance(assets, dict):
            return []
        names: List[str] = []
        if local_transcription_addon_is_installed() and LOCAL_TRANSCRIPTION_ADDON_ZIP_NAME in assets:
            names.append("Local Transcription")
        if media_tools_addon_is_installed() and MEDIA_TOOLS_ADDON_ZIP_NAME in assets:
            names.append("Media Tools")
        return names

    update_dialog_ref: Dict[str, Any] = {"window": None}

    def _show_update_dialog(info: Dict[str, Any]) -> None:
        tag = str(info.get("tag") or "").strip()
        if not tag:
            return
        existing = update_dialog_ref.get("window")
        try:
            if existing is not None and existing.winfo_exists():
                existing.lift()
                return
        except tk.TclError:
            pass
        prefs = load_preferences()
        prefs[UPDATE_LAST_SEEN_RELEASE_PREF_KEY] = tag
        save_preferences(prefs)

        t = th()
        dlg = tk.Toplevel(root)
        update_dialog_ref["window"] = dlg
        dlg.title(tr("updates.title"))
        dlg.configure(bg=t.surface)
        dlg.transient(root)
        dlg.geometry("560x460")
        dlg.minsize(500, 360)

        wrap = tk.Frame(dlg, bg=t.surface)
        wrap.pack(fill="both", expand=True, padx=18, pady=18)
        title_lbl = tk.Label(
            wrap,
            text=tr("updates.available"),
            bg=t.surface,
            fg=t.text,
            font=("Segoe UI", 14, "bold"),
            anchor="w",
        )
        title_lbl.pack(fill="x", pady=(0, 10))

        published = str(info.get("published_at") or "").strip()
        if "T" in published:
            published = published.split("T", 1)[0]
        detail_lines = [
            tr("updates.current").format(version=APP_VERSION),
            tr("updates.latest").format(version=str(info.get("version") or tag)),
        ]
        if published:
            detail_lines.append(tr("updates.published").format(date=published))
        assets = _release_asset_names(info)
        if assets:
            detail_lines.append(tr("updates.assets").format(assets=", ".join(assets)))
        addon_names = _installed_addon_release_names(info)
        if addon_names:
            detail_lines.append(tr("updates.addons").format(addons=", ".join(addon_names)))
        details_lbl = tk.Label(
            wrap,
            text="\n".join(detail_lines),
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 10),
            anchor="w",
            justify="left",
        )
        details_lbl.pack(fill="x", pady=(0, 10))

        notes = release_notes_preview(str(info.get("body") or "")) or tr("updates.notes_empty")
        notes_box = tk.Text(
            wrap,
            height=10,
            wrap="word",
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            relief="flat",
            highlightthickness=1,
            highlightbackground=t.border,
            highlightcolor=t.accent,
            font=("Segoe UI", 9),
        )
        notes_box.pack(fill="both", expand=True, pady=(0, 12))
        notes_box.insert("1.0", notes)
        notes_box.config(state="disabled")

        btn_row = tk.Frame(wrap, bg=t.surface)
        btn_row.pack(fill="x")

        def _close() -> None:
            try:
                dlg.destroy()
            except tk.TclError:
                pass

        def _open_release() -> None:
            webbrowser.open(str(info.get("html_url") or APP_RELEASE_PAGE_URL))

        def _download_portable() -> None:
            webbrowser.open(_release_download_url(info, APP_PORTABLE_ZIP_NAME))

        def _skip_version() -> None:
            prefs_inner = load_preferences()
            prefs_inner[UPDATE_DISMISSED_RELEASE_PREF_KEY] = tag
            save_preferences(prefs_inner)
            _close()

        def _make_dialog_button(label: str, command: Callable[[], None]) -> Any:
            btn = tk.Button(
                btn_row,
                text=label,
                command=command,
                bg=t.btn_secondary,
                fg=t.text,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
                relief="flat",
                font=("Segoe UI", 9, "bold"),
                padx=10,
                pady=7,
                cursor="hand2",
            )
            btn.pack(side="left", padx=(0, 8))
            return btn

        _make_dialog_button(tr("updates.open_release"), _open_release)
        _make_dialog_button(tr("updates.download_portable"), _download_portable)
        _make_dialog_button(tr("updates.remind_later"), _close)
        _make_dialog_button(tr("updates.skip_version"), _skip_version)

        dlg.protocol("WM_DELETE_WINDOW", _close)
        try:
            dlg.lift()
        except tk.TclError:
            pass

    def _set_updates_buttons_busy(busy: bool) -> None:
        updates_state["busy"] = busy
        state = "disabled" if busy else "normal"
        try:
            updates_toggle_btn.config(state=state)
            updates_check_btn.config(state=state)
        except tk.TclError:
            pass

    def _set_updates_status(text: str) -> None:
        try:
            updates_status_var.set((text or "").strip())
        except tk.TclError:
            pass

    def _finish_update_check(info: Optional[Dict[str, Any]], err_msg: str, *, manual: bool) -> None:
        _set_updates_buttons_busy(False)
        prefs = load_preferences()
        if err_msg:
            if manual:
                text = tr("status.update_failed").format(error=err_msg)
                _set_updates_status(text)
                _set_settings_status(text)
            else:
                _append_console_session_update(
                    tr("status.update_failed").format(error=err_msg),
                    key="updates:auto_failed",
                )
            return
        if not info:
            return
        prefs[UPDATE_LAST_CHECK_DATE_PREF_KEY] = today_update_check_key()
        save_preferences(prefs)
        tag = str(info.get("tag") or "").strip()
        is_new = release_tag_is_newer(tag, APP_VERSION)
        if not is_new:
            if manual:
                current_version = normalize_release_tag(APP_VERSION) or APP_VERSION
                status_text = tr("status.update_current").format(version=current_version)
                _set_updates_status(status_text)
                _set_settings_status(status_text)
            return
        if not manual and prefs.get(UPDATE_DISMISSED_RELEASE_PREF_KEY, "").strip() == tag:
            return
        status_text = tr("status.update_available").format(version=str(info.get("version") or tag))
        _set_updates_status(status_text)
        _set_settings_status(status_text)
        _show_update_dialog(info)

    def _check_updates_async(*, manual: bool = False, force: bool = False) -> None:
        if updates_state["busy"]:
            if manual:
                _set_updates_status(tr("status.update_checking"))
                _set_settings_status(tr("status.update_checking"))
            return
        prefs = load_preferences()
        if not manual and not update_check_enabled():
            return
        if (
            not manual
            and not force
            and prefs.get(UPDATE_LAST_CHECK_DATE_PREF_KEY, "").strip() == today_update_check_key()
        ):
            return
        _set_updates_buttons_busy(True)
        if manual:
            _set_updates_status(tr("status.update_checking"))
            _set_settings_status(tr("status.update_checking"))

        def _work() -> None:
            info, err_msg = fetch_latest_release_info(timeout_sec=8)
            try:
                root.after(0, lambda: _finish_update_check(info, err_msg, manual=manual))
            except tk.TclError:
                pass

        threading.Thread(target=_work, daemon=True).start()

    def _on_toggle_updates() -> None:
        should_enable = not updates_state["enabled"]
        if not save_update_check_enabled(should_enable):
            _set_settings_status(tr("status.update_failed").format(error="Could not save setting."))
            return
        updates_state["enabled"] = should_enable
        updates_toggle_btn.config(
            text=tr("settings.on") if should_enable else tr("settings.off")
        )
        status_text = tr("status.update_on") if should_enable else tr("status.update_off")
        _set_updates_status(status_text)
        _set_settings_status(status_text)

    def _set_fullscreen_enabled(
        should_enable: bool,
        *,
        persist: bool = True,
        announce: bool = True,
    ) -> bool:
        enabled = bool(should_enable)
        try:
            root.attributes("-fullscreen", enabled)
        except tk.TclError:
            if announce:
                _set_settings_status(tr("status.fullscreen_fail"))
            return False
        fullscreen_state["enabled"] = enabled
        try:
            display_fullscreen_btn.config(
                text=tr("settings.on") if enabled else tr("settings.off")
            )
        except tk.TclError:
            pass
        if persist:
            prefs = load_preferences()
            prefs[JOURNAL_PREF_FULLSCREEN_KEY] = "true" if enabled else "false"
            save_preferences(prefs)
        if announce:
            _set_settings_status(
                tr("status.fullscreen_on") if enabled else tr("status.fullscreen_off")
            )
        return True

    def _on_toggle_fullscreen(_evt: Optional[Any] = None) -> str:
        _set_fullscreen_enabled(not fullscreen_state["enabled"])
        return "break"

    def _persist_backup_mode(mode: str) -> None:
        prefs = load_preferences()
        if mode == "On":
            prefs["backup_enabled"] = "true"
            prefs["backup_limited"] = "false"
        elif mode == "Off":
            prefs["backup_enabled"] = "false"
            prefs["backup_limited"] = "false"
        else:
            prefs["backup_enabled"] = "true"
            prefs["backup_limited"] = "true"
        if save_preferences(prefs):
            _set_settings_status(tr("status.backup_mode", mode=_backup_mode_btn_label(mode)))
        else:
            _set_settings_status(tr("status.backup_save_fail"))

    def _on_cycle_backup_mode() -> None:
        order = ("On", "Off", "Limited")
        idx = order.index(backup_mode["value"])
        next_mode = order[(idx + 1) % len(order)]
        backup_mode["value"] = next_mode
        backup_mode_btn.config(text=_backup_mode_btn_label(next_mode))
        _persist_backup_mode(next_mode)

    def _on_manual_backup() -> None:
        prefs = load_preferences()
        evict_oldest_backup_if_limited_full(prefs)
        backup_path = run_backup_now()
        if backup_path is None:
            _set_settings_status(tr("status.backup_skip"))
            return
        trim_backups_if_limited(prefs)
        _set_settings_status(tr("status.backup_ok", name=backup_path.name))

    def _on_token_focus_in(_evt: Optional[Any] = None) -> None:
        if _is_token_mask(token_entry.get()):
            token_entry.delete(0, "end")

    def _on_token_save() -> None:
        typed = token_entry.get().strip()
        if _is_token_mask(typed):
            _set_settings_status(tr("status.token_same"))
            return
        if not typed:
            if delete_openai_api_key():
                token_saved["value"] = ""
                _refresh_token_entry_mask()
                _set_settings_status(tr("status.token_removed"))
                _ai_i18n = getattr(build_ai_recap_and_chatbot_pages, "_i18n", None)
                if callable(_ai_i18n):
                    _ai_i18n()
            else:
                _set_settings_status(tr("status.token_remove_fail"))
            return
        if save_openai_api_key(typed):
            token_saved["value"] = typed
            _refresh_token_entry_mask()
            _set_settings_status(tr("status.token_saved"))
            _ai_i18n = getattr(build_ai_recap_and_chatbot_pages, "_i18n", None)
            if callable(_ai_i18n):
                _ai_i18n()
        else:
            _set_settings_status(tr("status.token_save_fail"))

    def _on_token_copy() -> None:
        current = get_openai_api_key() or ""
        if not current:
            _set_settings_status(tr("status.token_no_copy"))
            return
        if copy_text_to_clipboard(current):
            _set_settings_status(tr("status.token_copied"))
        else:
            _set_settings_status(tr("status.token_copy_fail"))

    def _on_start_menu_button(selected: str) -> None:
        if selected == "journal":
            ok = sb_create_journal_search_shortcut()
        elif selected == "reader":
            ok = sb_create_reader_search_shortcut()
        else:
            ok = sb_create_bat_search_shortcut()
        if ok:
            _set_settings_status(tr("status.start_menu_ok", which=selected))
        else:
            _set_settings_status(tr("status.start_menu_fail", which=selected))

    rename_btn.config(command=_on_rename_apply)
    startup_toggle_btn.config(command=_on_toggle_startup)
    iphone_receive_toggle_btn.config(command=_on_toggle_iphone_receive)
    transcription_models_btn.config(command=_on_open_transcription_models)
    updates_toggle_btn.config(command=_on_toggle_updates)
    updates_check_btn.config(command=lambda: _check_updates_async(manual=True, force=True))
    display_fullscreen_btn.config(command=_on_toggle_fullscreen)
    backup_mode_btn.config(command=_on_cycle_backup_mode)
    backup_manual_btn.config(command=_on_manual_backup)
    token_save_btn.config(command=_on_token_save)
    token_copy_btn.config(command=_on_token_copy)
    start_menu_app_btn.config(command=lambda: _on_start_menu_button("app"))
    start_menu_journal_btn.config(command=lambda: _on_start_menu_button("journal"))
    start_menu_reader_btn.config(command=lambda: _on_start_menu_button("reader"))
    token_entry.bind("<FocusIn>", _on_token_focus_in, add="+")

    def _goto_settings_token_field() -> None:
        show_page("settings")

        def _focus_token() -> None:
            try:
                token_entry.focus_set()
            except tk.TclError:
                return
            try:
                token_entry.selection_range(0, "end")
            except tk.TclError:
                pass

        root.after(100, _focus_token)

    api_key_prompt_hooks["goto_token"] = _goto_settings_token_field

    for _btn in (
        rename_btn,
        startup_toggle_btn,
        iphone_receive_toggle_btn,
        transcription_models_btn,
        settings_theme_btn,
        backup_mode_btn,
        backup_manual_btn,
        token_save_btn,
        token_copy_btn,
        start_menu_app_btn,
        start_menu_journal_btn,
        start_menu_reader_btn,
    ):
        bind_button_hover_if_enabled(
            _btn,
            lambda b=_btn: (
                str(b.cget("state")),
                th().btn_secondary,
                th().text,
                th().secondary_hover,
                th().text,
            ),
            lambda: th().secondary_hover,
            lambda: th().text,
        )

    console_wrap = tk.Frame(console_page, bg=t_init.surface)
    console_wrap.pack(fill="both", expand=True, padx=20, pady=20)
    _register_page_toggle(console_page)
    console_title = tk.Label(
        console_wrap,
        text="Console",
        bg=t_init.surface,
        fg=t_init.text,
        font=("Segoe UI", 16, "bold"),
        anchor="w",
    )
    console_title.pack(anchor="w", pady=(0, 8))
    console_output = tk.Text(
        console_wrap,
        wrap="word",
        height=20,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        padx=12,
        pady=12,
        font=("Consolas", 10),
        state="disabled",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
    )
    console_output.pack(fill="both", expand=True, side="left")
    console_output_holder["widget"] = console_output
    _append_console_session_update("Session started.", key="session:start")
    console_scroll = tk.Scrollbar(
        console_wrap,
        command=console_output.yview,
        bg=t_init.panel,
        troughcolor=t_init.field,
        activebackground=t_init.accent,
        bd=0,
        highlightthickness=0,
        width=11,
    )
    console_scroll.pack(fill="y", side="right")
    console_output.configure(yscrollcommand=console_scroll.set)

    console_input_row = tk.Frame(content_host, bg=t_init.surface)
    console_input_row.place_forget()
    console_input_holder["row"] = console_input_row
    console_prompt = tk.Label(
        console_input_row,
        text="> ",
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Consolas", 11, "bold"),
        cursor="hand2",
    )
    console_prompt.pack(side="left")
    console_entry = tk.Entry(
        console_input_row,
        bg=t_init.field,
        fg=t_init.text,
        insertbackground=t_init.text,
        relief="flat",
        highlightthickness=1,
        highlightbackground=t_init.border,
        highlightcolor=t_init.accent,
        font=("Consolas", 10),
        width=1,
    )
    console_entry.pack(side="left", fill="x", expand=True, padx=(0, 0))
    console_insertwidth_normal = int(console_entry.cget("insertwidth") or 1)
    console_entry_state: Dict[str, bool] = {"placeholder": False}

    def _console_placeholder_fg(theme: Optional[JournalWindowThemeSpec] = None) -> str:
        current_theme = theme or th()
        return "#8A8FA8" if current_theme.is_dark else "#6F7480"

    def _set_console_placeholder() -> None:
        if console_entry.get():
            return
        console_entry_state["placeholder"] = True
        console_entry.config(
            fg=_console_placeholder_fg(t_init),
            font=("Consolas", 10, "italic"),
            insertwidth=0,
        )
        hint = str(console_hint_state.get("text", "")).strip()
        console_entry.insert(0, hint or tr("console.placeholder"))

    def _clear_console_placeholder() -> None:
        if not console_entry_state["placeholder"]:
            return
        console_entry.delete(0, "end")
        console_entry_state["placeholder"] = False
        console_entry.config(
            fg=th().text,
            font=("Consolas", 10),
            insertwidth=console_insertwidth_normal,
        )
        console_hint_state["text"] = ""

    def _show_console_hint_placeholder() -> None:
        if console_entry_state["placeholder"] or not console_entry.get().strip():
            console_entry.delete(0, "end")
            console_entry_state["placeholder"] = False
            _set_console_placeholder()

    console_hint_state["apply"] = _show_console_hint_placeholder
    _set_console_placeholder()
    _journal_find_state: Dict[str, Any] = {"widget": text_box}

    def _active_journal_text_widget() -> tk.Text:
        w = root.focus_get()
        if isinstance(w, tk.Text) and w in (text_box, stt_box, report_box):
            _journal_find_state["widget"] = w
            return w
        saved = _journal_find_state.get("widget")
        if isinstance(saved, tk.Text):
            return saved
        return text_box

    def _search_scope_widgets() -> Tuple[tk.Text, ...]:
        if find_scope_var.get() == "one":
            return (_active_journal_text_widget(),)
        return (text_box, stt_box, report_box)

    def _journal_select_all(_evt: Optional[Any] = None) -> str:
        w = _active_journal_text_widget()
        try:
            w.tag_add("sel", "1.0", "end-1c")
            w.mark_set("insert", "end-1c")
            w.see("insert")
        except tk.TclError:
            pass
        return "break"

    def _find_close() -> None:
        find_row.pack_forget()
        find_status.config(text="")
        _active_journal_text_widget().focus_set()

    def _find_all_ranges(w: tk.Text, query: str, case_sensitive: bool, whole_word: bool) -> List[Tuple[str, str]]:
        pattern = query if not whole_word else rf"\m{re.escape(query)}\M"
        ranges: List[Tuple[str, str]] = []
        idx = "1.0"
        while True:
            pos = w.search(
                pattern,
                idx,
                stopindex="end-1c",
                nocase=not case_sensitive,
                regexp=whole_word,
            )
            if not pos:
                break
            end = f"{pos}+{len(query)}c"
            ranges.append((pos, end))
            idx = end
        return ranges

    def _find_update_status_for_selection(
        w: tk.Text, query: str, sel_start: str, sel_widget: Optional[tk.Text] = None
    ) -> None:
        if not query:
            find_status.config(text="")
            return
        widgets = _search_scope_widgets()
        all_ranges: List[Tuple[tk.Text, str, str]] = []
        for _w in widgets:
            _ranges = _find_all_ranges(
                _w,
                query,
                case_sensitive=find_case_var.get(),
                whole_word=find_word_var.get(),
            )
            all_ranges.extend([(_w, _s, _e) for _s, _e in _ranges])
        if not all_ranges:
            find_status.config(text="No matches")
            return
        current = 1
        sw = sel_widget or w
        for i, (rw, s, _e) in enumerate(all_ranges, start=1):
            if rw is sw and s == sel_start:
                current = i
                break
        find_status.config(text=f"{current}/{len(all_ranges)}")

    def _find_next(direction: int = 1) -> str:
        query = find_var.get()
        if not query:
            find_status.config(text="Type text to find")
            return "break"
        widgets = _search_scope_widgets()
        ranges_by_widget: Dict[tk.Text, List[Tuple[str, str]]] = {}
        total_matches = 0
        for _w in widgets:
            _ranges = _find_all_ranges(
                _w,
                query,
                case_sensitive=find_case_var.get(),
                whole_word=find_word_var.get(),
            )
            ranges_by_widget[_w] = _ranges
            total_matches += len(_ranges)
        if total_matches == 0:
            find_status.config(text="No matches")
            return "break"
        w = _active_journal_text_widget()
        if w not in widgets:
            w = widgets[0]
        start = w.index("insert")
        if w.tag_ranges("sel"):
            start = w.index("sel.last") if direction > 0 else w.index("sel.first")
        start_widget_idx = widgets.index(w)

        def _pick_forward() -> Optional[Tuple[tk.Text, str, str]]:
            for wi in range(start_widget_idx, len(widgets)):
                _w = widgets[wi]
                _ranges = ranges_by_widget.get(_w, [])
                if not _ranges:
                    continue
                if wi == start_widget_idx:
                    for s, e in _ranges:
                        if _w.compare(s, ">", start):
                            return (_w, s, e)
                else:
                    return (_w, _ranges[0][0], _ranges[0][1])
            for wi in range(0, start_widget_idx + 1):
                _w = widgets[wi]
                _ranges = ranges_by_widget.get(_w, [])
                if not _ranges:
                    continue
                if wi == start_widget_idx:
                    return (_w, _ranges[0][0], _ranges[0][1])
                return (_w, _ranges[0][0], _ranges[0][1])
            return None

        def _pick_backward() -> Optional[Tuple[tk.Text, str, str]]:
            for wi in range(start_widget_idx, -1, -1):
                _w = widgets[wi]
                _ranges = ranges_by_widget.get(_w, [])
                if not _ranges:
                    continue
                if wi == start_widget_idx:
                    for s, e in reversed(_ranges):
                        if _w.compare(s, "<", start):
                            return (_w, s, e)
                else:
                    s, e = _ranges[-1]
                    return (_w, s, e)
            for wi in range(len(widgets) - 1, start_widget_idx - 1, -1):
                _w = widgets[wi]
                _ranges = ranges_by_widget.get(_w, [])
                if not _ranges:
                    continue
                if wi == start_widget_idx:
                    s, e = _ranges[-1]
                    return (_w, s, e)
                s, e = _ranges[-1]
                return (_w, s, e)
            return None

        picked = _pick_forward() if direction > 0 else _pick_backward()
        if not picked:
            find_status.config(text="No matches")
            return "break"
        pw, pos, end = picked
        for _w in (text_box, stt_box, report_box):
            _w.tag_remove("sel", "1.0", "end")
        pw.tag_add("sel", pos, end)
        pw.mark_set("insert", end if direction > 0 else pos)
        pw.see(pos)
        pw.focus_set()
        _journal_find_state["widget"] = pw
        _find_update_status_for_selection(pw, query, pos, sel_widget=pw)
        return "break"

    def _find_prev(_evt: Optional[Any] = None) -> str:
        return _find_next(-1)

    def _find_open(_evt: Optional[Any] = None) -> str:
        if str(find_row.winfo_manager()) == "pack":
            _find_close()
            return "break"
        if active_page["key"] != "journal":
            show_page("journal")
        w = _active_journal_text_widget()
        find_row.pack(fill="x", padx=t_init.pad_outer, pady=(0, 6), before=center)
        selected = ""
        try:
            selected = w.get("sel.first", "sel.last")
        except tk.TclError:
            selected = ""
        if selected.strip():
            find_var.set(selected)
        _find_update_status_for_selection(w, find_var.get(), "", sel_widget=w)
        find_entry.focus_set()
        find_entry.selection_range(0, "end")
        return "break"

    find_next_btn.config(command=lambda: _find_next(1))
    find_prev_btn.config(command=lambda: _find_next(-1))
    find_close_btn.config(command=_find_close)
    find_scope_all_rb.config(
        command=lambda: _find_update_status_for_selection(
            _active_journal_text_widget(), find_var.get(), "", sel_widget=_active_journal_text_widget()
        )
    )
    find_scope_one_rb.config(
        command=lambda: _find_update_status_for_selection(
            _active_journal_text_widget(), find_var.get(), "", sel_widget=_active_journal_text_widget()
        )
    )
    find_case_chk.config(
        command=lambda: _find_update_status_for_selection(
            _active_journal_text_widget(), find_var.get(), "", sel_widget=_active_journal_text_widget()
        )
    )
    find_word_chk.config(
        command=lambda: _find_update_status_for_selection(
            _active_journal_text_widget(), find_var.get(), "", sel_widget=_active_journal_text_widget()
        )
    )
    find_entry.bind("<Return>", lambda _e: _find_next(1), add="+")
    find_entry.bind("<Shift-Return>", _find_prev, add="+")
    find_entry.bind(
        "<KeyRelease>",
        lambda _e: _find_update_status_for_selection(_active_journal_text_widget(), find_var.get(), ""),
        add="+",
    )
    for _b in (find_prev_btn, find_next_btn, find_close_btn):
        bind_button_hover_if_enabled(
            _b,
            lambda: th().toolbar_bind_rest(),
            lambda: th().toolbar_hover()[0],
            lambda: th().toolbar_hover()[1],
        )

    def gen_rest_style() -> Tuple[str, str, str, str, str]:
        if str(gen_button.cget("state")) != "normal":
            return th().gen_bind_disabled()
        return th().gen_bind_rest()

    bind_button_hover_if_enabled(
        gen_button,
        gen_rest_style,
        lambda: th().hover_primary,
        lambda: "white",
    )

    save_entry_btn_holder: Dict[str, Any] = {"btn": None}

    def refresh_save_entry_state() -> None:
        btn = save_entry_btn_holder.get("btn")
        if btn is None:
            return
        has_any = bool(
            text_box.get("1.0", "end-1c").strip()
            or stt_box.get("1.0", "end-1c").strip()
            or report_box.get("1.0", "end-1c").strip()
        )
        t = th()
        if has_any:
            btn.config(
                state="normal",
                bg=t.accent,
                fg="white",
                activebackground=t.hover_save,
                activeforeground="white",
                cursor="hand2",
            )
        else:
            btn.config(
                state="disabled",
                bg=t.btn_disabled,
                fg=t.disabled_fg,
                disabledforeground=t.disabled_fg,
                cursor="arrow",
            )

    def load_latest_entry_into_current_journal(values: Dict[str, object]) -> None:
        """
        Load an existing journal record into the currently visible journal text boxes.

        Used by the GUI console command `JS -> EDITPREV` so we don't open a new journal editor window.
        """
        nonlocal edit_target_sheet, edit_target_row
        # Journal console helpers pass keys from `get_latest_journal_entry_for_edit()`.
        edit_target_sheet = str(values.get("sheet_name", "") or "")
        try:
            edit_target_row = int(values.get("row_index", 0) or 0)
        except (TypeError, ValueError):
            edit_target_row = 0
        is_edit_mode["v"] = bool(edit_target_sheet and edit_target_row > 0)

        new_text = normalize_journal_text_punctuation(str(values.get("text", "") or ""))
        new_speech = normalize_journal_text_punctuation(
            str(values.get("speech_transcript", "") or "")
        )
        new_report = normalize_journal_text_punctuation(str(values.get("ai_report", "") or ""))
        new_date = str(values.get("date", "") or "")
        new_time = str(values.get("time", "") or "")

        text_box.delete("1.0", "end")
        text_box.insert("1.0", new_text)
        stt_box.delete("1.0", "end")
        stt_box.insert("1.0", new_speech)
        report_box.delete("1.0", "end")
        report_box.insert("1.0", new_report)

        # DateEntry may be `DateEntry` or `tk.Entry`; try both.
        try:
            date_entry.delete(0, "end")
            date_entry.insert(0, new_date)
        except Exception:
            try:
                if new_date:
                    date_entry.set_date(new_date)  # type: ignore[attr-defined]
            except Exception:
                pass
        try:
            time_entry.delete(0, "end")
            time_entry.insert(0, new_time)
        except Exception:
            pass

        last_journal_wav["path"] = None
        _set_stt_saved_path_display("")
        stt_status.config(text="")
        report_status.config(text="")
        update_transcribe_ui()
        refresh_save_entry_state()
        try:
            text_box.focus_set()
        except tk.TclError:
            pass

    def load_draft_into_current_journal() -> bool:
        """Load saved journal window draft into the current journal editor widgets."""
        draft = load_journal_window_draft()
        if not draft:
            return False
        # Drafts should not be treated as "edit existing row".
        nonlocal edit_target_sheet, edit_target_row
        edit_target_sheet = ""
        edit_target_row = 0
        is_edit_mode["v"] = False

        text_box.delete("1.0", "end")
        text_box.insert("1.0", normalize_journal_text_punctuation(str(draft.get("text", "") or "")))
        stt_box.delete("1.0", "end")
        stt_box.insert(
            "1.0",
            normalize_journal_text_punctuation(str(draft.get("speech_transcript", "") or "")),
        )
        report_box.delete("1.0", "end")
        report_box.insert(
            "1.0",
            normalize_journal_text_punctuation(str(draft.get("ai_report", "") or "")),
        )

        draft_date = str(draft.get("date", "") or "").strip()
        draft_time = str(draft.get("time", "") or "").strip()
        try:
            date_entry.delete(0, "end")
            date_entry.insert(0, draft_date)
        except Exception:
            try:
                if draft_date:
                    date_entry.set_date(draft_date)  # type: ignore[attr-defined]
            except Exception:
                pass
        try:
            time_entry.delete(0, "end")
            time_entry.insert(0, draft_time)
        except Exception:
            pass

        last_journal_wav["path"] = None
        _set_stt_saved_path_display("")
        stt_status.config(text="")
        report_status.config(text="")
        update_transcribe_ui()
        refresh_save_entry_state()
        try:
            text_box.focus_set()
        except tk.TclError:
            pass
        return True

    def start_new_journal(discard_without_confirm: bool = False) -> bool:
        """
        Clear the current journal editor widgets to start a fresh page.
        If there is unsaved content, ask for confirmation unless `discard_without_confirm` is True.
        """
        if recording_ui_busy["v"] or transcribing_busy["v"]:
            messagebox.showinfo(
                "New Journal",
                "Finish recording/transcribing before starting a new journal.",
            )
            return False

        has_content = any(
            [
                text_box.get("1.0", "end-1c").strip(),
                stt_box.get("1.0", "end-1c").strip(),
                report_box.get("1.0", "end-1c").strip(),
            ]
        )

        should_discard = True
        if has_content and not discard_without_confirm:
            should_discard = messagebox.askyesno(
                "Start new journal",
                "Discard the current journal editor content and clear the saved draft? ",
            )
        if not should_discard:
            return False

        # Clear stored draft so we don't immediately bring it back from disk.
        try:
            clear_journal_window_draft()
        except Exception:
            pass

        # Reset editor state.
        edit_target_sheet = ""
        edit_target_row = 0
        is_edit_mode["v"] = False
        last_journal_wav["path"] = None
        _set_stt_saved_path_display("")
        stt_status.config(text="")
        report_status.config(text="")
        transcribing_progress["v"] = 0

        text_box.delete("1.0", "end")
        stt_box.delete("1.0", "end")
        report_box.delete("1.0", "end")

        # Reset date/time to now.
        now = datetime.now()
        current_now_date = now.strftime("%m/%d/%Y")
        current_now_time = now.strftime("%I:%M%p").lstrip("0")
        try:
            date_entry.delete(0, "end")
            date_entry.insert(0, current_now_date)
        except Exception:
            try:
                date_entry.set_date(current_now_date)  # type: ignore[attr-defined]
            except Exception:
                pass
        try:
            time_entry.delete(0, "end")
            time_entry.insert(0, current_now_time)
        except Exception:
            pass

        update_transcribe_ui()
        refresh_save_entry_state()

        try:
            text_box.focus_set()
        except tk.TclError:
            pass
        return True

    record_stop = threading.Event()
    record_pause = threading.Event()
    record_mic_enabled = threading.Event()
    record_computer_enabled = threading.Event()
    record_thread_holder: Dict[str, object] = {"thread": None}
    record_path_holder: Dict[str, object] = {"path": None}
    record_source_mode: Dict[str, str] = {"value": get_selected_record_source_mode()}
    recording_ui_busy = {"v": False}
    recording_background_mode = {"v": False}
    record_close_requested = {"v": False}
    last_journal_wav: Dict[str, Optional[Path]] = {"path": None}
    transcribing_busy = {"v": False}
    transcribing_progress: Dict[str, int] = {"v": 0}
    transcribing_job_state: Dict[str, int] = {"id": 0}
    wave_lock = threading.Lock()
    wave_holder: Dict[str, List[float]] = {"levels": []}
    wave_gate: Dict[str, Any] = {"rms": 0.0}
    wave_after: Dict[str, Optional[Any]] = {"id": None}

    def cancel_wave_tick() -> None:
        wid = wave_after["id"]
        if wid is not None:
            try:
                root.after_cancel(wid)
            except tk.TclError:
                pass
            wave_after["id"] = None

    def redraw_waveform_canvas() -> None:
        with wave_lock:
            pts = list(wave_holder["levels"])
        wave_canvas.update_idletasks()
        wpx = max(40, int(wave_canvas.winfo_width()))
        hpx = max(30, int(wave_canvas.winfo_height()))
        wave_canvas.delete("all")
        mid = hpx * 0.5
        t = th()
        base_color = t.muted if isinstance(t.muted, str) else "#888888"
        if len(pts) < 1:
            wave_canvas.create_line(4, mid, wpx - 4, mid, fill=base_color, width=1)
            return
        if len(pts) == 1:
            v = float(pts[0])
            y0 = mid - v * (hpx * 0.38)
            wave_canvas.create_line(4, y0, wpx - 4, y0, fill=t.waveform, width=1)
            return
        n = len(pts)
        coords: List[float] = []
        for i, v in enumerate(pts):
            x = 4.0 + (wpx - 8.0) * (i / max(n - 1, 1))
            y = mid - float(v) * (hpx * 0.38)
            coords.extend([x, y])
        wave_canvas.create_line(*coords, fill=t.waveform, width=1)

    def wave_tick() -> None:
        if not recording_ui_busy["v"]:
            wave_after["id"] = None
            return
        redraw_waveform_canvas()
        wave_after["id"] = root.after(33, wave_tick)

    def start_wave_tick() -> None:
        cancel_wave_tick()
        wave_after["id"] = root.after(33, wave_tick)

    def reset_waveform_session() -> None:
        cancel_wave_tick()
        with wave_lock:
            wave_holder["levels"].clear()
        wave_gate["rms"] = 0.0
        redraw_waveform_canvas()

    def on_pcm_block_journal(block: Any) -> None:
        try:
            import numpy as np

            flat = np.asarray(block, dtype=np.float64).reshape(-1)
            if flat.size == 0:
                return
            rms = float(np.sqrt(np.mean(flat * flat)))
            wave_gate["rms"] = rms
            adj = max(0.0, rms - WAVEFORM_RMS_NOISE_FLOOR)
            denom = max(WAVEFORM_RMS_NORM - WAVEFORM_RMS_NOISE_FLOOR, 1.0)
            peak = min(1.0, adj / denom)
            with wave_lock:
                levels = wave_holder["levels"]
                levels.append(peak)
                over = len(levels) - WAVEFORM_MAX_DRAW_SAMPLES
                if over > 0:
                    del levels[:over]
        except Exception:
            pass

    iphone_pending_paths: List[Path] = []
    iphone_pending_keys: set[str] = set()
    iphone_receiver_state: Dict[str, Any] = {
        "active": False,
        "passive": False,
        "server": None,
        "thread": None,
        "url": "",
        "urls": {},
        "token": get_or_create_iphone_import_token(),
        "port": IPHONE_IMPORT_DEFAULT_PORT,
    }

    def _iphone_path_key(path: Path) -> str:
        try:
            return str(path.resolve()).casefold()
        except OSError:
            return str(path).casefold()

    def _set_iphone_status(
        text: str,
        *,
        temp: bool = True,
        log: Optional[bool] = None,
    ) -> None:
        _publish_console_update(text, key="iphone_status", temp=temp, log=log)

    iphone_upload_status_by_id: Dict[str, Tuple[float, str]] = {}
    iphone_upload_id_by_path: Dict[str, str] = {}

    def publish_iphone_upload_status(upload_id: str, text: str) -> None:
        if upload_id:
            iphone_upload_status_by_id[upload_id] = (time.time(), text)
            cutoff = time.time() - 3600
            for old_id, (stamp, _msg) in list(iphone_upload_status_by_id.items()):
                if stamp < cutoff:
                    iphone_upload_status_by_id.pop(old_id, None)
        root.after(0, lambda msg=text: _set_iphone_status(msg))

    def _update_iphone_receive_button() -> None:
        t = th()
        bg, fg, abg, afg = t.side_action_config()
        active = bool(iphone_receiver_state.get("active"))
        passive = bool(iphone_receiver_state.get("passive"))
        receive_iphone_btn.config(
            text=tr("journal.iphone_receive") if passive else (
                tr("journal.iphone_stop") if active else tr("journal.iphone_receive")
            ),
            state="normal",
            width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH,
            bg=bg,
            fg=fg,
            activebackground=abg,
            activeforeground=afg,
            cursor="hand2",
        )

    def _iphone_pending_status_text() -> str:
        count = len(iphone_pending_paths)
        if count <= 0:
            return ""
        return tr("journal.iphone_pending").format(count=count)

    def _refresh_iphone_pending_status() -> None:
        text = _iphone_pending_status_text()
        if text and not transcribing_busy["v"]:
            _set_iphone_status(text)

    def _move_iphone_batch_to_processed(success: bool, paths: List[Path]) -> None:
        if success:
            for path in paths:
                upload_id = iphone_upload_id_by_path.get(_iphone_path_key(path), "")
                if upload_id:
                    publish_iphone_upload_status(
                        upload_id,
                        f"Processed {path.name}. Transcription complete.",
                    )
                    iphone_upload_id_by_path.pop(_iphone_path_key(path), None)
                _append_console_session_update(f"iPhone file transcribed: {path.name}", key=f"iphone:done:{_iphone_path_key(path)}")
            mark_iphone_inbox_files_processed(paths)

    def _rescan_and_drain_iphone_pending() -> None:
        receive_iphone_incoming_files(list_incoming_iphone_files())
        enqueue_iphone_imports(list_pending_iphone_inbox_files())
        drain_iphone_pending()

    def _schedule_iphone_pending_drain_after_install() -> None:
        for delay in (100, 1000, 3000):
            root.after(delay, _rescan_and_drain_iphone_pending)

    def drain_iphone_pending() -> None:
        if transcribing_busy["v"] or recording_ui_busy["v"]:
            _refresh_iphone_pending_status()
            _append_console_session_update(
                "iPhone transcription is pending until the current recording/transcription finishes.",
                key="iphone:pending:busy",
            )
            return
        if not iphone_pending_paths:
            if bool(iphone_receiver_state.get("active")):
                _set_iphone_status(
                    tr("journal.iphone_waiting"),
                    temp=not bool(iphone_receiver_state.get("passive")),
                )
            return
        paths = sorted(iphone_pending_paths, key=_transcription_file_sort_key)
        media_waiting_paths: List[Path] = []
        if _find_ffmpeg_executable() is None:
            ready_paths: List[Path] = []
            for path in paths:
                if transcription_path_needs_media_tools(path):
                    media_waiting_paths.append(path)
                else:
                    ready_paths.append(path)
            if media_waiting_paths:
                iphone_pending_paths.clear()
                iphone_pending_paths.extend(media_waiting_paths)
                iphone_pending_keys.clear()
                iphone_pending_keys.update(_iphone_path_key(path) for path in media_waiting_paths)
                first_waiting = media_waiting_paths[0]
                _set_iphone_status(_iphone_pending_status_text())
                _append_console_session_update(
                    f"Pending iPhone file needs Media Tools: {first_waiting.name}",
                    key="iphone:pending:media_tools",
                )
                if messagebox.askyesno(
                    tr("download_manager.title"),
                    tr("download_manager.media_required_prompt"),
                ):
                    _install_media_tools_addon()
                if not ready_paths:
                    return
                paths = ready_paths
            else:
                iphone_pending_paths.clear()
                iphone_pending_keys.clear()
        else:
            iphone_pending_paths.clear()
            iphone_pending_keys.clear()
        if not paths:
            _set_iphone_status(_iphone_pending_status_text())
            return
        pending_choice = normalize_transcription_model_choice(transcription_model_choice["value"])
        if transcription_model_is_cloud(pending_choice) and not get_openai_api_key():
            _set_iphone_status(_iphone_pending_status_text())
            _append_console_session_update(
                "iPhone transcription is pending because no cloud API key is saved.",
                key="iphone:pending:model",
            )
            return
        if transcription_model_is_local(pending_choice):
            pending_model_name = transcription_local_model_name(pending_choice)
            local_ready = False
            local_reason = ""
            if not local_transcription_model_is_downloaded(pending_model_name):
                local_reason = "the selected local model is not downloaded"
            else:
                runtime_ok, runtime_err = ensure_local_transcription_runtime_loaded()
                local_ready = runtime_ok
                local_reason = runtime_err or "the local transcription add-on is not ready"
            if not local_ready:
                if get_openai_api_key():
                    _append_console_session_update(
                        f"iPhone upload switched to cloud transcription because {local_reason}.",
                        key="iphone:cloud_fallback",
                    )
                    _set_transcription_model_choice(TRANSCRIPTION_MODEL_CLOUD)
                    pending_choice = normalize_transcription_model_choice(transcription_model_choice["value"])
                else:
                    _set_iphone_status(_iphone_pending_status_text())
                    _append_console_session_update(
                        f"iPhone transcription is pending because {local_reason}.",
                        key="iphone:pending:model",
                    )
                    _open_transcription_downloads_manager("transcription")
                    return
        _set_stt_saved_path_display(tr("journal.iphone_inbox"))
        _append_console_session_update(
            f"Starting iPhone transcription queue: {len(paths)} file(s).",
            key=f"iphone:queue:start:{int(time.time())}",
        )
        begin_transcribe_paths(
            paths,
            display_label=tr("journal.iphone_inbox"),
            after_done=_move_iphone_batch_to_processed,
        )

    def enqueue_iphone_imports(paths: List[Path]) -> None:
        added = 0
        for path in paths:
            if not path.exists() or not is_transcription_media_file(path):
                continue
            key = _iphone_path_key(path)
            if key in iphone_pending_keys:
                continue
            iphone_pending_keys.add(key)
            iphone_pending_paths.append(path)
            added += 1
        if added:
            _set_iphone_status(tr("journal.iphone_received").format(count=added))
            _append_console_session_update(
                f"Added {added} iPhone file(s) to pending transcription.",
                key=f"iphone:pending:add:{int(time.time())}",
            )
        root.after(100, drain_iphone_pending)

    iphone_incoming_paths: List[Path] = []
    iphone_incoming_keys: set[str] = set()
    iphone_incoming_prompt_after = {"id": None}

    def _refresh_settings_iphone_toggle() -> None:
        fn = globals().get("_daily_logger_refresh_iphone_settings_toggle")
        if callable(fn):
            try:
                fn()
            except Exception:
                pass

    def _accept_iphone_incoming(paths: List[Path]) -> None:
        accepted: List[Path] = []
        for src in paths:
            upload_id = iphone_upload_id_by_path.pop(_iphone_path_key(src), "")
            moved = move_iphone_file_to_inbox(src)
            if moved is not None:
                if upload_id:
                    iphone_upload_id_by_path[_iphone_path_key(moved)] = upload_id
                    publish_iphone_upload_status(
                        upload_id,
                        f"Accepted {moved.name} on PC. Waiting for transcription to start.",
                    )
                accepted.append(moved)
        if accepted:
            enqueue_iphone_imports(accepted)

    def _decline_iphone_incoming(paths: List[Path]) -> None:
        declined = 0
        for src in paths:
            if decline_iphone_file(src):
                declined += 1
        if declined:
            _set_iphone_status(tr("journal.iphone_declined").format(count=declined))

    def _show_iphone_incoming_prompt() -> None:
        iphone_incoming_prompt_after["id"] = None
        paths = sorted(
            [p for p in iphone_incoming_paths if p.exists() and is_transcription_media_file(p)],
            key=_transcription_file_sort_key,
        )
        iphone_incoming_paths.clear()
        iphone_incoming_keys.clear()
        if not paths:
            return
        win = tk.Toplevel(root)
        win.title(tr("journal.iphone_incoming_title"))
        win.transient(root)
        win.geometry("430x260")
        win.resizable(False, False)
        t = th()
        win.configure(bg=t.surface)
        wrap = tk.Frame(win, bg=t.surface, padx=18, pady=16)
        wrap.pack(fill="both", expand=True)
        tk.Label(
            wrap,
            text=tr("journal.iphone_incoming_title"),
            bg=t.surface,
            fg=t.text,
            font=("Segoe UI", 13, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))
        names = "\n".join(p.name for p in paths[:6])
        if len(paths) > 6:
            names += f"\n...and {len(paths) - 6} more"
        tk.Label(
            wrap,
            text=tr("journal.iphone_incoming_body").format(count=len(paths), names=names),
            bg=t.surface,
            fg=t.muted,
            justify="left",
            anchor="w",
            wraplength=390,
            font=("Segoe UI", 9),
        ).pack(fill="x", pady=(0, 16))
        button_row = tk.Frame(wrap, bg=t.surface)
        button_row.pack(fill="x", side="bottom")

        def _close_after(action: Callable[[], None]) -> None:
            action()
            try:
                win.destroy()
            except tk.TclError:
                pass

        accept_btn = tk.Button(
            button_row,
            text=tr("journal.iphone_accept"),
            command=lambda: _close_after(lambda: _accept_iphone_incoming(paths)),
            bg=t.accent,
            fg="white",
            activebackground=t.hover_primary,
            activeforeground="white",
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=12,
            pady=7,
            cursor="hand2",
        )
        accept_btn.pack(side="left", padx=(0, 8))
        decline_btn = tk.Button(
            button_row,
            text=tr("journal.iphone_decline"),
            command=lambda: _close_after(lambda: _decline_iphone_incoming(paths)),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=12,
            pady=7,
            cursor="hand2",
        )
        decline_btn.pack(side="left", padx=(0, 8))

        def _turn_off() -> None:
            save_iphone_passive_receive_enabled(False)
            _decline_iphone_incoming(paths)
            stop_iphone_receiver()
            _refresh_settings_iphone_toggle()

        off_btn = tk.Button(
            button_row,
            text=tr("journal.iphone_turn_off"),
            command=lambda: _close_after(_turn_off),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=12,
            pady=7,
            cursor="hand2",
        )
        off_btn.pack(side="left")

    def receive_iphone_incoming_files(paths: List[Path]) -> None:
        added = 0
        for path in paths:
            if not path.exists() or not is_transcription_media_file(path):
                continue
            key = _iphone_path_key(path)
            if key in iphone_incoming_keys:
                continue
            iphone_incoming_keys.add(key)
            iphone_incoming_paths.append(path)
            added += 1
        if added:
            _set_iphone_status(tr("journal.iphone_incoming_status").format(count=added))
            if iphone_incoming_prompt_after["id"] is None:
                iphone_incoming_prompt_after["id"] = root.after(700, _show_iphone_incoming_prompt)

    def _build_iphone_setup_message(url: str, token: str) -> str:
        return tr("journal.iphone_setup_guide").format(
            url=url,
            token=token,
            inbox=str(IPHONE_INBOX_DIR),
        )

    def show_iphone_setup_window(url: str, token: str) -> None:
        win = tk.Toplevel(root)
        win.title(tr("journal.iphone_inbox"))
        win.transient(root)
        win.geometry("520x690")
        win.minsize(460, 600)
        t = th()
        win.configure(bg=t.surface)
        wrap = tk.Frame(win, bg=t.surface, padx=20, pady=18)
        wrap.pack(fill="both", expand=True)
        wrap.grid_columnconfigure(0, weight=1)

        title = tk.Label(
            wrap,
            text=tr("journal.iphone_waiting"),
            bg=t.surface,
            fg=t.text,
            font=("Segoe UI", 13, "bold"),
            anchor="w",
        )
        title.grid(row=0, column=0, sticky="ew", pady=(0, 12))

        hint = tk.Label(
            wrap,
            text=tr("journal.iphone_qr_hint"),
            bg=t.surface,
            fg=t.muted,
            justify="left",
            anchor="w",
            wraplength=420,
            font=("Segoe UI", 9),
        )
        hint.grid(row=1, column=0, sticky="ew", pady=(0, 12))

        def _copy(value: str) -> None:
            try:
                win.clipboard_clear()
                win.clipboard_append(value)
                win.update()
            except tk.TclError:
                copy_text_to_clipboard(value)

        qr_box = tk.Frame(wrap, bg="white", padx=10, pady=10)
        qr_box.grid(row=2, column=0, pady=(0, 14))
        qr_label = tk.Label(qr_box, bg="white")
        qr_label.pack()
        qr_error = tk.Label(
            qr_box,
            text="",
            bg="white",
            fg="#111827",
            justify="center",
            wraplength=300,
            font=("Segoe UI", 9),
        )

        shortcut_frame = tk.Frame(wrap, bg=t.surface)
        shortcut_frame.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        shortcut_frame.grid_columnconfigure(0, weight=1)
        tk.Label(
            shortcut_frame,
            text=tr("journal.iphone_shortcut_url"),
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 9, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew")
        shortcut_entry = tk.Entry(
            shortcut_frame,
            readonlybackground=t.field,
            fg=t.text,
            relief="flat",
            font=("Segoe UI", 9),
        )
        shortcut_entry.grid(row=1, column=0, sticky="ew", pady=(4, 0), ipady=4)

        note = tk.Label(
            wrap,
            text=tr("journal.iphone_simple_note"),
            bg=t.surface,
            fg=t.muted,
            justify="left",
            anchor="w",
            wraplength=420,
            font=("Segoe UI", 9),
        )
        note.grid(row=4, column=0, sticky="ew", pady=(0, 14))

        guide = tk.Label(
            wrap,
            text=tr("journal.iphone_shortcut_guide_compact"),
            bg=t.field,
            fg=t.text,
            justify="left",
            anchor="w",
            wraplength=450,
            font=("Segoe UI", 9),
            padx=12,
            pady=10,
        )
        guide.grid(row=5, column=0, sticky="ew", pady=(0, 14))

        def _set_entry_value(value: str) -> None:
            shortcut_entry.config(state="normal")
            shortcut_entry.delete(0, "end")
            shortcut_entry.insert(0, value)
            shortcut_entry.config(state="readonly")

        def _render_qr(qr_url: str) -> None:
            try:
                import qrcode
                from PIL import ImageTk

                qr = qrcode.QRCode(border=2, box_size=8)
                qr.add_data(qr_url)
                qr.make(fit=True)
                image = qr.make_image(fill_color="black", back_color="white").convert("RGB")
                photo = ImageTk.PhotoImage(image)
                qr_label.configure(image=photo, text="")
                qr_label.image = photo  # type: ignore[attr-defined]
                qr_error.pack_forget()
            except Exception:
                qr_label.configure(image="", text="")
                qr_label.image = None  # type: ignore[attr-defined]
                qr_error.configure(text=tr("journal.iphone_qr_failed"))
                qr_error.pack(padx=16, pady=16)

        def _refresh_urls() -> None:
            try:
                port = int(iphone_receiver_state.get("port") or IPHONE_IMPORT_DEFAULT_PORT)
            except (TypeError, ValueError):
                port = IPHONE_IMPORT_DEFAULT_PORT
            token_now = str(iphone_receiver_state.get("token") or token)
            refreshed = build_iphone_receiver_urls(port, token_now)
            iphone_receiver_state["urls"] = refreshed
            iphone_receiver_state["url"] = refreshed.get("wifi", "")
            if refreshed.get("wifi"):
                _set_stt_saved_path_display(tr("journal.iphone_inbox"))
                _publish_console_update(
                    f"iPhone receiver URL ready: {refreshed['wifi']}",
                    key="iphone:url",
                )
            _set_entry_value(refreshed.get("shortcut", refreshed.get("wifi", "")))
            _render_qr(refreshed.get("mobile", refreshed.get("wifi", "")))

        _refresh_urls()

        button_row = tk.Frame(wrap, bg=t.surface)
        button_row.grid(row=6, column=0, sticky="e")
        copy_btn = tk.Button(
            button_row,
            text=tr("journal.iphone_copy_shortcut"),
            command=lambda: _copy(shortcut_entry.get()),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            padx=18,
            pady=6,
            cursor="hand2",
        )
        copy_btn.pack(side="left", padx=(0, 8))
        refresh_btn = tk.Button(
            button_row,
            text=tr("journal.iphone_refresh_qr"),
            command=_refresh_urls,
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            padx=18,
            pady=6,
            cursor="hand2",
        )
        refresh_btn.pack(side="left", padx=(0, 8))
        close_btn = tk.Button(
            button_row,
            text=tr("find.close"),
            command=win.destroy,
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            relief="flat",
            padx=18,
            pady=6,
            cursor="hand2",
        )
        close_btn.pack(side="left")

    def _make_iphone_upload_handler(token: str) -> type[BaseHTTPRequestHandler]:
        class IPhoneUploadHandler(BaseHTTPRequestHandler):
            server_version = "DailyLoggerIPhone/1.0"

            def log_message(self, _fmt: str, *_args: Any) -> None:
                return

            def _send_text(self, status: int, text: str, content_type: str = "text/plain; charset=utf-8") -> None:
                data = text.encode("utf-8", errors="replace")
                self.send_response(status)
                self.send_header("Content-Type", content_type)
                self.send_header("Content-Length", str(len(data)))
                self.send_header("Access-Control-Allow-Origin", "*")
                self.send_header("Cache-Control", "no-store")
                self.send_header("Connection", "close")
                self.end_headers()
                self.wfile.write(data)
                try:
                    self.wfile.flush()
                except Exception:
                    pass
                self.close_connection = True

            def _authorized(self) -> bool:
                parsed = urlparse(self.path)
                qs = parse_qs(parsed.query)
                provided = (
                    (qs.get("token") or [""])[0]
                    or self.headers.get("X-DailyLogger-Token", "")
                    or self.headers.get("Authorization", "").removeprefix("Bearer ").strip()
                )
                return hmac.compare_digest(str(provided), token)

            def do_GET(self) -> None:
                parsed = urlparse(self.path)
                if parsed.path == "/api/health":
                    self._send_text(200, "ok")
                    return
                if parsed.path == "/api/upload-status":
                    if not self._authorized():
                        self._send_text(403, "Invalid Daily Logger upload token.")
                        return
                    qs = parse_qs(parsed.query)
                    upload_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", (qs.get("uploadId") or [""])[0])[:80]
                    _stamp, message = iphone_upload_status_by_id.get(upload_id, (0.0, "Daily Logger is waiting for this upload."))
                    self._send_text(200, message)
                    return
                if parsed.path.rstrip("/") == "/iphone":
                    if not self._authorized():
                        body = (
                            "<!doctype html><meta name='viewport' content='width=device-width,initial-scale=1'>"
                            "<title>Daily Logger iPhone Inbox</title>"
                            "<style>body{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;"
                            "margin:36px auto;max-width:560px;padding:0 22px;line-height:1.45;color:#111}</style>"
                            "<h1>Daily Logger iPhone Inbox</h1>"
                            "<p>This link is missing or has an invalid upload token. Start Receive from iPhone again and scan the QR code.</p>"
                        )
                        self._send_text(403, body, "text/html; charset=utf-8")
                        return
                    body = (
                        "<!doctype html><html><head><meta name='viewport' content='width=device-width,initial-scale=1'>"
                        "<title>Daily Logger iPhone Inbox</title>"
                        "<style>"
                        ":root{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;color:#111;background:#f6f7fb}"
                        "body{margin:0;padding:22px}.wrap{max-width:560px;margin:0 auto}"
                        "h1{font-size:26px;margin:10px 0 8px}.hint{color:#4b5563;line-height:1.45}"
                        ".panel{background:white;border:1px solid #e5e7eb;border-radius:14px;padding:18px;margin-top:18px;box-shadow:0 8px 28px rgba(15,23,42,.08)}"
                        "input[type=file]{box-sizing:border-box;width:100%;padding:14px;border:1px dashed #94a3b8;border-radius:12px;background:#f8fafc}"
                        "button{width:100%;margin-top:14px;padding:14px 16px;border:0;border-radius:12px;background:#2563eb;color:white;font-size:17px;font-weight:700}"
                        "button:disabled{background:#94a3b8}.bar{height:10px;background:#e5e7eb;border-radius:999px;overflow:hidden;margin-top:14px}"
                        ".fill{height:100%;width:0;background:#2563eb;transition:width .15s}.status{white-space:pre-wrap;margin-top:14px;color:#1f2937;line-height:1.45}"
                        ".ok{color:#047857}.err{color:#b91c1c}"
                        "</style></head><body><main class='wrap'>"
                        "<h1>Daily Logger iPhone Inbox</h1>"
                        "<p class='hint'>Choose iPhone videos, Voice Memos, or audio files. For very long videos, the Share Sheet Shortcut should send audio only first; this page can upload raw video in chunks, but it is slower.</p>"
                        "<section class='panel'>"
                        "<input id='files' type='file' multiple accept='video/*,audio/*,.mov,.mp4,.m4a,.aac,.caf,.wav,.mp3,.webm,.ogg,.flac,.mpeg,.mpga,.qt'>"
                        "<button id='upload'>Upload to Daily Logger</button>"
                        "<div class='bar'><div id='fill' class='fill'></div></div>"
                        "<div id='status' class='status'>Waiting for files.</div>"
                        "</section></main>"
                        "<script>"
                        f"const CHUNK_SIZE={IPHONE_IMPORT_BROWSER_CHUNK_BYTES};"
                        "const UPLOAD_TIMEOUT_MS=300000;const MAX_UPLOAD_RETRIES=3;"
                        "const token=new URLSearchParams(location.search).get('token')||'';"
                        "const input=document.getElementById('files');const btn=document.getElementById('upload');"
                        "const statusEl=document.getElementById('status');const fill=document.getElementById('fill');"
                        "function setStatus(text,cls=''){statusEl.className='status '+cls;statusEl.textContent=text}"
                        "function endpoint(file,uploadId){const q=new URLSearchParams({token,uploadId,filename:file.name,lastModified:String(file.lastModified||0)});return '/upload?'+q.toString()}"
                        "function chunkEndpoint(file,uploadId,chunkIndex,chunkTotal){const q=new URLSearchParams({token,uploadId,filename:file.name,lastModified:String(file.lastModified||0),contentType:file.type||'application/octet-stream',index:String(chunkIndex),total:String(chunkTotal)});return '/upload-chunk?'+q.toString()}"
                        "function wait(ms){return new Promise(resolve=>setTimeout(resolve,ms))}"
                        "function uploadIdFromUrl(url){try{return new URL(url,location.href).searchParams.get('uploadId')||''}catch(_err){return ''}}"
                        "async function pollPcStatus(uploadId){if(!uploadId){return ''}const q=new URLSearchParams({token,uploadId});"
                        "try{const r=await fetch('/api/upload-status?'+q.toString(),{cache:'no-store'});return r.ok?await r.text():''}catch(_err){return ''}}"
                        "function pcStatusMeansChunkDone(text){return /^Saved\\b/.test(text||'')||/Waiting for the next chunk/i.test(text||'')||/Waiting for PC accept/i.test(text||'')}"
                        "function pcStatusMeansDone(text){return /Transcription complete/i.test(text||'')||/Processed/i.test(text||'')}"
                        "function postBlobOnce(url,blob,file,basePct,spanPct,label,attempt){return new Promise((resolve,reject)=>{"
                        "let uploaded=false;let settled=false;let lastUpdate=Date.now();let polling=false;const uploadId=uploadIdFromUrl(url);const xhr=new XMLHttpRequest();"
                        "function finish(){clearInterval(timer)}"
                        "function succeed(text,abortRequest=false){if(settled){return}settled=true;finish();if(abortRequest){try{xhr.abort()}catch(_err){}}fill.style.width=(basePct+spanPct).toFixed(1)+'%';resolve(text||'Received by Daily Logger.')}"
                        "function fail(err){if(settled){return}settled=true;finish();reject(err)}"
                        "const timer=setInterval(()=>{const seconds=Math.max(1,Math.round((Date.now()-lastUpdate)/1000));"
                        "const fallback=uploaded?`${label}: ${file.name} uploaded. Waiting for PC to save/respond (${seconds}s). Attempt ${attempt}/${MAX_UPLOAD_RETRIES}.`:`${label}: ${file.name} still uploading (${seconds}s since last progress). Attempt ${attempt}/${MAX_UPLOAD_RETRIES}.`;"
                        "setStatus(fallback);if(uploadId&&!polling){polling=true;pollPcStatus(uploadId).then(pc=>{if(pc){setStatus(`${fallback}\\nPC: ${pc}`);if(uploaded&&pcStatusMeansChunkDone(pc)){succeed(pc,true)}}}).finally(()=>{polling=false})}},1000);"
                        "xhr.open('POST',url);xhr.timeout=UPLOAD_TIMEOUT_MS;xhr.setRequestHeader('Content-Type',file.type||'application/octet-stream');"
                        "xhr.upload.onprogress=e=>{lastUpdate=Date.now();if(e.lengthComputable){const part=e.loaded/e.total;const overall=basePct+(part*spanPct);fill.style.width=overall.toFixed(1)+'%';"
                        "uploaded=e.loaded>=e.total;if(uploaded){setStatus(`${label}: ${file.name} (100%). Waiting for PC to save/respond...`)}"
                        "else{setStatus(`${label}: ${file.name} (${Math.round(part*100)}%)`)}}};"
                        "xhr.onload=()=>{if(xhr.status>=200&&xhr.status<300){succeed(xhr.responseText||'Received by Daily Logger.')}else{fail(new Error(xhr.responseText||`Upload failed (${xhr.status})`))}};"
                        "xhr.onerror=()=>{fail(new Error('Could not reach Daily Logger. Keep the app open and stay on the same Wi-Fi.'))};"
                        "xhr.ontimeout=async()=>{const pc=await pollPcStatus(uploadId);if(uploaded&&pcStatusMeansChunkDone(pc)){succeed(pc,true);return}fail(new Error(`${label} timed out while waiting for Daily Logger to respond.`))};"
                        "xhr.send(blob);});}"
                        "async function postBlob(url,blob,file,basePct,spanPct,label){let lastError=null;"
                        "for(let attempt=1;attempt<=MAX_UPLOAD_RETRIES;attempt++){try{if(attempt>1){setStatus(`${label}: retry ${attempt}/${MAX_UPLOAD_RETRIES} for ${file.name}...`)}"
                        "return await postBlobOnce(url,blob,file,basePct,spanPct,label,attempt)}catch(err){lastError=err;if(attempt>=MAX_UPLOAD_RETRIES){break}await wait(1200*attempt)}}"
                        "throw lastError||new Error('Upload failed.')}"
                        "function makeUploadId(){return Date.now().toString(36)+'-'+Math.random().toString(36).slice(2)}"
                        "async function monitorPc(uploadId,fileName){if(!uploadId){return}for(let i=0;i<900;i++){const pc=await pollPcStatus(uploadId);if(pc){setStatus(`Uploaded ${fileName}.\\nPC: ${pc}`,'ok');if(pcStatusMeansDone(pc)){return}}await wait(2000)}}"
                        "async function uploadRaw(file,index,total){const id=makeUploadId();const reply=await postBlob(endpoint(file,id),file,file,(index/total)*100,(1/total)*100,`Uploading ${index+1}/${total}`);return {id,reply}}"
                        "async function uploadChunked(file,index,total){const chunks=Math.ceil(file.size/CHUNK_SIZE);const id=makeUploadId();"
                        "setStatus(`${file.name}: uploading ${chunks} larger chunks. Daily Logger starts transcription after the full media file is saved and accepted on the PC.`);"
                        "for(let c=0;c<chunks;c++){const start=c*CHUNK_SIZE;const blob=file.slice(start,Math.min(file.size,start+CHUNK_SIZE));"
                        "const base=((index+(c/chunks))/total)*100;const span=(1/chunks/total)*100;const label=(c===chunks-1)?`Uploading final chunk ${c+1}/${chunks} for ${index+1}/${total}; PC assembles after this`:`Uploading chunk ${c+1}/${chunks} for ${index+1}/${total}`;"
                        "const reply=await postBlob(chunkEndpoint(file,id,c,chunks),blob,file,base,span,label);setStatus(`${file.name}: Daily Logger received chunk ${c+1}/${chunks}.\\n${reply}`)}return {id,reply:'Saved to Daily Logger.'}}"
                        "function uploadOne(file,index,total){if(file.size>CHUNK_SIZE*1.5){return uploadChunked(file,index,total)}return uploadRaw(file,index,total)}"
                        "btn.addEventListener('click',async()=>{const files=Array.from(input.files||[]).sort((a,b)=>(a.lastModified||0)-(b.lastModified||0)||a.name.localeCompare(b.name));"
                        "if(!files.length){setStatus('Choose one or more files first.','err');return}"
                        "btn.disabled=true;fill.style.width='0';"
                        "try{let last=null;for(let i=0;i<files.length;i++){last=await uploadOne(files[i],i,files.length)}setStatus(`Uploaded ${files.length} file(s). Accept on the PC to start transcription. This page will keep showing PC progress.`, 'ok');if(last&&last.id){monitorPc(last.id,files[files.length-1].name)}}"
                        "catch(err){setStatus(err.message||String(err),'err')}"
                        "finally{btn.disabled=false}});"
                        "</script></body></html>"
                    )
                    self._send_text(200, body, "text/html; charset=utf-8")
                    return
                body = (
                    "<!doctype html><meta name='viewport' content='width=device-width,initial-scale=1'>"
                    "<title>Daily Logger iPhone Inbox</title>"
                    "<style>body{font-family:-apple-system,BlinkMacSystemFont,Segoe UI,sans-serif;"
                    "margin:36px auto;max-width:560px;padding:0 22px;line-height:1.45;color:#111}</style>"
                    "<h1>Daily Logger iPhone Inbox</h1>"
                    "<p>Receiver is running. Open Daily Logger and scan the iPhone Inbox QR code to upload files.</p>"
                )
                self._send_text(200, body, "text/html; charset=utf-8")

            def _query_filename(self, parsed: Any, qs: Dict[str, List[str]]) -> str:
                filename = unquote((qs.get("filename") or [""])[0]).strip()
                if not filename:
                    filename = self.headers.get("X-Filename", "").strip()
                if not filename:
                    filename = _filename_from_content_disposition(
                        self.headers.get("Content-Disposition", "")
                    ).strip()
                return filename

            def _read_upload_body(
                self,
                target: Path,
                length: int,
                progress: Optional[Callable[[int, int], None]] = None,
            ) -> Optional[str]:
                remaining = length
                target.parent.mkdir(parents=True, exist_ok=True)
                written = 0
                next_progress_at = time.monotonic()
                try:
                    self.connection.settimeout(300)
                except Exception:
                    pass
                try:
                    with target.open("wb") as out:
                        while remaining > 0:
                            chunk = self.rfile.read(min(IPHONE_IMPORT_CHUNK_BYTES, remaining))
                            if not chunk:
                                break
                            out.write(chunk)
                            written += len(chunk)
                            remaining -= len(chunk)
                            now = time.monotonic()
                            if progress is not None and (now >= next_progress_at or remaining <= 0):
                                progress(written, length)
                                next_progress_at = now + 1.0
                except (OSError, TimeoutError) as exc:
                    try:
                        target.unlink(missing_ok=True)
                    except OSError:
                        pass
                    return f"Upload stalled while Daily Logger was saving the file: {exc}"
                if remaining != 0:
                    try:
                        target.unlink(missing_ok=True)
                    except OSError:
                        pass
                    return "Upload ended before the full file was received."
                return None

            def _finish_received_file(self, dest: Path, last_modified_raw: str, upload_id: str = "") -> None:
                apply_iphone_last_modified(dest, last_modified_raw)
                if upload_id:
                    iphone_upload_id_by_path[_iphone_path_key(dest)] = upload_id
                root.after(0, lambda p=dest: receive_iphone_incoming_files([p]))

            def _handle_chunk_upload(self, parsed: Any, qs: Dict[str, List[str]], length: int) -> None:
                if length <= 0:
                    self._send_text(400, "No file body was uploaded.")
                    return
                if length > IPHONE_IMPORT_CHUNK_MAX_BYTES:
                    self._send_text(413, "That upload chunk is too large.")
                    return
                upload_id = re.sub(r"[^A-Za-z0-9_.-]+", "_", (qs.get("uploadId") or [""])[0])[:80]
                if not upload_id:
                    self._send_text(400, "Missing upload id.")
                    return
                try:
                    index = int((qs.get("index") or [""])[0])
                    total = int((qs.get("total") or [""])[0])
                except (TypeError, ValueError):
                    self._send_text(400, "Invalid chunk index.")
                    return
                if total < 1 or total > 10000 or index < 0 or index >= total:
                    self._send_text(400, "Invalid chunk index.")
                    return
                content_type = unquote((qs.get("contentType") or [""])[0]).strip() or self.headers.get_content_type()
                filename = self._query_filename(parsed, qs)
                stem, suffix = _sanitize_iphone_upload_name(filename, content_type)
                if not suffix:
                    allowed = ", ".join(sorted(TRANSCRIPTION_MEDIA_SUFFIXES))
                    self._send_text(415, f"Unsupported upload type. Use: {allowed}")
                    return
                chunk_dir = ensure_iphone_incoming_dir() / ".chunks" / upload_id
                chunk_path = chunk_dir / f"{index:06d}.chunk"
                tmp_chunk = chunk_path.with_suffix(".part")
                try:
                    label_name = filename or "iPhone upload"
                    publish_iphone_upload_status(
                        upload_id,
                        f"Receiving {label_name}: chunk {index + 1}/{total} ({_media_size_mb(length)}).",
                    )

                    def _chunk_progress(done: int, expected: int) -> None:
                        publish_iphone_upload_status(
                            upload_id,
                            (
                                f"Receiving {label_name}: chunk {index + 1}/{total} "
                                f"({_media_size_mb(done)} / {_media_size_mb(expected)})."
                            ),
                        )

                    err_msg = self._read_upload_body(tmp_chunk, length, _chunk_progress)
                    if err_msg:
                        publish_iphone_upload_status(upload_id, err_msg)
                        self._send_text(400, err_msg)
                        return
                    tmp_chunk.replace(chunk_path)
                    ready = all((chunk_dir / f"{i:06d}.chunk").is_file() for i in range(total))
                    if not ready:
                        publish_iphone_upload_status(
                            upload_id,
                            f"Saved {label_name}: chunk {index + 1}/{total}. Waiting for the next chunk.",
                        )
                        self._send_text(202, f"Received chunk {index + 1}/{total}.")
                        return
                    dest, err = _unique_iphone_inbox_path(filename, content_type, folder=ensure_iphone_incoming_dir())
                    if dest is None:
                        publish_iphone_upload_status(upload_id, err)
                        self._send_text(415, err)
                        return
                    tmp_dest = dest.with_suffix(dest.suffix + ".part")
                    publish_iphone_upload_status(
                        upload_id,
                        f"All chunks received for {label_name}. Assembling file on PC...",
                    )
                    with tmp_dest.open("wb") as out:
                        for i in range(total):
                            part = chunk_dir / f"{i:06d}.chunk"
                            with part.open("rb") as inp:
                                shutil.copyfileobj(inp, out, length=1024 * 1024)
                            if i == 0 or (i + 1) == total or (i + 1) % 10 == 0:
                                publish_iphone_upload_status(
                                    upload_id,
                                    f"Assembling {label_name}: {i + 1}/{total} chunks.",
                                )
                    tmp_dest.replace(dest)
                    publish_iphone_upload_status(upload_id, f"Saved {label_name} to Daily Logger. Waiting for PC accept.")
                    self._finish_received_file(dest, (qs.get("lastModified") or [""])[0], upload_id)
                    shutil.rmtree(chunk_dir, ignore_errors=True)
                    self._send_text(200, f"Saved to Daily Logger iPhone Inbox: {dest.name}")
                except OSError as exc:
                    try:
                        tmp_chunk.unlink(missing_ok=True)
                    except OSError:
                        pass
                    publish_iphone_upload_status(upload_id, f"Could not save upload chunk: {exc}")
                    self._send_text(500, f"Could not save upload chunk: {exc}")

            def do_POST(self) -> None:
                parsed = urlparse(self.path)
                path_key = parsed.path.rstrip("/")
                if path_key not in {"", "/upload", "/iphone-upload", "/upload-chunk"}:
                    self._send_text(404, "Upload endpoint not found.")
                    return
                if not self._authorized():
                    self._send_text(403, "Invalid Daily Logger upload token.")
                    return
                try:
                    length = int(self.headers.get("Content-Length", "0") or "0")
                except ValueError:
                    length = 0
                if length <= 0:
                    self._send_text(400, "No file body was uploaded.")
                    return
                qs = parse_qs(parsed.query)
                if path_key == "/upload-chunk":
                    self._handle_chunk_upload(parsed, qs, length)
                    return
                if length > IPHONE_IMPORT_MAX_UPLOAD_BYTES:
                    self._send_text(413, "That file is too large for the iPhone Inbox receiver.")
                    return

                content_type = self.headers.get_content_type()
                if content_type.startswith("multipart/"):
                    self._send_text(
                        415,
                        "Use iPhone Shortcuts request body: File, not Form Data.",
                    )
                    return
                filename = self._query_filename(parsed, qs)
                incoming_dir = ensure_iphone_incoming_dir()
                tmp = incoming_dir / f".upload_{secrets.token_hex(8)}.part"
                try:
                    raw_upload_id = re.sub(
                        r"[^A-Za-z0-9_.-]+",
                        "_",
                        (qs.get("uploadId") or [secrets.token_hex(8)])[0],
                    )[:80]
                    label_name = filename or "iPhone upload"
                    publish_iphone_upload_status(
                        raw_upload_id,
                        f"Receiving {label_name} ({_media_size_mb(length)}).",
                    )

                    def _raw_progress(done: int, expected: int) -> None:
                        publish_iphone_upload_status(
                            raw_upload_id,
                            f"Receiving {label_name}: {_media_size_mb(done)} / {_media_size_mb(expected)}.",
                        )

                    err_msg = self._read_upload_body(tmp, length, _raw_progress)
                    if err_msg:
                        publish_iphone_upload_status(raw_upload_id, err_msg)
                        self._send_text(400, err_msg)
                        return
                    suffix = _infer_iphone_upload_suffix_from_file(tmp, filename, content_type)
                    dest, err_msg = _unique_iphone_inbox_path_for_suffix(
                        filename,
                        suffix,
                        folder=incoming_dir,
                    )
                    if dest is None:
                        try:
                            tmp.unlink(missing_ok=True)
                        except OSError:
                            pass
                        publish_iphone_upload_status(raw_upload_id, err_msg)
                        self._send_text(415, err_msg)
                        return
                    tmp.replace(dest)
                    publish_iphone_upload_status(raw_upload_id, f"Saved {label_name} to Daily Logger. Waiting for PC accept.")
                    self._finish_received_file(dest, (qs.get("lastModified") or [""])[0], raw_upload_id)
                except OSError as exc:
                    try:
                        tmp.unlink(missing_ok=True)
                    except OSError:
                        pass
                    publish_iphone_upload_status(raw_upload_id if "raw_upload_id" in locals() else "", f"Could not save upload: {exc}")
                    self._send_text(500, f"Could not save upload: {exc}")
                    return

                self._send_text(200, f"Received by Daily Logger: {dest.name}. Accept it on the PC to transcribe.")

        return IPhoneUploadHandler

    def stop_iphone_receiver() -> None:
        server = iphone_receiver_state.get("server")
        iphone_receiver_state["active"] = False
        iphone_receiver_state["passive"] = False
        iphone_receiver_state["server"] = None
        iphone_receiver_state["thread"] = None
        iphone_receiver_state["url"] = ""
        iphone_receiver_state["urls"] = {}
        if server is not None:
            def _shutdown() -> None:
                try:
                    server.shutdown()
                    server.server_close()
                except Exception:
                    pass

            threading.Thread(target=_shutdown, daemon=True).start()
        _update_iphone_receive_button()

    def start_iphone_receiver(*, show_setup: bool = True, passive: bool = False) -> None:
        if bool(iphone_receiver_state.get("active")):
            if show_setup:
                urls = iphone_receiver_state.get("urls") or {}
                url = str((urls or {}).get("wifi") or iphone_receiver_state.get("url") or "")
                token_now = str(iphone_receiver_state.get("token") or get_or_create_iphone_import_token())
                show_iphone_setup_window(url, token_now)
            return
        token = str(iphone_receiver_state.get("token") or get_or_create_iphone_import_token())
        handler = _make_iphone_upload_handler(token)
        server = None
        bound_port = 0
        last_exc: Optional[BaseException] = None
        for port in range(IPHONE_IMPORT_DEFAULT_PORT, IPHONE_IMPORT_DEFAULT_PORT + 25):
            try:
                server = ThreadingHTTPServer(("0.0.0.0", port), handler)
                server.daemon_threads = True
                bound_port = port
                break
            except OSError as exc:
                last_exc = exc
        if server is None:
            if show_setup:
                messagebox.showerror(
                    tr("journal.iphone_inbox"),
                    tr("journal.iphone_start_failed").format(exc=last_exc or "port unavailable"),
                )
            return
        urls = build_iphone_receiver_urls(bound_port, token)
        url = urls.get("wifi", "")
        iphone_receiver_state.update({
            "active": True,
            "passive": passive,
            "server": server,
            "port": bound_port,
            "url": url,
            "urls": urls,
            "token": token,
        })
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        iphone_receiver_state["thread"] = thread
        thread.start()
        if show_setup:
            _set_stt_saved_path_display(tr("journal.iphone_inbox"))
            if url:
                _publish_console_update(f"iPhone receiver URL ready: {url}", key="iphone:url")
        _set_iphone_status(tr("journal.iphone_waiting"), temp=bool(show_setup))
        _update_iphone_receive_button()
        if show_setup:
            show_iphone_setup_window(url, token)

    def toggle_iphone_receiver() -> None:
        if bool(iphone_receiver_state.get("active")):
            if bool(iphone_receiver_state.get("passive")):
                start_iphone_receiver(show_setup=True, passive=True)
                return
            stop_iphone_receiver()
            return
        start_iphone_receiver(show_setup=True, passive=False)

    journal_cleanup_callbacks.append(stop_iphone_receiver)

    transcription_model_choice = {"value": get_selected_transcription_model_choice()}
    transcription_model_download_busy = {"v": False}
    transcription_model_download_job = {"id": ""}

    def _set_download_manager_status(text: str, *, console_key: str = "") -> None:
        status_var = transcription_model_manager_window.get("status")
        if status_var is not None:
            try:
                status_var.set(text)
            except tk.TclError:
                pass
        if console_key:
            _publish_console_update(text, key=console_key)

    def _set_download_manager_status_from_worker(text: str, *, console_key: str = "") -> None:
        try:
            root.after(0, lambda: _set_download_manager_status(text, console_key=console_key))
        except Exception:
            pass

    def _addon_progress_text(prefix: str, downloaded: int, total: int) -> str:
        if total > 0:
            percent = max(0, min(100, int(downloaded * 100 / total)))
            return f"{prefix} {percent}% ({format_size_short(downloaded)} / {format_size_short(total)})"
        return f"{prefix} {format_size_short(downloaded)} downloaded"

    def _start_model_download_progress_monitor(model_name: str, job_id: str) -> None:
        stats = TRANSCRIPTION_LOCAL_MODEL_STATS.get(model_name, {})
        estimate = int(stats.get("bytes") or 0)
        target = local_transcription_model_path(model_name)
        started = time.monotonic()

        def _tick() -> None:
            if transcription_model_download_job.get("id") != job_id:
                return
            if not bool(transcription_model_download_busy["v"]):
                return
            current_size = directory_size_bytes(target)
            elapsed = max(1, int(time.monotonic() - started))
            if estimate > 0:
                percent = max(1, min(99, int(current_size * 100 / estimate)))
                text = (
                    f"Downloading Local - {model_name}: {percent}% "
                    f"({format_size_short(current_size)} / {format_size_short(estimate)}, {elapsed}s)"
                )
            else:
                text = (
                    f"Downloading Local - {model_name}: "
                    f"{format_size_short(current_size)} downloaded, {elapsed}s"
                )
            _set_download_manager_status(text, console_key=f"model:{model_name}:progress")
            root.after(1000, _tick)

        root.after(250, _tick)

    def _transcription_model_display(choice: str) -> str:
        normalized = normalize_transcription_model_choice(choice)
        if transcription_model_is_cloud(normalized):
            return f"Cloud - {transcription_cloud_model_name(normalized)}"
        model_name = transcription_local_model_name(normalized)
        return f"Local - {model_name}"

    def _set_transcription_model_choice(choice: str) -> None:
        normalized = normalize_transcription_model_choice(choice)
        transcription_model_choice["value"] = normalized
        save_selected_transcription_model_choice(normalized)
        _publish_console_update(
            tr("journal.transcription_model_selected").format(
                model=_transcription_model_display(normalized)
            ),
            key="transcription:model:selected",
        )
        try:
            _refresh_transcription_model_manager_window()
        except NameError:
            pass
        try:
            _refresh_transcription_model_selectors()
        except NameError:
            pass

    def _install_local_addon(on_success: Optional[Callable[[], None]] = None) -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("journal.transcription_model_title"),
                tr("journal.transcription_model_busy"),
            )
            return
        transcription_model_download_busy["v"] = True
        update_transcribe_ui()
        _set_download_manager_status(
            tr("journal.local_addon_installing"),
            console_key="addon:local:installing",
        )

        def _work() -> None:
            def _progress(downloaded: int, total: int) -> None:
                _set_download_manager_status_from_worker(
                    _addon_progress_text("Downloading local transcription add-on:", downloaded, total),
                    console_key="addon:local:download_progress",
                )

            zip_path, err_msg = resolve_or_download_addon_zip(
                LOCAL_TRANSCRIPTION_ADDON_ZIP_NAME,
                progress=_progress,
            )
            ok = False
            if zip_path is not None:
                _set_download_manager_status_from_worker(
                    "Installing local transcription add-on: extracting...",
                    console_key="addon:local:extracting",
                )
                ok, err_msg = install_local_transcription_addon(zip_path)

            def _done() -> None:
                transcription_model_download_busy["v"] = False
                update_transcribe_ui()
                if not ok:
                    _append_console_session_update(
                        f"Local transcription add-on install failed: {err_msg}",
                        key="addon:local:failed",
                    )
                    messagebox.showerror(tr("journal.transcription_model_title"), err_msg[:4000])
                    return
                _set_download_manager_status(
                    f"{tr('journal.local_addon_installed')} (100%)",
                    console_key="addon:local:installed",
                )
                _refresh_transcription_model_manager_window()
                if callable(on_success):
                    root.after(50, on_success)

            root.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    def _uninstall_local_addon() -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("journal.transcription_model_title"),
                tr("journal.transcription_model_busy"),
            )
            return
        if not local_transcription_addon_is_installed():
            return
        if not messagebox.askyesno(
            tr("journal.transcription_model_title"),
            tr("journal.local_addon_uninstall_confirm"),
        ):
            return
        ok, err_msg = uninstall_local_transcription_addon()
        if not ok:
            messagebox.showerror(tr("journal.transcription_model_title"), err_msg[:4000])
            return
        if transcription_model_is_local(transcription_model_choice["value"]):
            _set_transcription_model_choice(TRANSCRIPTION_MODEL_CLOUD)
        _publish_console_update(tr("journal.local_addon_uninstalled"), key="addon:local:uninstalled")
        update_transcribe_ui()
        _refresh_transcription_model_manager_window()

    def _start_local_model_download(model_name: str) -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("journal.transcription_model_title"),
                tr("journal.transcription_model_busy"),
            )
            return
        runtime_ok, runtime_err = ensure_local_transcription_runtime_loaded()
        if not runtime_ok:
            messagebox.showinfo(
                tr("journal.transcription_model_title"),
                f"Local transcription engine needs repair before models can download.\n\n{runtime_err}",
            )
            _append_console_session_update(
                f"Local transcription engine repair needed: {runtime_err}",
                key="addon:local:repair_needed",
            )
            return
        if model_name not in TRANSCRIPTION_LOCAL_MODEL_NAMES:
            return
        if local_transcription_model_is_downloaded(model_name):
            _set_transcription_model_choice(f"local:{model_name}")
            return
        if not messagebox.askyesno(
            tr("journal.transcription_model_title"),
            tr("journal.transcription_model_download_confirm").format(model=model_name),
        ):
            return
        transcription_model_download_busy["v"] = True
        job_id = secrets.token_hex(6)
        transcription_model_download_job["id"] = job_id
        update_transcribe_ui()
        _set_download_manager_status(
            tr("journal.transcription_model_downloading").format(model=model_name),
            console_key=f"model:{model_name}:downloading",
        )

        def _work() -> None:
            def _progress(pct: int) -> None:
                _set_download_manager_status_from_worker(
                    f"Downloading Local - {model_name}: {min(100, max(0, int(pct)))}%",
                    console_key=f"model:{model_name}:progress",
                )

            def _status(text: str) -> None:
                if text:
                    _set_download_manager_status_from_worker(
                        text,
                        console_key=f"model:{model_name}:status",
                    )

            ok, err_msg = download_local_transcription_model(
                model_name,
                progress=_progress,
                status=_status,
            )

            def _done() -> None:
                transcription_model_download_job["id"] = ""
                transcription_model_download_busy["v"] = False
                update_transcribe_ui()
                if not ok:
                    _append_console_session_update(
                        f"Local model download failed ({model_name}): {err_msg}",
                        key=f"model:{model_name}:failed",
                    )
                    messagebox.showerror(tr("journal.transcription_model_title"), err_msg[:4000])
                    _refresh_transcription_model_manager_window()
                    return
                _set_download_manager_status(
                    f"Local model downloaded: {model_name} (100%)",
                    console_key=f"model:{model_name}:downloaded",
                )
                _append_console_session_update(
                    f"Local model downloaded: {model_name}",
                    key=f"model:{model_name}:downloaded",
                )
                _set_transcription_model_choice(f"local:{model_name}")
                _refresh_transcription_model_manager_window()

            root.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    def _uninstall_local_model(model_name: str) -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("journal.transcription_model_title"),
                tr("journal.transcription_model_busy"),
            )
            return
        if not local_transcription_model_is_downloaded(model_name):
            return
        if not messagebox.askyesno(
            tr("journal.transcription_model_title"),
            tr("journal.transcription_model_uninstall_confirm").format(model=model_name),
        ):
            return
        ok, err_msg = uninstall_local_transcription_model(model_name)
        if not ok:
            messagebox.showerror(tr("journal.transcription_model_title"), err_msg[:4000])
            return
        _clear_local_transcription_model_cache(model_name)
        if transcription_model_choice["value"] == f"local:{model_name}":
            _set_transcription_model_choice(TRANSCRIPTION_MODEL_CLOUD)
        update_transcribe_ui()
        _refresh_transcription_model_manager_window()

    def _install_media_tools_addon() -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("download_manager.title"),
                tr("journal.transcription_model_busy"),
            )
            return
        transcription_model_download_busy["v"] = True
        update_transcribe_ui()
        _set_download_manager_status(
            tr("download_manager.media_installing"),
            console_key="addon:media:installing",
        )

        def _work() -> None:
            def _progress(downloaded: int, total: int) -> None:
                _set_download_manager_status_from_worker(
                    _addon_progress_text("Downloading Media Tools:", downloaded, total),
                    console_key="addon:media:download_progress",
                )

            zip_path, err_msg = resolve_or_download_addon_zip(
                MEDIA_TOOLS_ADDON_ZIP_NAME,
                progress=_progress,
            )
            ok = False
            if zip_path is not None:
                _set_download_manager_status_from_worker(
                    "Installing Media Tools: extracting...",
                    console_key="addon:media:extracting",
                )
                ok, err_msg = install_media_tools_addon(zip_path)

            def _done() -> None:
                transcription_model_download_busy["v"] = False
                update_transcribe_ui()
                if not ok:
                    _append_console_session_update(
                        f"Media Tools add-on install failed: {err_msg}",
                        key="addon:media:failed",
                    )
                    messagebox.showerror(tr("download_manager.title"), err_msg[:4000])
                    _refresh_transcription_model_manager_window()
                    return
                _set_download_manager_status(
                    f"{tr('download_manager.media_installed')} (100%)",
                    console_key="addon:media:installed",
                )
                _refresh_transcription_model_manager_window()
                _schedule_iphone_pending_drain_after_install()

            root.after(0, _done)

        threading.Thread(target=_work, daemon=True).start()

    def _uninstall_media_tools_addon() -> None:
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("download_manager.title"),
                tr("journal.transcription_model_busy"),
            )
            return
        if not media_tools_addon_is_installed():
            return
        if not messagebox.askyesno(
            tr("download_manager.title"),
            tr("download_manager.media_uninstall_confirm"),
        ):
            return
        ok, err_msg = uninstall_media_tools_addon()
        if not ok:
            messagebox.showerror(tr("download_manager.title"), err_msg[:4000])
            return
        _publish_console_update(tr("download_manager.media_uninstalled"), key="addon:media:uninstalled")
        update_transcribe_ui()
        _refresh_transcription_model_manager_window()

    transcription_model_manager_window: Dict[str, Any] = {
        "win": None,
        "body": None,
        "status": None,
        "canvas": None,
        "scroll": None,
        "tabs": {},
        "tab": "transcription",
        "size_label": None,
        "size_canvas": None,
        "size_preview": 0,
    }

    def _set_download_manager_tab(tab_name: str) -> None:
        if tab_name not in ("transcription", "media", "reader"):
            tab_name = "transcription"
        transcription_model_manager_window["tab"] = tab_name
        _refresh_transcription_model_manager_window()

    def _path_size_or_estimate(path: Path, estimate: int) -> int:
        actual = directory_size_bytes(path)
        return actual if actual > 0 else int(estimate)

    def _installed_addons_size_bytes() -> int:
        total = 0
        if local_transcription_addon_is_installed():
            total += directory_size_bytes(LOCAL_TRANSCRIPTION_ADDON_DIR)
        if media_tools_addon_is_installed():
            total += directory_size_bytes(MEDIA_TOOLS_ADDON_DIR)
        return total

    def _visible_model_size_bytes() -> int:
        total = 0
        for model_name in TRANSCRIPTION_LOCAL_MODEL_NAMES:
            path = local_transcription_model_path(model_name)
            if path.is_dir():
                total += directory_size_bytes(path)
        return total

    def _other_model_paths() -> List[Path]:
        visible = {name.casefold() for name in TRANSCRIPTION_LOCAL_MODEL_NAMES}
        try:
            if not LOCAL_TRANSCRIPTION_MODEL_DIR.is_dir():
                return []
            items = [
                path
                for path in LOCAL_TRANSCRIPTION_MODEL_DIR.iterdir()
                if path.name.casefold() not in visible
                and not path.name.lower().endswith(".part")
            ]
        except OSError:
            return []
        return sorted(items, key=lambda p: p.name.lower())

    def _other_model_size_bytes() -> int:
        return sum(directory_size_bytes(path) for path in _other_model_paths())

    def _uninstall_other_local_models() -> None:
        paths = _other_model_paths()
        if not paths:
            return
        if transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"]:
            messagebox.showinfo(
                tr("download_manager.title"),
                tr("journal.transcription_model_busy"),
            )
            return
        names = ", ".join(path.name for path in paths[:8])
        if len(paths) > 8:
            names += f", +{len(paths) - 8}"
        if not messagebox.askyesno(
            tr("download_manager.title"),
            tr("download_manager.other_models_confirm").format(names=names),
        ):
            return
        try:
            root_dir = LOCAL_TRANSCRIPTION_MODEL_DIR.resolve()
        except OSError as exc:
            messagebox.showerror(tr("download_manager.title"), str(exc)[:4000])
            return
        errors: List[str] = []
        for path in paths:
            try:
                target = path.resolve()
                target.relative_to(root_dir)
                if target.is_dir():
                    shutil.rmtree(target)
                elif target.exists():
                    target.unlink()
            except (OSError, ValueError) as exc:
                errors.append(f"{path.name}: {exc}")
        if errors:
            messagebox.showerror(tr("download_manager.title"), "\n".join(errors[:8])[:4000])
        update_transcribe_ui()
        _refresh_transcription_model_manager_window()

    def _base_package_size_bytes() -> int:
        try:
            if getattr(sys, "frozen", False):
                return directory_size_bytes(BASE_DIR)
            dist_root = BASE_DIR / "dist" / "DailyLogger"
            if dist_root.is_dir():
                return directory_size_bytes(dist_root)
        except OSError:
            pass
        return 0

    def _download_manager_size_parts() -> Dict[str, int]:
        return {
            "base": _base_package_size_bytes(),
            "addons": _installed_addons_size_bytes(),
            "visible_models": _visible_model_size_bytes(),
            "other_models": _other_model_size_bytes(),
        }

    def _update_download_manager_size_bar(preview_free_bytes: int = 0) -> None:
        label = transcription_model_manager_window.get("size_label")
        canvas = transcription_model_manager_window.get("size_canvas")
        if label is None or canvas is None:
            return
        t = th()
        parts = _download_manager_size_parts()
        total = max(sum(parts.values()), 1)
        text = tr("download_manager.size_summary").format(
            base=format_size_short(parts["base"]),
            addons=format_size_short(parts["addons"]),
            visible_models=format_size_short(parts["visible_models"]),
            other_models=format_size_short(parts["other_models"]),
            total=format_size_short(sum(parts.values())),
        )
        if preview_free_bytes > 0:
            text += "  " + tr("download_manager.free_preview").format(
                size=format_size_short(preview_free_bytes)
            )
        try:
            label.config(text=text, fg=("#F04438" if preview_free_bytes > 0 else t.muted))
            canvas.delete("all")
            width = max(1, int(canvas.winfo_width() or 1))
            height = max(1, int(canvas.winfo_height() or 10))
            x = 0
            colors = {
                "base": t.muted,
                "addons": t.accent,
                "visible_models": t.hover_primary,
                "other_models": "#9CA3AF",
            }
            for index, key in enumerate(("base", "addons", "visible_models", "other_models")):
                value = max(0, int(parts[key]))
                seg_w = int(width * value / total) if total else 0
                if index == 3:
                    seg_w = width - x
                if value > 0 and seg_w > 0:
                    canvas.create_rectangle(x, 0, x + seg_w, height, fill=colors[key], width=0)
                x += seg_w
            if preview_free_bytes > 0:
                preview_w = max(6, min(width, int(width * min(preview_free_bytes, total) / total)))
                canvas.create_rectangle(width - preview_w, 0, width, height, fill="#F04438", width=0)
            canvas.create_rectangle(0, 0, width, height, outline=t.border, width=1)
        except tk.TclError:
            pass

    def _transcription_model_state_text(choice: str) -> str:
        normalized = normalize_transcription_model_choice(choice)
        selected = normalize_transcription_model_choice(transcription_model_choice["value"])
        if transcription_model_is_cloud(normalized):
            if selected == normalized:
                if transcription_cloud_is_ready():
                    return tr("journal.transcription_state_default")
                return (
                    f"{tr('journal.transcription_state_default')} - "
                    f"{tr('journal.transcription_state_api_key_required')}"
                )
            return (
                tr("journal.transcription_state_ready")
                if transcription_cloud_is_ready()
                else tr("journal.transcription_state_api_key_required")
            )
        model_name = transcription_local_model_name(normalized)
        downloaded = local_transcription_model_is_downloaded(model_name)
        runtime_ready, _runtime_err = ensure_local_transcription_runtime_loaded()
        if selected == normalized and downloaded and runtime_ready:
            return tr("journal.transcription_state_default")
        if downloaded and runtime_ready:
            return tr("journal.transcription_state_downloaded")
        if downloaded and not runtime_ready:
            return tr("journal.transcription_state_addon_required")
        if not runtime_ready:
            return tr("journal.transcription_state_addon_required")
        return tr("journal.transcription_state_not_downloaded")

    def _style_manager_button(btn: Any, *, enabled: bool = True) -> None:
        t = th()
        if enabled:
            btn.config(
                state="normal",
                bg=t.btn_secondary,
                fg=t.text,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
                cursor="hand2",
            )
        else:
            db = t.transcribe_idle_disabled_config()
            btn.config(
                state="disabled",
                bg=db[0],
                fg=db[1],
                activebackground=db[2],
                activeforeground=db[3],
                disabledforeground=db[4],
                cursor="arrow",
            )

    def _make_manager_button(
        parent: Any,
        text: str,
        command: Callable[[], None],
        enabled: bool = True,
        *,
        free_bytes: int = 0,
    ) -> Any:
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10,
            pady=5,
            wraplength=150,
        )
        _style_manager_button(btn, enabled=enabled)
        btn.pack(side="left", padx=(0, 8), pady=(8, 0))
        bind_button_hover_if_enabled(
            btn,
            lambda: th().side_action_bind_rest(),
            lambda: th().accent,
            lambda: "white",
        )
        if free_bytes > 0:
            def _show_free_preview(_evt: Optional[Any] = None, b: int = free_bytes) -> None:
                transcription_model_manager_window["size_preview"] = b
                _update_download_manager_size_bar(b)

            def _clear_free_preview(_evt: Optional[Any] = None) -> None:
                transcription_model_manager_window["size_preview"] = 0
                _update_download_manager_size_bar(0)

            btn.bind(
                "<Enter>",
                _show_free_preview,
                add="+",
            )
            btn.bind(
                "<Leave>",
                _clear_free_preview,
                add="+",
            )
        return btn

    def _make_model_card(
        parent: Any,
        title: str,
        subtitle: str,
        note: str,
        state_text: str,
        actions: List[Tuple[str, Callable[[], None], bool, int]],
    ) -> None:
        t = th()
        card = tk.Frame(parent, bg=t.field, highlightthickness=1, highlightbackground=t.border)
        card.pack(fill="x", pady=(0, 8))
        card.grid_columnconfigure(0, weight=1)
        card.grid_columnconfigure(1, weight=0)
        info = tk.Frame(card, bg=t.field)
        info.grid(row=0, column=0, sticky="ew", padx=12, pady=9)
        title_lbl = tk.Label(
            info,
            text=title,
            bg=t.field,
            fg=t.text,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
            justify="left",
        )
        title_lbl.pack(fill="x")
        subtitle_lbl = tk.Label(
            info,
            text=f"{subtitle}  |  {state_text}",
            bg=t.field,
            fg=t.accent if state_text == tr("journal.transcription_state_default") else t.muted,
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
            wraplength=430,
        )
        subtitle_lbl.pack(fill="x", pady=(2, 0))
        note_lbl = tk.Label(
            info,
            text=note,
            bg=t.field,
            fg=t.muted,
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
            wraplength=430,
        )
        note_lbl.pack(fill="x", pady=(2, 0))
        buttons = tk.Frame(card, bg=t.field)
        buttons.grid(row=0, column=1, sticky="e", padx=12, pady=9)
        for label, command, enabled, free_bytes in actions:
            _make_manager_button(buttons, label, command, enabled, free_bytes=free_bytes)

    def _refresh_transcription_model_manager_window() -> None:
        win = transcription_model_manager_window.get("win")
        body = transcription_model_manager_window.get("body")
        status_var = transcription_model_manager_window.get("status")
        if win is None or body is None:
            return
        try:
            if not bool(win.winfo_exists()):
                return
        except tk.TclError:
            return
        t = th()
        try:
            win.configure(bg=t.surface)
            canvas = transcription_model_manager_window.get("canvas")
            scroll = transcription_model_manager_window.get("scroll")
            if canvas is not None:
                canvas.configure(bg=t.surface, highlightbackground=t.border)
            if scroll is not None:
                scroll.configure(bg=t.panel, troughcolor=t.field, activebackground=t.accent)
            body.configure(bg=t.surface)
            for child in body.winfo_children():
                child.destroy()
        except tk.TclError:
            return
        if status_var is not None:
            selected_label = _transcription_model_display(transcription_model_choice["value"])
            status_var.set(tr("journal.transcription_manager_current").format(model=selected_label))

        busy = bool(transcription_model_download_busy["v"] or transcribing_busy["v"] or recording_ui_busy["v"])
        active_tab = str(transcription_model_manager_window.get("tab") or "transcription")
        runtime_ready, runtime_err = ensure_local_transcription_runtime_loaded()
        if not runtime_ready and status_var is not None and active_tab == "transcription":
            selected_label = _transcription_model_display(transcription_model_choice["value"])
            status_var.set(
                f"{tr('journal.transcription_manager_current').format(model=selected_label)}"
                f" | {runtime_err}"
            )
        selected = normalize_transcription_model_choice(transcription_model_choice["value"])
        tab_buttons = transcription_model_manager_window.get("tabs") or {}
        for key, btn in dict(tab_buttons).items():
            try:
                if key == active_tab:
                    btn.config(
                        bg=t.accent,
                        fg="white",
                        activebackground=t.hover_primary,
                        activeforeground="white",
                    )
                else:
                    btn.config(
                        bg=t.btn_secondary,
                        fg=t.text,
                        activebackground=t.secondary_hover,
                        activeforeground=t.text,
                    )
            except tk.TclError:
                pass

        def _heading(parent: Any, key: str) -> None:
            tk.Label(
                parent,
                text=tr(key),
                bg=t.surface,
                fg=t.text,
                font=("Segoe UI", 12, "bold"),
                anchor="w",
            ).pack(fill="x", pady=(0, 8))

        def _render_transcription_tab() -> None:
            left = tk.Frame(body, bg=t.surface)
            right = tk.Frame(body, bg=t.surface)
            left.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
            right.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
            body.grid_columnconfigure(0, weight=1, uniform="download_manager")
            body.grid_columnconfigure(1, weight=1, uniform="download_manager")
            body.grid_rowconfigure(0, weight=1)
            _heading(left, "journal.transcription_manager_free_models")
            _heading(right, "journal.transcription_manager_cloud_models")

            for model_name in TRANSCRIPTION_LOCAL_MODEL_NAMES:
                choice = f"local:{model_name}"
                stats = TRANSCRIPTION_LOCAL_MODEL_STATS.get(model_name, {})
                downloaded = local_transcription_model_is_downloaded(model_name)
                recommended = (
                    f" {tr('journal.transcription_manager_recommended')}"
                    if model_name == TRANSCRIPTION_SUGGESTED_LOCAL_MODEL
                    else ""
                )
                actions: List[Tuple[str, Callable[[], None], bool, int]] = []
                if not runtime_ready:
                    actions.append((tr("journal.local_addon_repair_short"), _install_local_addon, not busy, 0))
                elif not downloaded:
                    actions.append((
                        tr("journal.transcription_model_download"),
                        lambda m=model_name: _start_local_model_download(m),
                        not busy,
                        0,
                    ))
                else:
                    model_size = directory_size_bytes(local_transcription_model_path(model_name))
                    actions.append((
                        tr("journal.transcription_model_use_default"),
                        lambda c=choice: _set_transcription_model_choice(c),
                        (not busy and selected != choice),
                        0,
                    ))
                    actions.append((
                        tr("journal.transcription_model_uninstall_short"),
                        lambda m=model_name: _uninstall_local_model(m),
                        not busy,
                        model_size,
                    ))
                _make_model_card(
                    left,
                    f"Local - {model_name}{recommended}",
                    tr("journal.transcription_manager_local_stats").format(
                        quality=stats.get("quality", ""),
                        speed=stats.get("speed", ""),
                        disk=stats.get("disk", ""),
                    ),
                    str(stats.get("note", "")),
                    _transcription_model_state_text(choice),
                    actions,
                )

            other_model_size = _other_model_size_bytes()
            if other_model_size > 0:
                _make_model_card(
                    left,
                    tr("download_manager.other_models_title"),
                    tr("download_manager.other_models_stats").format(
                        size=format_size_short(other_model_size)
                    ),
                    tr("download_manager.other_models_note"),
                    tr("journal.transcription_state_downloaded"),
                    [(
                        tr("journal.transcription_model_uninstall_short"),
                        _uninstall_other_local_models,
                        not busy,
                        other_model_size,
                    )],
                )

            if runtime_ready and local_transcription_addon_is_installed():
                footer = tk.Frame(left, bg=t.surface)
                footer.pack(fill="x", pady=(0, 8))
                _make_manager_button(
                    footer,
                    tr("journal.local_addon_uninstall_short"),
                    _uninstall_local_addon,
                    enabled=not busy,
                    free_bytes=directory_size_bytes(LOCAL_TRANSCRIPTION_ADDON_DIR),
                )

            for cloud_name in TRANSCRIPTION_CLOUD_MODEL_NAMES:
                cloud_choice = normalize_transcription_model_choice(f"cloud:{cloud_name}")
                cloud_stats = TRANSCRIPTION_CLOUD_MODEL_STATS.get(
                    cloud_name,
                    {
                        "quality": "Cloud",
                        "speed": "Fast",
                        "cost": "Uses API tokens",
                        "note": "OpenAI cloud transcription model.",
                    },
                )
                actions = [
                    (
                        tr("journal.transcription_model_use_default"),
                        lambda c=cloud_choice: _set_transcription_model_choice(c),
                        (not busy and selected != cloud_choice),
                        0,
                    )
                ]
                if not transcription_cloud_is_ready():
                    actions.append((
                        tr("journal.transcription_add_api_key"),
                        lambda: (_close_transcription_model_manager(), _goto_settings_token_field()),
                        not busy,
                        0,
                    ))
                _make_model_card(
                    right,
                    f"Cloud - {cloud_name}",
                    tr("journal.transcription_manager_cloud_stats").format(
                        quality=cloud_stats.get("quality", ""),
                        speed=cloud_stats.get("speed", ""),
                        cost=cloud_stats.get("cost", ""),
                    ),
                    str(cloud_stats.get("note", "")),
                    _transcription_model_state_text(cloud_choice),
                    actions,
                )

        def _render_media_tab() -> None:
            body.grid_columnconfigure(0, weight=1)
            body.grid_columnconfigure(1, weight=0)
            _heading(body, "download_manager.media_tab")
            installed = media_tools_addon_is_installed()
            media_size = directory_size_bytes(MEDIA_TOOLS_ADDON_DIR)
            state = (
                tr("journal.transcription_state_downloaded")
                if installed
                else tr("journal.transcription_state_not_downloaded")
            )
            actions: List[Tuple[str, Callable[[], None], bool, int]]
            if installed:
                actions = [(
                    tr("journal.transcription_model_uninstall_short"),
                    _uninstall_media_tools_addon,
                    not busy,
                    media_size,
                )]
            else:
                actions = [(
                    tr("download_manager.install"),
                    _install_media_tools_addon,
                    not busy,
                    0,
                )]
            _make_model_card(
                body,
                tr("download_manager.media_title"),
                tr("download_manager.media_stats").format(
                    size=format_size_short(media_size or MEDIA_TOOLS_ADDON_ESTIMATED_BYTES)
                ),
                tr("download_manager.media_note"),
                state,
                actions,
            )

        def _render_reader_tab() -> None:
            body.grid_columnconfigure(0, weight=1)
            body.grid_columnconfigure(1, weight=0)
            _heading(body, "download_manager.reader_tab")
            reader_path = BASE_DIR / "_internal" / "virtual-journal-reader"
            if not reader_path.exists():
                reader_path = BASE_DIR / "virtual-journal-reader"
            reader_size = directory_size_bytes(reader_path)
            _make_model_card(
                body,
                tr("download_manager.reader_title"),
                tr("download_manager.reader_stats").format(size=format_size_short(reader_size)),
                tr("download_manager.reader_note"),
                tr("download_manager.core_installed"),
                [],
            )

        if active_tab == "media":
            _render_media_tab()
        elif active_tab == "reader":
            _render_reader_tab()
        else:
            _render_transcription_tab()
        _update_download_manager_size_bar(int(transcription_model_manager_window.get("size_preview") or 0))

    def _close_transcription_model_manager() -> None:
        win = transcription_model_manager_window.get("win")
        if win is not None:
            try:
                win.destroy()
            except tk.TclError:
                pass
        transcription_model_manager_window.update(
            {
                "win": None,
                "body": None,
                "status": None,
                "canvas": None,
                "scroll": None,
                "tabs": {},
                "size_label": None,
                "size_canvas": None,
                "size_preview": 0,
            }
        )

    def _open_transcription_downloads_manager(tab_name: str = "") -> None:
        if tab_name:
            transcription_model_manager_window["tab"] = (
                tab_name if tab_name in ("transcription", "media", "reader") else "transcription"
            )
        win = transcription_model_manager_window.get("win")
        try:
            if win is not None and bool(win.winfo_exists()):
                win.lift()
                win.focus_force()
                _refresh_transcription_model_manager_window()
                return
        except tk.TclError:
            transcription_model_manager_window.update(
                {
                    "win": None,
                    "body": None,
                    "status": None,
                    "canvas": None,
                    "scroll": None,
                    "tabs": {},
                    "size_label": None,
                    "size_canvas": None,
                    "size_preview": 0,
                }
            )
        t = th()
        win = tk.Toplevel(root)
        win.title(tr("download_manager.title"))
        win.geometry("900x650")
        win.minsize(780, 540)
        win.configure(bg=t.surface)
        win.transient(root)
        win.protocol("WM_DELETE_WINDOW", _close_transcription_model_manager)
        transcription_model_manager_window["win"] = win

        outer = tk.Frame(win, bg=t.surface)
        outer.pack(fill="both", expand=True, padx=18, pady=16)
        title_lbl = tk.Label(
            outer,
            text=tr("download_manager.title"),
            bg=t.surface,
            fg=t.text,
            font=("Segoe UI", 16, "bold"),
            anchor="w",
        )
        title_lbl.pack(fill="x")
        status_var = tk.StringVar(value="")
        transcription_model_manager_window["status"] = status_var
        tk.Label(
            outer,
            textvariable=status_var,
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(fill="x", pady=(2, 12))
        tabs_frame = tk.Frame(outer, bg=t.surface)
        tabs_frame.pack(fill="x", pady=(0, 10))
        tab_specs = (
            ("transcription", "download_manager.transcription_tab"),
            ("media", "download_manager.media_tab"),
            ("reader", "download_manager.reader_tab"),
        )
        tab_buttons: Dict[str, Any] = {}
        for tab_name, label_key in tab_specs:
            tab_btn = tk.Button(
                tabs_frame,
                text=tr(label_key),
                command=lambda n=tab_name: _set_download_manager_tab(n),
                relief="flat",
                font=("Segoe UI", 9, "bold"),
                padx=12,
                pady=6,
                cursor="hand2",
            )
            tab_btn.pack(side="left", padx=(0, 8))
            tab_buttons[tab_name] = tab_btn
            bind_button_hover_if_enabled(
                tab_btn,
                lambda b=tab_btn, n=tab_name: (
                    "normal",
                    th().accent if transcription_model_manager_window.get("tab") == n else th().btn_secondary,
                    "white" if transcription_model_manager_window.get("tab") == n else th().text,
                    th().hover_primary if transcription_model_manager_window.get("tab") == n else th().secondary_hover,
                    "white" if transcription_model_manager_window.get("tab") == n else th().text,
                ),
                lambda: th().hover_primary,
                lambda: "white",
            )
        transcription_model_manager_window["tabs"] = tab_buttons
        body_holder = tk.Frame(outer, bg=t.surface)
        body_holder.pack(fill="both", expand=True)
        body_canvas = tk.Canvas(
            body_holder,
            bg=t.surface,
            highlightthickness=1,
            highlightbackground=t.border,
            bd=0,
        )
        body_scroll = tk.Scrollbar(
            body_holder,
            command=body_canvas.yview,
            bg=t.panel,
            troughcolor=t.field,
            activebackground=t.accent,
            bd=0,
            highlightthickness=0,
            width=11,
        )
        body_canvas.configure(yscrollcommand=body_scroll.set)
        body_canvas.pack(side="left", fill="both", expand=True)
        body_scroll.pack(side="right", fill="y")
        body = tk.Frame(body_canvas, bg=t.surface)
        body_window = body_canvas.create_window((0, 0), window=body, anchor="nw")

        def _sync_manager_scroll(_evt: Optional[Any] = None) -> None:
            try:
                body_canvas.configure(scrollregion=body_canvas.bbox("all"))
                body_canvas.itemconfigure(body_window, width=body_canvas.winfo_width())
            except tk.TclError:
                pass

        body.bind("<Configure>", _sync_manager_scroll, add="+")
        body_canvas.bind("<Configure>", _sync_manager_scroll, add="+")

        def _manager_mousewheel(evt: Any) -> str:
            try:
                body_canvas.yview_scroll(int(-1 * (evt.delta / 120)), "units")
            except tk.TclError:
                pass
            return "break"

        win.bind("<MouseWheel>", _manager_mousewheel, add="+")
        transcription_model_manager_window["body"] = body
        transcription_model_manager_window["canvas"] = body_canvas
        transcription_model_manager_window["scroll"] = body_scroll
        size_wrap = tk.Frame(outer, bg=t.surface)
        size_wrap.pack(fill="x", pady=(10, 0))
        size_label = tk.Label(
            size_wrap,
            text="",
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
            wraplength=820,
        )
        size_label.pack(fill="x", pady=(0, 4))
        size_canvas = tk.Canvas(
            size_wrap,
            height=12,
            bg=t.field,
            highlightthickness=1,
            highlightbackground=t.border,
            bd=0,
        )
        size_canvas.pack(fill="x")
        transcription_model_manager_window["size_label"] = size_label
        transcription_model_manager_window["size_canvas"] = size_canvas
        size_canvas.bind(
            "<Configure>",
            lambda _e: _update_download_manager_size_bar(
                int(transcription_model_manager_window.get("size_preview") or 0)
            ),
            add="+",
        )
        bottom = tk.Frame(outer, bg=t.surface)
        bottom.pack(fill="x", pady=(10, 0))
        close_btn = tk.Button(
            bottom,
            text=tr("find.close"),
            command=_close_transcription_model_manager,
            relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=14,
            pady=6,
        )
        _style_manager_button(close_btn)
        close_btn.pack(side="right")
        bind_button_hover_if_enabled(
            close_btn,
            lambda: th().side_action_bind_rest(),
            lambda: th().hover_primary,
            lambda: "white",
        )
        _refresh_transcription_model_manager_window()
        win.lift()
        win.focus_force()

    transcription_models_manager_hooks["open"] = _open_transcription_downloads_manager

    def _show_transcription_model_menu(anchor: Any) -> None:
        if transcription_model_download_busy["v"]:
            return
        menu = tk.Menu(anchor, tearoff=0)
        selected = normalize_transcription_model_choice(transcription_model_choice["value"])
        for cloud_name in TRANSCRIPTION_CLOUD_MODEL_NAMES:
            cloud_choice = normalize_transcription_model_choice(f"cloud:{cloud_name}")
            cloud_prefix = "[x] " if selected == cloud_choice else "    "
            menu.add_command(
                label=f"{cloud_prefix}{_transcription_model_display(cloud_choice)}",
                command=lambda c=cloud_choice: _set_transcription_model_choice(c),
            )
        downloaded_choices = [
            model_name
            for model_name in usable_local_transcription_model_names()
            if local_transcription_model_is_downloaded(model_name)
        ]
        if downloaded_choices:
            menu.add_separator()
        for model_name in downloaded_choices:
            choice = f"local:{model_name}"
            prefix = "[x] " if selected == choice else "    "
            menu.add_command(
                label=f"{prefix}{_transcription_model_display(choice)}",
                command=lambda c=choice: _set_transcription_model_choice(c),
            )
        menu.add_separator()
        menu.add_command(
            label=tr("journal.transcription_manage_models"),
            command=lambda: _open_transcription_downloads_manager("transcription"),
        )
        try:
            menu.tk_popup(anchor.winfo_rootx(), anchor.winfo_rooty() + anchor.winfo_height())
        finally:
            try:
                menu.grab_release()
            except tk.TclError:
                pass

    def _transcription_model_selector_label() -> str:
        return tr("journal.transcription_model_short")

    def _transcription_model_selector_entries() -> List[Tuple[str, str]]:
        selected = normalize_transcription_model_choice(transcription_model_choice["value"])
        entries: List[Tuple[str, str]] = []
        for cloud_name in TRANSCRIPTION_CLOUD_MODEL_NAMES:
            cloud_choice = normalize_transcription_model_choice(f"cloud:{cloud_name}")
            suffix = tr("journal.transcription_state_default") if selected == cloud_choice else ""
            entries.append(
                (
                    f"{_transcription_model_display(cloud_choice)} {suffix}".strip(),
                    cloud_choice,
                )
            )
        for model_name in downloaded_local_transcription_model_names():
            if not local_transcription_model_is_downloaded(model_name):
                continue
            choice = f"local:{model_name}"
            suffix = tr("journal.transcription_state_default") if selected == choice else ""
            entries.append(
                (
                    f"{_transcription_model_display(choice)} {suffix}".strip(),
                    choice,
                )
            )
        entries.append((tr("journal.transcription_manage_models"), "__manage__"))
        return entries

    def _refresh_transcription_model_selectors() -> None:
        if not transcription_model_selector_is_combo:
            return
        transcription_model_combo_map.clear()
        values: List[str] = []
        for label, action in _transcription_model_selector_entries():
            transcription_model_combo_map[label] = action
            values.append(label)
        if not values:
            values = [_transcription_model_selector_label()]
        for widget, var in (
            (transcribe_model_btn, transcribe_model_var),
            (transcribe_file_model_btn, transcribe_file_model_var),
        ):
            try:
                widget.config(values=tuple(values))
                var.set(_transcription_model_selector_label())
            except tk.TclError:
                pass

    def _on_transcription_model_selector_change(variable: Any) -> None:
        label = str(variable.get() or "")
        action = transcription_model_combo_map.get(label, "")
        variable.set(_transcription_model_selector_label())
        if not action:
            return
        if action == "__manage__":
            _open_transcription_downloads_manager("transcription")
            return
        _set_transcription_model_choice(action)

    transcription_setup_prompt_window: Dict[str, Any] = {"win": None}

    def _maybe_prompt_media_tools_from_error(text: str) -> bool:
        if "Media Tools add-on" not in (text or ""):
            return False
        if messagebox.askyesno(
            tr("download_manager.title"),
            tr("download_manager.media_required_prompt"),
        ):
            _install_media_tools_addon()
        return True

    def _close_transcription_setup_prompt() -> None:
        win = transcription_setup_prompt_window.get("win")
        if win is not None:
            try:
                win.destroy()
            except tk.TclError:
                pass
        transcription_setup_prompt_window["win"] = None

    def _show_transcription_setup_prompt() -> None:
        win = transcription_setup_prompt_window.get("win")
        try:
            if win is not None and bool(win.winfo_exists()):
                win.lift()
                win.focus_force()
                return
        except tk.TclError:
            transcription_setup_prompt_window["win"] = None
        t = th()
        win = tk.Toplevel(root)
        transcription_setup_prompt_window["win"] = win
        win.title(tr("journal.transcription_setup_title"))
        win.geometry("520x250")
        win.minsize(460, 230)
        win.configure(bg=t.surface)
        win.transient(root)
        win.protocol("WM_DELETE_WINDOW", _close_transcription_setup_prompt)
        wrap = tk.Frame(win, bg=t.surface)
        wrap.pack(fill="both", expand=True, padx=18, pady=16)
        tk.Label(
            wrap,
            text=tr("journal.transcription_setup_title"),
            bg=t.surface,
            fg=t.text,
            font=("Segoe UI", 14, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 8))
        tk.Label(
            wrap,
            text=tr("journal.transcription_setup_body").format(
                model=TRANSCRIPTION_SUGGESTED_LOCAL_MODEL
            ),
            bg=t.surface,
            fg=t.muted,
            font=("Segoe UI", 10),
            anchor="w",
            justify="left",
            wraplength=470,
        ).pack(fill="x", pady=(0, 16))
        buttons = tk.Frame(wrap, bg=t.surface)
        buttons.pack(fill="x", side="bottom")

        def _download_suggested() -> None:
            _close_transcription_setup_prompt()
            _start_local_model_download(TRANSCRIPTION_SUGGESTED_LOCAL_MODEL)

        def _enter_key() -> None:
            _close_transcription_setup_prompt()
            _goto_settings_token_field()

        _make_manager_button(
            buttons,
            tr("journal.transcription_setup_download").format(
                model=TRANSCRIPTION_SUGGESTED_LOCAL_MODEL
            ),
            _download_suggested,
            True,
        )
        _make_manager_button(buttons, tr("journal.transcription_add_api_key"), _enter_key, True)
        _make_manager_button(buttons, tr("find.close"), _close_transcription_setup_prompt, True)
        win.lift()
        win.focus_force()

    def _ensure_selected_transcription_model_ready() -> bool:
        choice = normalize_transcription_model_choice(transcription_model_choice["value"])
        if transcription_model_is_cloud(choice):
            if get_openai_api_key():
                return True
            fallback_choice = first_usable_local_transcription_choice()
            if fallback_choice:
                if messagebox.askyesno(
                    tr("journal.transcription_model_title"),
                    tr("journal.transcription_use_downloaded_confirm").format(
                        model=_transcription_model_display(fallback_choice)
                    ),
                ):
                    _set_transcription_model_choice(fallback_choice)
                    return True
            _show_transcription_setup_prompt()
            return False
        model_name = transcription_local_model_name(choice)
        if local_transcription_model_is_downloaded(model_name):
            runtime_ok, runtime_err = ensure_local_transcription_runtime_loaded()
            if runtime_ok:
                return True
            _append_console_session_update(
                f"Local transcription add-on is not ready: {runtime_err}",
                key="transcription:local_runtime_failed",
            )
            if get_openai_api_key() and messagebox.askyesno(
                tr("journal.transcription_model_title"),
                tr("journal.transcription_use_cloud_confirm"),
            ):
                _set_transcription_model_choice(TRANSCRIPTION_MODEL_CLOUD)
                return True
            _open_transcription_downloads_manager("transcription")
            return False
        if not transcription_has_any_ready_model():
            _show_transcription_setup_prompt()
        else:
            _open_transcription_downloads_manager()
        return False

    def update_transcribe_ui() -> None:
        t = th()
        _update_iphone_receive_button()
        _refresh_transcription_model_selectors()
        root.after_idle(_sync_journal_side_action_columns)

        for row in (
            transcribe_btn_row,
            transcribe_file_btn_row,
            receive_iphone_btn_row,
            gen_report_btn_row,
        ):
            row.configure(bg=t.panel)

        def _set_model_buttons_enabled(enabled: bool) -> None:
            state = "normal" if enabled else "disabled"
            bg, fg, abg, afg = t.side_action_config()
            if not enabled:
                bg, fg, abg, afg, _dfg = t.transcribe_idle_disabled_config()
            for btn in (transcribe_model_btn, transcribe_file_model_btn):
                if transcription_model_selector_is_combo:
                    btn.config(state="readonly" if enabled else "disabled")
                else:
                    btn.config(
                        state=state,
                        bg=bg,
                        fg=fg,
                        activebackground=abg,
                        activeforeground=afg,
                    )

        def _set_transcribe_file_button_enabled(enabled: bool) -> None:
            if enabled:
                bg, fg, abg, afg = t.side_action_config()
                transcribe_file_btn.config(
                    state="normal",
                    width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                    bg=bg,
                    fg=fg,
                    activebackground=abg,
                    activeforeground=afg,
                )
            else:
                tb = t.transcribe_idle_disabled_config()
                transcribe_file_btn.config(
                    state="disabled",
                    width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                    bg=tb[0],
                    fg=tb[1],
                    activebackground=tb[2],
                    activeforeground=tb[3],
                    disabledforeground=tb[4],
                )

        if transcribing_busy["v"]:
            tb = t.transcribe_busy_config()
            transcribe_btn.config(
                state="disabled",
                width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                bg=tb[0],
                fg=tb[1],
                activebackground=tb[2],
                activeforeground=tb[3],
                disabledforeground=tb[4],
            )
            _set_transcribe_file_button_enabled(False)
            _set_model_buttons_enabled(False)
            return
        if recording_ui_busy["v"]:
            tb = t.transcribe_idle_disabled_config()
            transcribe_btn.config(
                state="disabled",
                width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                bg=tb[0],
                fg=tb[1],
                activebackground=tb[2],
                activeforeground=tb[3],
                disabledforeground=tb[4],
            )
            _set_transcribe_file_button_enabled(False)
            _set_model_buttons_enabled(False)
            return
        _set_model_buttons_enabled(not transcription_model_download_busy["v"])
        _set_transcribe_file_button_enabled(filedialog is not None)
        p = last_journal_wav.get("path")
        has_session = p is not None and isinstance(p, Path) and p.exists()
        has_archived = latest_archived_journal_wav() is not None
        if has_session or has_archived:
            bg, fg, abg, afg = t.side_action_config()
            transcribe_btn.config(
                state="normal",
                width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                bg=bg,
                fg=fg,
                activebackground=abg,
                activeforeground=afg,
            )
        else:
            tb = t.transcribe_idle_disabled_config()
            transcribe_btn.config(
                state="disabled",
                width=JOURNAL_SIDE_ACTION_BTN_WIDTH_CH - 3,
                bg=tb[0],
                fg=tb[1],
                activebackground=tb[2],
                activeforeground=tb[3],
                disabledforeground=tb[4],
            )

    def transcribe_tooltip_text() -> str:
        if transcribing_busy["v"]:
            pct = int(transcribing_progress.get("v", 0))
            return tr("journal.transcribe_tooltip_busy_full").format(pct=pct)
        if recording_ui_busy["v"]:
            return tr("journal.transcribe_tooltip_wait_recording")
        p = last_journal_wav.get("path")
        if p is not None and isinstance(p, Path) and p.exists():
            return tr("journal.transcribe_tooltip_prev_session")
        if latest_archived_journal_wav() is not None:
            return tr("journal.transcribe_tooltip_archived")
        return tr("journal.transcribe_tooltip_no_recording").format(dir=str(RECORDING_DIR))

    def transcribe_file_tooltip_text() -> str:
        if transcribing_busy["v"]:
            pct = int(transcribing_progress.get("v", 0))
            return tr("journal.transcribe_tooltip_busy_full").format(pct=pct)
        if recording_ui_busy["v"]:
            return tr("journal.transcribe_tooltip_wait_recording")
        return tr("journal.transcribe_file_tooltip")

    def receive_iphone_tooltip_text() -> str:
        if bool(iphone_receiver_state.get("active")):
            if bool(iphone_receiver_state.get("passive")):
                return tr("journal.iphone_passive_tooltip")
            return tr("journal.iphone_active_tooltip")
        return tr("journal.iphone_receive_tooltip")

    def begin_transcribe_path(use_path: Path, *, display_path: Optional[Path] = None) -> None:
        if not use_path.exists():
            messagebox.showinfo("Speech to text", f"That file is no longer on disk:\n{use_path}")
            update_transcribe_ui()
            return
        if not _ensure_selected_transcription_model_ready():
            return
        if display_path is not None:
            _set_stt_saved_path_display(f"Selected: {display_path}")
        transcribing_progress["v"] = 0
        transcribing_job_state["id"] += 1
        job_id = transcribing_job_state["id"]

        def schedule_progress(pct: int) -> None:
            if job_id != transcribing_job_state["id"]:
                return
            p = min(100, max(0, int(pct)))
            transcribing_progress["v"] = p

            def _ui() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                _publish_console_update(f"Transcribing... ({p}%)", key="transcribe:single:progress", log=False)

            root.after(0, _ui)

        transcribing_busy["v"] = True
        update_transcribe_ui()
        schedule_progress(0)
        _append_console_session_update(f"Transcription started: {use_path.name}", key=f"transcribe:start:{_iphone_path_key(use_path)}:{int(time.time())}")
        lang_snap = _language_code_for_whisper()
        model_choice_snap = normalize_transcription_model_choice(transcription_model_choice["value"])

        def work() -> None:
            result = ""
            try:
                result = transcribe_audio_with_model(
                    use_path,
                    lang_snap,
                    model_choice_snap,
                    temperature=0.0,
                    progress=schedule_progress,
                )
            except BaseException as _tw_exc:
                result = f"Whisper request failed: {_tw_exc}"
            finally:
                try:
                    schedule_progress(100)
                except Exception:
                    pass

            def done() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                transcribing_job_state["id"] += 1
                transcribing_busy["v"] = False
                transcribing_progress["v"] = 0
                update_transcribe_ui()
                stt_status.config(text="")
                _clear_console_hint()
                if _is_likely_api_error_message(result):
                    if _maybe_prompt_media_tools_from_error(result):
                        return
                    _append_console_session_update(
                        f"Transcription failed: {use_path.name}: {result[:240]}",
                        key=f"transcribe:failed:{_iphone_path_key(use_path)}:{int(time.time())}",
                    )
                    messagebox.showerror("Speech to text", result[:4000])
                    return
                final_text = normalize_journal_text_punctuation(result.strip())
                if final_text:
                    if stt_box.get("1.0", "end-1c").strip():
                        stt_box.insert("end", " ")
                    stt_box.insert("end", final_text)
                _append_console_session_update(
                    f"Transcription finished: {use_path.name}",
                    key=f"transcribe:done:{_iphone_path_key(use_path)}:{int(time.time())}",
                )
                _publish_console_update(
                    f"Transcription complete: {use_path.name}",
                    key=f"transcribe:complete:{_iphone_path_key(use_path)}",
                    temp=True,
                    log=False,
                )
                save_draft()
                refresh_save_entry_state()

            root.after(0, done)

        threading.Thread(target=work, daemon=True).start()

    def run_transcribe() -> None:
        if transcribing_busy["v"] or recording_ui_busy["v"]:
            return
        p = last_journal_wav.get("path")
        cleared_stale_cache = False
        if p is not None and isinstance(p, Path) and not p.exists():
            last_journal_wav["path"] = None
            cleared_stale_cache = True
            update_transcribe_ui()
            p = None
        if p is not None and isinstance(p, Path) and p.exists():
            use_path = p
        else:
            archived = latest_archived_journal_wav()
            if archived is None:
                messagebox.showinfo(
                    "Speech to text",
                    "No recording is available. Record audio first, or save a journal recording "
                    f"to your Recording folder:\n{RECORDING_DIR}",
                )
                return
            if cleared_stale_cache:
                use_path = archived
                last_journal_wav["path"] = archived
            elif not messagebox.askyesno(
                "Speech to text",
                "There is no recording from this session.\n\n"
                "Would you like to transcribe the most recent saved file in your Recording folder?\n\n"
                f"{archived.name}",
            ):
                return
            else:
                use_path = archived
                last_journal_wav["path"] = archived
        if not use_path.exists():
            last_journal_wav["path"] = None
            alt = latest_archived_journal_wav()
            if alt is None:
                messagebox.showinfo(
                    "Speech to text",
                    "That recording file is no longer on disk. There are no other saved "
                    f"recordings in:\n{RECORDING_DIR}",
                )
                update_transcribe_ui()
                return
            use_path = alt
            last_journal_wav["path"] = alt
        if not _ensure_selected_transcription_model_ready():
            return
        transcribing_progress["v"] = 0
        transcribing_job_state["id"] += 1
        job_id = transcribing_job_state["id"]

        def schedule_progress(pct: int) -> None:
            if job_id != transcribing_job_state["id"]:
                return
            p = min(100, max(0, int(pct)))
            transcribing_progress["v"] = p

            def _ui() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                _publish_console_update(f"Transcribing... ({p}%)", key="transcribe:recording:progress", log=False)

            root.after(0, _ui)

        transcribing_busy["v"] = True
        update_transcribe_ui()
        schedule_progress(0)
        _append_console_session_update(f"Transcription started: {use_path.name}", key=f"transcribe:start:{_iphone_path_key(use_path)}:{int(time.time())}")
        lang_snap = _language_code_for_whisper()
        model_choice_snap = normalize_transcription_model_choice(transcription_model_choice["value"])

        def work() -> None:
            result = ""
            try:
                result = transcribe_audio_with_model(
                    use_path,
                    lang_snap,
                    model_choice_snap,
                    temperature=0.0,
                    progress=schedule_progress,
                )
            except BaseException as _tw_exc:
                result = f"Whisper request failed: {_tw_exc}"
            finally:
                try:
                    schedule_progress(100)
                except Exception:
                    pass

            def done() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                transcribing_job_state["id"] += 1
                transcribing_busy["v"] = False
                transcribing_progress["v"] = 0
                update_transcribe_ui()
                stt_status.config(text="")
                _clear_console_hint()
                if _is_likely_api_error_message(result):
                    if _maybe_prompt_media_tools_from_error(result):
                        return
                    _append_console_session_update(
                        f"Transcription failed: {use_path.name}: {result[:240]}",
                        key=f"transcribe:failed:{_iphone_path_key(use_path)}:{int(time.time())}",
                    )
                    messagebox.showerror("Speech to text", result[:4000])
                    return
                final_text = normalize_journal_text_punctuation(result.strip())
                if final_text:
                    if stt_box.get("1.0", "end-1c").strip():
                        stt_box.insert("end", " ")
                    stt_box.insert("end", final_text)
                _append_console_session_update(
                    f"Transcription finished: {use_path.name}",
                    key=f"transcribe:done:{_iphone_path_key(use_path)}:{int(time.time())}",
                )
                _publish_console_update(
                    f"Transcription complete: {use_path.name}",
                    key=f"transcribe:complete:{_iphone_path_key(use_path)}",
                    temp=True,
                    log=False,
                )
                save_draft()
                refresh_save_entry_state()

            root.after(0, done)

        threading.Thread(target=work, daemon=True).start()

    def _transcription_file_sort_key(path: Path) -> Tuple[float, str, str]:
        try:
            mtime = float(path.stat().st_mtime)
        except OSError:
            mtime = float("inf")
        return (mtime, path.name.lower(), str(path).lower())

    def begin_transcribe_paths(
        use_paths: List[Path],
        *,
        display_label: Optional[str] = None,
        after_done: Optional[Callable[[bool, List[Path]], None]] = None,
    ) -> None:
        ordered_paths = sorted(use_paths, key=_transcription_file_sort_key)
        if not ordered_paths:
            return
        missing = [p for p in ordered_paths if not p.exists()]
        if missing:
            messagebox.showinfo("Speech to text", f"That file is no longer on disk:\n{missing[0]}")
            update_transcribe_ui()
            return
        pre_skipped_results: List[str] = []
        if _find_ffmpeg_executable() is None:
            ready_paths = [path for path in ordered_paths if not transcription_path_needs_media_tools(path)]
            media_needed_paths = [path for path in ordered_paths if transcription_path_needs_media_tools(path)]
            if media_needed_paths:
                if messagebox.askyesno(
                    tr("download_manager.title"),
                    tr("download_manager.media_required_prompt"),
                ):
                    _install_media_tools_addon()
                if not ready_paths:
                    return
                pre_skipped_results = [
                    f"{path.name}: Media Tools add-on is required for this media type. "
                    "Install Media Tools, then transcribe it again."
                    for path in media_needed_paths
                ]
                ordered_paths = ready_paths
        if any_transcription_path_needs_media_tools(ordered_paths) and _find_ffmpeg_executable() is None:
            if messagebox.askyesno(
                tr("download_manager.title"),
                tr("download_manager.media_required_prompt"),
            ):
                _install_media_tools_addon()
            return
        if not _ensure_selected_transcription_model_ready():
            return
        if display_label:
            _set_stt_saved_path_display(display_label)
        elif len(ordered_paths) == 1:
            _set_stt_saved_path_display(f"Selected: {ordered_paths[0]}")
        else:
            _set_stt_saved_path_display(f"Selected: {len(ordered_paths)} files")
        transcribing_progress["v"] = 0
        transcribing_job_state["id"] += 1
        job_id = transcribing_job_state["id"]

        def _publish_transcribe_path_status(path: Optional[Path], text: str) -> None:
            if path is None:
                return
            upload_id = iphone_upload_id_by_path.get(_iphone_path_key(path), "")
            if upload_id:
                publish_iphone_upload_status(upload_id, text)
            else:
                _publish_console_update(text, key=f"transcribe:path:{path.name}", log=False)

        def schedule_progress(
            file_index: int,
            file_count: int,
            pct: int,
            filename: str,
            path: Optional[Path] = None,
        ) -> None:
            if job_id != transcribing_job_state["id"]:
                return
            part_pct = min(100, max(0, int(pct)))
            total_pct = min(
                100,
                max(0, int(((file_index + (part_pct / 100.0)) / max(file_count, 1)) * 100)),
            )
            transcribing_progress["v"] = total_pct
            message = f"Transcribing {file_index + 1}/{file_count}: {filename} ({part_pct}%)"

            def _ui() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                _publish_transcribe_path_status(path, message)
                _publish_console_update(message, key=f"transcribe:progress:{filename}", log=False)

            root.after(0, _ui)

        def schedule_status(
            file_index: int,
            file_count: int,
            filename: str,
            message: str,
            path: Optional[Path] = None,
        ) -> None:
            if job_id != transcribing_job_state["id"]:
                return
            detail = (message or "").strip()
            if not detail:
                return
            text = f"Transcribing {file_index + 1}/{file_count}: {filename}. {detail}"

            def _ui() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                _publish_transcribe_path_status(path, text)
                _publish_console_update(text, key=f"transcribe:status:{filename}")

            root.after(0, _ui)

        pending_transcript_parts: List[str] = []
        pending_transcript_lock = threading.Lock()

        def _insert_transcribed_text_now(final_text: str) -> None:
            if stt_box.get("1.0", "end-1c").strip():
                stt_box.insert("end", " ")
            stt_box.insert("end", final_text)
            save_draft()
            refresh_save_entry_state()

        def drain_pending_transcript_parts() -> None:
            with pending_transcript_lock:
                parts = list(pending_transcript_parts)
                pending_transcript_parts.clear()
            for final_text in parts:
                try:
                    _insert_transcribed_text_now(final_text)
                except tk.TclError:
                    pass

        def append_transcribed_text(text: str) -> None:
            final_text = normalize_journal_text_punctuation((text or "").strip())
            if not final_text:
                return
            with pending_transcript_lock:
                pending_transcript_parts.append(final_text)

            def _ui() -> None:
                drain_pending_transcript_parts()

            root.after(0, _ui)

        def _is_skippable_file_transcribe_error(text: str) -> bool:
            t = (text or "").strip()
            if not t:
                return False
            prefixes = (
                "No audio track found",
                "Unsupported transcription file type",
                "That media file is too large",
                "That iPhone video is too large",
                "The selected media is still too large",
                "Could not prepare iPhone video",
                "Could not extract audio",
                "No speech detected",
                "Empty audio.",
            )
            return any(t.startswith(prefix) for prefix in prefixes)

        transcribing_busy["v"] = True
        update_transcribe_ui()
        schedule_progress(0, len(ordered_paths), 0, ordered_paths[0].name, ordered_paths[0])
        _append_console_session_update(
            f"Transcription queue started: {len(ordered_paths)} file(s).",
            key=f"transcribe:queue:start:{int(time.time())}",
        )
        lang_snap = _language_code_for_whisper()
        model_choice_snap = normalize_transcription_model_choice(transcription_model_choice["value"])

        def work() -> None:
            error_result = ""
            skipped_results: List[str] = list(pre_skipped_results)
            final_fallback_results: List[Tuple[Path, str]] = []
            try:
                for file_index, use_path in enumerate(ordered_paths):
                    def _file_progress(pct: int, _idx: int = file_index, _path: Path = use_path) -> None:
                        schedule_progress(_idx, len(ordered_paths), pct, _path.name, _path)

                    def _file_status(text: str, _idx: int = file_index, _path: Path = use_path) -> None:
                        schedule_status(_idx, len(ordered_paths), _path.name, text, _path)

                    live_parts_inserted = {"count": 0}

                    def _file_part(text: str) -> None:
                        if (text or "").strip():
                            live_parts_inserted["count"] += 1
                        append_transcribed_text(text)

                    result = transcribe_audio_with_model(
                        use_path,
                        lang_snap,
                        model_choice_snap,
                        temperature=0.0,
                        on_part=_file_part,
                        progress=_file_progress,
                        status=_file_status,
                    )
                    if _is_likely_api_error_message(result):
                        if "Media Tools add-on" in result:
                            error_result = result
                            break
                        if _is_skippable_file_transcribe_error(result):
                            skipped_results.append(f"{use_path.name}: {result}")
                            schedule_progress(file_index, len(ordered_paths), 100, use_path.name, use_path)
                            continue
                        error_result = result
                        break
                    if (
                        result.strip()
                        and live_parts_inserted["count"] <= 0
                        and not _is_likely_api_error_message(result)
                    ):
                        final_fallback_results.append((use_path, result.strip()))
                    schedule_progress(file_index, len(ordered_paths), 100, use_path.name, use_path)
            except BaseException as _tw_exc:
                error_result = f"Whisper request failed: {_tw_exc}"
            finally:
                try:
                    if not error_result and ordered_paths:
                        schedule_progress(
                            len(ordered_paths) - 1,
                            len(ordered_paths),
                            100,
                            ordered_paths[-1].name,
                            ordered_paths[-1],
                        )
                except Exception:
                    pass

            def done() -> None:
                if job_id != transcribing_job_state["id"]:
                    return
                transcribing_job_state["id"] += 1
                transcribing_busy["v"] = False
                transcribing_progress["v"] = 0
                update_transcribe_ui()
                stt_status.config(text="")
                _clear_console_hint()
                if error_result:
                    if after_done is not None:
                        after_done(False, ordered_paths)
                    root.after(100, drain_iphone_pending)
                    if _maybe_prompt_media_tools_from_error(error_result):
                        return
                    _append_console_session_update(
                        f"Transcription queue failed: {error_result[:240]}",
                        key=f"transcribe:queue:failed:{int(time.time())}",
                    )
                    messagebox.showerror("Speech to text", error_result[:4000])
                    return
                if final_fallback_results:
                    for _path, _text in final_fallback_results:
                        append_transcribed_text(_text)
                        _append_console_session_update(
                            f"Inserted final transcript for {_path.name}.",
                            key=f"transcribe:fallback:{_iphone_path_key(_path)}:{int(time.time())}",
                        )
                drain_pending_transcript_parts()
                if skipped_results:
                    preview = "\n".join(skipped_results[:12])
                    if len(skipped_results) > 12:
                        preview += f"\n...and {len(skipped_results) - 12} more."
                    _append_console_session_update(
                        f"Transcription skipped {len(skipped_results)} file(s).",
                        key=f"transcribe:queue:skipped:{int(time.time())}",
                    )
                    messagebox.showinfo(
                        "Speech to text",
                        f"Skipped {len(skipped_results)} file(s) that could not be transcribed:\n{preview}"[:4000],
                    )
                save_draft()
                refresh_save_entry_state()
                if after_done is not None:
                    after_done(True, ordered_paths)
                _append_console_session_update(
                    f"Transcription queue finished: {len(ordered_paths)} file(s).",
                    key=f"transcribe:queue:done:{int(time.time())}",
                )
                _publish_console_update(
                    f"Transcription complete: {len(ordered_paths)} file(s).",
                    key=f"transcribe:queue:complete:{int(time.time())}",
                    temp=True,
                    log=False,
                )
                root.after(100, drain_iphone_pending)

            root.after(0, done)

        threading.Thread(target=work, daemon=True).start()

    def run_transcribe_file() -> None:
        if transcribing_busy["v"] or recording_ui_busy["v"]:
            return
        if filedialog is None:
            messagebox.showerror("Speech to text", "File selection is not available.")
            return
        try:
            RECORDING_DIR.mkdir(parents=True, exist_ok=True)
            initial_dir = str(RECORDING_DIR)
        except OSError:
            initial_dir = str(DATA_DIR)
        selected = filedialog.askopenfilenames(
            title="Choose files to transcribe",
            initialdir=initial_dir,
            filetypes=[
                (
                    "Transcribable media",
                    "*.wav *.mp3 *.m4a *.aac *.caf *.mp4 *.mov *.webm *.mpeg *.mpga *.flac *.ogg",
                ),
                ("iPhone videos", "*.mov *.mp4"),
                ("iPhone audio / Voice Memos", "*.m4a *.aac *.caf"),
                ("Audio files", "*.wav *.mp3 *.m4a *.aac *.caf *.mpeg *.mpga *.flac *.ogg"),
                ("Video files", "*.mp4 *.mov *.webm"),
                ("All files", "*.*"),
            ],
        )
        if not selected:
            return
        begin_transcribe_paths([Path(p) for p in selected])

    transcribe_btn.config(command=run_transcribe)
    transcribe_file_btn.config(command=run_transcribe_file)
    if transcription_model_selector_is_combo:
        transcribe_model_btn.bind(
            "<<ComboboxSelected>>",
            lambda _evt: _on_transcription_model_selector_change(transcribe_model_var),
        )
        transcribe_file_model_btn.bind(
            "<<ComboboxSelected>>",
            lambda _evt: _on_transcription_model_selector_change(transcribe_file_model_var),
        )
    else:
        transcribe_model_btn.config(
            command=lambda: _show_transcription_model_menu(transcribe_model_btn)
        )
        transcribe_file_model_btn.config(
            command=lambda: _show_transcription_model_menu(transcribe_file_model_btn)
        )
    receive_iphone_btn.config(command=toggle_iphone_receiver)
    bind_hover_tooltip(transcribe_btn, transcribe_tooltip_text)
    bind_hover_tooltip(transcribe_file_btn, transcribe_file_tooltip_text)
    bind_hover_tooltip(
        transcribe_model_btn,
        lambda: tr("journal.transcription_model_tooltip").format(
            model=_transcription_model_display(transcription_model_choice["value"])
        ),
    )
    bind_hover_tooltip(
        transcribe_file_model_btn,
        lambda: tr("journal.transcription_model_tooltip").format(
            model=_transcription_model_display(transcription_model_choice["value"])
        ),
    )
    bind_hover_tooltip(receive_iphone_btn, receive_iphone_tooltip_text)

    def transcribe_rest_style() -> Tuple[str, str, str, str, str]:
        t = th()
        if transcribing_busy["v"]:
            b0, b1, b2, b3, b4 = t.transcribe_busy_config()
            return ("disabled", b0, b1, b2, b3)
        if recording_ui_busy["v"]:
            b0, b1, b2, b3, b4 = t.transcribe_idle_disabled_config()
            return ("disabled", b0, b1, b2, b3)
        p = last_journal_wav.get("path")
        if (p is not None and isinstance(p, Path) and p.exists()) or (
            latest_archived_journal_wav() is not None
        ):
            return t.side_action_bind_rest()
        b0, b1, b2, b3, b4 = t.transcribe_idle_disabled_config()
        return ("disabled", b0, b1, b2, b3)

    def transcribe_file_rest_style() -> Tuple[str, str, str, str, str]:
        t = th()
        if transcribing_busy["v"] or recording_ui_busy["v"] or filedialog is None:
            b0, b1, b2, b3, b4 = t.transcribe_idle_disabled_config()
            return ("disabled", b0, b1, b2, b3)
        return t.side_action_bind_rest()

    def transcribe_model_rest_style() -> Tuple[str, str, str, str, str]:
        t = th()
        if transcribing_busy["v"] or recording_ui_busy["v"] or transcription_model_download_busy["v"]:
            b0, b1, b2, b3, b4 = t.transcribe_idle_disabled_config()
            return ("disabled", b0, b1, b2, b3)
        return t.side_action_bind_rest()

    def receive_iphone_rest_style() -> Tuple[str, str, str, str, str]:
        return th().side_action_bind_rest()

    bind_button_hover_if_enabled(
        transcribe_btn,
        transcribe_rest_style,
        lambda: th().hover_primary,
        lambda: "white",
    )
    bind_button_hover_if_enabled(
        transcribe_file_btn,
        transcribe_file_rest_style,
        lambda: th().hover_primary,
        lambda: "white",
    )
    if not transcription_model_selector_is_combo:
        for _model_btn in (transcribe_model_btn, transcribe_file_model_btn):
            bind_button_hover_if_enabled(
                _model_btn,
                transcribe_model_rest_style,
                lambda: th().hover_primary,
                lambda: "white",
            )
    bind_button_hover_if_enabled(
        receive_iphone_btn,
        receive_iphone_rest_style,
        lambda: th().hover_primary,
        lambda: "white",
    )
    update_transcribe_ui()

    def _record_source_label(mode: str) -> str:
        normalized = normalize_record_source_mode(mode)
        if normalized == RECORD_SOURCE_MIC:
            return tr("journal.rec.source.mic")
        if normalized == RECORD_SOURCE_COMPUTER:
            return tr("journal.rec.source.computer")
        return tr("journal.rec.source.both")

    def _record_source_mode_from_label(label: str) -> str:
        text = (label or "").strip()
        for mode in RECORD_SOURCE_CHOICES:
            if text == _record_source_label(mode):
                return mode
        return normalize_record_source_mode(text)

    def _record_mode_from_enabled_events() -> str:
        mic_on = record_mic_enabled.is_set()
        computer_on = record_computer_enabled.is_set()
        if mic_on and computer_on:
            return RECORD_SOURCE_BOTH
        if computer_on:
            return RECORD_SOURCE_COMPUTER
        return RECORD_SOURCE_MIC

    def _record_source_status_text() -> str:
        mode = _record_mode_from_enabled_events() if recording_ui_busy["v"] else record_source_mode["value"]
        return tr("journal.status.recording_sources").format(sources=_record_source_label(mode))

    def _refresh_record_source_selector() -> None:
        mode = _record_mode_from_enabled_events() if recording_ui_busy["v"] else record_source_mode["value"]
        try:
            record_source_var.set(_record_source_label(mode))
            record_source_combo.config(text="▼")
        except Exception:
            pass

    def _apply_record_source_mode(mode: str, *, persist: bool = True, update_status: bool = True) -> None:
        normalized = normalize_record_source_mode(mode)
        if normalized in (RECORD_SOURCE_BOTH, RECORD_SOURCE_MIC):
            record_mic_enabled.set()
        else:
            record_mic_enabled.clear()
        if normalized in (RECORD_SOURCE_BOTH, RECORD_SOURCE_COMPUTER):
            record_computer_enabled.set()
        else:
            record_computer_enabled.clear()
        if not record_mic_enabled.is_set() and not record_computer_enabled.is_set():
            record_mic_enabled.set()
            normalized = RECORD_SOURCE_MIC
            try:
                messagebox.showinfo("Speech to text", tr("journal.rec.source_guard"))
            except Exception:
                pass
        record_source_mode["value"] = normalized
        if persist:
            save_selected_record_source_mode(normalized)
        _refresh_record_source_selector()
        if update_status and recording_ui_busy["v"] and not record_pause.is_set():
            try:
                stt_status.config(text=_record_source_status_text())
            except tk.TclError:
                pass

    def _on_record_source_selected(_evt: Optional[Any] = None) -> None:
        _apply_record_source_mode(
            _record_source_mode_from_label(record_source_var.get()),
            persist=True,
            update_status=True,
        )

    def _show_record_source_menu(anchor: Any) -> None:
        menu = tk.Menu(anchor, tearoff=0)
        current = _record_mode_from_enabled_events() if recording_ui_busy["v"] else record_source_mode["value"]
        for mode in RECORD_SOURCE_CHOICES:
            prefix = "[x] " if normalize_record_source_mode(mode) == normalize_record_source_mode(current) else "    "
            menu.add_command(
                label=f"{prefix}{_record_source_label(mode)}",
                command=lambda m=mode: _apply_record_source_mode(
                    m,
                    persist=True,
                    update_status=True,
                ),
            )
        try:
            menu.tk_popup(anchor.winfo_rootx(), anchor.winfo_rooty() + anchor.winfo_height())
        finally:
            try:
                menu.grab_release()
            except tk.TclError:
                pass

    def _on_record_source_error_from_worker(source: str, detail: str) -> None:
        def _ui() -> None:
            if source == RECORD_SOURCE_COMPUTER:
                msg = tr("journal.status.computer_audio_unavailable")
            elif source == RECORD_SOURCE_MIC:
                msg = tr("journal.status.mic_unavailable")
            else:
                msg = tr("journal.status.source_unavailable").format(source=source)
            _refresh_record_source_selector()
            _publish_console_update(
                f"{msg} {detail}".strip(),
                key=f"record:source:{source}:failed",
            )
            if recording_ui_busy["v"] and not recording_background_mode["v"]:
                try:
                    stt_status.config(text=msg)
                except tk.TclError:
                    pass

        try:
            root.after(0, _ui)
        except tk.TclError:
            pass

    def _startup_iphone_receiver_and_pending() -> None:
        if iphone_passive_receive_enabled():
            start_iphone_receiver(show_setup=False, passive=True)
        receive_iphone_incoming_files(list_incoming_iphone_files())
        enqueue_iphone_imports(list_pending_iphone_inbox_files())

    root.after(300, _startup_iphone_receiver_and_pending)
    root.after(1500, lambda: _check_updates_async(manual=False))
    wave_canvas.bind("<Configure>", lambda _e: redraw_waveform_canvas())
    wave_canvas.after(80, redraw_waveform_canvas)

    def _editor_has_meaningful_body_content() -> bool:
        return bool(
            text_box.get("1.0", "end-1c").strip()
            or stt_box.get("1.0", "end-1c").strip()
            or report_box.get("1.0", "end-1c").strip()
        )

    def _draft_file_has_restorable_content(d: Dict[str, object]) -> bool:
        if str(d.get("text", "") or "").strip():
            return True
        if str(d.get("speech_transcript", "") or "").strip():
            return True
        if str(d.get("ai_report", "") or "").strip():
            return True
        wraw = d.get("journal_recording_wav")
        if not wraw:
            return False
        try:
            return Path(str(wraw)).is_file()
        except OSError:
            return False

    def apply_draft_dict_to_ui(d: Dict[str, object]) -> None:
        nonlocal edit_target_sheet, edit_target_row
        _txt = normalize_journal_text_punctuation(str(d.get("text", "") or ""))
        _sp = normalize_journal_text_punctuation(str(d.get("speech_transcript", "") or ""))
        _rp = normalize_journal_text_punctuation(str(d.get("ai_report", "") or ""))
        text_box.delete("1.0", "end")
        text_box.insert("1.0", _txt)
        stt_box.delete("1.0", "end")
        stt_box.insert("1.0", _sp)
        report_box.delete("1.0", "end")
        report_box.insert("1.0", _rp)
        _dt = str(d.get("date", "") or "").strip()
        if _dt:
            if DateEntry is not None and isinstance(date_entry, DateEntry):  # type: ignore[arg-type]
                try:
                    date_entry.set_date(_dt)  # type: ignore[attr-defined]
                except Exception:
                    try:
                        date_entry.delete(0, "end")  # type: ignore[attr-defined]
                    except Exception:
                        pass
                    date_entry.insert(0, _dt)  # type: ignore[attr-defined]
            else:
                date_entry.delete(0, "end")
                date_entry.insert(0, _dt)
        _tm = str(d.get("time", "") or "").strip()
        if _tm:
            time_entry.delete(0, "end")
            time_entry.insert(0, _tm)
        edit_target_sheet = str(d.get("edit_target_sheet", "") or "")
        try:
            edit_target_row = int(d.get("edit_target_row", 0) or 0)
        except (TypeError, ValueError):
            edit_target_row = 0
        is_edit_mode["v"] = bool(edit_target_sheet and edit_target_row > 0)
        _wav = d.get("journal_recording_wav")
        last_journal_wav["path"] = None
        if _wav:
            try:
                _wp = Path(str(_wav))
                if _wp.is_file():
                    last_journal_wav["path"] = _wp.resolve()
                    _set_stt_saved_path_display(tr("journal.saved_path", path=str(_wp)))
                else:
                    _set_stt_saved_path_display("")
            except OSError:
                _set_stt_saved_path_display("")
        else:
            _set_stt_saved_path_display("")
        stt_status.config(text="")
        report_status.config(text="")
        update_transcribe_ui()
        refresh_save_entry_state()
        try:
            text_box.focus_set()
        except tk.TclError:
            pass

    def on_restore_draft_click() -> None:
        d = load_journal_window_draft()
        if not isinstance(d, dict) or not _draft_file_has_restorable_content(d):
            messagebox.showinfo(
                tr("msg.journal_window"),
                tr("msg.no_draft_to_restore"),
            )
            return
        if _editor_has_meaningful_body_content():
            if not messagebox.askyesno(
                tr("journal.restore_confirm_title"),
                tr("journal.restore_confirm"),
            ):
                return
        apply_draft_dict_to_ui(d)
        save_draft()

    restore_draft_btn = tk.Button(
        top,
        text=tr("journal.restore_draft"),
        command=on_restore_draft_click,
        bg=_ut_bg,
        fg=_ut_fg,
        activebackground=_ut_abg,
        activeforeground=_ut_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    restore_draft_btn.grid(row=0, column=7, sticky="e", padx=(8, 0), pady=12)
    bind_button_hover_if_enabled(
        restore_draft_btn,
        lambda: th().toolbar_bind_rest(),
        lambda: th().toolbar_hover()[0],
        lambda: th().toolbar_hover()[1],
    )
    bind_hover_tooltip(restore_draft_btn, lambda: tr("tip.restore_draft"))

    def _language_code_for_whisper() -> Optional[str]:
        choice = lang_var.get().strip()
        if choice == "English":
            return "en"
        if choice in ("\u7b80\u4f53\u4e2d\u6587", "\u4e2d\u6587", "Chinese"):
            return "zh"
        return None

    def _is_likely_api_error_message(text: str) -> bool:
        t = text.strip()
        if not t:
            return False
        prefixes = (
            "OPENAI_API_KEY",
            "ChatGPT API error",
            "Failed to contact ChatGPT",
            "ChatGPT returned",
            "No response received",
            "Whisper API error",
            "Whisper request failed",
            "Whisper returned",
        "Whisper transcription rejected",
        "Local transcription failed",
        "Local transcription add-on",
        "Local transcription model",
        "Could not load local transcription add-on",
        "Could not load local transcription model",
            "Media Tools add-on",
            "Unsupported transcription file type",
            "That media file is too large",
            "That iPhone video is too large",
            "The selected media is still too large",
            "Could not prepare iPhone video",
            "Could not extract audio",
            "Could not read audio file",
            "No audio track found",
            "Recording needs optional packages",
            "No speech detected",
            "Empty audio.",
        )
        return any(t.startswith(p) for p in prefixes)

    def _journal_rec_btn_set(btn: Any, enabled: bool) -> None:
        t = th()
        if enabled:
            bg, fg, abg, afg = t.side_action_config()
            btn.config(
                state="normal",
                bg=bg,
                fg=fg,
                activebackground=abg,
                activeforeground=afg,
                cursor="hand2",
            )
        else:
            btn.config(
                state="disabled",
                bg=t.btn_disabled,
                fg=t.disabled_fg,
                disabledforeground=t.disabled_fg,
                cursor="arrow",
            )

    def _journal_rec_btn_set_disabled_look_clickable(btn: Any) -> None:
        t = th()
        btn.config(
            state="normal",
            bg=t.btn_disabled,
            fg=t.disabled_fg,
            activebackground=t.btn_disabled,
            activeforeground=t.disabled_fg,
            disabledforeground=t.disabled_fg,
            cursor="hand2",
        )

    def _archive_finished_recording(wav_path: Path) -> Optional[Path]:
        dest = archive_journal_recording(wav_path)
        try:
            wav_path.unlink(missing_ok=True)
        except OSError:
            pass
        return dest

    def _reset_recording_state_after_close() -> None:
        record_thread_holder["thread"] = None
        record_pause.clear()
        record_close_requested["v"] = False
        recording_background_mode["v"] = False
        recording_ui_busy["v"] = False

    def on_record_worker_finished(err: Optional[str], wav_path: Optional[Path]) -> None:
        was_background = bool(recording_background_mode["v"])
        record_thread_holder["thread"] = None
        record_pause.clear()
        cancel_wave_tick()
        if was_background:
            recording_ui_busy["v"] = False
            recording_background_mode["v"] = False
            record_close_requested["v"] = False
            _journal_rec_btn_set(start_rec_button, True)
            _journal_rec_btn_set(pause_rec_button, False)
            _journal_rec_btn_set(stop_rec_button, False)
            pause_rec_button.config(text=tr("journal.rec.pause"))
            if err:
                last_journal_wav["path"] = None
                if wav_path is not None:
                    try:
                        wav_path.unlink(missing_ok=True)
                    except OSError:
                        pass
                update_transcribe_ui()
                try:
                    console_append(f"Background recording failed: {err}")
                except Exception:
                    pass
                return
            if wav_path is None or not wav_path.exists():
                last_journal_wav["path"] = None
                update_transcribe_ui()
                try:
                    console_append("Background recording stopped without a saved file.")
                except Exception:
                    pass
                return
            dest = _archive_finished_recording(wav_path)
            if dest is not None:
                last_journal_wav["path"] = dest
                try:
                    console_append(f"Background recording saved: {dest}")
                except Exception:
                    pass
            else:
                last_journal_wav["path"] = None
                try:
                    console_append(f"Background recording finished but could not copy to {RECORDING_DIR}.")
                except Exception:
                    pass
            update_transcribe_ui()
            return
        if err:
            recording_ui_busy["v"] = False
            last_journal_wav["path"] = None
            update_transcribe_ui()
            stt_status.config(text="")
            _set_stt_saved_path_display("")
            _journal_rec_btn_set(start_rec_button, True)
            _journal_rec_btn_set(pause_rec_button, False)
            _journal_rec_btn_set(stop_rec_button, False)
            pause_rec_button.config(text=tr("journal.rec.pause"))
            if ttk is not None:
                lang_combo.config(state="readonly")
            else:
                lang_combo.config(state="normal")
            if wav_path is not None:
                try:
                    wav_path.unlink(missing_ok=True)
                except OSError:
                    pass
            reset_waveform_session()
            messagebox.showerror("Speech to text", err[:4000])
            return
        if wav_path is None or not wav_path.exists():
            recording_ui_busy["v"] = False
            last_journal_wav["path"] = None
            update_transcribe_ui()
            stt_status.config(text="")
            _set_stt_saved_path_display("")
            _journal_rec_btn_set(start_rec_button, True)
            _journal_rec_btn_set(pause_rec_button, False)
            _journal_rec_btn_set(stop_rec_button, False)
            pause_rec_button.config(text=tr("journal.rec.pause"))
            if ttk is not None:
                lang_combo.config(state="readonly")
            else:
                lang_combo.config(state="normal")
            reset_waveform_session()
            return
        dest = _archive_finished_recording(wav_path)
        recording_ui_busy["v"] = False
        _journal_rec_btn_set(start_rec_button, True)
        _journal_rec_btn_set(pause_rec_button, False)
        _journal_rec_btn_set(stop_rec_button, False)
        pause_rec_button.config(text=tr("journal.rec.pause"))
        if ttk is not None:
            lang_combo.config(state="readonly")
        else:
            lang_combo.config(state="normal")
        reset_waveform_session()
        if dest is not None:
            last_journal_wav["path"] = dest
            stt_status.config(text="")
            _set_stt_saved_path_display(tr("journal.saved_path", path=str(dest)))
        else:
            last_journal_wav["path"] = None
            _set_stt_saved_path_display("")
            stt_status.config(
                text=f"Recording finished (could not copy to {RECORDING_DIR}).",
            )
        update_transcribe_ui()

    def record_worker_main(wav_path: Path) -> None:
        err = record_sources_session_wav(
            wav_path,
            record_stop,
            source_enabled_events={
                RECORD_SOURCE_MIC: record_mic_enabled,
                RECORD_SOURCE_COMPUTER: record_computer_enabled,
            },
            chunk_interval_sec=LIVE_STT_CHUNK_INTERVAL_SEC,
            on_audio_chunk=None,
            on_pcm_block=(None if recording_background_mode["v"] else on_pcm_block_journal),
            pause_event=record_pause,
            on_source_error=_on_record_source_error_from_worker,
        )
        if record_close_requested["v"]:
            if err is None and wav_path.exists():
                _archive_finished_recording(wav_path)
            else:
                try:
                    wav_path.unlink(missing_ok=True)
                except OSError:
                    pass
            _reset_recording_state_after_close()
            return
        root.after(0, lambda: on_record_worker_finished(err, wav_path))

    def start_recording(*, background: bool = False) -> bool:
        if recording_ui_busy["v"]:
            return False
        _apply_record_source_mode(record_source_mode["value"], persist=False, update_status=False)
        fd, tmp_name = tempfile.mkstemp(suffix=".wav", prefix="journal_recording_")
        os.close(fd)
        tmp = Path(tmp_name)
        record_path_holder["path"] = tmp
        last_journal_wav["path"] = None
        record_stop.clear()
        record_pause.clear()
        record_close_requested["v"] = False
        recording_background_mode["v"] = bool(background)
        recording_ui_busy["v"] = True
        if background:
            _journal_rec_btn_set_disabled_look_clickable(stop_rec_button)
            th = threading.Thread(target=record_worker_main, args=(tmp,), daemon=True)
            record_thread_holder["thread"] = th
            th.start()
            return True
        update_transcribe_ui()
        _set_stt_saved_path_display("")
        reset_waveform_session()
        start_wave_tick()
        _journal_rec_btn_set(start_rec_button, False)
        _journal_rec_btn_set(pause_rec_button, True)
        _journal_rec_btn_set(stop_rec_button, True)
        pause_rec_button.config(text=tr("journal.rec.pause"))
        if ttk is not None:
            lang_combo.config(state="disabled")
        else:
            lang_combo.config(state="disabled")
        stt_status.config(text=_record_source_status_text())
        th = threading.Thread(target=record_worker_main, args=(tmp,), daemon=True)
        record_thread_holder["thread"] = th
        th.start()
        return True

    def stop_recording() -> bool:
        th = record_thread_holder["thread"]
        if not (
            recording_ui_busy["v"]
            and isinstance(th, threading.Thread)
            and th.is_alive()
        ):
            return False
        if not recording_background_mode["v"]:
            stt_status.config(text=tr("journal.status.stopping"))
        record_stop.set()
        record_pause.clear()
        if recording_background_mode["v"]:
            _journal_rec_btn_set(stop_rec_button, False)
        else:
            _journal_rec_btn_set(start_rec_button, False)
            _journal_rec_btn_set(pause_rec_button, False)
            _journal_rec_btn_set(stop_rec_button, False)
        return True

    def toggle_pause_recording() -> None:
        th = record_thread_holder["thread"]
        if not (
            recording_ui_busy["v"]
            and isinstance(th, threading.Thread)
            and th.is_alive()
        ):
            return
        if record_pause.is_set():
            record_pause.clear()
            pause_rec_button.config(text=tr("journal.rec.pause"))
            stt_status.config(text=_record_source_status_text())
        else:
            record_pause.set()
            pause_rec_button.config(text=tr("journal.rec.resume"))
            stt_status.config(text=tr("journal.status.paused"))

    _sr_bg, _sr_fg, _sr_abg, _sr_afg = t_init.side_action_config()
    record_source_var = tk.StringVar(value=_record_source_label(record_source_mode["value"]))
    start_rec_button = tk.Button(
        stt_top,
        text="Start recording",
        command=start_recording,
        bg=_sr_bg,
        fg=_sr_fg,
        activebackground=_sr_abg,
        activeforeground=_sr_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=10,
        pady=6,
        cursor="hand2",
    )
    record_source_combo = tk.Button(
        stt_top,
        text="▼",
        command=lambda: _show_record_source_menu(record_source_combo),
        width=3,
        bg=_sr_bg,
        fg=_sr_fg,
        activebackground=_sr_abg,
        activeforeground=_sr_afg,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=4,
        pady=6,
        cursor="hand2",
    )
    pause_rec_button = tk.Button(
        stt_top,
        text="Pause recording",
        command=toggle_pause_recording,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
    )
    stop_rec_button = tk.Button(
        stt_top,
        text="Stop recording",
        command=stop_recording,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=8,
        pady=6,
    )
    _journal_rec_btn_set(pause_rec_button, False)
    _journal_rec_btn_set(stop_rec_button, False)
    start_rec_button.grid(row=0, column=0, sticky="w", padx=(12, 4), pady=8)
    record_source_combo.grid(row=0, column=1, sticky="w", padx=(0, 4), pady=8)
    pause_rec_button.grid(row=0, column=2, sticky="w", padx=(0, 4), pady=8)
    stop_rec_button.grid(row=0, column=3, sticky="w", padx=(0, 8), pady=8)

    def rec_primary_rest(btn: Any) -> Tuple[str, str, str, str, str]:
        t = th()
        if btn is stop_rec_button and recording_background_mode["v"]:
            return ("disabled", t.btn_disabled, t.disabled_fg, t.btn_disabled, t.disabled_fg)
        if str(btn.cget("state")) == "normal":
            return t.side_action_bind_rest()
        return t.side_action_disabled()

    bind_button_hover_if_enabled(
        start_rec_button,
        lambda b=start_rec_button: rec_primary_rest(b),
        lambda: th().hover_primary,
        lambda: "white",
    )
    bind_hover_tooltip(
        record_source_combo,
        lambda: _record_source_label(
            _record_mode_from_enabled_events()
            if recording_ui_busy["v"]
            else record_source_mode["value"]
        ),
    )
    bind_button_hover_if_enabled(
        record_source_combo,
        lambda b=record_source_combo: rec_primary_rest(b),
        lambda: th().hover_primary,
        lambda: "white",
    )
    bind_button_hover_if_enabled(
        pause_rec_button,
        lambda b=pause_rec_button: rec_primary_rest(b),
        lambda: th().hover_primary,
        lambda: "white",
    )
    bind_button_hover_if_enabled(
        stop_rec_button,
        lambda b=stop_rec_button: rec_primary_rest(b),
        lambda: th().hover_primary,
        lambda: "white",
    )
    stt_status.grid(row=0, column=4, sticky="ew", padx=(4, 12), pady=8)
    stt_lang_lbl = tk.Label(
        stt_top,
        text="Language:",
        bg=t_init.panel,
        fg=t_init.muted,
        font=("Segoe UI", 9),
    )
    stt_lang_lbl.grid(row=0, column=5, sticky="w", padx=(4, 0), pady=8)
    lang_combo.grid(row=0, column=6, sticky="ew", padx=(8, 12), pady=8)

    report_progress_state: Dict[str, object] = {
        "busy": False,
        "started_at": 0.0,
        "message": "",
        "job_id": 0,
    }

    def _normalize_text_widget_punctuation(widget: tk.Text) -> None:
        try:
            raw = widget.get("1.0", "end-1c")
            fixed = normalize_journal_text_punctuation(raw)
            if fixed == raw:
                return
            insert_index = widget.index("insert")
            widget.delete("1.0", "end")
            widget.insert("1.0", fixed)
            try:
                widget.mark_set("insert", insert_index)
                widget.see("insert")
            except tk.TclError:
                pass
        except tk.TclError:
            pass

    def _refresh_report_progress_status(job_id: int) -> None:
        if int(report_progress_state.get("job_id", 0) or 0) != job_id:
            return
        if not bool(report_progress_state.get("busy", False)):
            return
        started_at = float(report_progress_state.get("started_at", 0.0) or 0.0)
        elapsed = max(0, int(time.time() - started_at)) if started_at else 0
        message = str(report_progress_state.get("message", "") or "Generating report")
        try:
            report_status.config(text=f"{message} ({elapsed}s)")
        except tk.TclError:
            pass

    def _tick_report_progress(job_id: int) -> None:
        _refresh_report_progress_status(job_id)
        if (
            int(report_progress_state.get("job_id", 0) or 0) == job_id
            and bool(report_progress_state.get("busy", False))
        ):
            root.after(1000, lambda: _tick_report_progress(job_id))

    def _set_report_progress(message: str, job_id: int) -> None:
        clean_message = (message or "").strip() or "Generating report"

        def _ui() -> None:
            if int(report_progress_state.get("job_id", 0) or 0) != job_id:
                return
            if not bool(report_progress_state.get("busy", False)):
                return
            report_progress_state["message"] = clean_message
            _refresh_report_progress_status(job_id)

        root.after(0, _ui)

    def run_generate_report() -> None:
        if not get_openai_api_key():
            messagebox.showerror(
                "Journal Window",
                "No OpenAI API key. Use TOKEN ADD in the main menu or set OPENAI_API_KEY.",
            )
            return
        t = th()
        gen_button.config(
            state="disabled",
            bg=t.btn_disabled,
            fg=t.disabled_fg,
            disabledforeground=t.disabled_fg,
            cursor="arrow",
        )
        _normalize_text_widget_punctuation(text_box)
        _normalize_text_widget_punctuation(stt_box)
        journal_snapshot = text_box.get("1.0", "end-1c")
        speech_snapshot = stt_box.get("1.0", "end-1c")
        job_id = int(report_progress_state.get("job_id", 0) or 0) + 1
        report_progress_state.update(
            {
                "busy": True,
                "started_at": time.time(),
                "message": "Preparing report",
                "job_id": job_id,
            }
        )
        _refresh_report_progress_status(job_id)
        root.after(1000, lambda: _tick_report_progress(job_id))

        def work() -> None:
            body = generate_journal_report_from_sources(
                journal_snapshot,
                speech_snapshot,
                progress=lambda msg: _set_report_progress(msg, job_id),
            )
            root.after(0, lambda b=body, jid=job_id: on_generate_report_done(b, jid))

        threading.Thread(target=work, daemon=True).start()

    def on_generate_report_done(body: str, job_id: int) -> None:
        t = th()
        bg, fg, abg, afg = t.side_action_config()
        gen_button.config(
            state="normal",
            bg=bg,
            fg=fg,
            activebackground=abg,
            activeforeground=afg,
            cursor="hand2",
        )
        if int(report_progress_state.get("job_id", 0) or 0) == job_id:
            report_progress_state["busy"] = False
            report_status.config(text="")
        if _is_likely_api_error_message(body):
            messagebox.showerror("AI report", body[:4000])
            return
        body = normalize_journal_text_punctuation(body)
        report_box.delete("1.0", "end")
        report_box.insert("1.0", body.strip())
        save_draft()
        refresh_save_entry_state()

    gen_button.config(command=run_generate_report)

    def generate_report_tooltip_text() -> str:
        return tr("tip.generate_report")

    bind_hover_tooltip(gen_button, generate_report_tooltip_text)

    saved = {"value": False}
    autosave_id = {"value": None}

    def build_draft_dict() -> Dict[str, object]:
        return {
            "text": normalize_journal_text_punctuation(text_box.get("1.0", "end-1c")),
            "speech_transcript": normalize_journal_text_punctuation(
                stt_box.get("1.0", "end-1c")
            ),
            "ai_report": normalize_journal_text_punctuation(report_box.get("1.0", "end-1c")),
            "date": date_entry.get().strip(),
            "time": time_entry.get().strip(),
            "edit_target_sheet": edit_target_sheet,
            "edit_target_row": edit_target_row,
            "updated_at": datetime.now().isoformat(),
        }

    def save_draft() -> None:
        save_journal_window_draft(build_draft_dict())

    def autosave() -> None:
        save_draft()
        autosave_id["value"] = root.after(1500, autosave)

    def do_save() -> None:
        if not (
            text_box.get("1.0", "end-1c").strip()
            or stt_box.get("1.0", "end-1c").strip()
            or report_box.get("1.0", "end-1c").strip()
        ):
            return
        raw_date = date_entry.get().strip()
        parsed_date = parse_flexible_date(raw_date, now.year)
        if parsed_date is None:
            messagebox.showerror("Journal Window", "Invalid date. Example: 04/20/2026 or Apr 20")
            return
        date_value = parsed_date.strftime("%m/%d/%Y")
        normalized_time = normalize_window_time_input(time_entry.get().strip())
        if normalized_time is None:
            messagebox.showerror("Journal Window", "Invalid time. Example: 2:03PM")
            return
        _normalize_text_widget_punctuation(text_box)
        _normalize_text_widget_punctuation(stt_box)
        _normalize_text_widget_punctuation(report_box)
        text_value = normalize_journal_text_punctuation(text_box.get("1.0", "end-1c")).strip()
        if not text_value and is_edit_mode["v"]:
            should_delete = messagebox.askyesno(
                "Clear Entry",
                "Text is empty. Saving now will delete the previous entry. Are you sure?",
            )
            if not should_delete:
                return
            deleted_ok = delete_journal_entry_at(edit_target_sheet, edit_target_row)
            if not deleted_ok:
                messagebox.showerror(
                    "Journal Window",
                    "Could not delete previous entry. It may have changed. Try again.",
                )
                return
            clear_journal_window_draft()
            saved["value"] = True
            if autosave_id["value"] is not None:
                root.after_cancel(autosave_id["value"])
            _finish_or_background_journal_window()
            return
        if not text_value:
            text_value = "(no details entered)"
        speech_value = normalize_journal_text_punctuation(stt_box.get("1.0", "end-1c")).strip()
        report_value = normalize_journal_text_punctuation(report_box.get("1.0", "end-1c")).strip()
        _feed_desktop_pet_on_save(text_value, speech_value, report_value)
        row_payload = [date_value, normalized_time, text_value, speech_value, report_value]
        if edit_target_sheet and edit_target_row > 0:
            saved_ok = update_journal_entry_at(
                edit_target_sheet,
                edit_target_row,
                row_payload,
            )
            if not saved_ok:
                messagebox.showerror(
                    "Journal Window",
                    "Could not update previous entry. It may have changed. Try again.",
                )
                return
        else:
            append_row(MODULES["J"], row_payload)
        clear_journal_window_draft()
        saved["value"] = True
        if autosave_id["value"] is not None:
            root.after_cancel(autosave_id["value"])
        _finish_or_background_journal_window()

    def stop_active_recording_for_close() -> None:
        th = record_thread_holder.get("thread")
        if not (
            recording_ui_busy["v"]
            and isinstance(th, threading.Thread)
            and th.is_alive()
        ):
            return
        record_close_requested["v"] = True
        record_stop.set()
        record_pause.clear()
        try:
            th.join(timeout=12)
        except RuntimeError:
            pass

    def close_after_uninstall() -> None:
        uninstall_shutdown_requested["v"] = True
        stop_active_recording_for_close()
        if autosave_id["value"] is not None:
            try:
                root.after_cancel(autosave_id["value"])
            except tk.TclError:
                pass
            autosave_id["value"] = None
        try:
            destroy_journal_window()
        except tk.TclError:
            pass

    def on_close(event=None) -> None:
        if uninstall_shutdown_requested["v"]:
            close_after_uninstall()
            return
        stop_active_recording_for_close()
        has_content = any(
            [
                text_box.get("1.0", "end-1c").strip(),
                stt_box.get("1.0", "end-1c").strip(),
                report_box.get("1.0", "end-1c").strip(),
            ]
        )
        if not has_content:
            clear_journal_window_draft()
            if autosave_id["value"] is not None:
                root.after_cancel(autosave_id["value"])
            _finish_or_background_journal_window()
            return
        save_choice = messagebox.askyesnocancel(
            "Close Journal Window",
            "Do you want to save this journal entry before closing?",
        )
        if save_choice is None:
            return
        if save_choice:
            do_save()
            return
        should_discard = messagebox.askyesno(
            "Discard Changes",
            "Are you sure you want to close without saving to journal? Draft backup is kept.",
        )
        if not should_discard:
            return
        save_draft()
        if autosave_id["value"] is not None:
            root.after_cancel(autosave_id["value"])
        _finish_or_background_journal_window()

    console_history: List[str] = []
    console_hist_index = {"value": 0}
    prefs_for_console = load_preferences()
    console_app_name = {"value": prefs_for_console.get("app_name", "Daily Logger") or "Daily Logger"}
    # GUI console "JS" state: when user types JS, we enter a non-freezing sub-mode.
    js_gui_state: Dict[str, object] = {"active": False}

    def console_append(text: str) -> None:
        if not text:
            return
        console_output.config(state="normal")
        stamp = _console_update_timestamp()
        lines = (text.rstrip("\n") or "").splitlines() or [text.rstrip("\n")]
        console_output.insert("end", "\n".join(f"[{stamp}] {line}" for line in lines) + "\n")
        console_output.see("end")
        console_output.config(state="disabled")

    def run_console_command() -> None:
        if console_entry_state["placeholder"]:
            return
        raw = console_entry.get().strip()
        console_entry.delete(0, "end")
        if not raw:
            return
        console_append(f"> {raw}")
        if not console_history or console_history[-1] != raw:
            console_history.append(raw)
        console_hist_index["value"] = len(console_history)
        cmd = raw.upper()
        cmd_parts = cmd.split()
        if cmd in {"RS", "RECORD STOP", "RECORDSTOP"}:
            if stop_recording():
                console_append("Stopping recording; it will save to the Recording folder.")
            else:
                console_append("No active recording to stop.")
            return

        record_start_mode = ""
        if cmd == "RC":
            record_start_mode = record_source_mode["value"]
        elif cmd == "RECORD":
            record_start_mode = record_source_mode["value"]
        elif len(cmd_parts) == 2 and cmd_parts[0] == "RECORD":
            requested_source = cmd_parts[1]
            if requested_source in {"MIC", "MICROPHONE"}:
                record_start_mode = RECORD_SOURCE_MIC
            elif requested_source in {"COMPUTER", "COMPUTERAUDIO", "SYSTEM", "SYSTEMAUDIO", "PC"}:
                record_start_mode = RECORD_SOURCE_COMPUTER
            elif requested_source in {"BOTH", "ALL"}:
                record_start_mode = RECORD_SOURCE_BOTH
        if record_start_mode:
            if transcribing_busy["v"]:
                console_append("Finish the current transcription before starting a background recording.")
                return
            if recording_ui_busy["v"]:
                _apply_record_source_mode(record_start_mode, persist=True, update_status=True)
                console_append(f"Recording source changed to {_record_source_label(record_start_mode)}.")
                return
            _apply_record_source_mode(record_start_mode, persist=True, update_status=False)
            if start_recording(background=True):
                console_append(
                    f"Background recording started: {_record_source_label(record_start_mode)}. "
                    "Type RECORD STOP or rs to stop and save it."
                )
            else:
                console_append("Could not start background recording.")
            return
        if bool(js_gui_state.get("active")):
            # First token after JS is treated as "journal_settings_menu" choice.
            # This mirrors the old CLI menu, but stays non-blocking for the Tk GUI.
            try:
                choice = raw.strip()
                choice_key = choice.upper()
                help_text = (
                    "Journal settings:\n"
                    "  WINDOW               - open window editor\n"
                    "  CONSOLE              - type journal text in console\n"
                    "  EDITPREV             - edit latest entry in window\n"
                    "  DP                   - delete latest entry\n"
                    "  RESTORE              - reopen latest unsaved draft\n"
                    "  HELP                 - show this list\n"
                    "  Enter                - return to main menu\n"
                    "  DEFAULT WINDOWS     - set preferred journal input to window\n"
                    "  DEFAULT CONSOLE      - set preferred journal input to console"
                )
                if is_enter_equivalent(choice_key) or not choice:
                    js_gui_state["active"] = False
                    console_append("JS menu closed.")
                    return
                if choice_key == "HELP":
                    console_append(help_text)
                    return
                if choice_key in ("W", "WINDOW", "WINDOWS"):
                    js_gui_state["active"] = False
                    show_page("journal")
                    open_journal_window_editor()
                    console_append("Opened Journal window editor.")
                    return
                if choice_key in ("C", "CONSOLE", "CONSOLE", "COINSOLE"):
                    # Multi-step: typed note + date/time.
                    js_gui_state["active"] = False
                    show_page("journal")
                    typed_note = _ask_typed_note_gui(root)
                    if typed_note is None:
                        console_append("Journal CONSOLE cancelled.")
                        return
                    dt = ask_entry_date_time_gui(root)
                    if dt is None:
                        console_append("Journal date/time cancelled.")
                        return
                    date_value, time_value = dt
                    append_row(MODULES["J"], [date_value, time_value, typed_note, "", ""])
                    console_append(f'Journal saved to: {DATA_DIR / MODULES["J"].workbook_name}')
                    return
                if choice_key in ("EDITPREV", "EDIT PREV", "EDIT PREVIOUS", "OPENPREV", "OPEN PREV", "OPENPREVIOUS", "OPEN PREVIOUS"):
                    js_gui_state["active"] = False
                    show_page("journal")
                    latest = get_latest_journal_entry_for_edit()
                    if not latest:
                        console_append("No previous journal entry found to edit.")
                        return
                    try:
                        load_latest_entry_into_current_journal(latest)
                        console_append("Loaded latest journal entry into current textbox (edit mode on).")
                    except Exception as exc:
                        console_append(f"EDITPREV failed: {exc}")
                    return
                if choice_key == "DP":
                    js_gui_state["active"] = False
                    show_page("journal")
                    latest = get_latest_journal_entry_for_delete()
                    if not latest:
                        console_append("No previous journal entry found to delete.")
                        return
                    date_label = str(latest.get("date", "")).strip() or "(unknown date)"
                    time_label = str(latest.get("time", "")).strip() or "(unknown time)"
                    if messagebox.askyesno(
                        "Delete previous journal entry",
                        f"Delete previous journal entry at {date_label} {time_label}?",
                    ):
                        delete_latest_journal_entry()
                        console_append("Deleted previous journal entry.")
                    else:
                        console_append("Delete cancelled.")
                    return
                if choice_key == "RESTORE":
                    js_gui_state["active"] = False
                    show_page("journal")
                    draft = load_journal_window_draft()
                    if not draft:
                        console_append("No journal draft to restore.")
                        return
                    open_journal_window_editor(draft)
                    console_append("Restored draft opened.")
                    return
                if choice_key in ("DEFAULT WINDOWS", "DEFAULT CONSOLE"):
                    js_gui_state["active"] = False
                    show_page("journal")
                    prefs = load_preferences()
                    default_mode = "windows" if choice_key == "DEFAULT WINDOWS" else "console"
                    prefs["journal_input_default"] = default_mode
                    if save_preferences(prefs):
                        console_append(f"Default set to {default_mode}.")
                    else:
                        console_append("Could not save default journal input preference.")
                    return
                if choice_key == "J":
                    js_gui_state["active"] = False
                    show_page("journal")
                    console_append("Switched to Journal page.")
                    return
                console_append("Unknown JS choice. Type HELP for options.")
            except Exception as exc:
                console_append(f"JS menu error: {exc}")
            return

        # Non-JS direct commands in GUI console:
        # - RESTORE: load saved journal draft into the existing editor widgets
        # - EDITPREV/OPENPREV: load latest journal entry into the existing editor widgets
        if cmd in {"RESTORE"}:
            js_gui_state["active"] = False
            show_page("journal")
            ok = load_draft_into_current_journal()
            if not ok:
                console_append("No journal draft to restore.")
            else:
                console_append("Restored draft into current journal textbox.")
            return
        if cmd in {"EDITPREV", "EDIT PREV", "EDIT PREVIOUS", "OPENPREV", "OPEN PREV", "OPEN PREVIOUS"}:
            js_gui_state["active"] = False
            show_page("journal")
            latest = get_latest_journal_entry_for_edit()
            if not latest:
                console_append("No previous journal entry found to edit.")
                return
            load_latest_entry_into_current_journal(latest)
            console_append("Loaded latest journal entry into current textbox (edit mode on).")
            return
        if cmd in {"NEW", "NEW JOURNAL"}:
            js_gui_state["active"] = False
            show_page("journal")
            ok = start_new_journal()
            if ok:
                console_append("Started new journal. Editor cleared.")
            else:
                console_append("New journal cancelled.")
            return
        # Prevent GUI freeze: "JS" triggers an interactive CLI prompt (`input()`),
        # which can block the Tk main thread.
        # GUI version uses Tk dialogs instead of ``input()``.
        if cmd in {"J SETTINGS", "J SETTING", "JOURNAL SETTINGS", "JS"}:
            show_page("journal")
            js_gui_state["active"] = True
            console_append("JS menu opened. Type HELP for available choices, then submit one choice.")
            console_append(
                "Journal settings: "
                "WINDOW | CONSOLE | EDITPREV | DP | RESTORE | DEFAULT WINDOWS | DEFAULT CONSOLE | HELP"
            )
            return
        if cmd in {"J", "JOURNAL", "WINDOW"}:
            show_page("journal")
            console_append("Switched to Journal page.")
            return
        if cmd in {"R", "RT"}:
            show_page("ai_recap")
            console_append("Switched to AI Recap page.")
            return
        if cmd in {"C", "CT"}:
            show_page("chatbot")
            console_append("Switched to Chatbot page.")
            return
        if cmd in {"PET", "DESKTOP PET", "桌宠", "COMPANION"}:
            console_append("Companion has been removed from this build. The saved version is on the companion archive branch.")
            return
        if cmd == "CONSOLE":
            if sys.platform != "win32":
                console_append("Native console show is supported on Windows only.")
                return
            try:
                console_hwnd = ctypes.windll.kernel32.GetConsoleWindow()
                if console_hwnd:
                    ctypes.windll.user32.ShowWindow(console_hwnd, 5)  # SW_SHOW
                    console_append("Native console window shown.")
                else:
                    launch_cmd = (
                        f'Set-Location -LiteralPath "{str(BASE_DIR)}"; '
                        "Write-Host \"Daily Logger on-demand console\"; "
                        "Write-Host \"You can run commands here.\"; "
                        "Write-Host \"Close this window when finished.\""
                    )
                    subprocess.Popen(
                        ["powershell", "-NoExit", "-Command", launch_cmd],
                        creationflags=getattr(subprocess, "CREATE_NEW_CONSOLE", 0),
                    )
                    console_append("Opened a new on-demand console window.")
            except Exception:
                console_append("Could not show native console window.")
            return
        capture = io.StringIO()
        keep_running = True
        try:
            with contextlib.redirect_stdout(capture):
                keep_running, next_name = handle_choice(raw, console_app_name["value"])
            console_app_name["value"] = next_name
        except Exception as exc:
            console_append(f"Error: {exc}")
            return
        output = capture.getvalue().strip()
        if output:
            console_append(output)
        if not keep_running:
            if cmd == "CONFIRM UNINSTALL":
                close_after_uninstall()
            else:
                on_close()

    def _console_history_up(_evt: Optional[Any] = None) -> str:
        if not console_history:
            return "break"
        _clear_console_placeholder()
        console_hist_index["value"] = max(0, console_hist_index["value"] - 1)
        console_entry.delete(0, "end")
        console_entry.insert(0, console_history[console_hist_index["value"]])
        return "break"

    def _console_history_down(_evt: Optional[Any] = None) -> str:
        if not console_history:
            return "break"
        _clear_console_placeholder()
        console_hist_index["value"] = min(len(console_history), console_hist_index["value"] + 1)
        console_entry.delete(0, "end")
        if console_hist_index["value"] < len(console_history):
            console_entry.insert(0, console_history[console_hist_index["value"]])
        else:
            _set_console_placeholder()
        return "break"

    def _console_entry_focus_in(_evt: Optional[Any] = None) -> None:
        _clear_console_placeholder()

    def _console_entry_focus_out(_evt: Optional[Any] = None) -> None:
        if not console_entry.get().strip():
            console_entry.delete(0, "end")
            _set_console_placeholder()

    def _console_entry_keypress(evt: Optional[Any] = None) -> Optional[str]:
        if evt is None:
            return None
        if not console_entry_state["placeholder"]:
            return None
        if evt.keysym in {"Left", "Right", "Home", "End"}:
            return "break"
        if evt.keysym in {"BackSpace", "Delete"}:
            _clear_console_placeholder()
            return "break"
        if evt.char and evt.char >= " ":
            _clear_console_placeholder()
        return None

    def _console_tab_complete(_evt: Optional[Any] = None) -> str:
        current = "" if console_entry_state["placeholder"] else console_entry.get()
        if not current.strip():
            capture = io.StringIO()
            try:
                with contextlib.redirect_stdout(capture):
                    print_main_help()
            except Exception:
                capture = io.StringIO()
            output = capture.getvalue().strip()
            if output:
                console_append(output)
            return "break"
        completed, extended = _line_tab_extend(current, MAIN_MENU_COMPLETIONS)
        if extended:
            _clear_console_placeholder()
            console_entry.delete(0, "end")
            console_entry.insert(0, completed)
            console_entry.icursor("end")
        return "break"

    button_row = tk.Frame(journal_page, bg=t_init.surface, bd=0, highlightthickness=0)
    save_entry_btn = tk.Button(
        button_row,
        text="Save Entry",
        command=do_save,
        bg=t_init.btn_disabled,
        fg=t_init.disabled_fg,
        activebackground=t_init.hover_save,
        activeforeground="white",
        disabledforeground=t_init.disabled_fg,
        relief="flat",
        font=("Segoe UI", 10, "bold"),
        padx=18,
        pady=8,
        state="disabled",
        cursor="arrow",
    )
    save_entry_btn.pack(side="right")
    save_entry_btn_holder["btn"] = save_entry_btn
    root.after_idle(lambda: _layout_console_row(active_page_frame.get("frame") or journal_page))

    def save_rest_style() -> Tuple[str, str, str, str, str]:
        t = th()
        if str(save_entry_btn.cget("state")) != "normal":
            return t.save_bind_disabled()
        return ("normal", t.accent, "white", t.hover_save, "white")

    bind_button_hover_if_enabled(
        save_entry_btn,
        save_rest_style,
        lambda: th().hover_save,
        lambda: "white",
    )

    def _on_journal_text_changed(_evt: Optional[Any] = None) -> None:
        refresh_save_entry_state()
        _feed_desktop_pet_from_current_entry()

    def _journal_delete_prev_word(_evt: Optional[Any] = None) -> str:
        w = root.focus_get()
        if not isinstance(w, tk.Text):
            return "break"
        try:
            if w.tag_ranges("sel"):
                w.delete("sel.first", "sel.last")
                refresh_save_entry_state()
                return "break"
            left = w.get("1.0", "insert")
            if not left:
                return "break"
            end_non_ws = len(left)
            while end_non_ws > 0 and left[end_non_ws - 1].isspace():
                end_non_ws -= 1
            start = end_non_ws
            while start > 0 and not left[start - 1].isspace():
                start -= 1
            delete_chars = len(left) - start
            if delete_chars > 0:
                w.delete(f"insert-{delete_chars}c", "insert")
        except tk.TclError:
            pass
        refresh_save_entry_state()
        return "break"

    def _journal_delete_next_word(_evt: Optional[Any] = None) -> str:
        w = root.focus_get()
        if not isinstance(w, tk.Text):
            return "break"
        try:
            if w.tag_ranges("sel"):
                w.delete("sel.first", "sel.last")
                refresh_save_entry_state()
                return "break"
            right = w.get("insert", "end-1c")
            if not right:
                return "break"
            i = 0
            n = len(right)
            while i < n and right[i].isspace():
                i += 1
            while i < n and not right[i].isspace():
                i += 1
            if i > 0:
                w.delete("insert", f"insert+{i}c")
        except tk.TclError:
            pass
        refresh_save_entry_state()
        return "break"

    def _journal_undo(_evt: Optional[Any] = None) -> str:
        w = root.focus_get()
        if not isinstance(w, tk.Text):
            return "break"
        try:
            w.edit_undo()
        except tk.TclError:
            pass
        refresh_save_entry_state()
        return "break"

    def _journal_redo(_evt: Optional[Any] = None) -> str:
        w = root.focus_get()
        if not isinstance(w, tk.Text):
            return "break"
        try:
            w.edit_redo()
        except tk.TclError:
            pass
        refresh_save_entry_state()
        return "break"

    for _tb in (text_box, stt_box, report_box):
        _tb.bind("<KeyRelease>", _on_journal_text_changed, add="+")
        _tb.bind("<FocusIn>", lambda _e, _w=_tb: _journal_find_state.__setitem__("widget", _w), add="+")
        _tb.bind("<ButtonRelease-1>", _on_journal_text_changed, add="+")
        _tb.bind("<Control-BackSpace>", _journal_delete_prev_word, add="+")
        _tb.bind("<Control-w>", _journal_delete_prev_word, add="+")
        _tb.bind("<Control-Delete>", _journal_delete_next_word, add="+")
        _tb.bind("<Control-z>", _journal_undo, add="+")
        _tb.bind("<Control-y>", _journal_redo, add="+")
        _tb.bind("<Control-Z>", _journal_redo, add="+")
        _tb.bind("<Control-a>", _journal_select_all, add="+")
        _tb.bind("<Control-f>", _find_open, add="+")
    root.bind("<Control-f>", _find_open, add="+")

    def apply_journal_window_i18n() -> None:
        try:
            splash_title.config(text=tr("splash.title", app=window_app_name))
        except tk.TclError:
            pass
        settings_title.config(text=tr("settings.title"))
        for _lbl, _key in settings_label_keys:
            _lbl.config(text=tr(_key))
        try:
            if lang_ui_combo is not None and ttk is not None:
                lang_ui_combo.config(
                    values=(tr("settings.lang.english"), tr("settings.lang.chinese"))
                )
        except tk.TclError:
            pass
        _want_ui_lang = (
            tr("settings.lang.chinese")
            if ui_lang_holder[0] == "zh"
            else tr("settings.lang.english")
        )
        if ui_lang_var.get().strip() != _want_ui_lang.strip():
            ui_lang_var.set(_want_ui_lang)
        rename_btn.config(text=tr("settings.rename_btn"))
        startup_toggle_btn.config(
            text=tr("settings.on") if startup_state["enabled"] else tr("settings.off")
        )
        iphone_receive_toggle_btn.config(
            text=tr("settings.on") if iphone_receive_state["enabled"] else tr("settings.off")
        )
        transcription_models_btn.config(text=tr("settings.manage"))
        updates_toggle_btn.config(
            text=tr("settings.on") if updates_state["enabled"] else tr("settings.off")
        )
        updates_check_btn.config(text=tr("settings.check_now"))
        settings_theme_btn.config(
            text=tr(journal_window_theme_current_label_key(th().id))
        )
        display_fullscreen_btn.config(
            text=tr("settings.on") if fullscreen_state["enabled"] else tr("settings.off")
        )
        display_f11_lbl.config(text=tr("settings.display_f11"))
        backup_mode_btn.config(text=_backup_mode_btn_label(backup_mode["value"]))
        backup_manual_btn.config(text=tr("settings.manual"))
        token_save_btn.config(text=tr("settings.save"))
        token_copy_btn.config(text=tr("settings.copy"))
        start_menu_app_btn.config(text=tr("settings.start_menu_app"))
        start_menu_journal_btn.config(text=tr("settings.start_menu_journal"))
        start_menu_reader_btn.config(text=tr("settings.start_menu_reader"))
        _refresh_sidebar_module_styles()
        _render_module_library()
        date_lbl.config(text=tr("journal.date"))
        time_lbl.config(text=tr("journal.time"))
        update_time_btn.config(text=tr("journal.update_time"))
        restore_draft_btn.config(text=tr("journal.restore_draft"))
        find_lbl.config(text=tr("find.label"))
        find_scope_all_rb.config(text=tr("find.all"))
        find_scope_one_rb.config(text=tr("find.current_box"))
        find_case_chk.config(text=tr("find.case"))
        find_word_chk.config(text=tr("find.word"))
        find_prev_btn.config(text=tr("find.prev"))
        find_next_btn.config(text=tr("find.next"))
        find_close_btn.config(text=tr("find.close"))
        journal_title_lbl.config(text=tr("journal.section.journal"))
        stt_title_lbl.config(text=tr("journal.section.stt"))
        report_title_lbl.config(text=tr("journal.section.report"))
        pet_title_lbl.config(text=tr("pet.title"))
        pet_body_lbl.config(text=tr("pet.body"))
        pet_name_lbl.config(text=tr("pet.name_label"))
        pet_name_save_btn.config(text=tr("pet.name_save"))
        pet_feed_btn.config(text=tr("pet.feed_now"))
        _refresh_pet_art_button()
        pet_desktop_btn.config(text=tr("pet.desktop_mode"))
        pet_export_btn.config(text=tr("pet.export_profile"))
        pet_import_btn.config(text=tr("pet.import_profile"))
        pet_folder_btn.config(text=tr("pet.open_folder"))
        pet_delete_btn.config(text=tr("pet.delete_profile"))
        pet_modules_lbl.config(text=tr("pet.modules_title"))
        pet_memory_title_lbl.config(text=tr("pet.module_memory_title"))
        pet_tts_title_lbl.config(text=tr("pet.module_tts_title"))
        pet_chat_title_lbl.config(text=tr("pet.module_chat_title"))
        pet_chat_ask_btn.config(text=tr("pet.chat_ask"))
        _refresh_pet_module_ui()
        _refresh_desktop_pet_ui()
        open_recording_btn.config(text=tr("journal.open"))
        stt_lang_lbl.config(text=tr("journal.lang_label"))
        transcribe_btn.config(text=tr("journal.transcribe"))
        transcribe_file_btn.config(text=tr("journal.transcribe_file"))
        receive_iphone_btn.config(
            text=tr("journal.iphone_receive")
            if bool(iphone_receiver_state.get("passive"))
            else (
                tr("journal.iphone_stop")
                if bool(iphone_receiver_state.get("active"))
                else tr("journal.iphone_receive")
            )
        )
        gen_button.config(text=tr("journal.generate_report"))
        save_entry_btn.config(text=tr("journal.save_entry"))
        start_rec_button.config(text=tr("journal.rec.start"))
        stop_rec_button.config(text=tr("journal.rec.stop"))
        _refresh_record_source_selector()
        _refresh_transcription_model_selectors()
        if recording_ui_busy["v"]:
            pause_rec_button.config(
                text=tr("journal.rec.resume")
                if record_pause.is_set()
                else tr("journal.rec.pause")
            )
        else:
            pause_rec_button.config(text=tr("journal.rec.pause"))
        console_title.config(text=tr("console.title"))
        _show_console_hint_placeholder()
        _ai_i18n = getattr(build_ai_recap_and_chatbot_pages, "_i18n", None)
        if callable(_ai_i18n):
            _ai_i18n()

    def _apply_console_ui_language(new_lang: str) -> None:
        ui_lang_holder[0] = new_lang
        apply_journal_window_i18n()

    set_journal_ui_language_changed_hook(_apply_console_ui_language)

    def _on_journal_root_destroy(event: Any) -> None:
        if getattr(event, "widget", None) is root:
            set_journal_ui_language_changed_hook(None)

    root.bind("<Destroy>", _on_journal_root_destroy, add="+")

    def apply_journal_window_colors() -> None:
        t = th()
        root.configure(bg=t.surface)
        shell.configure(bg=t.surface)
        nav_rail.configure(bg=t.panel)
        nav_title.configure(bg=t.panel, fg=t.muted)
        nav_modules_frame.configure(bg=t.panel)
        nav_bottom_frame.configure(bg=t.panel)
        nav_add_canvas.configure(bg=t.panel)
        for _placeholder in nav_drop_placeholders:
            _placeholder.configure(bg=t.panel)
        nav_summon_btn.configure(
            bg=t.toolbar_btn_config()[0],
            fg=t.toolbar_btn_config()[1],
            activebackground=t.toolbar_btn_config()[2],
            activeforeground=t.toolbar_btn_config()[3],
        )
        content_host.configure(bg=t.surface)
        journal_page.configure(bg=t.surface)
        ai_recap_page.configure(bg=t.surface)
        chatbot_page.configure(bg=t.surface)
        desktop_pet_page.configure(bg=t.surface)
        console_page.configure(bg=t.surface)
        settings_page.configure(bg=t.surface)
        module_library_page.configure(bg=t.surface)
        _refresh_sidebar_module_styles()
        _render_module_library()
        for _w in placeholder_frames:
            _w.configure(bg=t.surface)
        for _w in placeholder_title_labels:
            _w.configure(bg=t.surface, fg=t.text)
        for _w in placeholder_body_labels:
            _w.configure(bg=t.surface, fg=t.muted)
        _ai_theme_fn = getattr(build_ai_recap_and_chatbot_pages, "_apply_theme", None)
        if callable(_ai_theme_fn):
            _ai_theme_fn()
        settings_wrap.configure(bg=t.surface)
        settings_title.configure(bg=t.surface, fg=t.text)
        settings_status_lbl.configure(bg=t.surface, fg=t.muted)
        updates_status_lbl.configure(bg=t.surface, fg=t.muted)
        display_f11_lbl.configure(bg=t.surface, fg=t.muted)
        for _w in settings_rows:
            _w.configure(bg=t.surface)
        for _w in settings_labels:
            _w.configure(bg=t.surface, fg=t.muted)
        rename_entry.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        token_entry.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        for _btn in (
            rename_btn,
            startup_toggle_btn,
            iphone_receive_toggle_btn,
            transcription_models_btn,
            updates_toggle_btn,
            updates_check_btn,
            settings_theme_btn,
            display_fullscreen_btn,
            backup_mode_btn,
            backup_manual_btn,
            token_save_btn,
            token_copy_btn,
            start_menu_app_btn,
            start_menu_journal_btn,
            start_menu_reader_btn,
        ):
            _btn.configure(
                bg=t.btn_secondary,
                fg=t.text,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
            )
        for _btn in page_toggle_buttons:
            _btn.configure(
                bg=t.toolbar_btn_config()[0],
                fg=t.toolbar_btn_config()[1],
                activebackground=t.toolbar_btn_config()[2],
                activeforeground=t.toolbar_btn_config()[3],
            )
        _refresh_sidebar_module_styles()
        top.configure(bg=t.panel)
        top.pack_configure(padx=t.pad_outer, pady=t.pad_top_y)
        find_row.configure(bg=t.panel)
        if str(find_row.winfo_manager()) == "pack":
            find_row.pack_configure(padx=t.pad_outer, pady=(0, 6))
        find_lbl.configure(bg=t.panel, fg=t.muted)
        find_entry.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        find_status.configure(bg=t.panel, fg=t.muted)
        find_scope_all_rb.configure(
            bg=t.panel,
            fg=t.muted,
            activebackground=t.panel,
            activeforeground=t.text,
            selectcolor=t.panel,
        )
        find_scope_one_rb.configure(
            bg=t.panel,
            fg=t.muted,
            activebackground=t.panel,
            activeforeground=t.text,
            selectcolor=t.panel,
        )
        find_case_chk.configure(
            bg=t.panel,
            fg=t.muted,
            activebackground=t.panel,
            activeforeground=t.text,
            selectcolor=t.panel,
        )
        find_word_chk.configure(
            bg=t.panel,
            fg=t.muted,
            activebackground=t.panel,
            activeforeground=t.text,
            selectcolor=t.panel,
        )
        console_wrap.configure(bg=t.surface)
        console_title.configure(bg=t.surface, fg=t.text)
        console_output.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        console_scroll.config(bg=t.panel, troughcolor=t.field, activebackground=t.accent)
        console_input_row.configure(bg=t.surface)
        console_prompt.configure(bg=t.surface, fg=t.muted)
        console_entry.config(
            bg=t.field,
            fg=(_console_placeholder_fg(t) if console_entry_state["placeholder"] else t.text),
            insertbackground=t.text,
            insertwidth=(0 if console_entry_state["placeholder"] else console_insertwidth_normal),
            highlightbackground=t.border,
            highlightcolor=t.accent,
            font=("Consolas", 10, "italic") if console_entry_state["placeholder"] else ("Consolas", 10),
        )
        date_lbl.configure(bg=t.panel, fg=t.muted, font=t.date_label_font)
        time_lbl.configure(bg=t.panel, fg=t.muted, font=t.date_label_font)
        try:
            date_entry.config(background=t.field, foreground=t.text)
        except tk.TclError:
            try:
                date_entry.config(
                    bg=t.field,
                    fg=t.text,
                    insertbackground=t.text,
                    highlightbackground=t.border,
                    highlightcolor=t.accent,
                )
            except tk.TclError:
                pass
        time_entry.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        tbg, tfg, tabg, tafg = t.toolbar_btn_config()
        update_time_btn.config(
            bg=tbg, fg=tfg, activebackground=tabg, activeforeground=tafg
        )
        restore_draft_btn.config(bg=tbg, fg=tfg, activebackground=tabg, activeforeground=tafg)
        for _b in (find_prev_btn, find_next_btn, find_close_btn):
            _b.config(bg=tbg, fg=tfg, activebackground=tabg, activeforeground=tafg)
        settings_theme_btn.config(
            text=tr(journal_window_theme_current_label_key(t.id)),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
        )
        center.configure(bg=t.surface)
        center.pack_configure(
            padx=t.pad_outer,
            pady=(0, t.pad_center_y + JOURNAL_WINDOW_CONSOLE_RESERVE_BOTTOM),
        )
        left_col.configure(bg=t.surface)
        journal_title_lbl.configure(
            bg=t.surface, fg=t.muted, font=t.section_label_font
        )
        editor_frame.configure(bg=t.panel)
        text_box.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        scroll_bar.configure_theme(t)
        right_col.configure(bg=t.surface)
        stt_outer.configure(bg=t.surface)
        stt_header.configure(bg=t.surface)
        stt_title_lbl.configure(
            bg=t.surface, fg=t.muted, font=t.section_label_font
        )
        stt_saved_path_entry.config(
            readonlybackground=t.surface,
            fg=t.muted,
            highlightbackground=t.surface,
            selectbackground=t.accent,
        )
        open_recording_btn.config(
            bg=tbg, fg=tfg, activebackground=tabg, activeforeground=tafg
        )
        stt_top.configure(bg=t.panel)
        stt_status.configure(bg=t.panel, fg=t.muted)
        stt_frame.configure(bg=t.panel)
        transcribe_hover.configure(bg=t.panel)
        wave_canvas.config(
            bg=t.field, highlightbackground=t.border, highlightcolor=t.accent
        )
        stt_box.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        stt_scroll.configure_theme(t)
        _update_iphone_receive_button()
        report_outer.configure(bg=t.surface)
        report_header.configure(bg=t.surface)
        report_title_lbl.configure(
            bg=t.surface, fg=t.muted, font=t.section_label_font
        )
        report_status.configure(bg=t.surface, fg=t.muted)
        report_frame.configure(bg=t.panel)
        gen_report_hover.configure(bg=t.panel)
        gen_report_btn_row.configure(bg=t.panel)
        report_box.config(
            bg=t.field,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        report_scroll.configure_theme(t)
        pet_wrap.configure(bg=t.surface)
        pet_title_lbl.configure(bg=t.surface, fg=t.text)
        pet_body_lbl.configure(bg=t.surface, fg=t.muted)
        pet_info_holder.configure(bg=t.panel, highlightbackground=t.border)
        pet_info_canvas.configure(bg=t.panel)
        pet_info_scroll.configure(
            bg=t.panel,
            troughcolor=t.field,
            activebackground=t.accent,
        )
        pet_info.configure(bg=t.panel, highlightbackground=t.border)
        pet_level_lbl.configure(bg=t.panel, fg=t.text)
        pet_xp_lbl.configure(bg=t.panel, fg=t.muted)
        pet_mood_lbl.configure(bg=t.panel, fg=t.muted)
        pet_name_frame.configure(bg=t.panel)
        pet_name_lbl.configure(bg=t.panel, fg=t.muted)
        pet_name_entry.configure(
            bg=t.field,
            fg=t.text,
            disabledbackground=t.btn_disabled,
            disabledforeground=t.disabled_fg,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        pet_profile_frame.configure(bg=t.panel)
        pet_modules_lbl.configure(bg=t.panel, fg=t.text)
        pet_bubble_module.configure(bg=t.field, highlightbackground=t.border)
        pet_bubble_title_lbl.configure(bg=t.field, fg=t.text)
        pet_bubble_status_lbl.configure(bg=t.field, fg=t.muted)
        pet_memory_module.configure(bg=t.field, highlightbackground=t.border)
        pet_memory_title_lbl.configure(bg=t.field, fg=t.text)
        pet_memory_status_lbl.configure(bg=t.field, fg=t.muted)
        pet_tts_module.configure(bg=t.field, highlightbackground=t.border)
        pet_tts_title_lbl.configure(bg=t.field, fg=t.text)
        pet_tts_status_lbl.configure(bg=t.field, fg=t.muted)
        pet_chat_module.configure(bg=t.field, highlightbackground=t.border)
        pet_chat_title_lbl.configure(bg=t.field, fg=t.text)
        pet_chat_status_lbl.configure(bg=t.field, fg=t.muted)
        pet_chat_entry_row.configure(bg=t.field)
        pet_chat_entry.configure(
            bg=t.panel,
            fg=t.text,
            insertbackground=t.text,
            highlightbackground=t.border,
            highlightcolor=t.accent,
        )
        pet_feed_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
        )
        pet_art_ref_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
        )
        pet_desktop_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        pet_journal_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        for _btn in (pet_name_save_btn, pet_export_btn, pet_import_btn, pet_folder_btn):
            _btn.configure(
                bg=t.btn_secondary,
                fg=t.text,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
                cursor="hand2",
            )
        pet_delete_btn.configure(
            bg="#8B2F3A",
            fg="white",
            activebackground="#A43B48",
            activeforeground="white",
            cursor="hand2",
        )
        pet_bubble_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        pet_memory_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        pet_tts_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        pet_chat_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        pet_chat_ask_btn.configure(
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
            cursor="hand2",
        )
        _refresh_pet_art_button()
        _refresh_desktop_pet_ui()
        journal_pet_canvas.configure(bg=journal_pet_transparent)
        root.after_idle(_sync_journal_side_action_columns)
        stt_lang_lbl.configure(bg=t.panel, fg=t.muted)
        if ttk is not None and _jw_style is not None:
            _jw_style.configure("Journal.TCombobox", **t.ttk_combobox_kwargs())
            if t.is_dark:
                _jw_style.map(
                    "Journal.TCombobox",
                    fieldbackground=[
                        ("readonly", t.field),
                        ("disabled", t.btn_disabled),
                    ],
                    selectbackground=[("readonly", t.accent)],
                    selectforeground=[("readonly", "white")],
                )
            else:
                _jw_style.map(
                    "Journal.TCombobox",
                    fieldbackground=[
                        ("readonly", t.field),
                        ("disabled", t.btn_disabled),
                    ],
                )
        else:
            try:
                lang_combo.config(bg=t.panel, fg=t.text)
            except tk.TclError:
                pass
            try:
                record_source_combo.config(bg=t.panel, fg=t.text)
            except tk.TclError:
                pass
        if str(gen_button.cget("state")) == "normal":
            _gs, gb, gf, gab, gaf = t.gen_bind_rest()
            gen_button.config(
                bg=gb,
                fg=gf,
                activebackground=gab,
                activeforeground=gaf,
                cursor="hand2",
            )
        else:
            _ds, gb, gf, _dab, _daf = t.gen_bind_disabled()
            gen_button.config(
                bg=gb,
                fg=gf,
                disabledforeground=gf,
                cursor="arrow",
            )
        update_transcribe_ui()
        for _b in (start_rec_button, record_source_combo, pause_rec_button, stop_rec_button):
            if _b is stop_rec_button and recording_background_mode["v"]:
                _journal_rec_btn_set_disabled_look_clickable(_b)
            else:
                _journal_rec_btn_set(_b, str(_b.cget("state")) == "normal")
        button_row.configure(bg=t.surface)
        _layout_console_row(active_page_frame.get("frame") or journal_page)
        refresh_save_entry_state()
        redraw_waveform_canvas()
        apply_journal_window_i18n()

    ui_lang_var.trace_add("write", lambda *_a: root.after_idle(_on_ui_language_selected))

    _startup_step("splash.detail.pages")

    def toggle_journal_window_theme() -> None:
        prefs = load_preferences()
        cur = normalize_journal_window_theme_key(th().id)
        nxt = next_journal_window_theme_key(cur)
        prefs[JOURNAL_PREF_THEME_KEY] = nxt
        save_preferences(prefs)
        theme_holder[0] = journal_window_theme_spec_for_key(nxt)
        apply_journal_window_colors()

    def on_virtual_reader_nav_clicked() -> None:
        ok, err = open_virtual_reader_nav_action()
        if not ok and err:
            messagebox.showerror(tr("msg.virtual_reader_title"), err)

    sidebar_module_defs: Dict[str, Dict[str, Any]] = {
        "journal": {"label_key": "nav.journal", "fallback_en": "Journal", "fallback_zh": "日记", "page": "journal"},
        "ai_recap": {"label_key": "nav.ai_recap", "fallback_en": "AI Recap", "fallback_zh": "AI 回顾", "page": "ai_recap"},
        "chatbot": {"label_key": "nav.chatbot", "fallback_en": "Chatbot", "fallback_zh": "聊天助手", "page": "chatbot"},
        "console": {"label_key": "nav.console", "fallback_en": "Console", "fallback_zh": "控制台", "page": "console"},
        "reader": {
            "label_key": "nav.virtual_reader",
            "short_label_key": "nav.reader_short",
            "fallback_en": "Virtual Reader",
            "fallback_zh": "阅读器",
            "short_fallback_en": "Reader",
            "short_fallback_zh": "阅读器",
            "action": on_virtual_reader_nav_clicked,
        },
        "settings": {"label_key": "nav.settings", "fallback_en": "Settings", "fallback_zh": "设置", "page": "settings"},
    }
    sidebar_all_module_ids = tuple(
        mid for mid in SIDEBAR_DEFAULT_MODULE_ORDER if mid in sidebar_module_defs
    )

    def _sidebar_module_label(module_id: str, *, short: bool = False) -> str:
        info = sidebar_module_defs.get(module_id, {})
        key = str(info.get("short_label_key" if short else "label_key") or info.get("label_key") or module_id)
        label = tr(key)
        if label == key or label.startswith("nav."):
            lang_suffix = "zh" if ui_lang_holder[0] == "zh" else "en"
            fallback_key = f"{'short_' if short else ''}fallback_{lang_suffix}"
            label = str(info.get(fallback_key) or info.get(f"fallback_{lang_suffix}") or module_id)
        return label

    def _default_sidebar_slots() -> List[str]:
        slots = [""] * SIDEBAR_TOTAL_SLOT_COUNT
        for index, module_id in enumerate(("journal", "ai_recap", "chatbot", "console")):
            if index < SIDEBAR_TOP_SLOT_COUNT:
                slots[index] = module_id
        slots[SIDEBAR_TOP_SLOT_COUNT] = "reader"
        slots[SIDEBAR_TOP_SLOT_COUNT + 1] = "settings"
        return slots

    def _slot_is_bottom(slot_index: int) -> bool:
        return slot_index >= SIDEBAR_TOP_SLOT_COUNT

    def _normalize_sidebar_module_order(order: List[str]) -> List[str]:
        slots = [""] * SIDEBAR_TOTAL_SLOT_COUNT
        used: set[str] = set()
        saw_empty = False
        legacy_modules: List[str] = []
        for item in order:
            module_id = str(item or "").strip()
            if not module_id:
                saw_empty = True
                legacy_modules.append("")
            elif module_id in sidebar_module_defs and module_id not in used:
                used.add(module_id)
                legacy_modules.append(module_id)
        if not saw_empty and [mid for mid in legacy_modules if mid] == list(sidebar_all_module_ids):
            return _default_sidebar_slots()
        if saw_empty:
            for index, module_id in enumerate(legacy_modules[:SIDEBAR_TOTAL_SLOT_COUNT]):
                slots[index] = module_id if module_id in sidebar_module_defs else ""
            return slots
        for index, module_id in enumerate(mid for mid in legacy_modules if mid):
            if index >= SIDEBAR_TOTAL_SLOT_COUNT:
                break
            slots[index] = module_id
        return slots

    def _insert_sidebar_module_once(slots: List[str], module_id: str, *, after: str = "") -> bool:
        if module_id in slots or module_id not in sidebar_module_defs:
            return False
        target_index: Optional[int] = None
        if after and after in slots:
            after_index = slots.index(after)
            for index in range(after_index + 1, SIDEBAR_TOP_SLOT_COUNT):
                if not slots[index]:
                    target_index = index
                    break
        if target_index is None:
            for index in range(SIDEBAR_TOP_SLOT_COUNT):
                if not slots[index]:
                    target_index = index
                    break
        if target_index is None:
            return False
        slots[target_index] = module_id
        return True

    def _top_sidebar_module_ids() -> List[str]:
        return list(nav_module_order_state["ids"][:SIDEBAR_TOP_SLOT_COUNT])

    def _bottom_sidebar_module_ids() -> List[str]:
        return list(nav_module_order_state["ids"][SIDEBAR_TOP_SLOT_COUNT:SIDEBAR_TOTAL_SLOT_COUNT])

    def _visible_sidebar_module_ids() -> List[str]:
        return [mid for mid in nav_module_order_state["ids"] if mid]

    def _module_slot_index(module_id: str) -> Optional[int]:
        for index, slot_module_id in enumerate(nav_module_order_state["ids"]):
            if slot_module_id == module_id:
                return index
        return None

    def _sidebar_is_bottom_module(module_id: str) -> bool:
        index = _module_slot_index(module_id)
        return bool(index is not None and _slot_is_bottom(index))

    def _load_sidebar_module_order() -> List[str]:
        prefs = load_preferences()
        if SIDEBAR_MODULE_ORDER_PREF_KEY not in prefs:
            return _default_sidebar_slots()
        raw = str(prefs.get(SIDEBAR_MODULE_ORDER_PREF_KEY, "") or "").strip()
        try:
            parsed = json.loads(raw)
            source = parsed if isinstance(parsed, list) else []
        except Exception:
            source = [part.strip() for part in raw.split(",")] if raw else []
        order: List[str] = []
        used: set[str] = set()
        for item in source:
            module_id = str(item).strip()
            if not module_id:
                order.append("")
            elif module_id in sidebar_module_defs and module_id not in used:
                order.append(module_id)
                used.add(module_id)
        normalized = _normalize_sidebar_module_order(order)
        if not bool(prefs.get(SIDEBAR_DESKTOP_PET_MIGRATION_PREF_KEY, False)):
            prefs[SIDEBAR_DESKTOP_PET_MIGRATION_PREF_KEY] = True
            save_preferences(prefs)
        return normalized

    def _save_sidebar_module_order() -> None:
        prefs = load_preferences()
        prefs[SIDEBAR_MODULE_ORDER_PREF_KEY] = json.dumps(nav_module_order_state["ids"])
        save_preferences(prefs)

    nav_modules_frame = tk.Frame(nav_rail, bg=t_init.panel, bd=0, highlightthickness=0)
    nav_modules_frame.grid(row=1, column=0, sticky="new", padx=0, pady=(0, 0))
    nav_modules_frame.grid_columnconfigure(0, weight=1)

    nav_bottom_frame = tk.Frame(nav_rail, bg=t_init.panel, bd=0, highlightthickness=0)
    nav_bottom_frame.grid(row=101, column=0, sticky="ew", padx=10, pady=(0, 10))
    nav_bottom_frame.grid_columnconfigure(0, weight=1, uniform="nav_bottom")
    nav_bottom_frame.grid_columnconfigure(1, weight=1, uniform="nav_bottom")

    nav_drop_placeholders: List[Any] = []
    nav_slot_spacers: List[Any] = []
    nav_drag_preview: Dict[str, Any] = {"window": None, "canvas": None, "module": ""}

    nav_add_canvas = tk.Canvas(
        nav_rail,
        height=44,
        bg=t_init.panel,
        bd=0,
        highlightthickness=0,
        cursor="hand2",
    )
    nav_add_canvas.grid(row=99, column=0, sticky="ew", padx=10, pady=(6, 12))

    module_library_wrap = tk.Frame(module_library_page, bg=t_init.surface)
    module_library_wrap.pack(fill="both", expand=True, padx=28, pady=28)
    module_library_title = tk.Label(
        module_library_wrap,
        text=tr("module_library.title"),
        bg=t_init.surface,
        fg=t_init.text,
        font=("Segoe UI", 18, "bold"),
        anchor="w",
    )
    module_library_title.pack(anchor="w")
    module_library_body = tk.Label(
        module_library_wrap,
        text=tr("module_library.body"),
        bg=t_init.surface,
        fg=t_init.muted,
        font=("Segoe UI", 10),
        anchor="w",
        justify="left",
        wraplength=720,
    )
    module_library_body.pack(anchor="w", pady=(6, 18))
    module_library_restore_btn = tk.Button(
        module_library_wrap,
        text=tr("module_library.restore_defaults"),
        bg=t_init.btn_secondary,
        fg=t_init.text,
        activebackground=t_init.secondary_hover,
        activeforeground=t_init.text,
        relief="flat",
        font=("Segoe UI", 9, "bold"),
        padx=12,
        pady=8,
        cursor="hand2",
    )
    module_library_restore_btn.pack(anchor="e", pady=(0, 14))
    module_library_cards = tk.Frame(module_library_wrap, bg=t_init.surface)
    module_library_cards.pack(fill="both", expand=True)

    def _draw_plus_canvas() -> None:
        t = th()
        nav_add_canvas.delete("all")
        w = max(120, nav_add_canvas.winfo_width())
        h = max(40, nav_add_canvas.winfo_height())
        nav_add_canvas.configure(bg=t.panel)
        nav_add_canvas.create_rectangle(
            2,
            2,
            w - 2,
            h - 2,
            outline=t.muted,
            dash=(4, 3),
            width=1,
            fill=t.panel,
        )
        nav_add_canvas.create_text(
            w / 2,
            h / 2,
            text="+",
            fill=t.muted,
            font=("Segoe UI", 18, "bold"),
        )

    def _draw_drop_placeholder(canvas: Any, index: int, *, active: bool = False) -> None:
        t = th()
        canvas.delete("all")
        w = max(72, canvas.winfo_width())
        h = max(38, canvas.winfo_height())
        canvas.configure(bg=t.panel)
        outline = t.accent if active else t.muted
        fill = t.field if active else t.panel
        canvas.create_rectangle(
            1,
            1,
            w - 1,
            h - 1,
            outline=outline,
            dash=(4, 3),
            width=1,
            fill=fill,
        )

    def _make_drop_placeholder(index: int) -> Any:
        parent = nav_bottom_frame if _slot_is_bottom(index) else nav_modules_frame
        canvas = tk.Canvas(
            parent,
            height=38 if _slot_is_bottom(index) else 42,
            bg=t_init.panel,
            bd=0,
            highlightthickness=0,
        )
        setattr(canvas, "_sidebar_placeholder_index", index)
        canvas.bind(
            "<Configure>",
            lambda _e, c=canvas, i=index: _draw_drop_placeholder(
                c, i, active=(nav_edit_state.get("placeholder_index") == i)
            ),
            add="+",
        )
        return canvas

    def _make_slot_spacer(index: int) -> Any:
        parent = nav_bottom_frame if _slot_is_bottom(index) else nav_modules_frame
        spacer = tk.Frame(
            parent,
            height=38 if _slot_is_bottom(index) else 42,
            bg=t_init.panel,
            bd=0,
            highlightthickness=0,
        )
        setattr(spacer, "_sidebar_spacer_index", index)
        return spacer

    def _ensure_drop_placeholders(count: int) -> None:
        while len(nav_drop_placeholders) < count:
            nav_drop_placeholders.append(_make_drop_placeholder(len(nav_drop_placeholders)))
        for index, placeholder in enumerate(nav_drop_placeholders):
            if index >= count:
                try:
                    placeholder.grid_forget()
                except tk.TclError:
                    pass

    def _ensure_slot_spacers(count: int) -> None:
        while len(nav_slot_spacers) < count:
            nav_slot_spacers.append(_make_slot_spacer(len(nav_slot_spacers)))
        for index, spacer in enumerate(nav_slot_spacers):
            if index >= count:
                try:
                    spacer.grid_forget()
                except tk.TclError:
                    pass

    def _normal_slot_should_leave_gap(slots: List[str], index: int) -> bool:
        return any(bool(module_id) for module_id in slots[index + 1 :])

    def _sidebar_edit_placeholder_count() -> int:
        try:
            rail_height = int(nav_rail.winfo_height())
        except tk.TclError:
            rail_height = 0
        if rail_height <= 1:
            return 4
        top_rows = len(_top_sidebar_module_ids())
        bottom_rows = 1 if _bottom_sidebar_module_ids() else 0
        used_height = 86 + (top_rows * 52) + (bottom_rows * 48) + 58
        available = max(56, rail_height - used_height)
        return max(2, min(6, available // 50))

    def _hide_drag_preview() -> None:
        win = nav_drag_preview.get("window")
        if win is not None:
            try:
                win.destroy()
            except tk.TclError:
                pass
        nav_drag_preview.update({"window": None, "canvas": None, "module": ""})

    def _draw_drag_preview_canvas(canvas: Any, module_id: str) -> None:
        t = th()
        width = max(96, int(canvas.winfo_width() or 140))
        height = max(38, int(canvas.winfo_height() or 44))
        canvas.delete("all")
        canvas.configure(bg=t.panel)
        canvas.create_rectangle(
            1,
            1,
            width - 1,
            height - 1,
            fill=t.btn_secondary,
            outline=t.accent,
            dash=(4, 3),
            width=1,
        )
        canvas.create_text(
            14,
            height / 2,
            text=_sidebar_module_label(module_id, short=(module_id == "reader")),
            fill=t.text,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
        )

    def _move_drag_preview(
        module_id: str,
        x_root: int,
        y_root: int,
        *,
        snap_index: Optional[int] = None,
    ) -> None:
        if not module_id:
            _hide_drag_preview()
            return
        win = nav_drag_preview.get("window")
        canvas = nav_drag_preview.get("canvas")
        if win is None or canvas is None or nav_drag_preview.get("module") != module_id:
            _hide_drag_preview()
            win = tk.Toplevel(root)
            win.overrideredirect(True)
            try:
                win.attributes("-topmost", True)
            except tk.TclError:
                pass
            canvas = tk.Canvas(
                win,
                width=140,
                height=44,
                bg=th().panel,
                bd=0,
                highlightthickness=0,
            )
            canvas.pack(fill="both", expand=True)
            nav_drag_preview.update({"window": win, "canvas": canvas, "module": module_id})
        width = 140
        height = 44
        target_x = int(x_root - (width / 2))
        target_y = int(y_root - (height / 2))
        if snap_index is not None and 0 <= snap_index < len(nav_drop_placeholders):
            placeholder = nav_drop_placeholders[snap_index]
            try:
                if placeholder.winfo_ismapped():
                    target_x = placeholder.winfo_rootx()
                    target_y = placeholder.winfo_rooty()
                    width = max(96, placeholder.winfo_width())
                    height = max(38, placeholder.winfo_height())
            except tk.TclError:
                pass
        try:
            canvas.configure(width=width, height=height)
            _draw_drag_preview_canvas(canvas, module_id)
            win.geometry(f"{width}x{height}+{target_x}+{target_y}")
        except tk.TclError:
            pass

    def _module_canvas_active(module_id: str) -> bool:
        page = str(sidebar_module_defs.get(module_id, {}).get("page") or "")
        return bool(page and active_page["key"] == page)

    def _draw_module_canvas(module_id: str) -> None:
        data = nav_module_widgets.get(module_id)
        if not data:
            return
        canvas = data["canvas"]
        t = th()
        edit_active = bool(nav_edit_state.get("active"))
        is_active = _module_canvas_active(module_id)
        bg = t.accent if is_active and not edit_active else t.btn_secondary
        fg = "white" if is_active and not edit_active else t.text
        outline = t.accent if edit_active else bg
        dash = (4, 3) if edit_active else None
        w = max(48, canvas.winfo_width())
        h = max(34, canvas.winfo_height())
        canvas.delete("all")
        canvas.configure(bg=t.panel)
        canvas.create_rectangle(
            1,
            1,
            w - 1,
            h - 1,
            fill=bg,
            outline=outline,
            dash=dash,
            width=1,
        )
        label = _sidebar_module_label(module_id, short=(module_id == "reader"))
        text_x = 14
        if edit_active:
            canvas.create_text(
                12,
                h / 2,
                text="⋮⋮",
                fill=t.muted,
                font=("Segoe UI", 8, "bold"),
                anchor="w",
            )
            text_x = 30
        canvas.create_text(
            text_x,
            h / 2,
            text=label,
            fill=fg,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
        )
        if edit_active:
            x0 = max(6, w - 25)
            canvas.create_oval(x0, 9, x0 + 18, 27, fill=t.surface, outline=t.border)
            canvas.create_text(x0 + 9, 18, text="×", fill=t.text, font=("Segoe UI", 10, "bold"))

    def _refresh_sidebar_module_styles() -> None:
        for module_id in list(nav_module_widgets):
            _draw_module_canvas(module_id)
        _draw_plus_canvas()
        active_index = nav_edit_state.get("placeholder_index")
        for index, placeholder in enumerate(nav_drop_placeholders):
            _draw_drop_placeholder(placeholder, index, active=(active_index == index))
        for spacer in nav_slot_spacers:
            try:
                spacer.configure(bg=th().panel)
            except tk.TclError:
                pass

    def _layout_sidebar_modules() -> None:
        for data in nav_module_widgets.values():
            try:
                data["canvas"].grid_forget()
            except tk.TclError:
                pass
        for placeholder in nav_drop_placeholders:
            try:
                placeholder.grid_forget()
            except tk.TclError:
                pass
        for spacer in nav_slot_spacers:
            try:
                spacer.grid_forget()
            except tk.TclError:
                pass
        edit_active = bool(nav_edit_state.get("active"))
        row = 0
        _ensure_drop_placeholders(SIDEBAR_TOTAL_SLOT_COUNT)
        _ensure_slot_spacers(SIDEBAR_TOTAL_SLOT_COUNT)
        top_slots = nav_module_order_state["ids"][:SIDEBAR_TOP_SLOT_COUNT]
        for slot_index, module_id in enumerate(top_slots):
            if module_id:
                data = nav_module_widgets.get(module_id)
                if data:
                    data["canvas"].grid(row=row, column=0, sticky="ew", padx=10, pady=(0, 8))
                    row += 1
            elif edit_active and slot_index < len(nav_drop_placeholders):
                nav_drop_placeholders[slot_index].grid(
                    row=row,
                    column=0,
                    sticky="ew",
                    padx=10,
                    pady=(0, 8),
                )
                row += 1
            elif (
                not edit_active
                and slot_index < len(nav_slot_spacers)
                and _normal_slot_should_leave_gap(top_slots, slot_index)
            ):
                nav_slot_spacers[slot_index].grid(
                    row=row,
                    column=0,
                    sticky="ew",
                    padx=10,
                    pady=(0, 8),
                )
                row += 1
        if edit_active or not _visible_sidebar_module_ids():
            nav_add_canvas.grid(row=99, column=0, sticky="ew", padx=10, pady=(6, 12))
        else:
            nav_add_canvas.grid_forget()
        bottom_slots = _bottom_sidebar_module_ids()
        show_bottom = any(bottom_slots) or edit_active
        if show_bottom:
            nav_bottom_frame.grid(row=101, column=0, sticky="ew", padx=10, pady=(0, 10))
            for col in (0, 1):
                nav_bottom_frame.grid_columnconfigure(col, weight=1, uniform="nav_bottom")
            for bottom_index, module_id in enumerate(bottom_slots[:SIDEBAR_BOTTOM_SLOT_COUNT]):
                slot_index = SIDEBAR_TOP_SLOT_COUNT + bottom_index
                padx = (0, 6) if bottom_index == 0 else (0, 0)
                if module_id:
                    data = nav_module_widgets.get(module_id)
                    if not data:
                        continue
                    canvas = data["canvas"]
                    canvas.configure(height=38)
                    canvas.grid(row=0, column=bottom_index, sticky="ew", padx=padx)
                elif edit_active and slot_index < len(nav_drop_placeholders):
                    nav_drop_placeholders[slot_index].grid(
                        row=0,
                        column=bottom_index,
                        sticky="ew",
                        padx=padx,
                    )
                elif (
                    not edit_active
                    and slot_index < len(nav_slot_spacers)
                    and _normal_slot_should_leave_gap(bottom_slots, bottom_index)
                ):
                    nav_slot_spacers[slot_index].grid(
                        row=0,
                        column=bottom_index,
                        sticky="ew",
                        padx=padx,
                    )
        else:
            nav_bottom_frame.grid_forget()
        _refresh_sidebar_module_styles()

    def _target_sidebar_module_from_pointer(x_root: int, y_root: int) -> str:
        try:
            widget = root.winfo_containing(x_root, y_root)
        except tk.TclError:
            return ""
        while widget is not None:
            module_id = getattr(widget, "_sidebar_module_id", "")
            if module_id:
                return str(module_id)
            widget = getattr(widget, "master", None)
        return ""

    def _drop_index_from_root_y(y_root: int) -> int:
        order = _top_sidebar_module_ids()
        for index, module_id in enumerate(order):
            data = nav_module_widgets.get(module_id)
            if not data:
                continue
            canvas = data["canvas"]
            try:
                mid_y = canvas.winfo_rooty() + (canvas.winfo_height() / 2)
            except tk.TclError:
                continue
            if y_root < mid_y:
                return index
        return len(order)

    def _placeholder_index_from_pointer(x_root: int, y_root: int) -> Optional[int]:
        for index, placeholder in enumerate(nav_drop_placeholders):
            try:
                if not placeholder.winfo_ismapped():
                    continue
                left = int(placeholder.winfo_rootx())
                top = int(placeholder.winfo_rooty())
                width = int(placeholder.winfo_width())
                height = int(placeholder.winfo_height())
            except tk.TclError:
                continue
            margin = 12
            if (
                left - margin <= x_root <= left + width + margin
                and top - margin <= y_root <= top + height + margin
            ):
                return index
        return None

    def _insert_index_for_placeholder(module_id: str, placeholder_index: int) -> int:
        return max(0, min(placeholder_index, SIDEBAR_TOTAL_SLOT_COUNT - 1))

    def _first_empty_sidebar_slot() -> Optional[int]:
        for index, module_id in enumerate(nav_module_order_state["ids"]):
            if not module_id:
                return index
        return None

    def _target_insert_index_from_pointer(module_id: str, x_root: int, y_root: int) -> Optional[int]:
        target_module = _target_sidebar_module_from_pointer(x_root, y_root)
        if not target_module or target_module == module_id:
            return None
        return _module_slot_index(target_module)

    def _sidebar_close_hit(canvas: Any, event: Any) -> bool:
        try:
            x = int(getattr(event, "x", 0))
            y = int(getattr(event, "y", 0))
            width = int(canvas.winfo_width())
            height = int(canvas.winfo_height())
        except (tk.TclError, TypeError, ValueError):
            return False
        close_left = max(0, width - 32)
        close_top = 4
        close_bottom = min(height, 32)
        return close_left <= x <= width and close_top <= y <= close_bottom

    def _pointer_inside_sidebar(x_root: int, y_root: int) -> bool:
        try:
            left = nav_rail.winfo_rootx()
            top = nav_rail.winfo_rooty()
            return (
                left <= x_root <= left + nav_rail.winfo_width()
                and top <= y_root <= top + nav_rail.winfo_height()
            )
        except tk.TclError:
            return False

    def _enter_sidebar_edit_mode(module_id: str = "") -> None:
        nav_edit_state["active"] = True
        nav_edit_state["long_hold"] = True
        nav_edit_state["drag_module"] = module_id
        nav_edit_state["placeholder_index"] = None
        _layout_sidebar_modules()

    def _exit_sidebar_edit_mode() -> None:
        _hide_drag_preview()
        after_id = nav_edit_state.get("press_after")
        if after_id is not None:
            try:
                root.after_cancel(after_id)
            except Exception:
                pass
        nav_edit_state.update(
            {
                "active": False,
                "press_module": "",
                "press_after": None,
                "long_hold": False,
                "dragging": False,
                "drag_module": "",
                "placeholder_index": None,
                "library_drag_module": "",
            }
        )
        _layout_sidebar_modules()

    def _show_first_sidebar_page_or_library() -> None:
        for module_id in _visible_sidebar_module_ids():
            page = str(sidebar_module_defs.get(module_id, {}).get("page") or "")
            if page:
                show_page(page)
                return
        show_page("module_library")

    def _show_module_library() -> None:
        _render_module_library()
        show_page("module_library")

    def _set_sidebar_order(order: List[str], *, persist: bool = True) -> None:
        cleaned = _normalize_sidebar_module_order(order)
        nav_module_order_state["ids"] = cleaned
        if persist:
            _save_sidebar_module_order()
        _rebuild_sidebar_module_widgets()
        _render_module_library()
        if not _visible_sidebar_module_ids() and active_page["key"] != "module_library":
            show_page("module_library")

    def _remove_sidebar_module(module_id: str) -> None:
        if module_id not in nav_module_order_state["ids"]:
            return
        removed_page = str(sidebar_module_defs.get(module_id, {}).get("page") or "")
        slots = list(nav_module_order_state["ids"])
        for index, slot_module_id in enumerate(slots):
            if slot_module_id == module_id:
                slots[index] = ""
        _set_sidebar_order(slots)
        if removed_page and active_page["key"] == removed_page:
            _show_first_sidebar_page_or_library()

    def _move_sidebar_module(module_id: str, index: int) -> None:
        slots = list(nav_module_order_state["ids"])
        if not slots:
            slots = [""] * SIDEBAR_TOTAL_SLOT_COUNT
        index = max(0, min(index, SIDEBAR_TOTAL_SLOT_COUNT - 1))
        source_index = _module_slot_index(module_id)
        if source_index == index:
            _layout_sidebar_modules()
            return
        target_module = slots[index] if index < len(slots) else ""
        if source_index is not None and source_index < len(slots):
            slots[source_index] = target_module if target_module != module_id else ""
        slots[index] = module_id
        _set_sidebar_order(slots)

    def _drop_library_module(module_id: str, *, index: Optional[int] = None, replace_id: str = "") -> None:
        if module_id not in sidebar_module_defs:
            return
        slots = list(nav_module_order_state["ids"])
        for slot_index, slot_module_id in enumerate(slots):
            if slot_module_id == module_id:
                slots[slot_index] = ""
        target_index = _module_slot_index(replace_id) if replace_id else index
        if target_index is None:
            target_index = _first_empty_sidebar_slot()
        if target_index is None:
            target_index = SIDEBAR_TOTAL_SLOT_COUNT - 1
        target_index = max(0, min(target_index, SIDEBAR_TOTAL_SLOT_COUNT - 1))
        slots[target_index] = module_id
        _set_sidebar_order(slots)

    def _handle_sidebar_module_action(module_id: str) -> None:
        if nav_edit_state.get("active"):
            return
        info = sidebar_module_defs.get(module_id, {})
        page = str(info.get("page") or "")
        if page:
            show_page(page)
            return
        action = info.get("action")
        if callable(action):
            action()

    def _cancel_sidebar_long_hold() -> None:
        after_id = nav_edit_state.get("press_after")
        if after_id is not None:
            try:
                root.after_cancel(after_id)
            except Exception:
                pass
        nav_edit_state["press_after"] = None

    def _on_sidebar_module_press(module_id: str, event: Any) -> str:
        nav_edit_state["press_module"] = module_id
        nav_edit_state["long_hold"] = False
        nav_edit_state["dragging"] = False
        nav_edit_state["drag_module"] = module_id
        nav_edit_state["press_y_root"] = int(getattr(event, "y_root", 0))
        _cancel_sidebar_long_hold()
        nav_edit_state["press_after"] = root.after(
            600, lambda mid=module_id: _enter_sidebar_edit_mode(mid)
        )
        return "break"

    def _on_sidebar_module_motion(module_id: str, event: Any) -> str:
        if not nav_edit_state.get("active"):
            return "break"
        was_dragging = bool(nav_edit_state.get("dragging"))
        nav_edit_state["dragging"] = True
        nav_edit_state["drag_module"] = module_id
        x_root = int(getattr(event, "x_root", 0))
        y_root = int(getattr(event, "y_root", 0))
        index = _placeholder_index_from_pointer(x_root, y_root)
        if (not was_dragging) or nav_edit_state.get("placeholder_index") != index:
            nav_edit_state["placeholder_index"] = index
            _layout_sidebar_modules()
        _move_drag_preview(
            module_id,
            x_root,
            y_root,
            snap_index=index if isinstance(index, int) else None,
        )
        return "break"

    def _on_sidebar_module_release(module_id: str, event: Any) -> str:
        _cancel_sidebar_long_hold()
        canvas = nav_module_widgets.get(module_id, {}).get("canvas")
        if nav_edit_state.get("active"):
            if canvas is not None and _sidebar_close_hit(canvas, event):
                _hide_drag_preview()
                _remove_sidebar_module(module_id)
                return "break"
            if nav_edit_state.get("dragging"):
                _hide_drag_preview()
                index = nav_edit_state.get("placeholder_index")
                nav_edit_state["placeholder_index"] = None
                if isinstance(index, int):
                    _move_sidebar_module(module_id, _insert_index_for_placeholder(module_id, index))
                else:
                    x_root = int(getattr(event, "x_root", 0))
                    y_root = int(getattr(event, "y_root", 0))
                    target_index = _target_insert_index_from_pointer(module_id, x_root, y_root)
                    if isinstance(target_index, int):
                        _move_sidebar_module(module_id, target_index)
                    else:
                        _layout_sidebar_modules()
                return "break"
            _hide_drag_preview()
            _refresh_sidebar_module_styles()
            return "break"
        if not nav_edit_state.get("long_hold"):
            _handle_sidebar_module_action(module_id)
        return "break"

    def _make_sidebar_module_canvas(module_id: str) -> Any:
        parent = nav_bottom_frame if _sidebar_is_bottom_module(module_id) else nav_modules_frame
        canvas = tk.Canvas(
            parent,
            height=44,
            bg=t_init.panel,
            bd=0,
            highlightthickness=0,
            cursor="hand2",
        )
        setattr(canvas, "_sidebar_module_id", module_id)
        canvas.bind("<ButtonPress-1>", lambda e, mid=module_id: _on_sidebar_module_press(mid, e), add="+")
        canvas.bind("<B1-Motion>", lambda e, mid=module_id: _on_sidebar_module_motion(mid, e), add="+")
        canvas.bind("<ButtonRelease-1>", lambda e, mid=module_id: _on_sidebar_module_release(mid, e), add="+")
        canvas.bind("<Configure>", lambda _e, mid=module_id: _draw_module_canvas(mid), add="+")
        return canvas

    def _rebuild_sidebar_module_widgets() -> None:
        for data in list(nav_module_widgets.values()):
            try:
                data["canvas"].destroy()
            except tk.TclError:
                pass
        nav_module_widgets.clear()
        nav_buttons.clear()
        for module_id in nav_module_order_state["ids"]:
            if not module_id:
                continue
            canvas = _make_sidebar_module_canvas(module_id)
            nav_module_widgets[module_id] = {"canvas": canvas}
            nav_buttons[module_id] = canvas
        _layout_sidebar_modules()

    def _on_plus_clicked(_event: Optional[Any] = None) -> str:
        if not nav_edit_state.get("active"):
            nav_edit_state["active"] = True
            nav_edit_state["long_hold"] = True
        nav_edit_state["placeholder_index"] = None
        _layout_sidebar_modules()
        _show_module_library()
        return "break"

    def _library_card_bind_drag(widget: Any, module_id: str) -> None:
        def _press(event: Any, mid: str = module_id) -> str:
            nav_edit_state["library_drag_module"] = mid
            nav_edit_state["active"] = True
            nav_edit_state["placeholder_index"] = None
            _refresh_sidebar_module_styles()
            _move_drag_preview(mid, int(event.x_root), int(event.y_root))
            return "break"

        def _motion(event: Any, mid: str = module_id) -> str:
            if not nav_edit_state.get("library_drag_module"):
                return "break"
            nav_edit_state["placeholder_index"] = _placeholder_index_from_pointer(
                int(event.x_root),
                int(event.y_root),
            )
            _layout_sidebar_modules()
            index = nav_edit_state.get("placeholder_index")
            _move_drag_preview(
                mid,
                int(event.x_root),
                int(event.y_root),
                snap_index=index if isinstance(index, int) else None,
            )
            return "break"

        def _release(event: Any, mid: str = module_id) -> str:
            _hide_drag_preview()
            replace_id = _target_sidebar_module_from_pointer(int(event.x_root), int(event.y_root))
            drop_index = nav_edit_state.get("placeholder_index")
            nav_edit_state["library_drag_module"] = ""
            nav_edit_state["placeholder_index"] = None
            if replace_id:
                target_index = _target_insert_index_from_pointer(mid, int(event.x_root), int(event.y_root))
                _drop_library_module(mid, index=target_index if isinstance(target_index, int) else None)
            elif isinstance(drop_index, int):
                _drop_library_module(mid, index=_insert_index_for_placeholder(mid, drop_index))
            else:
                _layout_sidebar_modules()
                _render_module_library()
            return "break"

        widget.bind("<ButtonPress-1>", _press, add="+")
        widget.bind("<B1-Motion>", _motion, add="+")
        widget.bind("<ButtonRelease-1>", _release, add="+")

    def _render_module_library() -> None:
        for child in list(module_library_cards.winfo_children()):
            child.destroy()
        t = th()
        visible = set(_visible_sidebar_module_ids())
        module_library_wrap.configure(bg=t.surface)
        module_library_cards.configure(bg=t.surface)
        module_library_title.configure(bg=t.surface, fg=t.text, text=tr("module_library.title"))
        module_library_body.configure(bg=t.surface, fg=t.muted, text=tr("module_library.body"))
        module_library_restore_btn.configure(
            text=tr("module_library.restore_defaults"),
            bg=t.btn_secondary,
            fg=t.text,
            activebackground=t.secondary_hover,
            activeforeground=t.text,
        )
        for index, module_id in enumerate(sidebar_all_module_ids):
            card = tk.Frame(
                module_library_cards,
                bg=t.panel,
                highlightthickness=1,
                highlightbackground=t.border,
                bd=0,
            )
            card.grid(row=index // 2, column=index % 2, sticky="ew", padx=(0, 12), pady=(0, 12))
            module_library_cards.grid_columnconfigure(index % 2, weight=1)
            setattr(card, "_library_module_id", module_id)
            _library_card_bind_drag(card, module_id)
            title = tk.Label(
                card,
                text=_sidebar_module_label(module_id),
                bg=t.panel,
                fg=t.text,
                font=("Segoe UI", 11, "bold"),
                anchor="w",
            )
            title.grid(row=0, column=0, sticky="ew", padx=12, pady=(10, 2))
            status_key = "module_library.visible" if module_id in visible else "module_library.hidden"
            status = tk.Label(
                card,
                text=tr(status_key),
                bg=t.panel,
                fg=t.muted,
                font=("Segoe UI", 9),
                anchor="w",
            )
            status.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
            add_btn = tk.Button(
                card,
                text=tr("module_library.add"),
                command=lambda mid=module_id: _drop_library_module(mid),
                bg=t.btn_secondary,
                fg=t.text,
                activebackground=t.secondary_hover,
                activeforeground=t.text,
                relief="flat",
                font=("Segoe UI", 9, "bold"),
                padx=10,
                pady=6,
                cursor="hand2",
            )
            add_btn.grid(row=0, column=1, rowspan=2, padx=(0, 12), pady=10)
            card.grid_columnconfigure(0, weight=1)
            if module_id in visible:
                add_btn.config(state="disabled", cursor="arrow")
            for widget in (title, status):
                _library_card_bind_drag(widget, module_id)

    def _restore_default_sidebar_modules() -> None:
        _exit_sidebar_edit_mode()
        _set_sidebar_order(list(sidebar_all_module_ids))
        _show_first_sidebar_page_or_library()

    nav_add_canvas.bind("<Button-1>", _on_plus_clicked, add="+")
    nav_add_canvas.bind("<Configure>", lambda _e: _draw_plus_canvas(), add="+")
    module_library_restore_btn.config(command=_restore_default_sidebar_modules)
    nav_module_order_state["ids"] = _load_sidebar_module_order()
    _rebuild_sidebar_module_widgets()

    nav_summon_btn.config(command=lambda: set_nav_visible(True))
    def _on_content_host_configure(_e: Optional[Any] = None) -> None:
        if nav_collapsed["value"] and not nav_animating["value"]:
            _place_nav_summon()
        frame = active_page_frame.get("frame")
        if frame is not None:
            _layout_console_row(frame)

    content_host.bind("<Configure>", _on_content_host_configure, add="+")
    bind_button_hover_if_enabled(
        nav_summon_btn,
        lambda: (
            "normal",
            th().toolbar_btn_config()[0],
            th().toolbar_btn_config()[1],
            th().toolbar_btn_config()[2],
            th().toolbar_btn_config()[3],
        ),
        lambda: th().toolbar_hover()[0],
        lambda: th().toolbar_hover()[1],
    )

    console_entry.bind("<Return>", lambda _e: (run_console_command(), "break")[1], add="+")
    console_entry.bind("<Up>", _console_history_up, add="+")
    console_entry.bind("<Down>", _console_history_down, add="+")
    console_entry.bind("<Tab>", _console_tab_complete, add="+")
    console_entry.bind("<KeyPress>", _console_entry_keypress, add="+")
    console_entry.bind("<FocusIn>", _console_entry_focus_in, add="+")
    console_entry.bind("<FocusOut>", _console_entry_focus_out, add="+")
    console_prompt.bind("<Button-1>", lambda _e: run_console_command(), add="+")

    def _unfocus_console_on_button_click(evt: Optional[Any] = None) -> None:
        if evt is None:
            return
        w = getattr(evt, "widget", None)
        if nav_edit_state.get("active"):
            current = w
            inside_nav = False
            inside_library = False
            while current is not None:
                if current is nav_rail:
                    inside_nav = True
                    break
                if current is module_library_page:
                    inside_library = True
                    break
                current = getattr(current, "master", None)
            if inside_nav:
                target_module = _target_sidebar_module_from_pointer(
                    int(getattr(evt, "x_root", 0)),
                    int(getattr(evt, "y_root", 0)),
                )
                is_placeholder = any(w is placeholder for placeholder in nav_drop_placeholders)
                if not target_module and w is not nav_add_canvas and not is_placeholder:
                    _exit_sidebar_edit_mode()
            elif not inside_library:
                _exit_sidebar_edit_mode()
        if isinstance(w, tk.Button) and w is not console_entry:
            root.focus_set()

    root.bind_all("<Button-1>", _unfocus_console_on_button_click, add="+")
    root.bind("<F11>", _on_toggle_fullscreen, add="+")
    _startup_step("splash.detail.finalize")
    if "journal" in nav_module_order_state["ids"]:
        show_page("journal")
    else:
        _show_first_sidebar_page_or_library()

    def _on_escape(event=None) -> None:
        if nav_edit_state.get("active"):
            _exit_sidebar_edit_mode()
            return
        if str(find_row.winfo_manager()) == "pack":
            _find_close()
            return
        on_close(event)

    root.bind("<Escape>", _on_escape)
    root.protocol("WM_DELETE_WINDOW", on_close)
    _startup_step("splash.detail.journal_ready")
    startup_overlay.destroy()
    root.lift()
    apply_journal_window_colors()

    def _background_post_init() -> None:
        _startup_step("splash.detail.other_pages")
        set_nav_visible(True)
        _startup_step("splash.detail.autosave")
        autosave()
        refresh_save_entry_state()
        _startup_step("splash.detail.ready")
        if start_auto_backup:
            start_daily_auto_backup_in_background()

    root.after(1, _background_post_init)
    root.mainloop()
    return saved["value"]


def maybe_prompt_startup_on_first_run() -> None:
    prefs = load_preferences()
    if prefs.get("startup_prompt_done", "").lower() == "true":
        return
    print("Open logger automatically when computer starts? (y/N): ", end="")
    try:
        answer = input().strip().lower()
    except (EOFError, RuntimeError):
        # Windowed launches (pythonw / console=False EXE) may not provide stdin.
        prefs["startup_enabled"] = "true" if is_startup_enabled() else "false"
        prefs["startup_prompt_done"] = "true"
        save_preferences(prefs)
        return
    if answer in ("y", "yes"):
        if create_startup_shortcut():
            print("Startup enabled.")
            prefs["startup_enabled"] = "true"
        else:
            print("Could not enable startup shortcut.")
            prefs["startup_enabled"] = "false"
    else:
        prefs["startup_enabled"] = "false"
        print("Startup remains disabled.")
    prefs["startup_prompt_done"] = "true"
    if not save_preferences(prefs):
        print("Warning: could not save startup preference.")


def setup_first_time_preferences() -> str:
    prefs = load_preferences()
    if prefs.get("initial_setup_done", "").lower() == "true":
        app_name = prefs.get("app_name", "").strip()
        if app_name:
            return app_name
        return get_or_create_app_name()

    print("First time setup: use default settings? (y/N): ", end="")
    try:
        answer = input().strip().lower()
    except (EOFError, RuntimeError):
        # Non-interactive/windowed launch: apply safe defaults without prompting.
        app_name = prefs.get("app_name", "").strip() or "Daily Logger"
        prefs["app_name"] = app_name
        prefs["startup_enabled"] = "true" if is_startup_enabled() else "false"
        prefs["startup_prompt_done"] = "true"
        prefs["initial_setup_done"] = "true"
        if not save_preferences(prefs):
            print("Warning: could not save initial preferences.")
        return app_name
    use_default = answer in ("y", "yes")

    if use_default:
        prefs["app_name"] = "Daily Logger"
        if create_startup_shortcut():
            prefs["startup_enabled"] = "true"
            print("Default setup applied: startup enabled.")
        else:
            prefs["startup_enabled"] = "false"
            print("Default setup applied: could not enable startup.")
        prefs["startup_prompt_done"] = "true"
    else:
        prefs["app_name"] = prompt_for_app_name()
        print("Open logger automatically when computer starts? (y/N): ", end="")
        try:
            startup_answer = input().strip().lower()
        except (EOFError, RuntimeError):
            startup_answer = ""
        if startup_answer in ("y", "yes"):
            if create_startup_shortcut():
                prefs["startup_enabled"] = "true"
                print("Startup enabled.")
            else:
                prefs["startup_enabled"] = "false"
                print("Could not enable startup shortcut.")
        else:
            prefs["startup_enabled"] = "false"
            print("Startup remains disabled.")
        prefs["startup_prompt_done"] = "true"

    prefs["initial_setup_done"] = "true"
    if not save_preferences(prefs):
        print("Warning: could not save initial preferences.")
    return prefs.get("app_name", "Daily Logger")


def get_chat_completions_url() -> str:
    return os.getenv("OPENAI_CHAT_COMPLETIONS_URL", OPENAI_CHAT_COMPLETIONS_URL).strip()


def _image_mime_for_path(path: Path) -> str:
    suf = path.suffix.lower()
    if suf in (".jpg", ".jpeg"):
        return "image/jpeg"
    if suf == ".gif":
        return "image/gif"
    if suf == ".webp":
        return "image/webp"
    return "image/png"


def build_user_message_with_attachments(
    question: str,
    image_paths: List[Path],
    file_paths: List[Path],
) -> Dict[str, object]:
    """Build a user message with optional images (vision) and text file excerpts."""
    text_chunks: List[str] = []
    q = (question or "").strip()
    if q:
        text_chunks.append(q)
    for fp in file_paths:
        ctx, _resolved, err = load_recap_context_from_file(str(fp))
        if err:
            text_chunks.append(f"[Attachment {fp.name}: {err}]")
        elif ctx:
            label = str(fp.resolve())
            clip = ctx if len(ctx) <= 48000 else ctx[:48000] + "\n\n[Truncated attachment]"
            text_chunks.append(f"[Attached file: {label}]\n{clip}")
    combined_text = "\n\n".join(text_chunks).strip() or "(no text)"
    parts: List[Dict[str, object]] = [{"type": "text", "text": combined_text}]
    for img_path in image_paths:
        try:
            raw = img_path.read_bytes()
        except OSError as exc:
            parts.append({"type": "text", "text": f"[Image {img_path.name} unreadable: {exc}]"})
            continue
        try:
            mime = _image_mime_for_path(img_path)
            b64 = base64.b64encode(raw).decode("ascii")
            parts.append(
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime};base64,{b64}"},
                }
            )
        except Exception as exc:
            parts.append(
                {"type": "text", "text": f"[Image {img_path.name} could not attach: {exc}]"}
            )

    if len(parts) == 1 and parts[0].get("type") == "text":
        return {"role": "user", "content": combined_text}
    return {"role": "user", "content": parts}


def chat_completion(
    messages: List[Dict[str, object]],
    model: str = OPENAI_MODEL,
    reasoning_effort: Optional[str] = None,
    *,
    progress: Optional[Callable[[str], None]] = None,
    timeout_sec: float = 60,
    attempts: int = 3,
) -> str:
    api_key = get_openai_api_key()
    if not api_key:
        return "OPENAI_API_KEY is not set. Set it, then try again."

    def _notify(message: str) -> None:
        if progress is None:
            return
        try:
            progress(message)
        except Exception:
            pass

    _notify("Preparing request")
    chat_url = get_chat_completions_url()
    payload_data: Dict[str, object] = {
        "model": model,
        "messages": messages,
    }
    if reasoning_effort:
        payload_data["reasoning_effort"] = reasoning_effort
    payload = json.dumps(payload_data).encode("utf-8")

    body = None
    last_exception: Optional[Exception] = None
    attempt_count = max(1, int(attempts))
    for attempt in range(attempt_count):
        req = request.Request(
            chat_url,
            data=payload,
            method="POST",
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            },
        )
        try:
            _notify(f"Contacting AI ({attempt + 1}/{attempt_count})")
            with request.urlopen(req, timeout=timeout_sec) as response:
                _notify("Reading response")
                body = response.read().decode("utf-8")
                break
        except error.HTTPError as exc:
            try:
                details = exc.read().decode("utf-8")
            except Exception:
                details = str(exc)
            return f"ChatGPT API error ({exc.code}): {details}"
        except Exception as exc:
            last_exception = exc
            if attempt < attempt_count - 1:
                delay = 2 * (attempt + 1)
                _notify(f"Connection issue, retrying in {delay}s ({attempt + 2}/{attempt_count})")
                time.sleep(delay)
            continue

    if body is None:
        return (
            "Failed to contact ChatGPT API after retries. "
            f"Last error: {last_exception}. "
            f"URL: {chat_url}. "
            "Check internet, firewall, VPN, or proxy settings."
        )

    try:
        _notify("Formatting report")
        parsed = json.loads(body)
        return parsed["choices"][0]["message"]["content"].strip()
    except Exception:
        return "ChatGPT returned an unexpected response format."


def chat_completion_with_spinner(
    messages: List[Dict[str, object]],
    model: str = OPENAI_MODEL,
    reasoning_effort: Optional[str] = None,
) -> str:
    holder: Dict[str, str] = {}

    def worker() -> None:
        holder["response"] = chat_completion(messages, model=model, reasoning_effort=reasoning_effort)

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    spinner = ["-", "\\", "|", "/"]
    spinner_colors = [
        "\033[31m",  # red
        "\033[33m",  # yellow
        "\033[32m",  # green
        "\033[36m",  # cyan
        "\033[34m",  # blue
        "\033[35m",  # magenta
    ]
    color_index = 0
    color_enabled = False
    index = 0
    cancelled = False
    start_time = time.time()
    print("AI is thinking... (press Enter to cancel)", end="", flush=True)
    while thread.is_alive():
        if msvcrt is not None and msvcrt.kbhit():
            char = msvcrt.getwch()
            if char in ("\r", "\n"):
                cancelled = True
                break
            if char == " ":
                color_enabled = True
                color_index = (color_index + 1) % len(spinner_colors)
        elapsed = time.time() - start_time
        spinner_char = spinner[index % len(spinner)]
        if color_enabled:
            spinner_char = f"{spinner_colors[color_index]}{spinner_char}\033[0m"
        sys.stdout.write(
            f"\rAI is thinking... {spinner_char} ({elapsed:.1f}s, press Enter to cancel)"
        )
        sys.stdout.flush()
        index += 1
        time.sleep(0.12)

    sys.stdout.write("\r" + (" " * 80) + "\r")
    sys.stdout.flush()
    if cancelled:
        return "Response cancelled by user."
    thread.join()
    return holder.get("response", "No response received.")


def print_chat_help() -> None:
    print("Chat commands:")
    print("  help - show this help")
    print("  ts   - take screenshot and attach to next AI message")
    print("  rs   - remove pending screenshot attachment")
    print("  Tab  - complete help / ts / rs; empty line + Tab shows this help")
    print("  Enter on empty line - exit chat")


def take_chat_screenshot_hidden_console() -> Optional[Path]:
    SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = SCREENSHOT_DIR / f"chat_{timestamp}.png"
    console_hwnd = None
    try:
        console_hwnd = ctypes.windll.kernel32.GetConsoleWindow()
    except Exception:
        console_hwnd = None

    if console_hwnd:
        try:
            ctypes.windll.user32.ShowWindow(console_hwnd, 0)  # SW_HIDE
            time.sleep(0.6)
        except Exception:
            console_hwnd = None

    try:
        import mss.tools
        from mss import MSS

        with MSS() as sct:
            monitor = sct.monitors[0]
            shot = sct.grab(monitor)
            mss.tools.to_png(shot.rgb, shot.size, output=str(output_path))
        return output_path
    except Exception as exc:
        print(f"Could not capture screenshot: {exc}")
        return None
    finally:
        if console_hwnd:
            try:
                ctypes.windll.user32.ShowWindow(console_hwnd, 5)  # SW_SHOW
            except Exception:
                pass


def build_user_message(question: str, screenshot_path: Optional[Path]) -> Dict[str, object]:
    if screenshot_path is None:
        return {"role": "user", "content": question}
    try:
        image_b64 = base64.b64encode(screenshot_path.read_bytes()).decode("ascii")
    except OSError as exc:
        print(f"Could not read screenshot for attachment: {exc}")
        return {"role": "user", "content": question}
    return {
        "role": "user",
        "content": [
            {"type": "text", "text": question},
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{image_b64}"}},
        ],
    }


def run_chat_mode(
    with_journal_context: bool,
    use_thinking_model: bool = False,
    recap_date_range: Optional[Tuple[datetime, datetime]] = None,
    recap_context_override: Optional[str] = None,
    recap_context_label: Optional[str] = None,
) -> None:
    base_mode_label = "Recap" if with_journal_context else "Chatbot"
    if use_thinking_model and with_journal_context:
        base_mode_label = "Recap (Thinking)"
    if use_thinking_model and not with_journal_context:
        base_mode_label = "Chatbot(Thinking)"
    maybe_warn_for_current_wifi()
    if not ensure_openai_api_key_for_ai():
        return
    model_name = OPENAI_THINKING_MODEL if use_thinking_model else OPENAI_MODEL

    system_message = "You are a helpful assistant."
    if with_journal_context:
        if recap_context_override is not None:
            journal_context = recap_context_override
            if recap_context_label:
                print(f"Recap source: {recap_context_label}")
        else:
            journal_context = build_journal_context_for_range(recap_date_range)
        if recap_context_override is None and recap_date_range is not None:
            included_dates = list_journal_dates_in_range(recap_date_range)
            if included_dates:
                print("Recap includes journal dates:")
                print("  " + ", ".join(included_dates))
            else:
                print("Recap includes journal dates: (none found in selected range)")
        system_message = (
            "You answer questions only using the user's journal context. "
            "If the answer is not in the journal, say you do not know based on the journal."
        )
        context_message = f"Journal context:\n{journal_context}"
        messages: List[Dict[str, object]] = [
            {"role": "system", "content": system_message},
            {"role": "system", "content": context_message},
        ]
    else:
        messages = [{"role": "system", "content": system_message}]

    pending_screenshot: Optional[Path] = None

    def format_mode_label() -> str:
        if pending_screenshot is None:
            return base_mode_label
        if base_mode_label == "Chatbot(Thinking)":
            return "Chatbot(Thinking, Screenshot Attached)"
        if base_mode_label == "Chatbot":
            return "Chatbot (Screenshot Attached)"
        return base_mode_label

    print(f"\n=== {format_mode_label()} ===")
    if not with_journal_context:
        print('GPT: Hello, how can I help you? If you are stuck type "help"')
    while True:
        question_prompt = "Recap: " if with_journal_context else "You: "
        question = input_line_with_tab_completions(
            question_prompt, CHAT_LINE_COMPLETIONS, on_empty_tab=print_chat_help
        )
        if is_enter_equivalent(question):
            print(f"Leaving {base_mode_label}.")
            return
        if question.lower() == "help":
            print_chat_help()
            continue
        if not with_journal_context and question.lower() == "ts":
            print("Taking screenshot...")
            pending_screenshot = take_chat_screenshot_hidden_console()
            if pending_screenshot:
                print(f"Screenshot ready: {pending_screenshot}")
                print(f"=== {format_mode_label()} ===")
            continue
        if not with_journal_context and question.lower() == "rs":
            if pending_screenshot is None:
                print("No pending screenshot to remove.")
            else:
                pending_screenshot = None
                print("Pending screenshot removed.")
                print(f"=== {format_mode_label()} ===")
            continue

        had_screenshot = pending_screenshot is not None
        user_message = build_user_message(question, pending_screenshot)
        messages.append(user_message)
        pending_screenshot = None
        effort = "high" if use_thinking_model else None
        answer = chat_completion_with_spinner(messages, model=model_name, reasoning_effort=effort)
        print(f"GPT: {answer}")
        if had_screenshot:
            print(f"=== {format_mode_label()} ===")
        messages.append({"role": "assistant", "content": answer})


def is_row_empty(values: List[object]) -> bool:
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def find_first_empty_data_row(ws, column_count: int) -> int:
    for row_index in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_index, column=col).value for col in range(1, column_count + 1)]
        if is_row_empty(values):
            return row_index
    return ws.max_row + 1


def ask_entry_date_time() -> Optional[Tuple[str, str]]:
    now = datetime.now()
    default_date = now.strftime("%m/%d/%Y")

    while True:
        date_input = input(
            f"Entry date (mm/dd/yyyy, Enter for today {default_date}): "
        ).strip()
        if not date_input:
            date_value = default_date
            break
        if date_input.upper() == "X":
            confirm = input("Return to main menu? (y/N): ").strip().lower()
            if confirm in ("y", "yes"):
                return None
            continue
        parsed_date = parse_flexible_date(date_input, now.year)
        if parsed_date:
            date_value = parsed_date.strftime("%m/%d/%Y")
            break
        print("Invalid date. Try 04/20/2026, 4/26, Apr 26, or April 26.")

    while True:
        time_input = input(
            "Entry time (example: 11:00AM, type rn for now, Enter for N/A): "
        ).strip()
        if not time_input:
            time_value = "N/A"
            break
        if time_input.upper() == "X":
            confirm = input("Return to main menu? (y/N): ").strip().lower()
            if confirm in ("y", "yes"):
                return None
            continue
        if time_input.lower() == "rn":
            time_value = datetime.now().strftime("%I:%M%p").lstrip("0")
            break

        normalized = time_input.upper().replace(" ", "")
        try:
            parsed = datetime.strptime(normalized, "%I:%M%p")
            time_value = parsed.strftime("%I:%M%p").lstrip("0")
            break
        except ValueError:
            print("Invalid time format. Use hh:mmAM/PM (example: 2:03PM), or rn for current time.")

    return date_value, time_value


def ask_entry_date_time_gui(parent: Optional[Any] = None) -> Optional[Tuple[str, str]]:
    """
    GUI version of ``ask_entry_date_time()``.
    The CLI version uses ``input()`` which freezes the Tk main thread in the GUI console.
    """
    if tk is None:
        return ask_entry_date_time()

    _parent = parent if parent is not None else tk._default_root  # type: ignore[attr-defined]
    if _parent is None:
        # Last-resort fallback: no Tk root to attach to.
        return ask_entry_date_time()

    now = datetime.now()
    default_date = now.strftime("%m/%d/%Y")

    dlg = tk.Toplevel(_parent)
    dlg.title("Entry date & time")
    dlg.transient(_parent)

    # Make modal without blocking background threads; Tk will run its own nested event loop.
    result: Dict[str, Optional[Tuple[str, str]]] = {"v": None}
    dlg.grab_set()

    tk.Label(dlg, text=f"Entry date (mm/dd/yyyy, Enter for today {default_date}):").pack(
        padx=12, pady=(12, 4)
    )
    date_var = tk.StringVar(value=default_date)
    date_entry = tk.Entry(dlg, textvariable=date_var, width=22)
    date_entry.pack(padx=12, pady=(0, 8))

    tk.Label(dlg, text="Entry time (example: 11:00AM, type rn for now, Enter for N/A):").pack(
        padx=12, pady=(0, 4)
    )
    time_var = tk.StringVar(value="")
    time_entry = tk.Entry(dlg, textvariable=time_var, width=22)
    time_entry.pack(padx=12, pady=(0, 8))

    def _parse_date(date_input: str) -> Optional[str]:
        di = date_input.strip()
        if not di:
            return default_date
        if di.upper() == "X":
            return None
        parsed_date = parse_flexible_date(di, now.year)
        if parsed_date:
            return parsed_date.strftime("%m/%d/%Y")
        return None

    def _parse_time(time_input: str) -> Optional[str]:
        ti = time_input.strip()
        if not ti:
            return "N/A"
        if ti.upper() == "X":
            return None
        if ti.lower() == "rn":
            return datetime.now().strftime("%I:%M%p").lstrip("0")
        normalized = ti.upper().replace(" ", "")
        try:
            parsed = datetime.strptime(normalized, "%I:%M%p")
            return parsed.strftime("%I:%M%p").lstrip("0")
        except ValueError:
            return None

    def on_ok() -> None:
        date_input = date_var.get()
        time_input = time_var.get()
        parsed_date = _parse_date(date_input)
        if parsed_date is None:
            # "X" or invalid date -> treat X as cancel, invalid as error.
            if date_input.strip().upper() == "X":
                result["v"] = None
                dlg.destroy()
                return
            messagebox.showerror("Invalid date", "Enter a valid date like 04/20/2026 or April 26.")
            return

        parsed_time = _parse_time(time_input)
        if parsed_time is None:
            if time_input.strip().upper() == "X":
                result["v"] = None
                dlg.destroy()
                return
            messagebox.showerror(
                "Invalid time",
                "Enter time like 11:00AM or type rn for now, or leave blank for N/A.",
            )
            return

        result["v"] = (parsed_date, parsed_time)
        dlg.destroy()

    def on_cancel() -> None:
        result["v"] = None
        dlg.destroy()

    btn_row = tk.Frame(dlg)
    btn_row.pack(padx=12, pady=12)
    tk.Button(btn_row, text="OK", command=on_ok, width=10).pack(side="left", padx=(0, 8))
    tk.Button(btn_row, text="Cancel", command=on_cancel, width=10).pack(side="left")

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    # Focus defaults for quick keyboard entry.
    try:
        date_entry.focus_set()
    except tk.TclError:
        pass
    dlg.wait_window()
    return result["v"]


def _ask_typed_note_gui(parent: Optional[Any] = None) -> Optional[str]:
    if tk is None:
        return None
    _parent = parent if parent is not None else tk._default_root  # type: ignore[attr-defined]
    if _parent is None:
        return None
    dlg = tk.Toplevel(_parent)
    dlg.title("What happened today?")
    dlg.transient(_parent)
    dlg.grab_set()

    result: Dict[str, Optional[str]] = {"v": None}
    tk.Label(dlg, text="Type what happened today:").pack(padx=12, pady=(12, 4))
    box = tk.Text(dlg, height=8, width=52)
    box.pack(padx=12, pady=(0, 8))

    def on_ok() -> None:
        result["v"] = box.get("1.0", "end-1c").strip()
        dlg.destroy()

    def on_cancel() -> None:
        result["v"] = None
        dlg.destroy()

    btn_row = tk.Frame(dlg)
    btn_row.pack(padx=12, pady=12)
    tk.Button(btn_row, text="OK", command=on_ok, width=10).pack(side="left", padx=(0, 8))
    tk.Button(btn_row, text="Cancel", command=on_cancel, width=10).pack(side="left")
    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    try:
        box.focus_set()
    except tk.TclError:
        pass
    dlg.wait_window()
    return result["v"]


def journal_settings_menu_gui(parent: Optional[Any] = None) -> Optional[List[str]]:
    """
    GUI replacement for ``journal_settings_menu()``.
    Mirrors the same choices, but avoids blocking ``input()`` used by the CLI menu.
    """
    if tk is None:
        return journal_settings_menu()

    _parent = parent if parent is not None else tk._default_root  # type: ignore[attr-defined]
    if _parent is None:
        return journal_settings_menu()

    result: Dict[str, Optional[List[str]]] = {"v": None}
    dlg = tk.Toplevel(_parent)
    dlg.title("Journal settings")
    dlg.transient(_parent)
    dlg.grab_set()

    def _show_help() -> None:
        help_text = (
            "Journal choices:\n"
            "  WINDOW               - open window editor\n"
            "  CONSOLE              - type journal text in console\n"
            "  EDITPREV             - edit latest entry in window\n"
            "  DP                   - delete latest entry\n"
            "  RESTORE              - reopen latest unsaved draft\n"
            "  HELP                 - show this list\n"
            "  Enter                - return to main menu\n"
            "  DEFAULT WINDOWS     - set preferred journal input to window\n"
            "  DEFAULT CONSOLE      - set preferred journal input to console\n"
        )
        messagebox.showinfo("Journal settings help", help_text)

    tk.Label(dlg, text="Journal choice (type HELP for options):").pack(padx=12, pady=(12, 4))
    cmd_var = tk.StringVar(value="")
    entry = tk.Entry(dlg, textvariable=cmd_var, width=44)
    entry.pack(padx=12, pady=(0, 10))

    # For multi-step flows (CONSOLE), keep a pointer so we can close the menu once they finish.
    def on_submit() -> None:
        note = (cmd_var.get() or "").strip()
        key = note.lower()

        if is_enter_equivalent(note.upper()):
            dlg.destroy()
            return
        if key == "help":
            _show_help()
            return

        if key in ("c", "console", "coinsole"):
            typed = _ask_typed_note_gui(dlg)
            if typed is None:
                dlg.destroy()
                return
            dt = ask_entry_date_time_gui(dlg)
            if dt is None:
                dlg.destroy()
                return
            date_value, time_value = dt
            result["v"] = [date_value, time_value, typed, "", ""]
            dlg.destroy()
            return

        if key in (
            "editprev",
            "edit prev",
            "edit previous",
            "openprev",
            "open prev",
            "openprevious",
            "open previous",
        ):
            latest = get_latest_journal_entry_for_edit()
            if not latest:
                messagebox.showinfo("Edit previous", "No previous journal entry found to edit.")
                return
            open_journal_window_editor(
                {
                    "text": str(latest.get("text", "")),
                    "speech_transcript": str(latest.get("speech_transcript", "")),
                    "ai_report": str(latest.get("ai_report", "")),
                    "date": str(latest.get("date", "")),
                    "time": str(latest.get("time", "")),
                    "images": [],
                    "edit_target_sheet": str(latest.get("sheet_name", "")),
                    "edit_target_row": int(latest.get("row_index", 0) or 0),
                }
            )
            dlg.destroy()
            return

        if key in ("w", "window", "windows"):
            open_journal_window_editor()
            dlg.destroy()
            return

        if key in ("default windows", "default console"):
            prefs = load_preferences()
            default_mode = "windows" if key.endswith("windows") else "console"
            prefs["journal_input_default"] = default_mode
            if save_preferences(prefs):
                messagebox.showinfo(
                    "Default updated",
                    f"Default journal input set to {default_mode}.",
                )
            dlg.destroy()
            return

        if key == "restore":
            draft = load_journal_window_draft()
            if not draft:
                messagebox.showinfo("Restore", "No journal draft to restore.")
                dlg.destroy()
                return
            open_journal_window_editor(draft)
            dlg.destroy()
            return

        if key.upper() == "DP":
            latest = get_latest_journal_entry_for_delete()
            if not latest:
                messagebox.showinfo("Delete previous", "No previous journal entry found to delete.")
                return
            date_label = str(latest.get("date", "")).strip() or "(unknown date)"
            time_label = str(latest.get("time", "")).strip() or "(unknown time)"
            should_delete = messagebox.askyesno(
                "Delete previous journal entry",
                f"Delete previous journal entry at {date_label} {time_label}?",
            )
            if should_delete:
                delete_latest_journal_entry()
                dlg.destroy()
            return

        messagebox.showerror("Unknown choice", "Unknown journal choice. Type HELP to see valid options.")

    def on_cancel() -> None:
        result["v"] = None
        dlg.destroy()

    entry.bind("<Return>", lambda _e: on_submit())
    btn_row = tk.Frame(dlg)
    btn_row.pack(padx=12, pady=(0, 12))
    tk.Button(btn_row, text="Submit", command=on_submit, width=10).pack(side="left", padx=(0, 8))
    tk.Button(btn_row, text="Cancel", command=on_cancel, width=10).pack(side="left")

    dlg.protocol("WM_DELETE_WINDOW", on_cancel)
    try:
        entry.focus_set()
    except tk.TclError:
        pass
    dlg.wait_window()
    return result["v"]


def parse_flexible_date(raw: str, default_year: int):
    cleaned = " ".join(raw.strip().split())
    if not cleaned:
        return None

    slash_parts = cleaned.split("/")
    if len(slash_parts) in (2, 3):
        try:
            month = int(slash_parts[0])
            day = int(slash_parts[1])
            if len(slash_parts) == 3:
                year = int(slash_parts[2])
                if year < 100:
                    year += 2000
            else:
                year = default_year
            return datetime(year, month, day)
        except ValueError:
            pass

    text_parts = cleaned.replace(",", "").split()
    if len(text_parts) in (2, 3):
        month_token = text_parts[0].lower()
        month_map = {
            "jan": 1,
            "january": 1,
            "feb": 2,
            "february": 2,
            "mar": 3,
            "march": 3,
            "apr": 4,
            "april": 4,
            "may": 5,
            "jun": 6,
            "june": 6,
            "jul": 7,
            "july": 7,
            "aug": 8,
            "august": 8,
            "sep": 9,
            "sept": 9,
            "september": 9,
            "oct": 10,
            "october": 10,
            "nov": 11,
            "november": 11,
            "dec": 12,
            "december": 12,
        }
        month = month_map.get(month_token)
        if month is not None:
            try:
                day = int(text_parts[1])
                if len(text_parts) == 3:
                    year = int(text_parts[2])
                    if year < 100:
                        year += 2000
                else:
                    year = default_year
                return datetime(year, month, day)
            except ValueError:
                pass

    return None


def sync_journal_workbook() -> None:
    module = MODULES["J"]
    workbook_path = ensure_workbook(module)
    wb = load_workbook_with_retry(workbook_path)
    rebuild_master_journal_from_daily_pages(wb, module)
    reorder_journal_sheets(wb)
    save_workbook_with_retry(wb, workbook_path)


def is_journal_workbook_write_locked() -> bool:
    journal_path = DATA_DIR / MODULES["J"].workbook_name
    if not journal_path.exists():
        return False
    try:
        with open(journal_path, "r+b"):
            return False
    except PermissionError:
        return True
    except OSError:
        return False


def journal_settings_menu() -> Optional[List[str]]:
    def print_journal_choice_help() -> None:
        print("Journal choices:")
        print("  WINDOW               - open window editor")
        print("  COINSOLE             - type journal text in console")
        print("  EDITPREV             - edit latest entry in window")
        print("  DP                   - delete latest entry")
        print("  RESTORE              - reopen latest unsaved draft")
        print("  HELP                 - show this list")
        print("  Enter                - return to main menu")
        print("  DEFAULT WINDOWS      - set preferred journal input to window")
        print("  DEFAULT CONSOLE      - set preferred journal input to console")

    while True:
        print_journal_choice_help()
        note = input_line_with_tab_completions(
            "Journal choice: ",
            (
                "help",
                "c",
                "console",
                "coinsole",
                "dp",
                "w",
                "window",
                "windows",
                "editprev",
                "edit previous",
                "openprev",
                "open previous",
                "restore",
                "default windows",
                "default console",
            ),
        )
        if is_enter_equivalent(note):
            return None
        if note.lower() == "help":
            print_journal_choice_help()
            continue
        if note.lower() in ("c", "console", "coinsole"):
            typed_note = input("What happened today? ").strip()
            if is_enter_equivalent(typed_note):
                return None
            date_time = ask_entry_date_time()
            if date_time is None:
                return None
            date_value, time_value = date_time
            return [date_value, time_value, typed_note, "", ""]
        if note.lower() in (
            "editprev",
            "edit prev",
            "edit previous",
            "openprev",
            "open prev",
            "openprevious",
            "open previous",
        ):
            latest = get_latest_journal_entry_for_edit()
            if not latest:
                print("No previous journal entry found to edit.")
                return None
            open_journal_window_editor(
                {
                    "text": str(latest.get("text", "")),
                    "speech_transcript": str(latest.get("speech_transcript", "")),
                    "ai_report": str(latest.get("ai_report", "")),
                    "date": str(latest.get("date", "")),
                    "time": str(latest.get("time", "")),
                    "images": [],
                    "edit_target_sheet": str(latest.get("sheet_name", "")),
                    "edit_target_row": int(latest.get("row_index", 0) or 0),
                }
            )
            return None
        if note.lower() in ("w", "window", "windows"):
            open_journal_window_editor()
            return None
        if note.lower() in ("default windows", "default console"):
            prefs = load_preferences()
            default_mode = "windows" if note.lower().endswith("windows") else "console"
            prefs["journal_input_default"] = default_mode
            if save_preferences(prefs):
                if default_mode == "windows":
                    print("Default set to windows. Typing J opens the window editor.")
                else:
                    print("Default set to console. Typing J shows journal choices.")
            else:
                print("Could not save default journal input preference.")
            return None
        if note.lower() == "restore":
            draft = load_journal_window_draft()
            if not draft:
                print("No journal draft to restore.")
                return None
            restored = open_journal_window_editor(draft)
            if restored:
                print("Restored draft saved.")
            else:
                print("Draft restore opened. Unsaved draft remains available.")
            return None
        if note.upper() == "DP":
            latest = get_latest_journal_entry_for_delete()
            if not latest:
                print("No previous journal entry found to delete.")
                return None
            date_label = str(latest.get("date", "")).strip() or "(unknown date)"
            time_label = str(latest.get("time", "")).strip() or "(unknown time)"
            text_label = str(latest.get("text", "")).strip()
            while True:
                confirm = input(
                    f'Delete previous journal entry at {date_label} {time_label}? (y/N or type "expand"): '
                ).strip().lower()
                if confirm == "expand":
                    if text_label:
                        print("Entry text:")
                        print(text_label)
                    else:
                        print("(Entry text is empty.)")
                    continue
                if confirm in ("y", "yes"):
                    deleted = delete_latest_journal_entry()
                    if deleted:
                        print("Previous journal entry deleted.")
                    else:
                        print("No previous journal entry found to delete.")
                    break
                print("Delete cancelled.")
                break
            return None
        print(
            "Unknown journal choice. Type HELP to see commands, or use C/CONSOLE to write in console."
        )


def journal_prompts() -> Optional[List[str]]:
    prefs = load_preferences()
    default_mode = prefs.get("journal_input_default", "").strip().lower() or "windows"
    if default_mode == "windows":
        open_journal_window_editor()
        return None
    if default_mode == "console":
        typed_note = input("What happened today? ").strip()
        if is_enter_equivalent(typed_note):
            return None
        date_time = ask_entry_date_time()
        if date_time is None:
            return None
        date_value, time_value = date_time
        return [date_value, time_value, typed_note, "", ""]
    return journal_settings_menu()


MODULES: Dict[str, ModuleConfig] = {
    "J": ModuleConfig(
        name="Journal",
        workbook_name="Journal.xlsx",
        sheet_name="Journal",
        headers=["Date", "Time", "Journal", "Speech to text", "AI report"],
        prompt_builder=journal_prompts,
    ),
}


def print_main_help() -> None:
    print("Main commands:")
    print("  J      - Journal")
    print("  R      - Recap")
    print("  RT     - Recap (thinking)")
    print("  R [date range] / RT [date range] - recap only within date range")
    print("      Examples: 4.27 4.30 | 4/27 4/30 | 4/27 - 4/30 | 4/27/2026 - 4/30/2026")
    print("  R [file] / RT [file] - recap using file text as context")
    print("      Examples: R notes.txt | RT daily_logs/meeting.md")
    print("  C      - Chatbot")
    print("  CT     - Chatbot (thinking)")
    print("  RC / RECORD - start background recording with the selected source")
    print("  RECORD MIC / RECORD COMPUTER / RECORD BOTH - start or switch recording source")
    print("  RS / RECORD STOP - stop background recording and save it")
    print("  H/HELP - show this help")
    print("  J SETTINGS / J SETTING / JOURNAL SETTINGS / JS - open journal command menu")
    print("  RENAME - change app name")
    print("  STARTUP TRUE  - enable startup shortcut")
    print("  STARTUP FALSE - disable startup shortcut")
    print("  DEFAULT WINDOWS - typing J opens journal window directly")
    print("  DEFAULT CONSOLE - typing J shows journal command choices")
    print("  OPEN DIRECTORY   - open app data folder")
    print("  OPEN JOURNAL     - open Journal.xlsx")
    print("  OPEN SCREENSHOTS - open chat_screenshots folder")
    print("  DIRECTOR OPEN - open app data folder in File Explorer")
    print("  BACKUP START   - create backup zip in daily_logs/backup")
    print("  BACKUP TRUE    - auto backup once on each new day (default)")
    print("  BACKUP FALSE   - disable auto backup")
    print("  BACKUP LIMITED - keep at most 3 zip files; remove latest when adding")
    print("  UNINSTALL - request uninstall (requires CONFIRM UNINSTALL)")
    print("  CONFIRM UNINSTALL - remove user data, add-ons, downloads, shortcuts, and the current portable app folder")
    print("  WIFI WARN [name] - warn when connected to that Wi-Fi")
    print("  RESTORE - reopen latest unsaved journal window draft")
    print("  TOKEN ADD [token] - save API token")
    print("  TOKEN RESET - delete saved API token")
    print("  TOKEN COPY - copy current API token")
    print("  LAN cn | LAN en | LANGUAGE Chinese | LANGUAGE English - UI language")
    print("  SB bat     - Start Menu shortcut so Windows Search finds the .bat launcher")
    print("  SB journal - Start Menu shortcut so Windows Search finds Journal.xlsx")
    print("  SB reader  - Start Menu shortcut for Virtual Journal Reader")
    print("  Enter  - Continue/Exit")
    print("  X      - Exit")
    print("  TS     - take screenshot now (not attached outside chat mode)")
    print("  Tab    - complete a command; empty line + Tab shows this help")


def print_menu(app_name: str) -> None:
    print(f"\n=== {app_name} ===")
    print("Select an option below:")
    has_api_key = get_openai_api_key() is not None
    recap_label = "R = AI Recap" if has_api_key else "R = AI Recap (No API Key)"
    chat_label = "C = Chatbot" if has_api_key else "C = Chatbot (No API Key)"
    print("J = Journal")
    print(recap_label)
    print(chat_label)
    print("H = Commands")
    print("Enter = Skip/Exit")


def handle_choice(choice: str, app_name: str) -> Tuple[bool, str]:
    global PENDING_UNINSTALL_CONFIRM
    raw = choice.strip()
    key = raw.upper()
    if is_enter_equivalent(key):
        print("Skipped. See you next time.")
        return False, app_name
    if key in ("H", "HELP"):
        print_main_help()
        return True, app_name
    if key == "UNINSTALL":
        PENDING_UNINSTALL_CONFIRM = True
        print(
            'Uninstall requested. This cleans Daily Logger data, add-ons, downloads, '
            'shortcuts, and the current portable app folder. Type "CONFIRM UNINSTALL" to continue.'
        )
        return True, app_name
    if key == "CONFIRM UNINSTALL":
        if not PENDING_UNINSTALL_CONFIRM:
            print('Type "UNINSTALL" first, then "CONFIRM UNINSTALL".')
            return True, app_name
        run_clean_uninstall()
        return False, app_name
    if key in ("J SETTINGS", "J SETTING", "JOURNAL SETTINGS", "JS"):
        values = journal_settings_menu()
        if values is None:
            if load_journal_window_draft():
                print("Draft saved without journal entry. Use RESTORE to reopen it.")
            return True, app_name
        append_row(MODULES["J"], values)
        print(f'Journal saved to: {DATA_DIR / MODULES["J"].workbook_name}')
        return True, app_name
    if key == "X":
        print("Exit requested.")
        return False, app_name
    if key == "RENAME":
        app_name = rename_app_name()
        return True, app_name
    if key.startswith("RENAME "):
        app_name = rename_app_name_to(raw[7:].strip())
        return True, app_name
    if key.startswith("REANAME "):
        app_name = rename_app_name_to(raw[8:].strip())
        return True, app_name
    if key.startswith("WIFI WARN "):
        wifi_name = raw[10:].strip()
        if not wifi_name:
            print("Usage: wifi warn [wifi name]")
            return True, app_name
        if add_wifi_warn_name(wifi_name):
            print(f'Wi-Fi warning added for "{wifi_name}".')
        else:
            print("Could not save Wi-Fi warning list.")
        return True, app_name
    if key == "RESTORE":
        draft = load_journal_window_draft()
        if not draft:
            print("No journal draft to restore.")
            return True, app_name
        restored = open_journal_window_editor(draft)
        if restored:
            print("Restored draft saved.")
        else:
            print("Draft restore opened. Unsaved draft remains available.")
        return True, app_name
    _menu_parts = raw.split(None, 1)
    if _menu_parts and _menu_parts[0].upper() in ("LAN", "LANGUAGE"):
        _arg = _menu_parts[1].strip() if len(_menu_parts) > 1 else ""
        if not _arg:
            print("Usage: LAN cn | LAN en | LANGUAGE Chinese | LANGUAGE English")
            return True, app_name
        new_lang = normalize_ui_language(_arg)
        prefs = load_preferences()
        cur = normalize_ui_language(str(prefs.get(UI_LANGUAGE_PREF_KEY, "en")))
        if new_lang == cur:
            print(f"UI language is already {'Chinese' if new_lang == 'zh' else 'English'}.")
            return True, app_name
        prefs[UI_LANGUAGE_PREF_KEY] = new_lang
        if not save_preferences(prefs):
            print("Could not save language preference.")
            return True, app_name
        hook = _journal_ui_language_changed_hook
        if hook is not None:
            hook(new_lang)
        print(f"UI language set to {'Chinese' if new_lang == 'zh' else 'English'}.")
        return True, app_name
    if key.startswith("TOKEN ADD "):
        token_value = raw[10:].strip()
        if not token_value:
            print("Usage: token add [token]")
            return True, app_name
        if save_openai_api_key(token_value):
            print("API token saved.")
        else:
            print("Could not save API token.")
        return True, app_name
    if key == "TOKEN RESET":
        confirm = input("Are you sure you want to delete saved API token? (y/N): ").strip().lower()
        if confirm in ("y", "yes"):
            if delete_openai_api_key():
                print("Saved API token deleted.")
            else:
                print("Could not delete saved API token.")
        else:
            print("Token reset cancelled.")
        return True, app_name
    if key == "TOKEN COPY":
        token_value = get_openai_api_key()
        if not token_value:
            print("No current API token found.")
            return True, app_name
        if copy_text_to_clipboard(token_value):
            print("Current API token copied to clipboard.")
        else:
            print("Could not copy token to clipboard.")
        return True, app_name
    if key == "STARTUP TRUE":
        if is_startup_enabled():
            print("Startup is already enabled.")
            return True, app_name
        if create_startup_shortcut():
            print("Startup enabled.")
            prefs = load_preferences()
            prefs["startup_enabled"] = "true"
            prefs["startup_prompt_done"] = "true"
            save_preferences(prefs)
        else:
            print("Could not enable startup shortcut.")
        return True, app_name
    if key == "STARTUP FALSE":
        if remove_startup_shortcut():
            print("Startup disabled.")
            prefs = load_preferences()
            prefs["startup_enabled"] = "false"
            prefs["startup_prompt_done"] = "true"
            save_preferences(prefs)
        else:
            print("Could not disable startup shortcut.")
        return True, app_name
    if key == "DEFAULT WINDOWS":
        prefs = load_preferences()
        prefs["journal_input_default"] = "windows"
        if save_preferences(prefs):
            print("Default set to windows. Typing J opens the window editor.")
        else:
            print("Could not save default journal input preference.")
        return True, app_name
    if key == "DEFAULT CONSOLE":
        prefs = load_preferences()
        prefs["journal_input_default"] = "console"
        if save_preferences(prefs):
            print("Default set to console. Typing J shows journal choices.")
        else:
            print("Could not save default journal input preference.")
        return True, app_name
    if key == "DIRECTOR OPEN":
        if open_current_directory_in_explorer():
            print(f"Opened folder: {USER_DATA_ROOT}")
        else:
            print("Could not open current folder in File Explorer.")
        return True, app_name
    if key.startswith("OPEN "):
        open_target = raw[5:].strip().upper()
        if open_target == "DIRECTORY":
            if open_current_directory_in_explorer():
                print(f"Opened folder: {USER_DATA_ROOT}")
            else:
                print("Could not open current folder in File Explorer.")
            return True, app_name
        if open_target == "JOURNAL":
            journal_path = ensure_workbook(MODULES["J"])
            if open_path_with_default_app(journal_path):
                print(f"Opened journal file: {journal_path}")
            else:
                print("Could not open Journal.xlsx.")
            return True, app_name
        if open_target == "SCREENSHOTS":
            SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
            if open_path_with_default_app(SCREENSHOT_DIR):
                print(f"Opened screenshots folder: {SCREENSHOT_DIR}")
            else:
                print("Could not open screenshots folder.")
            return True, app_name
        print("Usage: OPEN DIRECTORY | OPEN JOURNAL | OPEN SCREENSHOTS")
        return True, app_name
    if key == "TS":
        print("Taking screenshot...")
        screenshot_path = take_chat_screenshot_hidden_console()
        if screenshot_path:
            print(f"Screenshot saved: {screenshot_path}")
        return True, app_name
    if key == "BACKUP START":
        prefs = load_preferences()
        evict_oldest_backup_if_limited_full(prefs)
        backup_path = run_backup_now()
        if backup_path is None:
            print("No files/folders in daily_logs to back up.")
            return True, app_name
        trim_backups_if_limited(prefs)
        prefs["last_backup_date"] = datetime.now().strftime("%Y-%m-%d")
        save_preferences(prefs)
        print(f"Backup created: {backup_path}")
        return True, app_name
    if key == "BACKUP TRUE":
        prefs = load_preferences()
        prefs["backup_enabled"] = "true"
        if save_preferences(prefs):
            print("Auto backup enabled.")
        else:
            print("Could not save backup preference.")
        return True, app_name
    if key == "BACKUP FALSE":
        prefs = load_preferences()
        prefs["backup_enabled"] = "false"
        if save_preferences(prefs):
            print("Auto backup disabled.")
        else:
            print("Could not save backup preference.")
        return True, app_name
    if key == "BACKUP LIMITED":
        prefs = load_preferences()
        prefs["backup_limited"] = "true"
        trim_backups_if_limited(prefs)
        if save_preferences(prefs):
            print("Backup limited mode enabled (max 3 zip files).")
        else:
            print("Could not save backup limit preference.")
        return True, app_name
    if key.startswith("SB "):
        sub = raw[3:].strip().upper()
        if sub == "BAT":
            if not sb_create_bat_search_shortcut():
                print("Could not create BAT search shortcut.")
        elif sub == "JOURNAL":
            if not sb_create_journal_search_shortcut():
                print("Could not create Journal search shortcut.")
        elif sub == "READER":
            if not sb_create_reader_search_shortcut():
                print("Could not create Reader search shortcut.")
        else:
            print('Usage: SB bat   or   SB journal   or   SB reader')
        return True, app_name
    if key.startswith("RT "):
        recap_range, file_context, file_path, recap_err = resolve_recap_target(
            raw[3:].strip(), datetime.now().year
        )
        if recap_err is not None:
            print(recap_err)
            return True, app_name
        run_chat_mode(
            with_journal_context=True,
            use_thinking_model=True,
            recap_date_range=recap_range,
            recap_context_override=file_context,
            recap_context_label=file_path,
        )
        return True, app_name
    if key.startswith("R "):
        recap_range, file_context, file_path, recap_err = resolve_recap_target(
            raw[2:].strip(), datetime.now().year
        )
        if recap_err is not None:
            print(recap_err)
            return True, app_name
        run_chat_mode(
            with_journal_context=True,
            recap_date_range=recap_range,
            recap_context_override=file_context,
            recap_context_label=file_path,
        )
        return True, app_name
    if key == "R":
        run_chat_mode(with_journal_context=True)
        return True, app_name
    if key == "RT":
        run_chat_mode(with_journal_context=True, use_thinking_model=True)
        return True, app_name
    if key == "C":
        run_chat_mode(with_journal_context=False)
        return True, app_name
    if key == "CT":
        run_chat_mode(with_journal_context=False, use_thinking_model=True)
        return True, app_name

    if key == "J" and is_journal_workbook_write_locked():
        print("Journal is currently open in another program.")
        input("Close Journal.xlsx and press Enter to return to main menu...")
        return True, app_name

    module = MODULES.get(key)
    if not module:
        print(
            "Unknown choice. Please enter J, J SETTINGS, J SETTING, JOURNAL SETTINGS, JS, R, RT, C, CT, H, HELP, RENAME, STARTUP TRUE/FALSE, DEFAULT WINDOWS/CONSOLE, OPEN DIRECTORY/JOURNAL/SCREENSHOTS, DIRECTOR OPEN, BACKUP START/TRUE/FALSE/LIMITED, TS, UNINSTALL, CONFIRM UNINSTALL, SB bat/journal, WIFI WARN [name], RESTORE, LAN/LANGUAGE, TOKEN ADD/RESET/COPY, or press Enter to skip."
        )
        return True, app_name

    values = module.prompt_builder()
    if values is None:
        if key == "J" and load_journal_window_draft():
            print("Draft saved without journal entry. Use RESTORE to reopen it.")
        return True, app_name
    append_row(module, values)
    print(f"{module.name} saved to: {DATA_DIR / module.workbook_name}")
    return True, app_name


# Full-line menu strings for Tab completion (canonical spelling).
MAIN_MENU_COMPLETIONS: Tuple[str, ...] = tuple(
    sorted(
        {
            "J",
            "J SETTINGS",
            "J SETTING",
            "JOURNAL SETTINGS",
            "JS",
            "R",
            "R ",
            "RC",
            "RECORD",
            "RECORD STOP",
            "RS",
            "RT",
            "RT ",
            "C",
            "CT",
            "H",
            "HELP",
            "X",
            "RENAME",
            "RESTORE",
            "STARTUP TRUE",
            "STARTUP FALSE",
            "DEFAULT WINDOWS",
            "DEFAULT CONSOLE",
            "OPEN DIRECTORY",
            "OPEN JOURNAL",
            "OPEN SCREENSHOTS",
            "DIRECTOR OPEN",
            "BACKUP START",
            "BACKUP TRUE",
            "BACKUP FALSE",
            "BACKUP LIMITED",
            "UNINSTALL",
            "CONFIRM UNINSTALL",
            "TS",
            "WIFI WARN ",
            "TOKEN ADD ",
            "TOKEN RESET",
            "TOKEN COPY",
            "SB bat",
            "SB journal",
            "REANAME ",
            "LAN ",
            "LAN CN",
            "LAN EN",
            "LANGUAGE ",
            "LANGUAGE CHINESE",
            "LANGUAGE ENGLISH",
        },
        key=lambda s: s.upper(),
    )
)


def _lcp_length_case_insensitive(strings: List[str]) -> int:
    if not strings:
        return 0
    upper = [s.upper() for s in strings]
    limit = min(len(s) for s in upper)
    i = 0
    while i < limit and all(s[i] == upper[0][i] for s in upper):
        i += 1
    return i


CHAT_LINE_COMPLETIONS: Tuple[str, ...] = ("help", "rs", "ts")
WINDOWS_CONSOLE_LINE_HISTORY: List[str] = []


def _apply_typing_casing(user_line: str, completed_canonical: str) -> str:
    """Match completion casing to how the user typed (see print_main_help Tab note)."""
    if not user_line:
        return completed_canonical
    if user_line.islower():
        return completed_canonical.lower()
    # Sentence-style: "Startup t", "Startup ", "Startup true" - first word Title, rest lowercase.
    if " " in user_line:
        first_sp = user_line.find(" ")
        first_word = user_line[:first_sp]
        after_last_sp = user_line[user_line.rfind(" ") + 1 :]
        if (
            first_word
            and first_word[0].isupper()
            and (len(first_word) == 1 or first_word[1:].islower())
            and (after_last_sp == "" or after_last_sp.islower())
        ):
            return completed_canonical.lower()
    if (
        len(user_line) >= 2
        and user_line[0].isupper()
        and user_line[1:].islower()
    ):
        return completed_canonical.lower()
    return completed_canonical.upper()


def _readline_completion_suffix(before: str, cased_full: str, raw_m: str) -> str:
    """Return the string readline should insert; align before/cased_full by case-insensitive prefix."""
    n = min(len(before), len(cased_full))
    i = 0
    while i < n and before[i].upper() == cased_full[i].upper():
        i += 1
    if i == len(before):
        return cased_full[i:]
    if cased_full.upper().startswith(before.upper()):
        return cased_full[len(before) :]
    return raw_m[len(before) :]


def _line_tab_extend(line: str, completions: Tuple[str, ...]) -> Tuple[str, bool]:
    """Return (new_line, extended) after one Tab press for a fixed completion list."""
    matches = [c for c in completions if c.upper().startswith(line.upper())]
    if not matches:
        return line, False
    if len(matches) == 1:
        m = matches[0]
        if line.upper() == m.upper():
            return line, False
        cased = _apply_typing_casing(line, m)
        return cased, True
    k = _lcp_length_case_insensitive(matches)
    unified_canon = matches[0][:k]
    cased = _apply_typing_casing(line, unified_canon)
    if cased != line:
        return cased, True
    return line, False


def _build_readline_line_completer(
    completions: Tuple[str, ...],
    on_empty_tab: Optional[Callable[[], None]] = None,
):
    def completer(text: str, state: int) -> Optional[str]:
        if _readline is None:
            return None
        if state == 0:
            line0 = _readline.get_line_buffer()
            if on_empty_tab and not line0.strip():
                on_empty_tab()
                completer._matches = []  # type: ignore[attr-defined]
                completer._empty_tab_only = True  # type: ignore[attr-defined]
            else:
                completer._empty_tab_only = False  # type: ignore[attr-defined]
                beg = _readline.get_begidx()
                before = line0[:beg]
                stem_u = (before + text).upper()
                completer._matches = sorted(  # type: ignore[attr-defined]
                    [m for m in completions if m.upper().startswith(stem_u)],
                    key=lambda s: (len(s), s.upper()),
                )
                completer._line0 = line0  # type: ignore[attr-defined]
                completer._beg0 = beg  # type: ignore[attr-defined]
        if getattr(completer, "_empty_tab_only", False):
            return None
        matches: List[str] = getattr(completer, "_matches", [])
        line0 = getattr(completer, "_line0", _readline.get_line_buffer())
        beg0 = getattr(completer, "_beg0", _readline.get_begidx())
        try:
            m = matches[state]
            before = line0[:beg0]
            cased_full = _apply_typing_casing(line0, m)
            suffix = _readline_completion_suffix(before, cased_full, m)
            return suffix
        except (IndexError, AttributeError):
            return None

    return completer


def input_line_with_tab_completions(
    prompt: str,
    completions: Tuple[str, ...],
    on_empty_tab: Optional[Callable[[], None]] = None,
) -> str:
    """Read one line with Tab completing against a fixed list (readline, Windows msvcrt, or plain input)."""
    if _readline is not None:
        old_completer = _readline.get_completer()
        old_delims = _readline.get_completer_delims()
        try:
            _readline.set_completer(
                _build_readline_line_completer(completions, on_empty_tab=on_empty_tab)
            )
            _readline.set_completer_delims(" \t\n`!@#$%^&*()-=+[{]}\\|;:'\",<>/?")
            _readline.parse_and_bind("tab: complete")
            return input(prompt).strip()
        finally:
            _readline.set_completer(old_completer)
            _readline.set_completer_delims(old_delims)

    if msvcrt is not None and sys.platform == "win32":
        sys.stdout.write(prompt)
        sys.stdout.flush()
        buf: List[str] = []
        cursor = 0
        history = WINDOWS_CONSOLE_LINE_HISTORY
        hist_index = len(history)

        def _move_left(count: int = 1) -> None:
            nonlocal cursor
            n = max(0, min(count, cursor))
            if n:
                sys.stdout.write("\b" * n)
                sys.stdout.flush()
                cursor -= n

        def _move_right(count: int = 1) -> None:
            nonlocal cursor
            n = max(0, min(count, len(buf) - cursor))
            if n:
                sys.stdout.write("".join(buf[cursor : cursor + n]))
                sys.stdout.flush()
                cursor += n

        def _replace_tail_after_cursor(old_tail_len: int) -> None:
            tail = "".join(buf[cursor:])
            sys.stdout.write(tail)
            if old_tail_len > len(tail):
                sys.stdout.write(" " * (old_tail_len - len(tail)))
            back = max(len(tail), old_tail_len)
            if back:
                sys.stdout.write("\b" * back)
            sys.stdout.flush()

        def _insert_text(text: str) -> None:
            nonlocal cursor
            if not text:
                return
            old_tail_len = len(buf) - cursor
            buf[cursor:cursor] = list(text)
            cursor += len(text)
            sys.stdout.write(text)
            _replace_tail_after_cursor(old_tail_len)

        def _delete_left(count: int = 1) -> None:
            nonlocal cursor
            n = max(0, min(count, cursor))
            if n == 0:
                return
            _move_left(n)
            del buf[cursor : cursor + n]
            _replace_tail_after_cursor((len(buf) - cursor) + n)

        def _delete_right(count: int = 1) -> None:
            n = max(0, min(count, len(buf) - cursor))
            if n == 0:
                return
            del buf[cursor : cursor + n]
            _replace_tail_after_cursor((len(buf) - cursor) + n)

        def _erase_previous_word() -> None:
            # Match common text-box behavior: delete spaces first, then one word.
            n = 0
            i = cursor - 1
            while i >= 0 and buf[i].isspace():
                n += 1
                i -= 1
            while i >= 0 and not buf[i].isspace():
                n += 1
                i -= 1
            _delete_left(n)

        def _replace_line(new_line: str) -> None:
            nonlocal cursor
            old_len = len(buf)
            buf.clear()
            buf.extend(list(new_line))
            cursor = len(buf)
            # Carriage-return redraw is more stable in PowerShell/Windows Terminal
            # than backspace-based in-place erasing for tab-completion replacement.
            sys.stdout.write("\r" + prompt + new_line)
            if old_len > len(new_line):
                sys.stdout.write(" " * (old_len - len(new_line)))
            # Put visual cursor at logical cursor location (always end after replace).
            sys.stdout.write("\r" + prompt + "".join(buf))
            sys.stdout.flush()

        while True:
            ch = msvcrt.getwch()
            if ch in ("\x00", "\xe0"):
                code = msvcrt.getwch()
                if code == "H":  # Up
                    if history and hist_index > 0:
                        hist_index -= 1
                        _replace_line(history[hist_index])
                    else:
                        sys.stdout.write("\a")
                        sys.stdout.flush()
                elif code == "P":  # Down
                    if hist_index < len(history):
                        hist_index += 1
                        line = history[hist_index] if hist_index < len(history) else ""
                        _replace_line(line)
                    else:
                        sys.stdout.write("\a")
                        sys.stdout.flush()
                elif code == "K":  # Left
                    _move_left(1)
                elif code == "M":  # Right
                    _move_right(1)
                elif code == "G":  # Home
                    _move_left(cursor)
                elif code == "O":  # End
                    _move_right(len(buf) - cursor)
                elif code == "S":  # Delete
                    _delete_right(1)
                # Ignore remaining arrows/function keys.
                continue
            if ch in "\r\n":
                line = "".join(buf).strip()
                if line:
                    if not history or history[-1] != line:
                        history.append(line)
                hist_index = len(history)
                sys.stdout.write("\n")
                sys.stdout.flush()
                return line
            if ch == "\x03":
                sys.stdout.write("\n")
                raise KeyboardInterrupt
            if ch == "\x08":
                _delete_left(1)
                continue
            if ch in ("\x7f", "\x17"):
                # Ctrl+Backspace often arrives as DEL (\x7f); Ctrl+W as ETB (\x17).
                _erase_previous_word()
                continue
            if ch == "\t":
                line = "".join(buf)
                if not line.strip() and on_empty_tab:
                    on_empty_tab()
                    sys.stdout.write("\n")
                    sys.stdout.write(prompt)
                    sys.stdout.flush()
                    continue
                new_line, extended = _line_tab_extend(line, completions)
                if extended:
                    _replace_line(new_line)
                else:
                    matches = [
                        c for c in completions if c.upper().startswith(line.upper())
                    ]
                    if len(matches) > 1:
                        sys.stdout.write("\n  " + "\n  ".join(matches) + "\n")
                        sys.stdout.write(prompt + "".join(buf))
                        _move_left(len(buf) - cursor)
                        sys.stdout.flush()
                    else:
                        sys.stdout.write("\a")
                        sys.stdout.flush()
                continue
            if ord(ch) >= 32:
                _insert_text(ch)

    return input(prompt).strip()


def input_menu_choice(prompt: str) -> str:
    """Read main menu input with Tab completing known commands."""
    return input_line_with_tab_completions(
        prompt, MAIN_MENU_COMPLETIONS, on_empty_tab=print_main_help
    )


def run() -> None:
    _deps_ok = ensure_runtime_dependencies()
    if not _deps_ok:
        return
    migrate_legacy_storage_if_needed()
    finalize_pending_local_transcription_addon()
    setup_first_time_preferences()
    ensure_backup_folder()
    if sys.platform == "win32":
        try:
            hwnd = ctypes.windll.kernel32.GetConsoleWindow()
            if hwnd:
                ctypes.windll.user32.ShowWindow(hwnd, 0)  # SW_HIDE
        except Exception:
            pass
    # Do NOT auto-restore on app launch; only restore when user types `restore`.
    # This prevents the journal editor from popping up with an old unsaved draft.
    open_journal_window_editor(None, start_auto_backup=True)


def main(argv: Optional[List[str]] = None) -> int:
    args = list(sys.argv[1:] if argv is None else argv)
    if args and args[0] == VIRTUAL_READER_SERVER_ARG:
        return run_virtual_reader_server_from_cli(args[1:])
    run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

